"""Create local QA artifacts for eBay traffic experiment cover-priority group."""

from __future__ import annotations

import csv
import json
from pathlib import Path

from PIL import Image, ImageDraw
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
EXPERIMENT_CSV = DATABASE_DIR / "eBay_Traffic_Experiment.csv"
OUTPUT_DIR = DATABASE_DIR / "eBay_Cover_QA"
QA_CSV = DATABASE_DIR / "eBay_Cover_QA.csv"


def clean(value) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def load_experiment_ids() -> set[str]:
    ids = set()
    with EXPERIMENT_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if row.get("Group") == "B_COVER_QA_PRIORITY":
                ids.add(clean(row.get("ID")))
    return ids


def load_rows(ids: set[str]) -> list[dict]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or clean(row[cols["ID"]]) not in ids:
            continue
        rows.append({name: row[idx] for name, idx in cols.items()})
    wb.close()
    return rows


def image_info(path: str) -> dict:
    p = Path(path)
    if not path or not p.exists():
        return {"exists": False, "width": "", "height": "", "size_mb": "", "error": "missing"}
    try:
        with Image.open(p) as im:
            return {
                "exists": True,
                "width": im.width,
                "height": im.height,
                "size_mb": round(p.stat().st_size / (1024 * 1024), 2),
                "error": "",
            }
    except Exception as exc:  # noqa: BLE001
        return {"exists": False, "width": "", "height": "", "size_mb": "", "error": str(exc)}


def build_contact_sheet(rows: list[dict]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    thumb_w, thumb_h = 220, 220
    pad, label_h = 18, 62
    cols = 4
    sheet_rows = (len(rows) + cols - 1) // cols
    sheet = Image.new("RGB", (cols * (thumb_w + pad) + pad, sheet_rows * (thumb_h + label_h + pad) + pad), "white")
    draw = ImageDraw.Draw(sheet)
    for idx, row in enumerate(rows):
        c = idx % cols
        r = idx // cols
        x = pad + c * (thumb_w + pad)
        y = pad + r * (thumb_h + label_h + pad)
        cover = Path(clean(row.get("Cover_Path")))
        if cover.exists():
            with Image.open(cover) as im:
                im = im.convert("RGB")
                im.thumbnail((thumb_w, thumb_h), Image.Resampling.LANCZOS)
                ox = x + (thumb_w - im.width) // 2
                oy = y + (thumb_h - im.height) // 2
                sheet.paste(im, (ox, oy))
        else:
            draw.rectangle([x, y, x + thumb_w, y + thumb_h], fill=(245, 245, 245), outline=(200, 50, 50), width=3)
            draw.text((x + 20, y + 95), "MISSING COVER", fill=(160, 0, 0))
        draw.rectangle([x, y, x + thumb_w, y + thumb_h], outline=(170, 170, 170), width=1)
        label = clean(row.get("ID"))
        draw.text((x, y + thumb_h + 8), label, fill=(20, 20, 20))
        draw.text((x, y + thumb_h + 28), clean(row.get("Title"))[:34], fill=(60, 60, 60))
    out = OUTPUT_DIR / "Cover_QA_Contact_Sheet.jpg"
    sheet.save(out, "JPEG", quality=92, optimize=True)
    return out


def run() -> None:
    ids = load_experiment_ids()
    rows = load_rows(ids)
    qa_rows = []
    for row in rows:
        cover = clean(row.get("Cover_Path"))
        cover_info = image_info(cover)
        u1 = clean(row.get("Gallery_U1_Path"))
        u1_info = image_info(u1)
        qa_rows.append(
            {
                "ID": clean(row.get("ID")),
                "Title": clean(row.get("Title")),
                "eBay_Item_ID": clean(row.get("eBay_Item_ID")),
                "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
                "Cover_Path": cover,
                "Cover_Exists": cover_info["exists"],
                "Cover_Size": f"{cover_info['width']}x{cover_info['height']}" if cover_info["exists"] else "",
                "Cover_MB": cover_info["size_mb"],
                "Gallery_U1_Path": u1,
                "Gallery_U1_Exists": u1_info["exists"],
                "Gallery_U1_Size": f"{u1_info['width']}x{u1_info['height']}" if u1_info["exists"] else "",
                "Likely_Cover_OK": bool(cover_info["exists"] and cover_info["width"] >= 900 and cover_info["height"] >= 900),
                "Recommended_Action": "Online cover order spot-check before image sync",
            }
        )
    with QA_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = [
            "ID",
            "Title",
            "eBay_Item_ID",
            "Printify_Product_ID",
            "Cover_Path",
            "Cover_Exists",
            "Cover_Size",
            "Cover_MB",
            "Gallery_U1_Path",
            "Gallery_U1_Exists",
            "Gallery_U1_Size",
            "Likely_Cover_OK",
            "Recommended_Action",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(qa_rows)
    sheet = build_contact_sheet(rows)
    print(f"[COVER-QA] rows={len(qa_rows)} csv={QA_CSV}")
    print(f"[COVER-QA] contact_sheet={sheet}")
    print(json.dumps({"missing_cover": sum(1 for r in qa_rows if r["Cover_Exists"] != True)}, indent=2))


if __name__ == "__main__":
    run()
