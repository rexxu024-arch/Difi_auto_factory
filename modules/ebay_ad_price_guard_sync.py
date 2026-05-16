from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import Config

DATABASE = ROOT / "Database"
BOOK = DATABASE / "eBay_listing.xlsx"
INPUT_CSV = DATABASE / "eBay_Ad_Rate_Execution_Shortlist_NoSticker.csv"
PLAN_CSV = DATABASE / "eBay_Ad_Price_Guard_Apply_Plan.csv"
LOG_CSV = DATABASE / "eBay_Ad_Price_Guard_Sync_Log.csv"
NY = ZoneInfo("America/New_York")

PUBLISH_PRICE_ONLY = {
    "title": False,
    "description": False,
    "images": False,
    "variants": True,
    "tags": False,
    "keyFeatures": False,
    "shipping_template": False,
}


def clean(value) -> str:
    return str(value or "").strip()


def money(value) -> float:
    try:
        return float(re.sub(r"[^0-9.]", "", clean(value)) or "0")
    except ValueError:
        return 0.0


def now() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def load_book_map() -> dict[str, dict[str, str]]:
    wb = load_workbook(BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {clean(name): idx for idx, name in enumerate(headers) if name}
    out: dict[str, dict[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        local_id = clean(values[cols["ID"]]) if "ID" in cols else ""
        if not local_id:
            continue
        out[local_id] = {
            "ID": local_id,
            "Printify_Product_ID": clean(values[cols["Printify_Product_ID"]]) if "Printify_Product_ID" in cols else "",
            "Current_Workbook_Price": clean(values[cols["Price"]]) if "Price" in cols else "",
            "Title": clean(values[cols["Title"]]) if "Title" in cols else "",
            "Status": clean(values[cols["Status"]]) if "Status" in cols else "",
        }
    wb.close()
    return out


def write_plan(limit: int) -> list[dict[str, str]]:
    rows = list(csv.DictReader(INPUT_CSV.open("r", encoding="utf-8-sig", newline=""))) if INPUT_CSV.exists() else []
    book = load_book_map()
    out: list[dict[str, str]] = []
    for row in rows:
        if limit and len(out) >= limit:
            break
        local_id = clean(row.get("ID"))
        meta = book.get(local_id, {})
        target = money(row.get("Recommended_Printify_Source_Price_USD"))
        current = money(meta.get("Current_Workbook_Price") or row.get("Current_Listed_Price_USD"))
        result = "READY" if meta.get("Printify_Product_ID") and target > 0 else "HOLD_MISSING_PRODUCT_OR_PRICE"
        if target and current and abs(target - current) < 0.01:
            result = "SKIP_ALREADY_AT_TARGET"
        out.append(
            {
                "Timestamp": now(),
                "ID": local_id,
                "Product_Type": clean(row.get("Product_Type")),
                "Lane": clean(row.get("Lane")),
                "Ad_Rate_Pct": clean(row.get("Ad_Rate_Pct")),
                "Current_Price_USD": f"{current:.2f}" if current else "",
                "Target_Price_USD": f"{target:.2f}" if target else "",
                "Estimated_Profit_USD": clean(row.get("Estimated_Profit_USD")),
                "Printify_Product_ID": meta.get("Printify_Product_ID", ""),
                "eBay_Item_ID": clean(row.get("eBay_Item_ID")),
                "Result": result,
            }
        )
    fields = [
        "Timestamp",
        "ID",
        "Product_Type",
        "Lane",
        "Ad_Rate_Pct",
        "Current_Price_USD",
        "Target_Price_USD",
        "Estimated_Profit_USD",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Result",
    ]
    with PLAN_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(out)
    ready = sum(1 for row in out if row["Result"] == "READY")
    print(f"[AD-PRICE-GUARD] planned={len(out)} ready={ready} csv={PLAN_CSV}")
    return out


def update_workbook_price(local_id: str, target_price: str) -> None:
    """Keep the local workbook in lockstep after a confirmed Printify sync."""
    wb = load_workbook(BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {clean(name): idx + 1 for idx, name in enumerate(headers) if name}
    id_col = cols.get("ID")
    price_col = cols.get("Price")
    if not id_col or not price_col:
        wb.close()
        raise RuntimeError("Missing ID or Price column in eBay workbook")
    for row_idx in range(2, ws.max_row + 1):
        if clean(ws.cell(row_idx, id_col).value) == local_id:
            ws.cell(row_idx, price_col).value = f"${money(target_price):.2f}"
            wb.save(BOOK)
            wb.close()
            return
    wb.close()
    raise RuntimeError(f"Could not find local workbook row for {local_id}")


def request_json(method: str, url: str, **kwargs):
    headers = kwargs.pop("headers", {})
    headers["Authorization"] = f"Bearer {Config.Printify_API_KEY}"
    headers.setdefault("Content-Type", "application/json")
    for attempt in range(3):
        resp = requests.request(method, url, headers=headers, timeout=120, **kwargs)
        if resp.status_code < 500:
            return resp
        time.sleep(3 + attempt * 3)
    return resp


def sync_prices(limit: int, execute: bool, sleep_seconds: float) -> int:
    rows = list(csv.DictReader(PLAN_CSV.open("r", encoding="utf-8-sig", newline=""))) if PLAN_CSV.exists() else write_plan(0)
    rows = [row for row in rows if row.get("Result") == "READY"]
    if limit:
        rows = rows[:limit]
    base = Config.Printify_API_URL.rstrip("/")
    fields = ["Timestamp", "ID", "Printify_Product_ID", "Target_Price_USD", "HTTP_Get", "HTTP_Update", "HTTP_Publish", "Result", "Error"]
    exists = LOG_CSV.exists()
    done = 0
    with LOG_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        if not exists:
            writer.writeheader()
        for row in rows:
            product_id = clean(row.get("Printify_Product_ID"))
            target_cents = int(round(money(row.get("Target_Price_USD")) * 100))
            get_code = update_code = publish_code = ""
            result = "DRY_RUN"
            error = ""
            try:
                if execute:
                    get_resp = request_json("GET", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json")
                    get_code = str(get_resp.status_code)
                    get_resp.raise_for_status()
                    product = get_resp.json()
                    variants = [
                        {
                            "id": variant["id"],
                            "price": target_cents if variant.get("is_enabled") else variant.get("price"),
                            "is_enabled": bool(variant.get("is_enabled")),
                        }
                        for variant in product.get("variants") or []
                    ]
                    update_resp = request_json(
                        "PUT",
                        f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
                        json={"variants": variants},
                    )
                    update_code = str(update_resp.status_code)
                    update_resp.raise_for_status()
                    publish_resp = request_json(
                        "POST",
                        f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
                        json=PUBLISH_PRICE_ONLY,
                    )
                    publish_code = str(publish_resp.status_code)
                    publish_resp.raise_for_status()
                    update_workbook_price(row["ID"], row["Target_Price_USD"])
                    result = "SYNCED"
                    done += 1
                print(f"[AD-PRICE-GUARD] {result} {row['ID']} target=${row['Target_Price_USD']} product={product_id}")
            except Exception as exc:  # noqa: BLE001
                result = "FAILED"
                error = f"{type(exc).__name__}: {exc}"[:500]
                print(f"[AD-PRICE-GUARD-FAIL] {row.get('ID')}: {error}")
            writer.writerow(
                {
                    "Timestamp": now(),
                    "ID": row.get("ID", ""),
                    "Printify_Product_ID": product_id,
                    "Target_Price_USD": row.get("Target_Price_USD", ""),
                    "HTTP_Get": get_code,
                    "HTTP_Update": update_code,
                    "HTTP_Publish": publish_code,
                    "Result": result,
                    "Error": error,
                }
            )
            if execute:
                time.sleep(max(0.0, sleep_seconds))
    print(f"[AD-PRICE-GUARD-DONE] rows={len(rows)} synced={done} execute={execute}")
    return done


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare/sync Printify source prices before higher eBay ad-rate experiments.")
    parser.add_argument("--plan-limit", type=int, default=0)
    parser.add_argument("--sync-limit", type=int, default=0)
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--sleep-seconds", type=float, default=5.0)
    args = parser.parse_args()
    write_plan(args.plan_limit)
    sync_prices(args.sync_limit, execute=args.execute, sleep_seconds=args.sleep_seconds)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
