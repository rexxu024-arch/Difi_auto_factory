"""Build and optionally apply local title-length repairs.

The house rule for published eBay titles is 75-79 characters. This module does
not sync to eBay or Printify; it prepares workbook-safe local changes that can
be pushed later by the metadata sync path after cover gates are clear.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
LOCAL_QA = DATABASE_DIR / "Local_Listing_QA.csv"
COPY_PLAN = DATABASE_DIR / "Listing_Copy_Optimization.csv"
OUT_CSV = DATABASE_DIR / "eBay_Title_Repair_Queue.csv"
OUT_MD = DATABASE_DIR / "eBay_Title_Repair_Queue.md"

MIN_LEN = 75
MAX_LEN = 79
APPEND_TERMS = ["Gift", "Decor", "Study", "Decal", "Art"]

HEADERS = [
    "Timestamp",
    "ID",
    "Status",
    "Current_Title",
    "Current_Length",
    "Proposed_Title",
    "Proposed_Length",
    "Repair_Source",
    "Ready_For_Metadata_Sync",
]


def now_text() -> str:
    return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_workbook_rows() -> dict[str, dict[str, object]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    rows = {}
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            item_id = clean(values[cols["ID"]])
            if item_id:
                rows[item_id] = {header: values[index] for header, index in cols.items()}
    finally:
        wb.close()
    return rows


def title_issue_ids() -> set[str]:
    ids = set()
    for row in read_csv(LOCAL_QA):
        issues = clean(row.get("Issues"))
        if "title_length_" in issues:
            ids.add(clean(row.get("ID")))
    return ids


def copy_candidates() -> dict[str, dict[str, str]]:
    return {clean(row.get("ID")): row for row in read_csv(COPY_PLAN) if clean(row.get("ID"))}


def fit_title(title: str) -> tuple[str, str]:
    title = " ".join(clean(title).split())
    if MIN_LEN <= len(title) <= MAX_LEN:
        return title, "existing_ok"
    for term in APPEND_TERMS:
        if term.lower() in title.lower().split():
            continue
        candidate = f"{title} {term}".strip()
        if MIN_LEN <= len(candidate) <= MAX_LEN:
            return candidate, f"append_{term}"
    if len(title) > MAX_LEN:
        trimmed = title[:MAX_LEN].rsplit(" ", 1)[0]
        if len(trimmed) >= MIN_LEN:
            return trimmed, "trim_word_boundary"
        return title[:MAX_LEN].rstrip(), "trim_hard"
    return title, "no_safe_fit"


def build_rows() -> list[dict[str, object]]:
    workbook = load_workbook_rows()
    copies = copy_candidates()
    rows = []
    for item_id in sorted(title_issue_ids()):
        source = workbook.get(item_id, {})
        current = clean(source.get("Title"))
        copy_title = clean(copies.get(item_id, {}).get("Proposed_eBay_Title"))
        if copy_title and MIN_LEN <= len(copy_title) <= MAX_LEN:
            proposed, source_name = copy_title, "copy_plan"
        else:
            proposed, source_name = fit_title(current)
        rows.append(
            {
                "Timestamp": now_text(),
                "ID": item_id,
                "Status": clean(source.get("Status")),
                "Current_Title": current,
                "Current_Length": len(current),
                "Proposed_Title": proposed,
                "Proposed_Length": len(proposed),
                "Repair_Source": source_name,
                "Ready_For_Metadata_Sync": MIN_LEN <= len(proposed) <= MAX_LEN,
            }
        )
    return rows


def write_outputs(rows: list[dict[str, object]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    ready = sum(1 for row in rows if row["Ready_For_Metadata_Sync"])
    lines = [
        "# eBay Title Repair Queue",
        "",
        f"Generated: {now_text()} America/New_York",
        "",
        f"- Rows: {len(rows)}",
        f"- Ready for later metadata sync: {ready}",
        "- No online sync is performed by this module.",
        "",
        f"CSV: `{OUT_CSV}`",
        "",
    ]
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def apply_local(rows: list[dict[str, object]]) -> int:
    ready = {row["ID"]: row for row in rows if row["Ready_For_Metadata_Sync"]}
    if not ready:
        return 0
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    changed = 0
    try:
        for row_idx in range(2, ws.max_row + 1):
            item_id = clean(ws.cell(row_idx, cols["ID"]).value)
            repair = ready.get(item_id)
            if not repair:
                continue
            cell = ws.cell(row_idx, cols["Title"])
            if clean(cell.value) != clean(repair["Proposed_Title"]):
                cell.value = clean(repair["Proposed_Title"])
                changed += 1
        if changed:
            wb.save(EBAY_BOOK)
    finally:
        wb.close()
    return changed


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply-local", action="store_true")
    args = parser.parse_args()
    rows = build_rows()
    write_outputs(rows)
    changed = apply_local(rows) if args.apply_local else 0
    print(f"[TITLE-REPAIR] rows={len(rows)} ready={sum(1 for row in rows if row['Ready_For_Metadata_Sync'])} changed={changed} csv={OUT_CSV}")


if __name__ == "__main__":
    main()
