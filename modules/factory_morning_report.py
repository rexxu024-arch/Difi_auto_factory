import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
REPORTS_DIR = PROJECT_ROOT / "Reports"
GEMINI_DIR = PROJECT_ROOT / "Gemini_Advisor"
LATEST_DIR = PROJECT_ROOT / "Review_Packets" / "Latest"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
ETSY_PLAN = DATABASE_DIR / "Etsy_launch_plan.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
REGISTRY_LOG = DATABASE_DIR / "Unified_Listing_Registry.csv"
COPY_PLAN = DATABASE_DIR / "Listing_Copy_Optimization.csv"
PRICING_MATRIX = DATABASE_DIR / "Pricing_Strategy_Matrix.csv"
ETSY_DIGITAL_QUEUE = DATABASE_DIR / "Etsy_Digital_Upload_Queue.csv"
ETSY_DIGITAL_BUNDLES = DATABASE_DIR / "Etsy_Digital_Bundle_Queue.csv"
ETSY_DIGITAL_PREVIEWS = DATABASE_DIR / "Etsy_Digital_Preview_Assets.csv"
ETSY_DIGITAL_QA = DATABASE_DIR / "Etsy_Digital_QA.csv"
ETSY_DIGITAL_FINAL_PACKET = DATABASE_DIR / "Etsy_Digital_Final_Upload_Packet.csv"
ETSY_GRAY_QUEUE = DATABASE_DIR / "Etsy_Digital_Gray_Launch_Queue.csv"
ETSY_FEE_LEDGER = DATABASE_DIR / "Etsy_Fee_Ledger.csv"
ETSY_LIVE_AUDIT = DATABASE_DIR / "Etsy_Digital_Live_Audit.csv"
ETSY_LEGACY_STATUS = DATABASE_DIR / "Etsy_Legacy_Retirement_Status.csv"
EBAY_EXPERIMENT = DATABASE_DIR / "eBay_Traffic_Experiment.csv"
EBAY_EXPERIMENT_REPORT = DATABASE_DIR / "eBay_Traffic_Experiment_Report.csv"
EBAY_TRAFFIC_DIAGNOSIS = DATABASE_DIR / "eBay_Traffic_Diagnosis.csv"
EBAY_COVER_QA = DATABASE_DIR / "eBay_Cover_QA.csv"
EBAY_ONLINE_COVER_AUDIT = DATABASE_DIR / "eBay_Online_Cover_Audit.csv"
EBAY_ONLINE_COVER_FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"
EBAY_COVER_REPLACEMENT_QUEUE = DATABASE_DIR / "eBay_Cover_Replacement_Queue.csv"
PRINTIFY_IMAGE_DEFAULT_AUDIT = DATABASE_DIR / "Printify_Image_Default_Audit.csv"
PRINTIFY_GALLERY_DUPLICATE_AUDIT = DATABASE_DIR / "Printify_Gallery_Duplicate_Audit.csv"
FACTORY_BACKLOG = DATABASE_DIR / "Factory_Backlog.csv"


def _now():
    try:
        return datetime.now(ZoneInfo("America/New_York"))
    except Exception:
        return datetime.now().astimezone()


def _load_listing_state():
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {header: index for index, header in enumerate(headers)}
    status = Counter()
    product_counts = defaultdict(lambda: Counter())
    published = 0
    stable = 0
    ready = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        product_type = row[cols.get("Product_Type")] or "Unknown"
        row_status = str(row[cols.get("Status")] or "")
        status[row_status] += 1
        if row_status.startswith("Printify_Published"):
            published += 1
            stable += 1
            product_counts[product_type]["published"] += 1
            product_counts[product_type]["stable"] += 1
        elif row_status.startswith("Printify_UI_Mockups"):
            stable += 1
            product_counts[product_type]["stable"] += 1
        elif row_status == "Ready_for_Printify":
            ready += 1
            product_counts[product_type]["ready"] += 1
    workbook.close()
    return status, product_counts, published, stable, ready


def _load_etsy_plan():
    if not ETSY_PLAN.exists():
        return Counter(), 0
    workbook = load_workbook(ETSY_PLAN, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {header: index for index, header in enumerate(headers)}
    counts = Counter()
    total = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        counts[row[cols["Product_Type"]]] += 1
        total += 1
    workbook.close()
    return counts, total


def _latest_performance():
    if not PERFORMANCE_LOG.exists():
        return {}
    latest_ts = None
    rows = []
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            ts = row.get("Snapshot_Timestamp")
            if not latest_ts or ts > latest_ts:
                latest_ts = ts
                rows = [row]
            elif ts == latest_ts:
                rows.append(row)
    if not rows:
        return {}
    zero = sum(1 for row in rows if row.get("Views_30_Days") == "0")
    viewed = sum(1 for row in rows if str(row.get("Views_30_Days")).isdigit() and int(row["Views_30_Days"]) > 0)
    promoted = sum(1 for row in rows if row.get("General_Status") == "Promoted")
    return {
        "timestamp": latest_ts,
        "rows": len(rows),
        "zero_views": zero,
        "viewed": viewed,
        "promoted": promoted,
    }


def _registry_buckets():
    if not REGISTRY_LOG.exists():
        return Counter()
    counts = Counter()
    with REGISTRY_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            counts[row.get("Action_Bucket") or "Unknown"] += 1
    return counts


def _csv_count(path):
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def _count_by(path, column):
    if not path.exists():
        return Counter()
    counts = Counter()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            counts[row.get(column) or "Unknown"] += 1
    return counts


def _latest_count_by_id(path, column, id_column="ID"):
    if not path.exists():
        return Counter()
    latest = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            item_id = row.get(id_column) or ""
            if item_id:
                latest[item_id] = row
    counts = Counter()
    for row in latest.values():
        counts[row.get(column) or "Unknown"] += 1
    return counts


def _digital_summary():
    queue_rows = _csv_count(ETSY_DIGITAL_QUEUE)
    bundle_rows = _csv_count(ETSY_DIGITAL_BUNDLES)
    preview_rows = _csv_count(ETSY_DIGITAL_PREVIEWS)
    final_packet_rows = _csv_count(ETSY_DIGITAL_FINAL_PACKET)
    qa_bad = 0
    qa_missing = 0
    if ETSY_DIGITAL_QA.exists():
        with ETSY_DIGITAL_QA.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                ok = str(row.get("OK", "")).lower() == "true"
                exists = str(row.get("Exists", "")).lower() == "true"
                if not ok:
                    qa_bad += 1
                if not exists:
                    qa_missing += 1
    max_mb = 0.0
    all_ok = True
    if ETSY_DIGITAL_QUEUE.exists():
        with ETSY_DIGITAL_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                try:
                    max_mb = max(max_mb, float(row.get("Max_File_MB") or 0))
                except ValueError:
                    pass
                all_ok = all_ok and str(row.get("All_Under_20MB")).lower() == "true"
    gray_rows = []
    if ETSY_GRAY_QUEUE.exists():
        with ETSY_GRAY_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
            gray_rows = list(csv.DictReader(handle))
    gray_status = Counter(row.get("Launch_Status") or "Unknown" for row in gray_rows)
    gray_fee = Counter(row.get("Fee_Status") or "Unknown" for row in gray_rows)
    confirmed_spend = 0.0
    if ETSY_FEE_LEDGER.exists():
        with ETSY_FEE_LEDGER.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if str(row.get("Status") or "").startswith("CONFIRMED"):
                    try:
                        confirmed_spend += float(row.get("Confirmed_Spent_USD") or 0)
                    except ValueError:
                        pass
    live_counts = Counter()
    if ETSY_LIVE_AUDIT.exists():
        latest = {}
        with ETSY_LIVE_AUDIT.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if row.get("Etsy_Listing_ID"):
                    latest[row["Etsy_Listing_ID"]] = row
        for row in latest.values():
            live_counts[row.get("Status") or "Unknown"] += 1
    legacy_retired = 0
    if ETSY_LEGACY_STATUS.exists():
        with ETSY_LEGACY_STATUS.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if str(row.get("Status") or "").startswith("DELETED"):
                    legacy_retired += 1
    return {
        "queue_rows": queue_rows,
        "bundle_rows": bundle_rows,
        "preview_rows": preview_rows,
        "final_packet_rows": final_packet_rows,
        "qa_bad": qa_bad,
        "qa_missing": qa_missing,
        "max_mb": max_mb,
        "all_ok": all_ok,
        "gray_rows": len(gray_rows),
        "gray_status": gray_status,
        "gray_fee": gray_fee,
        "confirmed_spend": confirmed_spend,
        "live_counts": live_counts,
        "legacy_retired": legacy_retired,
    }


def _online_cover_summary():
    audit_counts = _latest_count_by_id(EBAY_ONLINE_COVER_AUDIT, "Result")
    fix_rows = _csv_count(EBAY_ONLINE_COVER_FIX_QUEUE)
    replacement_counts = _count_by(EBAY_COVER_REPLACEMENT_QUEUE, "Replacement_Status")
    default_counts = _count_by(PRINTIFY_IMAGE_DEFAULT_AUDIT, "Result")
    default_rows = _csv_count(PRINTIFY_IMAGE_DEFAULT_AUDIT)
    gallery_duplicate_counts = _count_by(PRINTIFY_GALLERY_DUPLICATE_AUDIT, "Result")
    gallery_duplicate_rows = _csv_count(PRINTIFY_GALLERY_DUPLICATE_AUDIT)
    return {
        "audit_counts": audit_counts,
        "fix_rows": fix_rows,
        "replacement_counts": replacement_counts,
        "default_counts": default_counts,
        "default_rows": default_rows,
        "gallery_duplicate_counts": gallery_duplicate_counts,
        "gallery_duplicate_rows": gallery_duplicate_rows,
    }


def _backlog_summary():
    status_counts = _count_by(FACTORY_BACKLOG, "Status")
    lane_counts = _count_by(FACTORY_BACKLOG, "Lane")
    top_tasks = []
    if FACTORY_BACKLOG.exists():
        with FACTORY_BACKLOG.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        rows.sort(key=lambda row: -int(row.get("Priority") or 0))
        for row in rows[:5]:
            top_tasks.append(
                f"- P{row.get('Priority')} {row.get('Lane')} / {row.get('Status')}: {row.get('Task')}"
            )
    return status_counts, lane_counts, top_tasks


def build():
    now = _now()
    REPORTS_DIR.mkdir(exist_ok=True)
    GEMINI_DIR.mkdir(exist_ok=True)
    status, product_counts, published, stable, ready = _load_listing_state()
    etsy_counts, etsy_total = _load_etsy_plan()
    performance = _latest_performance()
    registry = _registry_buckets()
    copy_rows = _csv_count(COPY_PLAN)
    pricing_rows = _csv_count(PRICING_MATRIX)
    digital = _digital_summary()
    online_cover = _online_cover_summary()
    backlog_status, backlog_lanes, backlog_top = _backlog_summary()
    experiment_groups = _count_by(EBAY_EXPERIMENT, "Group")
    experiment_deltas = _count_by(EBAY_EXPERIMENT_REPORT, "Group")
    traffic_diagnoses = _count_by(EBAY_TRAFFIC_DIAGNOSIS, "Diagnosis")
    cover_qa_rows = _csv_count(EBAY_COVER_QA)
    report_path = REPORTS_DIR / f"morning_report_{now:%Y%m%d_%H%M}.md"
    gemini_path = GEMINI_DIR / f"gemini_review_queue_{now:%Y%m%d_%H%M}.md"
    product_lines = []
    for product_type in sorted(product_counts):
        counts = product_counts[product_type]
        product_lines.append(
            f"- {product_type}: stable {counts['stable']}, published {counts['published']}, ready {counts['ready']}"
        )
    etsy_lines = [f"- {product_type}: {count}" for product_type, count in sorted(etsy_counts.items())]
    perf_lines = []
    if performance:
        perf_lines = [
            f"- Latest eBay snapshot: {performance['timestamp']}",
            f"- Rows read: {performance['rows']}",
            f"- 0-view rows in snapshot: {performance['zero_views']}",
            f"- Rows with at least 1 view: {performance['viewed']}",
            f"- General promoted rows in snapshot: {performance['promoted']}",
        ]
    else:
        perf_lines = ["- No Seller Hub performance snapshot logged yet."]
    registry_lines = [
        f"- {bucket}: {count}"
        for bucket, count in sorted(registry.items())
    ] or ["- Unified registry not generated yet."]
    experiment_lines = [
        f"- {group}: {count}"
        for group, count in sorted(experiment_groups.items())
    ] or ["- No eBay traffic experiment active yet."]
    diagnosis_lines = [
        f"- {diagnosis}: {count}"
        for diagnosis, count in sorted(traffic_diagnoses.items())
    ] or ["- No eBay traffic diagnosis generated yet."]
    online_cover_lines = [
        f"- Live eBay cover audit {result}: {count}"
        for result, count in sorted(online_cover["audit_counts"].items())
    ] or ["- Live eBay cover audit not run yet."]
    replacement_lines = [
        f"- Cover replacement queue {result}: {count}"
        for result, count in sorted(online_cover["replacement_counts"].items())
    ] or ["- Cover replacement queue not generated yet."]
    printify_default_lines = [
        f"- Printify image-default audit {result}: {count}"
        for result, count in sorted(online_cover["default_counts"].items())
    ] or ["- Printify image-default audit not run yet."]
    printify_gallery_lines = [
        f"- Printify gallery duplicate audit {result}: {count}"
        for result, count in sorted(online_cover["gallery_duplicate_counts"].items())
    ] or ["- Printify gallery duplicate audit not run yet."]
    backlog_status_lines = [
        f"- Backlog {status}: {count}"
        for status, count in sorted(backlog_status.items())
    ] or ["- Factory backlog not generated yet."]
    backlog_lane_lines = [
        f"- {lane}: {count}"
        for lane, count in sorted(backlog_lanes.items())
    ] or ["- Factory backlog lanes not available yet."]
    body = "\n".join(
        [
            "# OpenClaw Morning Report",
            "",
            f"Generated: {now:%Y-%m-%d %H:%M %z} America/New_York",
            "",
            "## Current Factory State",
            "",
            f"- Stable Printify-tracked products: {stable}",
            f"- Published through Printify/eBay tracking: {published}",
            f"- Ready for Printify: {ready}",
            "",
            "## Product Counts",
            "",
            *(product_lines or ["- No product counts available."]),
            "",
            "## Etsy Phase 1 Prep",
            "",
            f"- Draft-prepared Etsy launch candidates: {etsy_total}",
            *(etsy_lines or ["- Etsy launch plan not generated yet."]),
            "",
            "## Performance Snapshot",
            "",
            *perf_lines,
            "",
            "## Local Low-Bandwidth Work Completed",
            "",
            f"- Listing copy optimization candidates: {copy_rows}",
            f"- Pricing matrix scenarios: {pricing_rows}",
            f"- Unified registry rows bucketed: {sum(registry.values())}",
            f"- Etsy digital printable upload queue: {digital['queue_rows']} listings, max file {digital['max_mb']:.2f}MB, under 20MB limit: {digital['all_ok']}",
            f"- Etsy digital previews: {digital['preview_rows']} listings x 3 preview images",
            f"- Etsy digital final upload packet: {digital['final_packet_rows']} listings, QA bad={digital['qa_bad']}, missing={digital['qa_missing']}",
            f"- Etsy digital bundle concepts: {digital['bundle_rows']}",
            f"- Etsy Digital gray queue rows: {digital['gray_rows']}",
            f"- Etsy Digital live listings: {digital['gray_status'].get('PUBLISHED_UI_CONFIRMED', 0)}",
            f"- Etsy Digital confirmed listing-fee spend: ${digital['confirmed_spend']:.2f}",
            f"- Etsy Digital public audit active/readable: {digital['live_counts'].get('ACTIVE_READABLE', 0)}",
            f"- Etsy legacy listings retired/deleted: {digital['legacy_retired']}",
            f"- eBay cover QA rows: {cover_qa_rows}",
            "",
            "## Unified Registry Buckets",
            "",
            *registry_lines,
            "",
            "## eBay Traffic Experiment",
            "",
            *experiment_lines,
            "",
            "## eBay Traffic Diagnosis",
            "",
            *diagnosis_lines,
            "",
            "## Live Cover Integrity",
            "",
            *online_cover_lines,
            f"- Live cover fix queue rows: {online_cover['fix_rows']}",
            *replacement_lines,
            f"- Printify image-default audit rows: {online_cover['default_rows']}",
            *printify_default_lines,
            f"- Printify gallery duplicate audit rows: {online_cover['gallery_duplicate_rows']}",
            *printify_gallery_lines,
            "",
            "## Factory Backlog",
            "",
            *backlog_status_lines,
            "",
            "Top tasks:",
            *(backlog_top or ["- No backlog tasks available."]),
            "",
            "Lane counts:",
            *backlog_lane_lines,
            "",
            "## Current Guardrails",
            "",
            "- eBay rapid publish remains paused after Akamai/zero-size-object instability.",
            "- Wired LAN is fixed; online work may run normally, but marketplace/account-risk throttles still apply.",
            "- Etsy Digital first gray batch is live; do not spend beyond the next approved gray cell without traffic/signal logic.",
            "- eBay Promoted Listings Standard / General 2% is the only approved active ad mode; do not use Priority/PPC or suggested ad rates.",
            "- Sticker and non-sticker expansion remain paused until gallery duplicate risk is repaired or isolated.",
            "- Multiple Printify official/default mockups are allowed only when they are visually distinct; publish is blocked by missing custom design/cover, live buyer-page mismatch, zero default image, or repeated selected gallery images.",
            "",
            "## Operator Notes",
            "",
            "- Focus is Phase 1: data foundation, Etsy relaunch prep, and performance learning loop.",
            "- Printify storefront design is intentionally bypassed until Rex updates it or asks for drafts.",
            "",
        ]
    )
    report_path.write_text(body, encoding="utf-8")
    gemini_path.write_text(
        "\n".join(
            [
                "# Gemini Advisor Review Queue",
                "",
                f"Generated: {now:%Y-%m-%d %H:%M %z} America/New_York",
                "",
                "Rex is Commander, Gemini is Strategy Advisor, Codex is Executive Operator.",
                "",
                "Please review the current OpenClaw plan as a strategy advisor. Do not request API keys, account secrets, payment data, buyer private data, or direct account actions.",
                "",
                "## Report Summary",
                "",
                body,
                "",
                "## Questions for Gemini",
                "",
                "1. Given the first 10 Etsy Digital listings are live, what early signal should decide whether to spend the next $2 gray cell?",
                "2. If the first 10 get 0 views after indexing, which search-intent variable should be changed first: title/category angle, product format, or visual theme?",
                "3. Which three visual DNA themes should be expanded first if Etsy impressions appear but clicks remain low?",
                "4. What ad test would you run first with a $3-5/day Etsy Ads budget after 48-72 hours of organic data?",
                "5. Which product language sounds too mass-generated and should be softened before launch?",
                "",
                "## Codex Action Filter",
                "",
                "- Adopted:",
                "- Deferred:",
                "- Rejected:",
                "- Requires Rex confirmation:",
                "",
            ]
        ),
        encoding="utf-8",
    )
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    (LATEST_DIR / "morning_report_latest.md").write_text(body, encoding="utf-8")
    (LATEST_DIR / "gemini_review_queue_latest.md").write_text(gemini_path.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"[REPORT] {report_path}")
    print(f"[GEMINI] {gemini_path}")
    return report_path, gemini_path


if __name__ == "__main__":
    build()
