"""Patch local listing descriptions with the buyer-facing image note.

This is a local-data repair only. It does not call Etsy, eBay, or Printify.
The marketplace sync layer can later use the generated CSV to update live
listings in small, guarded batches.
"""

from __future__ import annotations

import csv
import shutil
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE = PROJECT_ROOT / "Database"
BOOK = DATABASE / "eBay_listing.xlsx"
QUEUE = DATABASE / "Listing_Description_Image_Note_Patch.csv"
PROGRESS = PROJECT_ROOT / "PROGRESS_LOG.md"
NY_TZ = ZoneInfo("America/New_York")

NEEDLE = "main image shows the actual product customers receive"
NOTE = (
    "Image note: the main image shows the actual product customers receive. "
    "Additional images may show close-up details, concept views, or room styling "
    "to help you understand the artwork atmosphere and finish."
)

HEADERS = [
    "ID",
    "Product_Type",
    "Status",
    "eBay_Item_ID",
    "Printify_Product_ID",
    "Patched_Local_Description",
    "Needs_Marketplace_Sync",
    "Note",
]


def now_et() -> str:
    return datetime.now(NY_TZ).strftime("%Y-%m-%d %H:%M ET")


def clean(value: object) -> str:
    return str(value or "").replace("\r", " ").strip()


def normalized_description(value: object) -> str:
    text = clean(value)
    return "\n".join(line.rstrip() for line in text.splitlines()).strip()


def should_patch(product_type: str, description: str) -> bool:
    return product_type in {"Poster", "Acrylic"} and NEEDLE not in description.lower()


def append_note(description: str) -> str:
    description = normalized_description(description)
    if not description:
        return NOTE
    return f"{description}\n\n{NOTE}"


def main() -> int:
    if not BOOK.exists():
        raise FileNotFoundError(BOOK)

    stamp = datetime.now(NY_TZ).strftime("%Y%m%d_%H%M%S")
    backup = DATABASE / f"eBay_listing.backup_image_note_{stamp}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    required = {"ID", "Product_Type", "Description", "Status", "Printify_Product_ID", "eBay_Item_ID"}
    missing = required.difference(cols)
    if missing:
        wb.close()
        raise RuntimeError(f"Missing workbook columns: {sorted(missing)}")

    records: list[dict[str, str]] = []
    for row_index in range(2, ws.max_row + 1):
        item_id = clean(ws.cell(row_index, cols["ID"]).value)
        if not item_id:
            continue
        product_type = clean(ws.cell(row_index, cols["Product_Type"]).value)
        description = normalized_description(ws.cell(row_index, cols["Description"]).value)
        if not should_patch(product_type, description):
            continue
        ws.cell(row_index, cols["Description"]).value = append_note(description)
        status = clean(ws.cell(row_index, cols["Status"]).value)
        ebay_item = clean(ws.cell(row_index, cols["eBay_Item_ID"]).value)
        records.append(
            {
                "ID": item_id,
                "Product_Type": product_type,
                "Status": status,
                "eBay_Item_ID": ebay_item,
                "Printify_Product_ID": clean(ws.cell(row_index, cols["Printify_Product_ID"]).value),
                "Patched_Local_Description": "yes",
                "Needs_Marketplace_Sync": "yes" if ebay_item or status.startswith("Printify_Published") else "no",
                "Note": "Local workbook patched; live marketplaces still need guarded sync where applicable.",
            }
        )

    wb.save(BOOK)
    wb.close()

    with QUEUE.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(records)

    with PROGRESS.open("a", encoding="utf-8") as handle:
        handle.write(
            f"\n## {now_et()} - Local description image-note patch\n"
            f"- Patched {len(records)} local Poster/Acrylic descriptions with the buyer-facing actual-product image note.\n"
            f"- Backup: `{backup}`.\n"
            f"- Sync queue: `{QUEUE}`.\n"
            "- No marketplace write, publish, fee, or ad action was taken.\n"
        )
    print(f"[IMAGE-NOTE-PATCH] patched={len(records)} backup={backup} queue={QUEUE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
