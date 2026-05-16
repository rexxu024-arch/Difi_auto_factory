from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from modules.ebay_inventory_category_audit import browse_item_group, flag_row, load_token


DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
OUT_CSV = DATABASE / "eBay_Clean_Clone_Readback_Gate.csv"
OUT_MD = REPORTS / "eBay_Clean_Clone_Readback_Gate.md"
NY = ZoneInfo("America/New_York")

BLOCK_FLAGS = {
    "SHIPPING_NOT_FREE",
    "BRAND_LOW_TRUST",
    "LOW_GALLERY_COUNT",
    "EXACT_DUPLICATE_IMAGE_URL",
    "CATEGORY_MISMATCH_ACRYLIC",
    "CATEGORY_MISMATCH_POSTER",
    "SHORT_DESCRIPTION_THIN",
}


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def now_et() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def load_clone_rows() -> list[dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    rows: list[dict[str, str]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {header: values[cols[header]] if header in cols and cols[header] < len(values) else "" for header in headers}
            item_id = clean(row.get("ID"))
            ebay_id = clean(row.get("eBay_Item_ID"))
            if "CLEAN_CLONE" in item_id and ebay_id:
                rows.append(row)
    finally:
        wb.close()
    return rows


def gate(flags: str) -> str:
    parts = {clean(part) for part in clean(flags).split("|") if clean(part)}
    blockers = sorted(parts.intersection(BLOCK_FLAGS))
    if blockers:
        return "HOLD_SOURCE_SYNC_FAILURE:" + ",".join(blockers)
    return "PASS_READY_FOR_TRAFFIC_TEST"


def run(limit: int = 0) -> int:
    token = load_token()
    rows = load_clone_rows()
    if limit:
        rows = rows[:limit]
    out: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        ebay_id = clean(row.get("eBay_Item_ID"))
        status, item, error = browse_item_group(ebay_id, token)
        audit = flag_row({"Item_ID": ebay_id, "Title": row.get("Title"), "Views_30_Days": "0"}, item, error)
        result = {
            "Timestamp": now_et(),
            "Workbook_ID": clean(row.get("ID")),
            "Product_Type": clean(row.get("Product_Type")),
            "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
            "eBay_Item_ID": ebay_id,
            "HTTP_Status": str(status),
            "API_Title": clean(audit.get("API_Title")),
            "Price": clean(audit.get("Price")),
            "Shipping_Cost": clean(audit.get("Shipping_Cost")),
            "Brand": clean(audit.get("Brand")),
            "Image_Count": clean(audit.get("Image_Count")),
            "Flags": clean(audit.get("Flags")),
            "Gate": gate(clean(audit.get("Flags"))),
        }
        out.append(result)
        print(
            f"[CLONE-READBACK] {index}/{len(rows)} {result['Workbook_ID']} "
            f"ship={result['Shipping_Cost']} brand={result['Brand']} gate={result['Gate']}",
            flush=True,
        )

    fields = [
        "Timestamp",
        "Workbook_ID",
        "Product_Type",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "HTTP_Status",
        "API_Title",
        "Price",
        "Shipping_Cost",
        "Brand",
        "Image_Count",
        "Flags",
        "Gate",
    ]
    DATABASE.mkdir(exist_ok=True)
    REPORTS.mkdir(exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(out)

    counts = Counter(row["Gate"].split(":", 1)[0] for row in out)
    flag_counts = Counter()
    for row in out:
        for flag in row["Flags"].split("|"):
            if flag:
                flag_counts[flag] += 1
    lines = [
        "# eBay Clean Clone Readback Gate",
        "",
        f"Generated: {now_et()}",
        f"Rows: {len(out)}",
        "",
        "## Gate Counts",
        "",
    ]
    for key, value in counts.most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Flag Counts", ""])
    for key, value in flag_counts.most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "- Clean clones are not eligible for ad spend until this gate returns PASS.",
            "- If `SHIPPING_NOT_FREE` persists, rebuild must start with Printify `sales_channel_properties.free_shipping=true` or a working channel shipping template, not ad-rate changes.",
            "- If `BRAND_LOW_TRUST` persists, marketplace-side metadata override is mandatory before traffic testing.",
            "",
            f"CSV: {OUT_CSV}",
        ]
    )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[CLONE-READBACK-DONE] rows={len(out)} csv={OUT_CSV} md={OUT_MD}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()
    return run(limit=args.limit)


if __name__ == "__main__":
    raise SystemExit(main())
