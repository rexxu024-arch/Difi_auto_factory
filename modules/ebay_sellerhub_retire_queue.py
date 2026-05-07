import argparse
import asyncio
import csv
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import requests
import websockets
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config

RETIRE_QUEUE = PROJECT_ROOT / "Database" / "eBay_Retire_Queue.csv"
EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
LOG_FILE = PROJECT_ROOT / "Database" / "eBay_Retire_Run_Log.csv"
CDP_PORT = int(os.getenv("OPENCLAW_EBAY_CDP_PORT") or os.getenv("OPENCLAW_CDP_PORT") or "9223")
CDP_BASE = f"http://127.0.0.1:{CDP_PORT}"


def _now():
    return datetime.now().isoformat(timespec="seconds")


def _read_queue():
    if not RETIRE_QUEUE.exists():
        return []
    with RETIRE_QUEUE.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _write_queue(rows):
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    for extra in ["Retire_Attempted_At", "Retire_Result"]:
        if extra not in fieldnames:
            fieldnames.append(extra)
    with RETIRE_QUEUE.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _log(row_id, ebay_id, result, note):
    new_file = not LOG_FILE.exists()
    with LOG_FILE.open("a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if new_file:
            writer.writerow(["Timestamp", "Old_ID", "Old_eBay_Item_ID", "Result", "Note"])
        writer.writerow([_now(), row_id, ebay_id, result, note])


def _edge_targets():
    return requests.get(f"{CDP_BASE}/json/list", timeout=5).json()


def _new_target(url):
    response = requests.put(f"{CDP_BASE}/json/new?{quote(url, safe=':/?=&')}", timeout=5)
    response.raise_for_status()
    return response.json()


def _close_target(target_id):
    if not target_id:
        return
    try:
        requests.get(f"{CDP_BASE}/json/close/{target_id}", timeout=5)
    except Exception:
        pass


async def _eval(ws_url, expression, wait=0):
    async with websockets.connect(ws_url, max_size=20_000_000) as sock:
        seq = 1

        async def send(method, params=None):
            nonlocal seq
            msg = {"id": seq, "method": method}
            if params is not None:
                msg["params"] = params
            await sock.send(json.dumps(msg))
            my = seq
            seq += 1
            while True:
                data = json.loads(await sock.recv())
                if data.get("id") == my:
                    return data

        await send("Runtime.enable")
        if wait:
            await asyncio.sleep(wait)
        result = await send("Runtime.evaluate", {"expression": expression, "returnByValue": True})
        return result["result"]["result"].get("value")


async def _retire_one_via_sellerhub(ebay_id, wait_seconds=8):
    url = f"https://www.ebay.com/sh/lst/active?keyword={ebay_id}"
    target = _new_target(url)
    ws_url = target["webSocketDebuggerUrl"]
    try:
        await _eval(ws_url, "document.title", wait=wait_seconds)

        page_state = await _eval(
            ws_url,
            r"""(() => ({
                url: location.href,
                title: document.title,
                text: document.body.innerText.slice(0, 6000)
            }))()""",
        )
        text = page_state.get("text") or ""
        page_url = page_state.get("url") or ""
        if "signin.ebay.com" in page_url or "Sign in" in text[:1000] and "eBay" in text[:1000]:
            return "LOGIN_REQUIRED", "Dedicated automation browser is not signed in to eBay Seller Hub."
        if "Looks like you don't have any active listings" in text or "Results:0" in text:
            return "ALREADY_NOT_ACTIVE", "Seller Hub active search returned 0 rows."
        if ebay_id not in text:
            return "SEARCH_MISMATCH", "Item id not found on Seller Hub active page."

        select_result = await _eval(
            ws_url,
            rf"""(() => {{
                const boxes=[...document.querySelectorAll('input[type=checkbox]')];
                const box=boxes.find(x => (x.value||'').includes('{ebay_id}') || (x.getAttribute('aria-label')||'').includes('{ebay_id}'));
                if (!box) return {{ok:false,msg:'checkbox not found'}};
                if (!box.checked) box.click();
                return {{ok:true, checked:box.checked, value:box.value}};
            }})()""",
        )
        if not select_result or not select_result.get("ok"):
            return "FAILED_SELECT", json.dumps(select_result, ensure_ascii=False)

        await _eval(
            ws_url,
            r"""(() => {
                const btn=[...document.querySelectorAll('button')].find(e => (e.innerText||'').trim()==='Actions' && !!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                if (!btn) return false;
                btn.click();
                return true;
            })()""",
            wait=1,
        )
        clicked_end = await _eval(
            ws_url,
            r"""(() => {
                const btn=[...document.querySelectorAll('button')].find(e => (e.innerText||'').trim()==='End listings' && !!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                if (!btn) return false;
                btn.click();
                return true;
            })()""",
            wait=1,
        )
        if not clicked_end:
            return "FAILED_OPEN_END_DIALOG", "End listings action not found."

        confirmed = await _eval(
            ws_url,
            r"""(() => {
                const buttons=[...document.querySelectorAll('button')].filter(e => (e.innerText||'').trim()==='End listing' && !!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                if (!buttons.length) return false;
                buttons[buttons.length - 1].click();
                return true;
            })()""",
            wait=2,
        )
        if not confirmed:
            return "FAILED_CONFIRM", "End listing confirmation button not found."

        final_state = await _eval(
            ws_url,
            r"""(() => ({
                url: location.href,
                title: document.title,
                text: document.body.innerText.slice(0, 6000)
            }))()""",
            wait=wait_seconds,
        )
        final_text = final_state.get("text") or ""
        if "listing was successfully ended" in final_text or "listings were successfully ended" in final_text:
            return "ENDED_CONFIRMED_SELLER_HUB", "Seller Hub success banner detected."
        if "Looks like you don't have any active listings" in final_text or "Results:0" in final_text:
            return "ENDED_PROBABLY_CONFIRMED", "Active search returned 0 rows after confirmation."
        return "UNKNOWN_AFTER_CONFIRM", final_text[:500]
    finally:
        _close_target(target.get("id"))


def _detach_printify(product_id):
    if not product_id or not str(product_id).strip():
        return "no_product_id"
    base = Config.Printify_API_URL.rstrip("/")
    headers = {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }
    response = requests.post(
        f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/unpublish.json",
        headers=headers,
        timeout=180,
    )
    if response.status_code not in {200, 201, 202, 204}:
        return f"printify_unpublish_http_{response.status_code}"
    return "printify_external_detached"


def _mark_workbook_retired(old_id, ebay_id, result):
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    cols = {v: i + 1 for i, v in enumerate(headers)}
    if "Retire_Timestamp" not in cols:
        ws.cell(1, ws.max_column + 1).value = "Retire_Timestamp"
        cols["Retire_Timestamp"] = ws.max_column
    if "Retire_Result" not in cols:
        ws.cell(1, ws.max_column + 1).value = "Retire_Result"
        cols["Retire_Result"] = ws.max_column
    for row in range(2, ws.max_row + 1):
        row_id = str(ws.cell(row, cols["ID"]).value or "").strip()
        row_ebay = str(ws.cell(row, cols.get("eBay_Item_ID", 0)).value or "").strip() if "eBay_Item_ID" in cols else ""
        if row_id == old_id or row_ebay == ebay_id:
            ws.cell(row, cols["Status"]).value = "Retired_Replaced"
            ws.cell(row, cols["Retire_Timestamp"]).value = datetime.now()
            ws.cell(row, cols["Retire_Result"]).value = result
            break
    wb.save(EBAY_BOOK)
    wb.close()


async def run(limit=1, delay=15, dry_run=False, detach_printify=True):
    rows = _read_queue()
    pending = [
        (idx, row)
        for idx, row in enumerate(rows)
        if str(row.get("Status") or "").strip() in {"WAIT_SAFE_END_LISTING_PATH", "RETIRE_FAILED_RETRY"}
    ]
    processed = 0
    for idx, row in pending[:limit]:
        old_id = str(row.get("Old_ID") or "").strip()
        ebay_id = str(row.get("Old_eBay_Item_ID") or "").strip()
        product_id = str(row.get("Old_Printify_Product_ID") or "").strip()
        if dry_run:
            print(f"[RETIRE-DRY] {old_id} ebay={ebay_id} product={product_id}")
            continue
        try:
            result, note = await _retire_one_via_sellerhub(ebay_id)
            detach_note = ""
            if result in {"ENDED_CONFIRMED_SELLER_HUB", "ENDED_PROBABLY_CONFIRMED", "ALREADY_NOT_ACTIVE"} and detach_printify:
                detach_note = _detach_printify(product_id)
            if result == "LOGIN_REQUIRED":
                row["Status"] = str(row.get("Status") or "WAIT_SAFE_END_LISTING_PATH")
            else:
                row["Status"] = "RETIRED_CONFIRMED" if result in {"ENDED_CONFIRMED_SELLER_HUB", "ENDED_PROBABLY_CONFIRMED", "ALREADY_NOT_ACTIVE"} else "RETIRE_FAILED_RETRY"
            row["Retire_Attempted_At"] = _now()
            row["Retire_Result"] = f"{result}; {note}; {detach_note}".strip("; ")
            if row["Status"] == "RETIRED_CONFIRMED":
                _mark_workbook_retired(old_id, ebay_id, row["Retire_Result"])
            _write_queue(rows)
            _log(old_id, ebay_id, result, f"{note}; {detach_note}".strip("; "))
            print(f"[RETIRE-{row['Status']}] {old_id} ebay={ebay_id} {result} {detach_note}")
            processed += 1
            if processed < limit:
                time.sleep(delay)
        except Exception as exc:
            row["Status"] = "RETIRE_FAILED_RETRY"
            row["Retire_Attempted_At"] = _now()
            row["Retire_Result"] = str(exc)
            _write_queue(rows)
            _log(old_id, ebay_id, "EXCEPTION", str(exc))
            print(f"[RETIRE-FAIL] {old_id} ebay={ebay_id}: {exc}")
    print(f"[RETIRE-DONE] attempted={processed} limit={limit}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=1)
    parser.add_argument("--delay", type=int, default=15)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-printify-detach", action="store_true")
    args = parser.parse_args()
    asyncio.run(run(args.limit, args.delay, args.dry_run, not args.no_printify_detach))


if __name__ == "__main__":
    main()
