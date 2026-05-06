"""Prepare replacement-listing queue for eBay cover failures.

This is a fallback plan only. It does not create, publish, revise, or end any
listing. It records which live items should be replaced if Printify source
repair plus re-sync cannot update eBay's inventory-managed variation pictures.
"""

from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
DECISIONS = DATABASE_DIR / "eBay_Cover_Repair_Decisions.csv"
RETIRE_QUEUE = DATABASE_DIR / "eBay_Retire_Queue.csv"
OUT_CSV = DATABASE_DIR / "eBay_Cover_Replacement_Queue.csv"
OUT_MD = DATABASE_DIR / "eBay_Cover_Replacement_Queue.md"

HEADERS = [
    "Priority",
    "ID",
    "Replacement_SKU",
    "Product_Type",
    "Category",
    "Old_eBay_Item_ID",
    "Old_Printify_Product_ID",
    "Repair_Method",
    "Replacement_Status",
    "Title",
    "Price",
    "Production_Path",
    "Cover_Path",
    "Action_Sequence",
    "Retire_Old_Only_After",
]


def now_text() -> str:
    return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def workbook_rows() -> dict[str, dict[str, object]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    rows: dict[str, dict[str, object]] = {}
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            item_id = values[cols["ID"]]
            if not item_id:
                continue
            rows[clean(item_id)] = {header: values[index] for header, index in cols.items()}
    finally:
        wb.close()
    return rows


def build_rows() -> list[dict[str, str]]:
    wb_rows = workbook_rows()
    retire_rows = {clean(row.get("Old_ID")): row for row in read_csv(RETIRE_QUEUE)}
    cover_only_proven = sum(1 for row in retire_rows.values() if clean(row.get("Replacement_eBay_Item_ID"))) >= 3
    output = []
    for decision in read_csv(DECISIONS):
        item_id = clean(decision.get("ID"))
        source = wb_rows.get(item_id, {})
        product_type = clean(source.get("Product_Type") or decision.get("Product_Type"))
        repair_method = clean(decision.get("Repair_Method"))
        repair_status = clean(decision.get("Status"))
        if repair_method not in {"SOURCE_REPAIR_REQUIRED", "NON_STICKER_REVIEW_REQUIRED"}:
            continue
        replacement_sku = f"{item_id}-FIX1"
        if item_id in retire_rows:
            retire_row = retire_rows[item_id]
            status = "REPLACEMENT_PUBLISHED_LIVE_PASS"
            priority = "110"
            replacement_sku = clean(retire_row.get("Replacement_ID")) or replacement_sku
            action = (
                "Replacement listing has passed live buyer-page audit. Keep in retire queue until a safe eBay "
                "end-listing path is confirmed, then retire the old item."
            )
        elif repair_method == "SOURCE_REPAIR_REQUIRED" and product_type == "Sticker" and cover_only_proven:
            status = "READY_TO_REPLACE_VERIFIED"
            priority = "100"
            action = (
                "Cover-only replacement path is proven by multiple live audits. Create a replacement listing using "
                "Cover-only custom art plus Printify official mockups, verify live buyer-page image, then retire the old listing."
            )
        elif repair_method == "SOURCE_REPAIR_REQUIRED":
            if repair_status == "SOURCE_REPAIR_DONE_LIVE_STILL_BAD":
                status = "READY_TO_REPLACE_VERIFIED"
                priority = "100"
                action = (
                    "Source repair plus live eBay audit already failed. Create a replacement listing from the same "
                    "local Production/Cover/Gallery assets, verify the new live buyer-page image, then retire the old listing."
                )
            else:
                status = "WAIT_SOURCE_REPAIR_RESULT"
                priority = "90"
                action = (
                    "First attempt Printify source mockup repair and re-sync. If live eBay audit still fails, "
                    "create replacement listing from the same local Production/Cover/Gallery assets using fixed image gate."
                )
        else:
            status = "REVIEW_BEFORE_REPLACE"
            priority = "60"
            action = (
                "Review because Poster/Acrylic cover can be visually similar to the main artwork. Replace only after "
                "a human or vision gate confirms buyer-facing image is materially wrong."
            )
        output.append(
            {
                "Priority": priority,
                "ID": item_id,
                "Replacement_SKU": replacement_sku,
                "Product_Type": product_type,
                "Category": clean(source.get("Category")),
                "Old_eBay_Item_ID": clean(source.get("eBay_Item_ID") or decision.get("eBay_Item_ID")),
                "Old_Printify_Product_ID": clean(source.get("Printify_Product_ID") or decision.get("Printify_Product_ID")),
                "Repair_Method": repair_method,
                "Replacement_Status": status,
                "Title": clean(source.get("Title")),
                "Price": clean(source.get("Price")),
                "Production_Path": clean(source.get("Production_Path")),
                "Cover_Path": clean(source.get("Cover_Path")),
                "Action_Sequence": action,
                "Retire_Old_Only_After": "New listing live audit returns LIKELY_COVER and local Production_Design exact/visual audit passes.",
            }
        )
    output.sort(key=lambda row: (-int(row["Priority"]), row["Product_Type"], row["ID"]))
    return output


def write_outputs(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    by_status = Counter(row["Replacement_Status"] for row in rows)
    by_type = Counter(row["Product_Type"] for row in rows)
    lines = [
        "# eBay Cover Replacement Queue",
        "",
        f"Generated: {now_text()} America/New_York",
        "",
        "This is a fallback queue only. It does not create, publish, revise, or end listings.",
        "",
        "## Counts",
        "",
    ]
    for key, count in sorted(by_status.items()):
        lines.append(f"- {key}: {count}")
    lines.extend(["", "## Product Types", ""])
    for key, count in sorted(by_type.items()):
        lines.append(f"- {key}: {count}")
    lines.extend(
        [
            "",
            "## Rule",
            "",
            "- Try Printify source repair and re-sync first.",
            "- If eBay inventory-managed variation pictures remain unchanged, create a verified replacement listing.",
            "- Retire the old listing only after the replacement passes live cover and production-design audits.",
            "",
            f"CSV: `{OUT_CSV}`",
            "",
        ]
    )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_outputs(rows)
    print(f"[COVER-REPLACEMENT-QUEUE] rows={len(rows)} csv={OUT_CSV}")
    for key, count in sorted(Counter(row["Replacement_Status"] for row in rows).items()):
        print(f"[COVER-REPLACEMENT-QUEUE] {key}={count}")


if __name__ == "__main__":
    main()
