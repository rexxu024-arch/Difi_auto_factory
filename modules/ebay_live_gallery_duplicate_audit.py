"""Read-only eBay buyer-page gallery duplicate audit through Edge CDP.

This complements the Printify-side audit. The only thing that ultimately
matters to buyer trust is what eBay shows publicly, so this script samples live
buyer pages and records repeated gallery picture sources.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import websockets
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
REPAIR_QUEUE = DATABASE_DIR / "Printify_Gallery_Repair_Queue.csv"
OUT_CSV = DATABASE_DIR / "eBay_Live_Gallery_Duplicate_Audit.csv"

HEADERS = [
    "Timestamp",
    "ID",
    "Product_Type",
    "eBay_Item_ID",
    "URL",
    "Picture_Count",
    "Unique_Source_Count",
    "Duplicate_Source_Count",
    "Duplicate_Source_Keys",
    "Picture_Slot_Count",
    "Unique_Slot_Source_Count",
    "Duplicate_Slot_Count",
    "Duplicate_Slot_Keys",
    "Result",
    "Title",
    "Error",
]


def now_text() -> str:
    return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def workbook_by_id() -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    rows: dict[str, dict[str, str]] = {}
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not values[cols["ID"]]:
                continue
            item_id = clean(values[cols["ID"]])
            rows[item_id] = {
                "ID": item_id,
                "Product_Type": clean(values[cols.get("Product_Type")]),
                "eBay_Item_ID": clean(values[cols.get("eBay_Item_ID")]),
                "Title": clean(values[cols.get("Title")]),
                "Status": clean(values[cols.get("Status")]),
            }
    finally:
        wb.close()
    return rows


def candidates(limit: int, ids: set[str] | None = None) -> list[dict[str, str]]:
    workbook = workbook_by_id()
    rows = []
    if ids:
        for item_id in ids:
            row = workbook.get(item_id)
            if row and row.get("eBay_Item_ID"):
                rows.append(row)
        return rows[:limit] if limit else rows
    for item in read_csv(REPAIR_QUEUE):
        row = workbook.get(clean(item.get("ID")))
        if not row:
            continue
        if not row.get("eBay_Item_ID") or row.get("Status").startswith("Retired"):
            continue
        rows.append(row)
        if limit and len(rows) >= limit:
            break
    return rows


def http_text(url: str, method: str = "GET") -> str:
    req = urllib.request.Request(url, method=method)
    with urllib.request.urlopen(req, timeout=10) as response:
        return response.read().decode("utf-8", errors="replace")


def http_json(url: str, method: str = "GET") -> dict:
    return json.loads(http_text(url, method=method))


def normalize_ebay_image(src: str) -> str:
    match = re.search(r"/images/g/([^/]+)/", src)
    return match.group(1) if match else src.split("?", 1)[0]


def picture_number(alt: str) -> int | None:
    match = re.search(r"Picture\s+(\d+)\s+of\s+\d+", alt or "")
    return int(match.group(1)) if match else None


def classify_live_gallery(pictures: list[dict], keys: list[str], duplicates: dict[str, int]) -> str:
    if not duplicates:
        return "OK"
    numbered = []
    for picture, key in zip(pictures, keys):
        number = picture_number(clean(picture.get("alt")))
        if number is not None:
            numbered.append((number, key))
    by_number: dict[int, str] = {}
    for number, key in sorted(numbered):
        by_number.setdefault(number, key)
    numbered_counts = Counter(by_number.values())
    numbered_duplicates = {key: count for key, count in numbered_counts.items() if count > 1}
    duplicate_slots = sum(count - 1 for count in numbered_duplicates.values())
    if duplicate_slots == 0:
        return "OK_DOM_DUPLICATE_ONLY"
    if duplicate_slots == 1 and len(numbered_duplicates) == 1:
        repeated_key = next(iter(numbered_duplicates))
        repeated_numbers = [number for number, key in by_number.items() if key == repeated_key]
        if repeated_numbers == [1, 2]:
            return "CHECK_LIVE_PRIMARY_DUPLICATE_REVIEW"
    return "CHECK_LIVE_DUPLICATE"


def numbered_slot_counts(pictures: list[dict], keys: list[str]) -> tuple[int, int, int, dict[str, int]]:
    by_number: dict[int, str] = {}
    for picture, key in zip(pictures, keys):
        number = picture_number(clean(picture.get("alt")))
        if number is not None:
            by_number.setdefault(number, key)
    counts = Counter(by_number.values())
    duplicates = {key: count for key, count in counts.items() if count > 1}
    duplicate_slots = sum(count - 1 for count in duplicates.values())
    return len(by_number), len(counts), duplicate_slots, duplicates


async def cdp_eval(ws, seq_state: dict[str, int], method: str, params: dict | None = None) -> dict:
    seq_state["seq"] += 1
    message = {"id": seq_state["seq"], "method": method}
    if params is not None:
        message["params"] = params
    await ws.send(json.dumps(message))
    while True:
        data = json.loads(await ws.recv())
        if data.get("id") == seq_state["seq"]:
            return data


async def audit_one(row: dict[str, str], port: int, wait_seconds: float) -> dict[str, str]:
    ebay_id = row["eBay_Item_ID"]
    url = f"https://www.ebay.com/itm/{ebay_id}"
    target = None
    try:
        encoded = urllib.parse.quote(url, safe="")
        target = http_json(f"http://127.0.0.1:{port}/json/new?{encoded}", method="PUT")
        async with websockets.connect(target["webSocketDebuggerUrl"], max_size=20_000_000) as ws:
            state = {"seq": 0}
            await cdp_eval(ws, state, "Page.enable")
            await cdp_eval(ws, state, "Runtime.enable")
            await cdp_eval(ws, state, "Page.navigate", {"url": url})
            await asyncio.sleep(wait_seconds)
            expression = r"""
(() => {
  const all=[...document.images].map((img,i)=>({
    i,
    src: img.currentSrc || img.src || '',
    alt: img.alt || '',
    visible: !!(img.offsetWidth || img.offsetHeight || img.getClientRects().length),
    w: img.naturalWidth || 0,
    h: img.naturalHeight || 0,
    box: [img.offsetWidth || 0, img.offsetHeight || 0]
  }));
  const pictures=all.filter(x => x.visible && /^Picture \d+ of \d+/.test(x.alt || '') && /i\.ebayimg\.com/.test(x.src || ''));
  const title=(document.querySelector('h1') || {}).innerText || document.title || '';
  return {url: location.href, title, pictures};
})()
"""
            result = await cdp_eval(ws, state, "Runtime.evaluate", {"expression": expression, "returnByValue": True})
            value = result.get("result", {}).get("result", {}).get("value") or {}
            pictures = value.get("pictures") or []
            keys = [normalize_ebay_image(clean(picture.get("src"))) for picture in pictures]
            counts = Counter(keys)
            duplicates = {key: count for key, count in counts.items() if count > 1}
            result = classify_live_gallery(pictures, keys, duplicates)
            slot_count, unique_slot_count, slot_duplicates, slot_duplicate_keys = numbered_slot_counts(pictures, keys)
            return {
                "Timestamp": now_text(),
                "ID": row["ID"],
                "Product_Type": row["Product_Type"],
                "eBay_Item_ID": ebay_id,
                "URL": value.get("url") or url,
                "Picture_Count": str(len(pictures)),
                "Unique_Source_Count": str(len(counts)),
                "Duplicate_Source_Count": str(sum(count - 1 for count in duplicates.values())),
                "Duplicate_Source_Keys": "|".join(f"{key}:{count}" for key, count in duplicates.items()),
                "Picture_Slot_Count": str(slot_count),
                "Unique_Slot_Source_Count": str(unique_slot_count),
                "Duplicate_Slot_Count": str(slot_duplicates),
                "Duplicate_Slot_Keys": "|".join(f"{key}:{count}" for key, count in slot_duplicate_keys.items()),
                "Result": result,
                "Title": clean(value.get("title")) or row.get("Title", ""),
                "Error": "",
            }
    except Exception as exc:  # noqa: BLE001
        return {
            "Timestamp": now_text(),
            "ID": row["ID"],
            "Product_Type": row["Product_Type"],
            "eBay_Item_ID": ebay_id,
            "URL": url,
            "Picture_Count": "",
            "Unique_Source_Count": "",
            "Duplicate_Source_Count": "",
            "Duplicate_Source_Keys": "",
            "Picture_Slot_Count": "",
            "Unique_Slot_Source_Count": "",
            "Duplicate_Slot_Count": "",
            "Duplicate_Slot_Keys": "",
            "Result": "ERROR",
            "Title": row.get("Title", ""),
            "Error": f"{type(exc).__name__}: {exc}"[:500],
        }
    finally:
        if target and target.get("id"):
            try:
                http_text(f"http://127.0.0.1:{port}/json/close/{target['id']}")
            except Exception:
                pass


async def run_async(limit: int, ids: set[str] | None, port: int, wait_seconds: float, sleep_seconds: float) -> list[dict[str, str]]:
    rows = candidates(limit=limit, ids=ids)
    records = []
    for row in rows:
        record = await audit_one(row, port=port, wait_seconds=wait_seconds)
        records.append(record)
        print(
            f"[EBAY-LIVE-GALLERY] {record['ID']} result={record['Result']} "
            f"pictures={record['Picture_Count']} unique={record['Unique_Source_Count']} dup={record['Duplicate_Source_Count']}",
            flush=True,
        )
        await asyncio.sleep(max(0.0, sleep_seconds))
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(records)
    checks = sum(1 for record in records if record["Result"] != "OK")
    print(f"[EBAY-LIVE-GALLERY-DONE] rows={len(records)} checks={checks} csv={OUT_CSV}")
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument("--ids", default="")
    parser.add_argument("--port", type=int, default=int(os.environ.get("OPENCLAW_EDGE_CDP_PORT", "9223")))
    parser.add_argument("--wait-seconds", type=float, default=6.0)
    parser.add_argument("--sleep-seconds", type=float, default=1.5)
    args = parser.parse_args()
    ids = {part.strip() for part in args.ids.split(",") if part.strip()} or None
    asyncio.run(
        run_async(
            limit=args.limit,
            ids=ids,
            port=args.port,
            wait_seconds=args.wait_seconds,
            sleep_seconds=args.sleep_seconds,
        )
    )


if __name__ == "__main__":
    main()
