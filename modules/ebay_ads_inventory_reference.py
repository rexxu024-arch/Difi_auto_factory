from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.ebay_ads_standard import AD_RATE, DEFAULT_CAMPAIGN_ID, _api, resolve_campaign_id

DATABASE = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
PENDING_ADS = DATABASE / "ebay_ads_pending_2pct.csv"
OUT_CSV = DATABASE / "ebay_ads_inventory_reference.csv"

FIELDS = [
    "Timestamp",
    "Action",
    "ID",
    "SKU",
    "Product_Type",
    "eBay_Item_ID",
    "Campaign_ID",
    "Inventory_Reference_ID",
    "Inventory_Reference_Type",
    "Bid_Percentage",
    "HTTP_Status",
    "Child_Status",
    "Ad_IDs",
    "Result",
    "Error",
]


def clean(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def pending_ids(status_prefix: str = "PENDING_API_RETRY") -> list[str]:
    if not PENDING_ADS.exists():
        return []
    out: list[str] = []
    with PENDING_ADS.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if clean(row.get("Status")).startswith(status_prefix):
                item_id = clean(row.get("ID"))
                if item_id and item_id not in out:
                    out.append(item_id)
    return out


def load_rows(limit: int = 1, ids: list[str] | None = None) -> list[dict[str, str]]:
    selected = set(ids or pending_ids())
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        data = {str(k): clean(v) for k, v in zip(headers, raw)}
        item_id = data.get("ID", "")
        if selected and item_id not in selected:
            continue
        if not data.get("eBay_Item_ID"):
            continue
        if "sticker" in data.get("Title", "").lower():
            continue
        rows.append(data)
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


def append_logs(rows: list[dict[str, Any]]) -> None:
    exists = OUT_CSV.exists()
    with OUT_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def parse_child(response, fallback_status: int) -> tuple[str, str, str, str]:
    try:
        body = response.json()
    except Exception:
        return str(fallback_status), "", clean(response.text)[:1500], clean(response.text)[:1500]
    responses = body.get("responses") or []
    first = responses[0] if responses else {}
    child_status = str(first.get("statusCode") or fallback_status)
    ads = first.get("ads") or []
    ad_ids = ",".join(clean(ad.get("adId")) for ad in ads if clean(ad.get("adId")))
    errors = first.get("errors") or body.get("errors") or []
    error_text = "; ".join(clean(err.get("message") or err.get("errorId") or err) for err in errors)
    return child_status, ad_ids, json.dumps(body, ensure_ascii=False)[:1800], error_text


def create_ads_by_inventory_reference(rows: list[dict[str, str]], campaign_id: str, execute: bool) -> list[dict[str, Any]]:
    logs: list[dict[str, Any]] = []
    timestamp = now_text()
    if not rows:
        return logs
    requests_payload = []
    for row in rows:
        # Printify-origin eBay listings show as one-variation group listings.
        # eBay's Marketing API accepts the multi-variation eBay item id as INVENTORY_ITEM_GROUP.
        requests_payload.append(
            {
                "bidPercentage": AD_RATE,
                "inventoryReferenceId": row["eBay_Item_ID"],
                "inventoryReferenceType": "INVENTORY_ITEM_GROUP",
            }
        )
    if execute:
        response = _api(
            "POST",
            f"/sell/marketing/v1/ad_campaign/{campaign_id}/bulk_create_ads_by_inventory_reference",
            json={"requests": requests_payload},
        )
        http_status = response.status_code
        try:
            response.raise_for_status()
        except Exception:
            # Multi-status or business errors are parsed per row below.
            if http_status not in {200, 201, 202, 207, 409}:
                raise
        try:
            body = response.json()
        except Exception:
            body = {"raw": response.text[:1800]}
        responses = body.get("responses") or []
    else:
        response = None
        http_status = ""
        body = {"dry_run": True, "requests": requests_payload}
        responses = []
    for index, row in enumerate(rows):
        if execute:
            child = responses[index] if index < len(responses) else {}
            child_status = str(child.get("statusCode") or http_status)
            ads = child.get("ads") or []
            ad_ids = ",".join(clean(ad.get("adId")) for ad in ads if clean(ad.get("adId")))
            errors = child.get("errors") or []
            error_text = "; ".join(clean(err.get("message") or err.get("errorId") or err) for err in errors)
            result = json.dumps(child or body, ensure_ascii=False)[:1800]
        else:
            child_status = ""
            ad_ids = ""
            error_text = ""
            result = "DRY_RUN only; no ad was created."
        logs.append(
            {
                "Timestamp": timestamp,
                "Action": "CREATE_AD_BY_INVENTORY_REFERENCE" if execute else "DRY_RUN_CREATE_AD_BY_INVENTORY_REFERENCE",
                "ID": row.get("ID", ""),
                "SKU": row.get("SKU", ""),
                "Product_Type": row.get("Product_Type", ""),
                "eBay_Item_ID": row.get("eBay_Item_ID", ""),
                "Campaign_ID": campaign_id,
                "Inventory_Reference_ID": row.get("eBay_Item_ID", ""),
                "Inventory_Reference_Type": "INVENTORY_ITEM_GROUP",
                "Bid_Percentage": AD_RATE,
                "HTTP_Status": http_status,
                "Child_Status": child_status,
                "Ad_IDs": ad_ids,
                "Result": result,
                "Error": error_text,
            }
        )
    append_logs(logs)
    ok = sum(1 for row in logs if str(row.get("Child_Status")).startswith("2") and row.get("Ad_IDs"))
    print(f"[EBAY-ADS-INVENTORY-REF] execute={execute} rows={len(rows)} ok={ok} campaign={campaign_id}")
    return logs


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=1)
    parser.add_argument("--id", action="append", default=[])
    parser.add_argument("--campaign-id", default="")
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args()

    campaign_id = resolve_campaign_id(args.campaign_id or DEFAULT_CAMPAIGN_ID)
    if not campaign_id:
        raise RuntimeError("No usable eBay campaign id found.")
    rows = load_rows(limit=args.limit, ids=args.id or None)
    create_ads_by_inventory_reference(rows, campaign_id=campaign_id, execute=args.execute)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
