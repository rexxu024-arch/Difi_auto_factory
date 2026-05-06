import argparse
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
SYNCABLE_STATUSES = {
    "Printify_UI_Mockups5",
    "Printify_Published_Mockups5",
    "Printify_BaseStaged_DefaultMockups3",
}


def _headers():
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def _canonical_product_type(value):
    value = str(value or "").strip().lower()
    if value.startswith("poster"):
        return "Poster"
    if value.startswith("acry"):
        return "Acrylic"
    if value.startswith("sticker"):
        return "Sticker"
    return str(value or "").strip()


def _load_rows(product_type=None, limit=0):
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cols = {header: index for index, header in enumerate(headers)}
        wanted_type = _canonical_product_type(product_type) if product_type else None
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[cols["ID"]]:
                continue
            product_id = str(row[cols.get("Printify_Product_ID")] or "").strip()
            if not product_id:
                continue
            status = str(row[cols["Status"]] or "").strip()
            if status not in SYNCABLE_STATUSES:
                continue
            row_type = _canonical_product_type(row[cols["Product_Type"]])
            if wanted_type and row_type != wanted_type:
                continue
            rows.append({
                "ID": row[cols["ID"]],
                "Product_Type": row_type,
                "Title": str(row[cols["Title"]] or "").strip(),
                "Description": str(row[cols["Description"]] or "").strip(),
                "Printify_Product_ID": product_id,
            })
            if limit and len(rows) >= limit:
                break
        return rows
    finally:
        wb.close()


def sync_metadata(product_type=None, limit=0, dry_run=False):
    if not Config.Printify_API_KEY:
        raise RuntimeError("Printify_API_KEY is missing")
    rows = _load_rows(product_type=product_type, limit=limit)
    base = Config.Printify_API_URL.rstrip("/")
    synced = 0
    for row in rows:
        payload = {"title": row["Title"], "description": row["Description"]}
        if dry_run:
            print(f"[DRY-RUN] {row['ID']} {row['Product_Type']} product={row['Printify_Product_ID']}")
            continue
        response = requests.put(
            f"{base}/shops/{Config.Printify_SHOP_ID}/products/{row['Printify_Product_ID']}.json",
            headers=_headers(),
            json=payload,
            timeout=120,
        )
        response.raise_for_status()
        synced += 1
        print(f"[SYNC] {row['ID']} product={row['Printify_Product_ID']} title_desc_updated")
    print(f"[DONE] Printify metadata sync: {synced}/{len(rows)}")
    return synced


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--product-type", choices=["Sticker", "Poster", "Acrylic"])
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    sync_metadata(product_type=args.product_type, limit=args.limit, dry_run=args.dry_run)
