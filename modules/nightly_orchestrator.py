import argparse
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
HANDOFF_LOG = PROJECT_ROOT / "Database" / "nightly_handoff_log.txt"


def _log(message):
    HANDOFF_LOG.parent.mkdir(parents=True, exist_ok=True)
    line = f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {message}"
    print(line, flush=True)
    with HANDOFF_LOG.open("a", encoding="utf-8") as handle:
        handle.write(line + "\n")


def _status_counts():
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cols = {header: idx for idx, header in enumerate(headers)}
        counts = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            status = row[cols["Status"]]
            counts[status] = counts.get(status, 0) + 1
        return counts
    finally:
        wb.close()


def _has_work():
    counts = _status_counts()
    return any(
        counts.get(status, 0)
        for status in ("Ready_for_Printify", "Printify_UI_Failed", "Printify_BaseStaged_DefaultMockups3", "Printify_PrimaryFix_Needed")
    )


def _run_step(args, timeout_seconds):
    started = time.time()
    proc = subprocess.run(
        [sys.executable, *args],
        cwd=PROJECT_ROOT,
        text=True,
        capture_output=True,
        timeout=timeout_seconds,
    )
    elapsed = time.time() - started
    if proc.stdout.strip():
        _log(proc.stdout.strip()[-1800:])
    if proc.stderr.strip():
        _log("[stderr] " + proc.stderr.strip()[-1200:])
    return proc.returncode, elapsed


def run_stickers(max_items=0, deadline_hours=0, publish=False):
    started = time.time()
    completed_before = _status_counts().get("Printify_UI_Mockups5", 0)
    processed = 0
    _log(f"Sticker overnight runner started. counts={_status_counts()}")
    while _has_work():
        if max_items and processed >= max_items:
            break
        if deadline_hours and time.time() - started > deadline_hours * 3600:
            _log("Sticker runner reached deadline.")
            break
        try:
            full_args = ["modules/printify_full_pipeline.py", "--limit", "1", "--batch-size", "0"]
            ui_args = ["modules/printify_mockup_ui_uploader.py", "--limit", "1", "--expected-count", "5"]
            if publish:
                full_args.append("--publish")
                ui_args.append("--publish")
            code, elapsed = _run_step(full_args, 720)
            if code:
                _log(f"Full pipeline returned code={code} elapsed={elapsed:.1f}s; attempting UI recovery.")
            _run_step(ui_args, 360)
        except subprocess.TimeoutExpired as exc:
            _log(f"Timeout: {' '.join(exc.cmd)}. Moving to recovery pass.")
            try:
                recovery_args = ["modules/printify_mockup_ui_uploader.py", "--limit", "1", "--expected-count", "5"]
                if publish:
                    recovery_args.append("--publish")
                _run_step(recovery_args, 360)
            except Exception as recovery_exc:
                _log(f"Recovery failed after timeout: {recovery_exc}")
        except Exception as exc:
            _log(f"Runner exception: {exc}")
            time.sleep(20)
        current = _status_counts()
        processed = max(0, current.get("Printify_UI_Mockups5", 0) - completed_before)
        _log(f"Sticker progress processed={processed} counts={current}")
        time.sleep(5)
    _log(f"Sticker overnight runner finished. counts={_status_counts()}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--stickers", action="store_true")
    parser.add_argument("--max-items", type=int, default=0)
    parser.add_argument("--deadline-hours", type=float, default=0)
    parser.add_argument("--publish", action="store_true")
    args = parser.parse_args()
    if args.stickers:
        run_stickers(max_items=args.max_items, deadline_hours=args.deadline_hours, publish=args.publish)
