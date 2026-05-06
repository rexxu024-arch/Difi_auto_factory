import csv
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
OUTPUT_CSV = DATABASE_DIR / "Pricing_Strategy_Matrix.csv"
OUTPUT_XLSX = DATABASE_DIR / "Pricing_Strategy_Matrix.xlsx"


@dataclass(frozen=True)
class ProductCost:
    product_type: str
    current_price: float
    product_cost: float
    shipping_cost: float
    notes: str


PRODUCTS = [
    ProductCost("Sticker", 11.99, 4.80, 4.69, "Sticker cost is a conservative placeholder; verify against current Printify variant before scaling."),
    ProductCost("Poster", 34.99, 6.00, 7.00, "User-provided Printify estimate: poster production about $6 plus $7 shipping."),
    ProductCost("Acrylic", 89.99, 35.43, 15.99, "User-provided Printify estimate: acrylic production $35.43 plus $15.99 shipping."),
]

PRICE_TESTS = {
    "Sticker": [9.99, 11.99, 12.99, 14.99],
    "Poster": [29.99, 34.99, 39.99, 44.99],
    "Acrylic": [79.99, 84.99, 89.99, 94.99],
}

ROWS = [
    "Platform",
    "Product_Type",
    "Scenario",
    "List_Price",
    "Buyer_Shipping_Charged",
    "Printify_Product_Cost",
    "Printify_Shipping_Cost",
    "Ad_Rate",
    "Marketplace_Percent_Fee",
    "Fixed_Fees",
    "Gross_Revenue",
    "Estimated_Fees",
    "Estimated_Profit",
    "Profit_Margin_On_Gross",
    "Notes",
]


def _money(value):
    return round(float(value), 2)


def _simulate(
    product: ProductCost,
    platform: str,
    price: float,
    ad_rate: float,
    buyer_shipping: float,
):
    gross = price + buyer_shipping
    if platform == "eBay":
        marketplace_percent = 0.136
        fixed_fees = 0.40
        notes = "Assumes eBay most-categories final value fee 13.6% plus $0.40 order fee; verify exact category/store fee before scaling. Promoted Listings Standard fee is pay-on-sale."
    else:
        marketplace_percent = 0.065 + 0.03
        fixed_fees = 0.25 + 0.20
        notes = "Assumes Etsy 6.5% transaction fee, US Etsy Payments 3% + $0.25, and $0.20 listing fee allocated to sale."
    fees = gross * (marketplace_percent + ad_rate) + fixed_fees
    profit = gross - product.product_cost - product.shipping_cost - fees
    margin = profit / gross if gross else 0
    return {
        "Gross_Revenue": _money(gross),
        "Estimated_Fees": _money(fees),
        "Estimated_Profit": _money(profit),
        "Profit_Margin_On_Gross": round(margin, 4),
        "Marketplace_Percent_Fee": round(marketplace_percent, 4),
        "Fixed_Fees": _money(fixed_fees),
        "Notes": notes + " " + product.notes,
    }


def build_rows():
    output = []
    for product in PRODUCTS:
        for price in PRICE_TESTS[product.product_type]:
            for platform in ("eBay", "Etsy"):
                for scenario, buyer_shipping in (
                    ("Buyer_Pays_Printify_Shipping", product.shipping_cost),
                    ("Free_Shipping_Embedded", 0.0),
                ):
                    for ad_rate in (0.02, 0.05):
                        result = _simulate(product, platform, price, ad_rate, buyer_shipping)
                        output.append(
                            {
                                "Platform": platform,
                                "Product_Type": product.product_type,
                                "Scenario": scenario,
                                "List_Price": _money(price),
                                "Buyer_Shipping_Charged": _money(buyer_shipping),
                                "Printify_Product_Cost": _money(product.product_cost),
                                "Printify_Shipping_Cost": _money(product.shipping_cost),
                                "Ad_Rate": ad_rate,
                                **result,
                            }
                        )
    return output


def write_outputs(rows):
    DATABASE_DIR.mkdir(exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ROWS)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Matrix"
    ws.append(ROWS)
    for row in rows:
        ws.append([row.get(header, "") for header in ROWS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 12,
        "B": 14,
        "C": 28,
        "D": 12,
        "E": 22,
        "F": 22,
        "G": 22,
        "H": 10,
        "I": 22,
        "J": 12,
        "K": 14,
        "L": 14,
        "M": 16,
        "N": 20,
        "O": 90,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=13):
        for cell in row:
            if isinstance(cell.value, (float, int)):
                cell.number_format = "$0.00"
    for cell in ws["H"][1:]:
        cell.number_format = "0.0%"
    for cell in ws["I"][1:]:
        cell.number_format = "0.0%"
    for cell in ws["N"][1:]:
        cell.number_format = "0.0%"
    wb.save(OUTPUT_XLSX)
    wb.close()


def main():
    rows = build_rows()
    write_outputs(rows)
    print(f"[PRICING] rows={len(rows)} csv={OUTPUT_CSV}")
    print(f"[PRICING] xlsx={OUTPUT_XLSX}")


if __name__ == "__main__":
    sys.exit(main())
