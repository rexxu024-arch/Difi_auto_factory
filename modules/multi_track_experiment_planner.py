from __future__ import annotations

import argparse
import csv
import json
import random
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.image_quality_gate import _metrics, _verdict
from modules.risk_guard import fee_kill_switch


DATABASE = PROJECT_ROOT / "Database"
REVIEW = PROJECT_ROOT / "Review_Packets"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
REGISTRY_PATH = DATABASE / "Unified_Listing_Registry.csv"
DIGITAL_META_PATH = DATABASE / "Digital_Etsy_Metadata.csv"
DIGITAL_QUEUE_PATH = DATABASE / "Etsy_Digital_Gray_Launch_Queue.csv"
FEE_LEDGER_PATH = DATABASE / "Etsy_Fee_Ledger.csv"
COVER_FIX_PATH = DATABASE / "eBay_Online_Cover_Fix_Queue.csv"
COVER_REPAIR_PATH = DATABASE / "eBay_Cover_Repair_Decisions.csv"
QUALITY_GATE_PATH = DATABASE / "Image_Quality_Gate.csv"
PERFORMANCE_LOG_PATH = DATABASE / "Performance_Log.csv"

PLAN_CSV = DATABASE / "Multi_Track_Experiment_Plan.csv"
STATE_JSON = DATABASE / "Multi_Track_Experiment_State.json"
REPORT_MD = REVIEW / f"MULTI_TRACK_EXPERIMENT_PLAN_{datetime.now():%Y%m%d}.md"
NY = ZoneInfo("America/New_York")

TRACKS = {
    "A_LOW_COMPETITION_NICHE": {
        "objective": "Force non-zero traffic through long-tail room-use and buyer-scene terms.",
        "platform": "eBay/Etsy",
        "slot_target": 55,
    },
    "B_HIGH_VOLUME_VALUE": {
        "objective": "Test broad-volume value terms with Rex-grade visuals and safe pricing.",
        "platform": "eBay",
        "slot_target": 55,
    },
    "C_DIGITAL_PURE_PROFIT": {
        "objective": "Use zero-production-cost Etsy digital downloads to test SEO templates cheaply.",
        "platform": "Etsy",
        "slot_target": 55,
    },
}

MOCKUP_MOODS = [
    "reading_nook_warm_lamp",
    "quiet_luxury_apartment",
    "meditation_corner_soft_shadow",
    "dorm_desk_clean_setup",
    "library_shelf_collector_scene",
    "deep_work_home_office",
]

ROOM_USE_TERMS = {
    "Zen": [
        "reading nook decor",
        "meditation room wall art",
        "tea room accent",
        "quiet corner decor",
        "dorm desk calm art",
        "deep work visual",
    ],
    "Academia": [
        "dark academia reading nook",
        "home library wall art",
        "book nook decor",
        "study room poster",
        "scholar desk decor",
        "dorm library print",
    ],
    "Grimdark": [
        "gothic shelf decor",
        "moody desk display",
        "occult library art",
        "dark study room decor",
        "collector shelf object",
        "smoky jade relic",
    ],
}

HIGH_VOLUME_TERMS = {
    "Sticker": ["vinyl sticker set", "laptop stickers", "water bottle decals", "journal stickers", "aesthetic sticker sheet"],
    "Poster": ["wall art poster", "gallery wall decor", "apartment wall art", "poster print", "room decor"],
    "Acrylic": ["acrylic photo block", "desk display", "shelf decor", "collector gift", "office decor"],
}

DIGITAL_TEMPLATES = [
    ("BUYER_PERSONA", "book lover gift, student dorm printable, writer office decor"),
    ("ROOM_USE", "reading nook printable, meditation room art, home office wall decor"),
    ("GIFT_INTENT", "instant download gift, printable art set, last minute decor"),
]

PLAN_FIELDS = [
    "Timestamp",
    "Slot",
    "Track",
    "Track_Objective",
    "Platform_Target",
    "ID",
    "Product_Type",
    "Category",
    "Current_Status",
    "Views_30d",
    "eBay_Item_ID",
    "Etsy_Listing_ID",
    "Price_Target",
    "SEO_Template",
    "Primary_Search_Intent",
    "Secondary_Keywords",
    "Mockup_Mood",
    "Jitter_Minutes",
    "QA_Status",
    "QA_Reason",
    "Risk_Guard",
    "Fee_Impact_USD",
    "Launch_Action",
    "Stop_Rule",
    "Source_Path",
    "Notes",
]


@dataclass
class Candidate:
    local_id: str
    product_type: str
    category: str
    status: str
    title: str = ""
    price: str = ""
    ebay_item_id: str = ""
    etsy_listing_id: str = ""
    views_30d: str = ""
    source_path: str = ""
    printify_product_id: str = ""
    source_kind: str = "physical"


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, rows: list[dict[str, object]], headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _latest_performance_by_item() -> dict[str, dict[str, str]]:
    rows = _read_csv(PERFORMANCE_LOG_PATH)
    latest: dict[str, dict[str, str]] = {}
    for row in rows:
        item_id = clean(row.get("Item_ID"))
        if item_id:
            latest[item_id] = row
    return latest


def _cover_gate_ids() -> set[str]:
    blocked = set()
    for path in [COVER_FIX_PATH, COVER_REPAIR_PATH]:
        for row in _read_csv(path):
            local_id = clean(row.get("ID"))
            status = clean(row.get("Status")).upper()
            result = clean(row.get("Result") or row.get("Online_Result")).upper()
            if not local_id:
                continue
            if "PENDING" in status or "MISMATCH" in result or "SOURCE_REPAIR_REQUIRED" in status:
                if "RETIRED_REPLACED_DONE" not in status:
                    blocked.add(local_id)
    return blocked


def _quality_cache() -> dict[str, list[dict[str, str]]]:
    by_id: dict[str, list[dict[str, str]]] = defaultdict(list)
    pattern = re.compile(r"(Sticker|Poster|Acrylic)-[A-Za-z]+-\d{4}(?:-FIX\d+)?")
    for row in _read_csv(QUALITY_GATE_PATH):
        path = clean(row.get("Path"))
        match = pattern.search(path)
        if match:
            by_id[match.group(0)].append(row)
    return by_id


def load_physical_candidates() -> list[Candidate]:
    if not EBAY_BOOK.exists():
        return []
    performance = _latest_performance_by_item()
    registry = {row.get("ID"): row for row in _read_csv(REGISTRY_PATH) if row.get("ID")}

    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers) if name}
    out: list[Candidate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        local_id = clean(row[cols.get("ID", 0)])
        if not local_id:
            continue
        product_type = clean(row[cols.get("Product_Type", 0)])
        if product_type not in {"Sticker", "Poster", "Acrylic"}:
            continue
        status = clean(row[cols.get("Status", 0)])
        ebay_item_id = clean(row[cols.get("eBay_Item_ID", 0)]) if "eBay_Item_ID" in cols else ""
        perf = performance.get(ebay_item_id, {})
        reg = registry.get(local_id, {})
        views = clean(perf.get("Views_30_Days") or reg.get("Latest_eBay_Views_30_Days"))
        source_path = clean(
            row[cols.get("Cover_Path", 0)] if "Cover_Path" in cols else ""
        ) or clean(reg.get("Cover_Path") or reg.get("Production_Path"))
        out.append(
            Candidate(
                local_id=local_id,
                product_type=product_type,
                category=clean(row[cols.get("Category", 0)]),
                status=status,
                title=clean(row[cols.get("Title", 0)]),
                price=clean(row[cols.get("Price", 0)]),
                ebay_item_id=ebay_item_id,
                views_30d=views,
                source_path=source_path,
                printify_product_id=clean(row[cols.get("Printify_Product_ID", 0)]) if "Printify_Product_ID" in cols else "",
                source_kind="physical",
            )
        )
    wb.close()
    return out


def load_digital_candidates() -> list[Candidate]:
    rows = _read_csv(DIGITAL_META_PATH)
    queue_by_id = {row.get("ID"): row for row in _read_csv(DIGITAL_QUEUE_PATH) if row.get("ID")}
    out: list[Candidate] = []
    for row in rows:
        local_id = clean(row.get("ID"))
        if not local_id:
            continue
        category = local_id.split("-")[1] if "-" in local_id else "Digital"
        queue = queue_by_id.get(local_id, {})
        out.append(
            Candidate(
                local_id=local_id,
                product_type="Digital",
                category=category,
                status=clean(row.get("Status") or queue.get("Launch_Status")),
                title=clean(row.get("Title")),
                price=clean(row.get("Price") or "6.99"),
                etsy_listing_id=clean(queue.get("Etsy_Listing_ID")),
                source_path=clean(row.get("Zip_Path")),
                source_kind="digital",
            )
        )
    return out


def qa_for(candidate: Candidate, quality_cache: dict[str, list[dict[str, str]]], cover_blocked: set[str], audit_images: bool) -> tuple[str, str]:
    status = candidate.status.upper()
    if any(marker in status for marker in ["HOLD", "QUARANTINED", "RETIRED", "BAD"]):
        return "HOLD", f"STATUS_{candidate.status}"

    # Digital products can reuse a physical product ID, so physical cover-gate
    # rows and physical Image_Quality_Gate rows must not bleed into Etsy ZIP QA.
    if candidate.source_kind == "digital":
        zip_path = Path(candidate.source_path) if candidate.source_path else None
        if not zip_path or not zip_path.exists():
            return "HOLD", "DIGITAL_ZIP_MISSING"
        if "PUBLISHED" in status:
            return "PASS", "LIVE_DIGITAL_ALREADY_CONFIRMED"
        if "RESERVED" in status:
            return "READY", "RESERVED_NOT_SPENT_RECONCILE_BEFORE_NEXT_FEE"
        return "READY", "DIGITAL_METADATA_READY"

    if candidate.local_id in cover_blocked:
        return "HOLD", "COVER_GATE_PENDING_OR_U_IMAGE_MISMATCH"

    cached = quality_cache.get(candidate.local_id) or []
    hard_reasons = []
    review_reasons = []
    for row in cached:
        verdict = clean(row.get("Verdict")).upper()
        reason = clean(row.get("Reason")).upper()
        if verdict == "HOLD" and any(token in reason for token in ["LOW_RESOLUTION", "SHADOW_CLIPPING", "HIGHLIGHT_CLIPPING"]):
            hard_reasons.append(reason or "QUALITY_GATE_HOLD")
        elif verdict and verdict != "PASS":
            review_reasons.append(reason or verdict)
    if hard_reasons:
        return "HOLD", ";".join(sorted(set(hard_reasons)))[:240]

    if not audit_images:
        return ("REVIEW", ";".join(review_reasons)[:240]) if review_reasons else ("READY", "IMAGE_AUDIT_SKIPPED")

    path = Path(candidate.source_path) if candidate.source_path else None
    if not path or not path.exists():
        return "HOLD", "SOURCE_IMAGE_MISSING"
    try:
        metrics = _metrics(path)
        verdict, reason = _verdict(metrics)
        reason = clean(reason)
        if verdict == "HOLD" and any(token in reason for token in ["LOW_RESOLUTION", "SHADOW_CLIPPING", "HIGHLIGHT_CLIPPING"]):
            return "HOLD", reason
        if verdict == "HOLD":
            return "HOLD", reason or "IMAGE_QUALITY_HOLD"
        if verdict != "PASS":
            return "REVIEW", reason or verdict
    except Exception as exc:
        return "HOLD", f"IMAGE_AUDIT_ERROR:{type(exc).__name__}"

    return ("REVIEW", ";".join(review_reasons)[:240]) if review_reasons else ("READY", "PASS")


def target_price(candidate: Candidate, track: str) -> str:
    if candidate.product_type == "Digital":
        return "$6.99"
    if candidate.product_type == "Sticker":
        return "$11.99" if track.startswith("A_") else "$10.99-$11.99"
    if candidate.product_type == "Poster":
        return "$34.99" if track.startswith("A_") else "$29.99-$34.99"
    if candidate.product_type == "Acrylic":
        return "$89.99" if track.startswith("A_") else "$79.99-$89.99"
    return candidate.price or ""


def primary_terms(candidate: Candidate, track: str, slot: int) -> tuple[str, str, str]:
    if track == "C_DIGITAL_PURE_PROFIT":
        template, terms = DIGITAL_TEMPLATES[(slot - 1) % len(DIGITAL_TEMPLATES)]
        primary = terms.split(",")[0]
        return template, primary, terms
    category_terms = ROOM_USE_TERMS.get(candidate.category) or ROOM_USE_TERMS.get("Zen")
    if track == "A_LOW_COMPETITION_NICHE":
        primary = category_terms[(slot - 1) % len(category_terms)]
        secondary = ", ".join(category_terms)
        return "ROOM_USE_LONG_TAIL", primary, secondary
    terms = HIGH_VOLUME_TERMS.get(candidate.product_type, ["room decor", "gift", "wall art"])
    primary = terms[(slot - 1) % len(terms)]
    secondary = ", ".join(terms + ["quiet jade", "premium aesthetic", "giftable decor"])
    return "HIGH_VOLUME_VALUE", primary, secondary


def launch_action(candidate: Candidate, track: str, qa_status: str, qa_reason: str) -> str:
    if qa_status == "HOLD":
        return "HOLD_DO_NOT_PUBLISH"
    if candidate.source_kind == "digital":
        if "PUBLISHED" in candidate.status:
            return "MONITOR_LIVE_DIGITAL_TRAFFIC"
        if "RESERVED" in candidate.status:
            return "RECONCILE_RESERVED_BEFORE_ANY_NEW_FEE"
        return "NEXT_ETSY_GRAY_BATCH_UNDER_FEE_CAP"
    if candidate.ebay_item_id:
        return "COPY_OR_PRICE_EXPERIMENT_ON_EXISTING_LISTING"
    if candidate.printify_product_id:
        return "PRINTIFY_SYNC_OR_PUBLISH_AFTER_QA"
    return "BUILD_OR_UPLOAD_AFTER_QA"


def _confirmed_etsy_spend_today() -> float:
    today = datetime.now(NY).strftime("%Y-%m-%d")
    total = 0.0
    for row in _read_csv(FEE_LEDGER_PATH):
        ts = clean(row.get("Timestamp"))
        if not ts.startswith(today):
            continue
        if clean(row.get("Status")).startswith("CONFIRMED"):
            try:
                total += float(row.get("Confirmed_Spent_USD") or 0)
            except ValueError:
                pass
    return total


def _stable_shuffle(rows: list[Candidate], salt: str) -> list[Candidate]:
    shuffled = list(rows)
    rng = random.Random(f"openclaw-{salt}-20260507")
    rng.shuffle(shuffled)
    return shuffled


def select_candidates(physical: list[Candidate], digital: list[Candidate], quality_cache: dict[str, list[dict[str, str]]], cover_blocked: set[str], audit_images: bool) -> list[dict[str, object]]:
    timestamp = now_text()
    qa_cache_result: dict[str, tuple[str, str]] = {}

    def get_qa(c: Candidate) -> tuple[str, str]:
        if c.local_id not in qa_cache_result:
            qa_cache_result[c.local_id] = qa_for(c, quality_cache, cover_blocked, audit_images)
        return qa_cache_result[c.local_id]

    physical_ready = [c for c in physical if get_qa(c)[0] in {"READY", "REVIEW"}]
    physical_holds = [c for c in physical if get_qa(c)[0] == "HOLD"]
    digital_ready = [c for c in digital if get_qa(c)[0] in {"READY", "PASS", "REVIEW"}]
    digital_holds = [c for c in digital if get_qa(c)[0] == "HOLD"]

    low_comp_pool = [
        c
        for c in physical_ready
        if c.product_type in {"Poster", "Acrylic", "Sticker"} and c.category in {"Zen", "Academia", "Grimdark"}
    ]
    high_value_pool = [
        c
        for c in physical_ready
        if c.product_type in {"Sticker", "Poster", "Acrylic"} and c not in low_comp_pool[:20]
    ]
    if len(high_value_pool) < TRACKS["B_HIGH_VOLUME_VALUE"]["slot_target"]:
        high_value_pool.extend([c for c in physical_ready if c not in high_value_pool])

    selected: list[tuple[str, Candidate | None]] = []
    used_physical_ids: set[str] = set()
    for track, pool in [
        ("A_LOW_COMPETITION_NICHE", _stable_shuffle(low_comp_pool, "track-a")),
    ]:
        target = TRACKS[track]["slot_target"]
        used = set()
        for candidate in pool:
            if len([1 for t, _ in selected if t == track]) >= target:
                break
            if candidate.local_id in used:
                continue
            used.add(candidate.local_id)
            used_physical_ids.add(candidate.local_id)
            selected.append((track, candidate))
        while len([1 for t, _ in selected if t == track]) < target:
            selected.append((track, None))

    track = "B_HIGH_VOLUME_VALUE"
    b_pool = [c for c in _stable_shuffle(high_value_pool, "track-b") if c.local_id not in used_physical_ids]
    target = TRACKS[track]["slot_target"]
    used_b = set()
    for candidate in b_pool:
        if len([1 for t, _ in selected if t == track]) >= target:
            break
        if candidate.local_id in used_b:
            continue
        used_b.add(candidate.local_id)
        selected.append((track, candidate))
    while len([1 for t, _ in selected if t == track]) < target:
        selected.append((track, None))

    digital_pool = _stable_shuffle(digital_ready, "track-c")
    for candidate in digital_pool[: TRACKS["C_DIGITAL_PURE_PROFIT"]["slot_target"]]:
        selected.append(("C_DIGITAL_PURE_PROFIT", candidate))
    while len([1 for t, _ in selected if t == "C_DIGITAL_PURE_PROFIT"]) < TRACKS["C_DIGITAL_PURE_PROFIT"]["slot_target"]:
        selected.append(("C_DIGITAL_PURE_PROFIT", None))

    rows: list[dict[str, object]] = []
    per_track_slot = Counter()
    for absolute_slot, (track, candidate) in enumerate(selected, start=1):
        per_track_slot[track] += 1
        slot = per_track_slot[track]
        mood = MOCKUP_MOODS[((slot - 1) // 5) % len(MOCKUP_MOODS)]
        jitter_rng = random.Random(f"{track}-{slot}-{candidate.local_id if candidate else 'backlog'}")
        jitter = jitter_rng.randint(9, 47)
        if candidate:
            qa_status, qa_reason = get_qa(candidate)
            seo_template, primary, secondary = primary_terms(candidate, track, slot)
            fee = 0.20 if track == "C_DIGITAL_PURE_PROFIT" and "PUBLISHED" not in candidate.status and "RESERVED" not in candidate.status else 0.0
            action = launch_action(candidate, track, qa_status, qa_reason)
            risk = "QA_HOLD" if qa_status == "HOLD" else "PASS_GUARDED"
            if track == "C_DIGITAL_PURE_PROFIT":
                risk = "ETSY_FEE_KILL_SWITCH_REQUIRED"
            rows.append(
                {
                    "Timestamp": timestamp,
                    "Slot": absolute_slot,
                    "Track": track,
                    "Track_Objective": TRACKS[track]["objective"],
                    "Platform_Target": TRACKS[track]["platform"],
                    "ID": candidate.local_id,
                    "Product_Type": candidate.product_type,
                    "Category": candidate.category,
                    "Current_Status": candidate.status,
                    "Views_30d": candidate.views_30d,
                    "eBay_Item_ID": candidate.ebay_item_id,
                    "Etsy_Listing_ID": candidate.etsy_listing_id,
                    "Price_Target": target_price(candidate, track),
                    "SEO_Template": seo_template,
                    "Primary_Search_Intent": primary,
                    "Secondary_Keywords": secondary,
                    "Mockup_Mood": mood,
                    "Jitter_Minutes": jitter,
                    "QA_Status": qa_status,
                    "QA_Reason": qa_reason,
                    "Risk_Guard": risk,
                    "Fee_Impact_USD": f"{fee:.2f}",
                    "Launch_Action": action,
                    "Stop_Rule": stop_rule(track),
                    "Source_Path": candidate.source_path,
                    "Notes": notes_for(candidate, track, qa_status, qa_reason),
                }
            )
        else:
            product_type = "Digital" if track == "C_DIGITAL_PURE_PROFIT" else "Physical"
            seo_template, primary, secondary = (
                ("BACKLOG_BUILD", "new candidate required", "generate additional vetted assets")
                if track != "C_DIGITAL_PURE_PROFIT"
                else ("DIGITAL_BACKLOG_BUILD", "build more printable packs", "no fee until QA and explicit gray batch")
            )
            rows.append(
                {
                    "Timestamp": timestamp,
                    "Slot": absolute_slot,
                    "Track": track,
                    "Track_Objective": TRACKS[track]["objective"],
                    "Platform_Target": TRACKS[track]["platform"],
                    "ID": f"BACKLOG-{track}-{slot:03d}",
                    "Product_Type": product_type,
                    "Category": "TBD",
                    "Current_Status": "BACKLOG_NEEDS_ASSET_OR_METADATA",
                    "Views_30d": "",
                    "eBay_Item_ID": "",
                    "Etsy_Listing_ID": "",
                    "Price_Target": "$6.99" if product_type == "Digital" else "",
                    "SEO_Template": seo_template,
                    "Primary_Search_Intent": primary,
                    "Secondary_Keywords": secondary,
                    "Mockup_Mood": mood,
                    "Jitter_Minutes": jitter,
                    "QA_Status": "NOT_READY",
                    "QA_Reason": "NO_CURRENT_CANDIDATE",
                    "Risk_Guard": "BUILD_ONLY_NO_PUBLISH",
                    "Fee_Impact_USD": "0.00",
                    "Launch_Action": "BUILD_ASSET_OR_METADATA_ONLY",
                    "Stop_Rule": stop_rule(track),
                    "Source_Path": "",
                    "Notes": "This slot reserves experiment capacity but must not publish until QA produces a concrete candidate.",
                }
            )

    hold_rows = []
    for c in (physical_holds + digital_holds)[:60]:
        qa_status, qa_reason = get_qa(c)
        hold_rows.append(
            {
                "Timestamp": timestamp,
                "Slot": f"HOLD-{len(hold_rows)+1}",
                "Track": "QA_HOLD_POOL",
                "Track_Objective": "Prevent low-quality or cover-unsafe assets from entering experiments.",
                "Platform_Target": "None",
                "ID": c.local_id,
                "Product_Type": c.product_type,
                "Category": c.category,
                "Current_Status": c.status,
                "Views_30d": c.views_30d,
                "eBay_Item_ID": c.ebay_item_id,
                "Etsy_Listing_ID": "",
                "Price_Target": "",
                "SEO_Template": "NONE",
                "Primary_Search_Intent": "",
                "Secondary_Keywords": "",
                "Mockup_Mood": "",
                "Jitter_Minutes": "",
                "QA_Status": qa_status,
                "QA_Reason": qa_reason,
                "Risk_Guard": "DO_NOT_PUBLISH",
                "Fee_Impact_USD": "0.00",
                "Launch_Action": "HOLD_DO_NOT_PUBLISH",
                "Stop_Rule": "Repair source asset or retire; never spend publish/ad budget.",
                "Source_Path": c.source_path,
                "Notes": "Listed outside the 165 slots so bad assets do not consume experiment capacity.",
            }
        )
    rows.extend(hold_rows)
    return rows


def stop_rule(track: str) -> str:
    if track == "A_LOW_COMPETITION_NICHE":
        return "If 10 live tests get 0 views after 48h, downgrade category terms again before adding volume."
    if track == "B_HIGH_VOLUME_VALUE":
        return "If broad terms get impressions but no clicks, change cover/price; if 0 views, treat account/channel as distribution bottleneck."
    if track == "C_DIGITAL_PURE_PROFIT":
        return "Hard stop at $2/batch and $6/day; if first 10 new paid tests get 0 views, pause fees and rewrite SEO before spending more."
    return "Do not publish until fixed."


def notes_for(candidate: Candidate, track: str, qa_status: str, qa_reason: str) -> str:
    bits = []
    if qa_status == "HOLD":
        bits.append("Blocked by QA before marketplace action.")
    if track == "C_DIGITAL_PURE_PROFIT":
        bits.append("No production/shipping cost; listing fee is the main paid risk.")
    if candidate.product_type == "Sticker":
        bits.append("Sticker Cover Gate rule applies: official/cover-safe mockups only, no U gallery as marketplace cover.")
    if candidate.product_type in {"Poster", "Acrylic"}:
        bits.append("Single artwork product: buyer receives the main design; extra imagery must remain concept/detail context.")
    if candidate.views_30d in {"0", "", None} and candidate.ebay_item_id:
        bits.append("Zero-view or unknown-view item; prioritize search-intent rewrite over more volume.")
    if qa_reason and qa_reason != "PASS":
        bits.append(f"QA: {qa_reason}")
    return " ".join(bits)


def write_state(rows: list[dict[str, object]]) -> dict[str, object]:
    fee = fee_kill_switch()
    today_spend = _confirmed_etsy_spend_today()
    track_counts = Counter(row["Track"] for row in rows if not str(row["Track"]).startswith("QA_"))
    action_counts = Counter(row["Launch_Action"] for row in rows)
    qa_counts = Counter(row["QA_Status"] for row in rows)
    digital_fee_pending = sum(float(row["Fee_Impact_USD"]) for row in rows if row["Track"] == "C_DIGITAL_PURE_PROFIT")
    next_paid_capacity = max(0.0, float(fee.get("daily_listing_fee_cap_usd", 6.0) or 6.0) - today_spend)
    state = {
        "timestamp": now_text(),
        "timezone": "America/New_York",
        "slot_target_total": 165,
        "track_targets": {track: config["slot_target"] for track, config in TRACKS.items()},
        "track_counts": dict(track_counts),
        "action_counts": dict(action_counts),
        "qa_counts": dict(qa_counts),
        "etsy_fee_caps": fee,
        "etsy_confirmed_spend_today_usd": round(today_spend, 2),
        "etsy_next_paid_capacity_today_usd": round(next_paid_capacity, 2),
        "digital_fee_impact_if_all_ready_spent_usd": round(digital_fee_pending, 2),
        "depatterning": {
            "jitter_minutes": "9-47 deterministic-random minutes per row; scheduler may add platform-specific jitter.",
            "mockup_mood_rotation": "Switch every 5 listing slots.",
            "description_tone": "Rex premium shop voice: quiet luxury, smoky jade, room-use scenes, no dry AI-fluff.",
        },
        "hard_blocks": [
            "SHADOW_CLIPPING/LOW_RESOLUTION/HIGHLIGHT_CLIPPING => HOLD",
            "Sticker Cover Gate mismatch => HOLD",
            "Etsy fee > $2/batch or > $6/day => STOP",
            "No Priority/PPC ads; only eBay Promoted Listings Standard / General fixed 2%.",
        ],
        "official_reference_links": {
            "etsy_fees": "https://www.etsy.com/legal/fees/",
            "etsy_seller_handbook": "https://www.etsy.com/seller-handbook",
            "ebay_promoted_listings_standard": "https://www.ebay.com/sellercenter/ebay-for-business/marketing/promoted-listings-standard",
        },
        "outputs": {
            "plan_csv": str(PLAN_CSV),
            "state_json": str(STATE_JSON),
            "report_md": str(REPORT_MD),
        },
    }
    STATE_JSON.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")
    return state


def write_report(rows: list[dict[str, object]], state: dict[str, object]) -> None:
    by_track = defaultdict(list)
    for row in rows:
        by_track[row["Track"]].append(row)

    def table_sample(track: str, limit: int = 8) -> str:
        lines = ["| ID | Action | Intent | QA | Price |", "|---|---|---|---|---|"]
        for row in by_track.get(track, [])[:limit]:
            lines.append(
                f"| {row['ID']} | {row['Launch_Action']} | {row['Primary_Search_Intent']} | {row['QA_Status']} | {row['Price_Target']} |"
            )
        return "\n".join(lines)

    report = f"""# Multi-Track Experiment Plan

Generated: {state['timestamp']}

## Guardrails Now Active
- 165 experiment slots are split equally: 55 low-competition niche, 55 high-volume value, 55 digital pure-profit.
- Etsy fee kill switch is active: $2/batch, $6/day. Confirmed Etsy spend today: ${state['etsy_confirmed_spend_today_usd']:.2f}.
- Image QA is active. `SHADOW_CLIPPING`, `LOW_RESOLUTION`, and `HIGHLIGHT_CLIPPING` are hard HOLD states.
- Sticker Cover Gate remains active: cover-safe official mockups only for marketplace publishing.
- eBay ads remain General / Promoted Listings Standard fixed 2%, never Priority/PPC.

## Track A - Low-Competition Niche
Objective: {TRACKS['A_LOW_COMPETITION_NICHE']['objective']}

{table_sample('A_LOW_COMPETITION_NICHE')}

## Track B - High-Volume Value
Objective: {TRACKS['B_HIGH_VOLUME_VALUE']['objective']}

{table_sample('B_HIGH_VOLUME_VALUE')}

## Track C - Digital Pure Profit
Objective: {TRACKS['C_DIGITAL_PURE_PROFIT']['objective']}

{table_sample('C_DIGITAL_PURE_PROFIT')}

## QA Hold Pool
- HOLD rows are excluded from the 165 experiment capacity and written as `QA_HOLD_POOL`.
- HOLD count in this run: {state['qa_counts'].get('HOLD', 0)}.

## Next Operator Move
1. Do not spend additional Etsy listing fees until the next gray cell is selected from Track C and the fee ledger is reconciled.
2. Use Track A first if eBay remains a 0-view channel: the goal is non-zero search entry, not immediate conversion.
3. Use Track B only after the cover/product image is clean, because broad-volume terms punish weak thumbnails faster.
4. If Track C first paid cells remain 0 views, stop fee spend and rewrite SEO using the Buyer Persona vs Room Use result split.

## References
- Etsy fees: https://www.etsy.com/legal/fees/
- Etsy Seller Handbook / marketplace insight workflow: https://www.etsy.com/seller-handbook
- eBay Promoted Listings Standard: https://www.ebay.com/sellercenter/ebay-for-business/marketing/promoted-listings-standard
"""
    REVIEW.mkdir(exist_ok=True)
    REPORT_MD.write_text(report, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Plan Rex/OpenClaw multi-track marketplace experiments without spending or publishing.")
    parser.add_argument("--slots", type=int, default=165)
    parser.add_argument("--skip-image-audit", action="store_true")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if args.slots != 165:
        per_track = max(1, args.slots // 3)
        for config in TRACKS.values():
            config["slot_target"] = per_track
        # Give leftovers to Track C because digital has the cleanest unit economics.
        remainder = args.slots - per_track * 3
        TRACKS["C_DIGITAL_PURE_PROFIT"]["slot_target"] += remainder

    physical = load_physical_candidates()
    digital = load_digital_candidates()
    rows = select_candidates(
        physical=physical,
        digital=digital,
        quality_cache=_quality_cache(),
        cover_blocked=_cover_gate_ids(),
        audit_images=not args.skip_image_audit,
    )
    _write_csv(PLAN_CSV, rows, PLAN_FIELDS)
    state = write_state(rows)
    write_report(rows, state)
    result = {
        "rows": len(rows),
        "plan": str(PLAN_CSV),
        "state": str(STATE_JSON),
        "report": str(REPORT_MD),
        "track_counts": state["track_counts"],
        "qa_counts": state["qa_counts"],
        "action_counts": state["action_counts"],
    }
    if args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(f"[MULTI-TRACK] rows={result['rows']} plan={PLAN_CSV}")
        print(json.dumps({k: result[k] for k in ["track_counts", "qa_counts", "action_counts"]}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
