"""Build a deterministic repair plan for live eBay cover mismatches.

This module does not modify eBay or Printify. It records the learned rules from
the cover audit so the next repair attempt can be automated instead of repeated
manually in a browser.
"""

from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"
WORKBOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUT_CSV = DATABASE_DIR / "eBay_Cover_Repair_Decisions.csv"
OUT_MD = DATABASE_DIR / "eBay_Cover_Repair_Decisions.md"


REPORTS_BLOCKER = (
    "eBay Reports cannot revise variation pictures for these Printify-synced "
    "inventory-managed listings. The failed result returned: Items that are "
    "managed by Inventory do not allow Add/Delete variation pictures."
)


def _workbook_rows() -> dict[str, dict[str, str]]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    rows: dict[str, dict[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        item_id = values[cols["ID"]]
        if not item_id:
            continue
        rows[str(item_id)] = {name: values[idx] for name, idx in cols.items()}
    wb.close()
    return rows


def _read_fix_queue() -> list[dict[str, str]]:
    if not FIX_QUEUE.exists():
        return []
    with FIX_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_existing_decisions() -> dict[str, dict[str, str]]:
    if not OUT_CSV.exists():
        return {}
    with OUT_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return {str(row.get("ID") or ""): row for row in csv.DictReader(handle) if row.get("ID")}


def _exists(value: object) -> bool:
    return bool(value) and Path(str(value)).exists()


def classify(row: dict[str, str], workbook_row: dict[str, object]) -> tuple[str, str]:
    product_type = str(workbook_row.get("Product_Type") or row["ID"].split("-")[0])
    status = str(workbook_row.get("Status") or "")
    has_cover = _exists(workbook_row.get("Cover_Path"))
    has_product = bool(row.get("Printify_Product_ID"))
    has_ebay = bool(row.get("eBay_Item_ID"))

    if status == "Retired_Replaced":
        return (
            "RETIRED_REPLACED_DONE",
            "Old live listing was ended after a replacement passed live buyer-page audit.",
        )
    if not has_ebay:
        return "LOCAL_ONLY_NO_EBAY_ITEM", "No live eBay item id is available."
    if not has_cover:
        return "BLOCKED_MISSING_LOCAL_COVER", "Local Cover_Mockup.png path is missing."
    if not has_product:
        return "BLOCKED_MISSING_PRINTIFY_PRODUCT", "Printify product id is missing."

    if product_type == "Sticker":
        return (
            "SOURCE_REPAIR_REQUIRED",
            "Repair Printify source mockups first, then re-sync from Printify and re-audit eBay. "
            + REPORTS_BLOCKER,
        )
    return (
        "NON_STICKER_REVIEW_REQUIRED",
        "Non-sticker cover mismatch is high-confidence but may be a legitimate single-image product. "
        "Review before relisting or source repair.",
    )


def build_decisions() -> list[dict[str, str]]:
    workbook = _workbook_rows()
    existing = _read_existing_decisions()
    decisions: list[dict[str, str]] = []
    for row in _read_fix_queue():
        item_id = row["ID"]
        wb_row = workbook.get(item_id, {})
        method, note = classify(row, wb_row)
        prior = existing.get(item_id, {})
        preserved_status = prior.get("Status") or "PENDING"
        preserved_note = prior.get("Repair_Note") or note
        last_attempt = prior.get("Last_Repair_Attempt", "")
        decisions.append(
            {
                "ID": item_id,
                "Product_Type": str(wb_row.get("Product_Type") or item_id.split("-")[0]),
                "eBay_Item_ID": row.get("eBay_Item_ID", ""),
                "Printify_Product_ID": row.get("Printify_Product_ID", ""),
                "Online_Result": row.get("Result", ""),
                "Best_U_Label": row.get("Best_U_Label", ""),
                "Repair_Method": method,
                "Repair_Note": preserved_note,
                "Cover_Path": str(wb_row.get("Cover_Path") or ""),
                "Status": preserved_status,
                "Last_Repair_Attempt": last_attempt,
            }
        )
    return decisions


def write_outputs(decisions: list[dict[str, str]]) -> None:
    fields = [
        "ID",
        "Product_Type",
        "eBay_Item_ID",
        "Printify_Product_ID",
        "Online_Result",
        "Best_U_Label",
        "Repair_Method",
        "Repair_Note",
        "Cover_Path",
        "Status",
        "Last_Repair_Attempt",
    ]
    with OUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(decisions)

    by_method = Counter(row["Repair_Method"] for row in decisions)
    by_type = Counter(row["Product_Type"] for row in decisions)
    lines = [
        "# eBay Cover Repair Decisions",
        "",
        f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S} America/New_York",
        "",
        "## Learned Rule",
        "",
        f"- {REPORTS_BLOCKER}",
        "- The current safe path is source repair in Printify, followed by a Printify re-sync and live eBay cover audit.",
        "- If source re-sync still cannot change a live Inventory-managed listing, create a correct replacement listing and retire the bad listing after verification.",
        "",
        "## Counts",
        "",
    ]
    for key, count in sorted(by_method.items()):
        lines.append(f"- {key}: {count}")
    lines.extend(["", "## Product Types", ""])
    for key, count in sorted(by_type.items()):
        lines.append(f"- {key}: {count}")
    lines.extend(["", f"CSV: `{OUT_CSV}`", ""])
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    decisions = build_decisions()
    write_outputs(decisions)
    print(f"[COVER-REPAIR-DECISIONS] rows={len(decisions)} csv={OUT_CSV}")
    for key, count in sorted(Counter(row["Repair_Method"] for row in decisions).items()):
        print(f"  {key}: {count}")


if __name__ == "__main__":
    main()
