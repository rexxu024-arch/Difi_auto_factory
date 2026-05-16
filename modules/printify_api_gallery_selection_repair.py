"""Repair selected Printify gallery images through the product API.

This is the second phase after UI source repair. Printify's mockup-library UI
can add official product context mockups, but it sometimes fails to deselect old
custom U/detail images. The API exposes the selected image list clearly enough
to keep only the intended buyer-facing gallery.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
BACKUP_DIR = DATABASE / "Printify_Product_Backups"
LOG_CSV = DATABASE / "Printify_API_Gallery_Selection_Repair_Log.csv"


def clean(value: object) -> str:
    return str(value or "").strip()


def headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def fetch_product(product_id: str) -> dict:
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers={"Authorization": f"Bearer {Config.Printify_API_KEY}"},
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def put_product(product_id: str, payload: dict) -> dict:
    response = requests.put(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=headers(),
        json=payload,
        timeout=180,
    )
    response.raise_for_status()
    return response.json() if response.content else {}


def workbook_rows(ids: set[str] | None = None, limit: int = 0) -> list[dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    heads = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(heads)}

    def value(values: tuple, name: str) -> str:
        idx = cols.get(name)
        if idx is None or idx >= len(values):
            return ""
        return clean(values[idx])

    rows: list[dict[str, str]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not value(values, "ID"):
                continue
            item_id = value(values, "ID")
            if ids and item_id not in ids:
                continue
            product_id = value(values, "Printify_Product_ID")
            if not product_id:
                continue
            rows.append(
                {
                    "ID": item_id,
                    "Product_Type": value(values, "Product_Type"),
                    "Printify_Product_ID": product_id,
                    "Cover_Path": value(values, "Cover_Path"),
                    "Status": value(values, "Status"),
                    "eBay_Item_ID": value(values, "eBay_Item_ID"),
                }
            )
            if limit and len(rows) >= limit:
                break
    finally:
        wb.close()
    return rows


def image_kind(image: dict) -> str:
    src = clean(image.get("src"))
    if "images.printify.com/mockup" in src:
        return "official"
    if "pfy-prod-products-mockup-media" in src:
        return "custom"
    return "other"


def camera_label(image: dict) -> str:
    src = clean(image.get("src"))
    if "camera_label=" not in src:
        return ""
    return src.split("camera_label=", 1)[1].split("&", 1)[0]


def ahash(image: Image.Image) -> str:
    image = image.convert("L").resize((16, 16), Image.Resampling.LANCZOS)
    pixels = list(image.getdata())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if pixel > avg else "0" for pixel in pixels)


def hamming(left: str, right: str) -> int:
    return sum(a != b for a, b in zip(left, right))


def score_against_cover(image: dict, cover_path: str) -> int:
    if not cover_path or not Path(cover_path).exists():
        return 999
    src = clean(image.get("src"))
    if not src:
        return 999
    try:
        local = ahash(Image.open(cover_path))
        response = requests.get(src, timeout=60)
        response.raise_for_status()
        remote = ahash(Image.open(io.BytesIO(response.content)))
        return hamming(local, remote)
    except Exception:
        return 999


def normalize_images(images: list[dict], product_type: str, cover_path: str) -> tuple[list[dict], str]:
    ptype = clean(product_type)
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    official = [image for image in images if image_kind(image) == "official"]
    custom = [image for image in images if image_kind(image) == "custom"]

    if ptype == "Sticker":
        if not official:
            raise RuntimeError("No official Printify mockups available; run UI source repair first.")
        if not custom:
            raise RuntimeError("No custom cover candidate available.")
        cover_scores = [(score_against_cover(image, cover_path), image) for image in custom]
        cover_score, cover = min(cover_scores, key=lambda item: item[0])
        # A high perceptual hash distance usually means the old gallery only
        # contains U/detail images, not the local cover mockup. Refuse to
        # promote a random U image to default; upload the cover again first.
        if cover_score > 24:
            raise RuntimeError(
                f"Best custom image does not match local cover closely enough: cover_score={cover_score}. "
                "Run cover-only UI upload, then retry API gallery repair."
            )
        official_order = ["front", "context-1", "context-2"]
        official_by_label = {camera_label(image): image for image in official}
        picked = [cover]
        for label in official_order:
            image = official_by_label.get(label)
            if image:
                picked.append(image)
        for image in official:
            if len(picked) >= 4:
                break
            if image not in picked:
                picked.append(image)
        if len(picked) < 4:
            raise RuntimeError(f"Sticker repair needs 1 cover + 3 official mockups, found {len(picked)}.")
        note = f"Sticker cover_score={cover_score} official={len(picked)-1}"
        return set_flags(images, picked[:4]), note

    if ptype in {"Poster", "Acrylic"}:
        if not official:
            raise RuntimeError("No official Printify mockups available.")
        default_image = next((image for image in official if image.get("is_default")), None)
        preferred = {
            "Poster": ["front", "close-up", "context-1", "context-2"],
            "Acrylic": ["front", "back", "side-1", "side-2", "context-1"],
        }.get(ptype, [])
        picked = []
        for label in preferred:
            for image in official:
                if image not in picked and camera_label(image) == label:
                    picked.append(image)
                    break
        for image in official:
            if len(picked) >= 4:
                break
            if image not in picked:
                picked.append(image)
        if len(picked) < 3:
            raise RuntimeError(f"{ptype} repair needs at least 3 official mockups, found {len(picked)}.")
        default_image = default_image or picked[0]
        gallery = [image for image in picked if image_key(image) != image_key(default_image)]
        if len(gallery) < 2:
            raise RuntimeError(f"{ptype} repair needs at least 2 non-default gallery mockups, found {len(gallery)}.")
        return set_flags_with_external_default(images, default_image, gallery[:3]), (
            f"{ptype} default_outside_gallery=1 gallery_official={len(gallery[:3])}"
        )

    raise RuntimeError(f"Unsupported product type: {ptype}")


def set_flags(all_images: list[dict], selected_images: list[dict]) -> list[dict]:
    selected_key_counts = Counter(image_key(image) for image in selected_images)
    repaired = []
    selected_index = 0
    for image in all_images:
        item = dict(image)
        key = image_key(image)
        is_selected = selected_key_counts.get(key, 0) > 0
        item["is_selected_for_publishing"] = is_selected
        item["is_default"] = is_selected and selected_index == 0
        if is_selected:
            selected_key_counts[key] -= 1
            item["order"] = selected_index
            selected_index += 1
        else:
            item["order"] = None
            item["is_default"] = False
        repaired.append(item)
    return repaired


def set_flags_with_external_default(all_images: list[dict], default_image: dict, gallery_images: list[dict]) -> list[dict]:
    """Keep one official default product image out of the publishing gallery.

    Printify/eBay renders the default product image as the buyer-facing primary
    photo. If the same default image is also selected for publishing, eBay often
    shows it again as Picture 2, creating the duplicate thumbnail pattern Rex
    flagged. For Poster/Acrylic products we keep the official default for
    physical-product clarity, but select only the non-default official mockups
    as the gallery.
    """
    default_key = image_key(default_image)
    gallery_key_counts = Counter(image_key(image) for image in gallery_images)
    repaired = []
    selected_index = 0
    for image in all_images:
        item = dict(image)
        key = image_key(image)
        is_default = key == default_key
        is_selected = gallery_key_counts.get(key, 0) > 0 and not is_default
        item["is_default"] = is_default
        item["is_selected_for_publishing"] = is_selected
        if is_selected:
            gallery_key_counts[key] -= 1
            item["order"] = selected_index
            selected_index += 1
        else:
            item["order"] = None
        repaired.append(item)
    return repaired


def image_key(image: dict) -> str:
    return clean(image.get("mockup_id")) or clean(image.get("id")) or clean(image.get("src"))


def backup_product(item_id: str, product_id: str, product: dict) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = BACKUP_DIR / f"{item_id}_{product_id}_before_gallery_repair_{stamp}.json"
    path.write_text(json.dumps(product, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_log(record: dict[str, str]) -> None:
    fields = [
        "Timestamp",
        "ID",
        "Product_Type",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Dry_Run",
        "Before_Selected",
        "Before_Official",
        "Before_Custom",
        "After_Selected",
        "After_Official",
        "After_Custom",
        "Result",
        "Note",
        "Backup_Path",
        "Error",
    ]
    exists = LOG_CSV.exists()
    with LOG_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        if not exists:
            writer.writeheader()
        writer.writerow(record)


def mix(images: list[dict]) -> tuple[int, int, int]:
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    official = sum(1 for image in selected if image_kind(image) == "official")
    custom = sum(1 for image in selected if image_kind(image) == "custom")
    return len(selected), official, custom


def selected_default_count(images: list[dict]) -> int:
    return sum(
        1
        for image in images
        if image.get("is_default") and image.get("is_selected_for_publishing") is not False
    )


def _is_live_or_external(row: dict[str, str]) -> bool:
    status = clean(row.get("Status"))
    return bool(clean(row.get("eBay_Item_ID"))) or status.startswith(("Printify_Published", "Printify_PublishExternalPending"))


def repair_row(row: dict[str, str], dry_run: bool, allow_unpublished_poster_acrylic: bool = False) -> dict[str, str]:
    item_id = row["ID"]
    product_id = row["Printify_Product_ID"]
    record = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ID": item_id,
        "Product_Type": row["Product_Type"],
        "Printify_Product_ID": product_id,
        "eBay_Item_ID": row.get("eBay_Item_ID", ""),
        "Dry_Run": str(bool(dry_run)),
        "Result": "ERROR",
        "Error": "",
    }
    try:
        if row["Product_Type"] in {"Poster", "Acrylic"} and not dry_run:
            if allow_unpublished_poster_acrylic and not _is_live_or_external(row):
                pass
            else:
                raise RuntimeError(
                    "API gallery repair disabled for live Poster/Acrylic after Printify expanded selected official images; "
                    "only unpublished, no-external draft products may be repaired with --allow-unpublished-poster-acrylic."
                )
        product = fetch_product(product_id)
        before = mix(product.get("images") or [])
        record["Before_Selected"], record["Before_Official"], record["Before_Custom"] = map(str, before)
        repaired_images, note = normalize_images(product.get("images") or [], row["Product_Type"], row.get("Cover_Path", ""))
        record["Note"] = note
        record["Backup_Path"] = str(backup_product(item_id, product_id, product).relative_to(PROJECT_ROOT))
        if dry_run:
            after = mix(repaired_images)
            record["After_Selected"], record["After_Official"], record["After_Custom"] = map(str, after)
            record["Result"] = "DRY_RUN_OK"
            return record
        payload = dict(product)
        payload["images"] = repaired_images
        put_product(product_id, payload)
        updated = fetch_product(product_id)
        after = mix(updated.get("images") or [])
        record["After_Selected"], record["After_Official"], record["After_Custom"] = map(str, after)
        expected = (4, 3, 1) if row["Product_Type"] == "Sticker" else None
        if expected and after != expected:
            raise RuntimeError(f"Post-API mix mismatch: expected={expected} actual={after}")
        if row["Product_Type"] in {"Poster", "Acrylic"}:
            selected_defaults = selected_default_count(updated.get("images") or [])
            if selected_defaults:
                raise RuntimeError(
                    "Printify kept the default image selected after API repair; "
                    f"selected_default_count={selected_defaults}. Use UI source repair/rebuild before marketplace publish."
                )
            if after[1] < 3:
                raise RuntimeError(f"Post-API official mockups too low: {after}")
            if after[2] != 0 or after[0] > 4:
                raise RuntimeError(
                    "Post-API gallery selection was not reduced to official-only mockups: "
                    f"selected/official/custom={after}. Printify ignored image selection flags; use rebuild or UI source repair."
                )
        record["Result"] = "REPAIRED"
    except Exception as exc:
        record["Error"] = str(exc)[:500]
    return record


def run(ids: set[str] | None, limit: int, dry_run: bool, allow_unpublished_poster_acrylic: bool = False) -> int:
    rows = workbook_rows(ids=ids, limit=limit)
    done = 0
    for row in rows:
        record = repair_row(
            row,
            dry_run=dry_run,
            allow_unpublished_poster_acrylic=allow_unpublished_poster_acrylic,
        )
        write_log(record)
        print(
            f"[PRINTIFY-API-GALLERY-REPAIR] {record['ID']} result={record['Result']} "
            f"before={record.get('Before_Selected')}/{record.get('Before_Official')}/{record.get('Before_Custom')} "
            f"after={record.get('After_Selected')}/{record.get('After_Official')}/{record.get('After_Custom')} "
            f"note={record.get('Note','')} err={record.get('Error','')}"
        )
        if record["Result"] in {"REPAIRED", "DRY_RUN_OK"}:
            done += 1
        time.sleep(0.5)
    print(f"[PRINTIFY-API-GALLERY-REPAIR-DONE] rows={len(rows)} ok={done} dry_run={dry_run} log={LOG_CSV}")
    return done


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ids", default="")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--allow-unpublished-poster-acrylic",
        action="store_true",
        help="Allow API gallery selection repair only for Poster/Acrylic rows that have no external marketplace item yet.",
    )
    args = parser.parse_args()
    ids = {item.strip() for item in args.ids.split(",") if item.strip()} or None
    run(
        ids=ids,
        limit=args.limit,
        dry_run=args.dry_run,
        allow_unpublished_poster_acrylic=args.allow_unpublished_poster_acrylic,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
