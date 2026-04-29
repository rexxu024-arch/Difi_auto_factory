import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import time
from copy import copy
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from config import BASE_URL, CLAUDE_API_KEY, DEEPSEEK_API_KEY
from modules.streaming_json import JsonObjectStream, iter_anthropic_text

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")

DATABASE_DIR = Path("Database")
MENTOR_FILE = DATABASE_DIR / "Mentor_Hub.xlsx"
PRODUCTION_FILE = DATABASE_DIR / "Production_Line.xlsx"
GENERATED_LOG = DATABASE_DIR / "product_line_generated.log"
PENDING_DESIGN_FILE = DATABASE_DIR / "Pending_design.txt"

MENTOR_COLUMNS = ["Category", "Layout", "Title", "Gold_Prompt_DNA", "Material_Keywords", "Design_Count", "Performance"]
PRODUCTION_COLUMNS = [
    "ID",
    "Timestamp",
    "Category",
    "Product_Type",
    "Style",
    "Title",
    "MJ_Prompt",
    "SEO_Hook",
    "Status",
]

MJ_SUFFIX = "--v 6.1 --style raw --no skin, person, text, watermark"
CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-5")
SYSTEM_PROMPT = 'Strictly output raw JSON array. Start directly with "[". No preamble, no markdown, no explanations.'
LIFECYCLE_LIMIT = 100
BATCH_LIMIT = 20

PRODUCT_TEMPLATES = {
    "Sticker": {
        "layout": "Isolated",
        "ar": "--ar 1:1",
        "suffix": "white contour border, vector clean edges, die-cut sticker style, solid white background, isolated on white background",
    },
    "Poster": {
        "layout": "Full_Frame",
        "ar": "--ar 2:3",
        "suffix": "premium matte vertical poster composition, full frame, cinematic lighting, edge-to-edge composition, immersive environment, 12x18 wall art format",
    },
    "Acrylic": {
        "layout": "Full_Frame",
        "ar": "--ar 5:7",
        "suffix": "premium vertical acrylic photo block composition, 3D depth, refractive light, internal glow, ray tracing, gallery collectible art object",
    },
    "T-shirt": {
        "layout": "Isolated",
        "ar": "--ar 2:3",
        "suffix": "centered design, graphic tee style, vector art, isolated on solid background",
    },
    "Mug": {
        "layout": "Full_Frame",
        "ar": "--ar 2:1",
        "suffix": "continuous seamless pattern, panoramic wrap-around design",
    },
}

PRODUCT_ALIASES = {
    "sticker": "Sticker",
    "stickers": "Sticker",
    "poster": "Poster",
    "posters": "Poster",
    "acrylic": "Acrylic",
    "acrylics": "Acrylic",
    "t-shirt": "T-shirt",
    "tshirt": "T-shirt",
    "tshirts": "T-shirt",
    "t-shirts": "T-shirt",
    "shirt": "T-shirt",
    "shirts": "T-shirt",
    "mug": "Mug",
    "mugs": "Mug",
}


class ProductLineError(RuntimeError):
    pass


def root_path(relative_path):
    return ROOT_DIR / relative_path


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_prompt(raw):
    return str(raw or "").replace("\n", " ").replace("\r", " ").strip()


def excel_timestamp():
    return datetime.now()


def header_map(sheet):
    return {cell.value: index + 1 for index, cell in enumerate(sheet[1]) if cell.value}


def validate_schema():
    for path, columns in ((MENTOR_FILE, MENTOR_COLUMNS), (PRODUCTION_FILE, PRODUCTION_COLUMNS)):
        workbook = load_workbook(root_path(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            missing = [column for column in columns if column not in headers]
            if missing:
                raise ProductLineError(f"{path} missing columns: {', '.join(missing)}")
        finally:
            workbook.close()


def canonical_product_type(product_type):
    token = clean_text(product_type).lower().replace("_", "-")
    token = re.sub(r"\s+", "-", token)
    canonical = PRODUCT_ALIASES.get(token)
    if not canonical:
        raise ProductLineError(f"[AUDIT] Unknown Product_Type template: {product_type}")
    if canonical not in PRODUCT_TEMPLATES:
        raise ProductLineError(f"[AUDIT] Missing exact Style template: {canonical}")
    return canonical


def read_pending_design():
    path = root_path(PENDING_DESIGN_FILE)
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("[]", encoding="utf-8")
        return []
    raw = path.read_text(encoding="utf-8-sig").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ProductLineError(f"Pending_design.txt must be a JSON Array: {exc}") from exc
    if not isinstance(parsed, list):
        raise ProductLineError("Pending_design.txt must be a JSON Array")
    tasks = []
    for item in parsed:
        if not isinstance(item, dict):
            print(f"[AUDIT] Invalid pending design item skipped: {item}")
            continue
        category = clean_text(item.get("Category"))
        product_type = clean_text(item.get("Product_Type"))
        try:
            count = int(item.get("Count") or item.get("Number_of_Designs") or 0)
        except (TypeError, ValueError):
            count = 0
        if not category or not product_type or count <= 0:
            print(f"[AUDIT] Invalid pending design item skipped: {item}")
            continue
        tasks.append({"Category": category, "Product_Type": product_type, "Count": count})
    return tasks


def write_pending_design(tasks):
    path = root_path(PENDING_DESIGN_FILE)
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized = []
    for task in tasks:
        count = int(task.get("Count") or 0)
        if count > 0:
            normalized.append({
                "Category": clean_text(task.get("Category")),
                "Product_Type": canonical_product_type(task.get("Product_Type")),
                "Count": count,
            })
    temp_path = path.with_suffix(".tmp")
    temp_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def load_mentor_rows():
    workbook = load_workbook(root_path(MENTOR_FILE), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        rows = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            item = {column: row[columns[column] - 1] if column in columns else None for column in MENTOR_COLUMNS}
            if item.get("Category") and item.get("Gold_Prompt_DNA"):
                item["_row_index"] = row_index
                rows.append(item)
        return rows
    finally:
        workbook.close()


def category_selector(task_category, product_type):
    category = clean_text(task_category)
    product = canonical_product_type(product_type)
    product_token = id_product_token(product)
    if category.lower().startswith(product_token.lower() + "-"):
        return category[len(product_token) + 1 :]
    if category.lower().startswith("wall-art-"):
        return category[len("wall-art-") :]
    return category


def mentor_matches_task(seed_category, selector):
    seed_category = clean_text(seed_category)
    selector = clean_text(selector)
    if not selector:
        return False
    if seed_category.lower() == selector.lower():
        return True
    return seed_category.lower().startswith(selector.lower() + "-")


def design_count_value(seed):
    try:
        return int(seed.get("Design_Count") or 0)
    except (TypeError, ValueError):
        return 0


def select_mentor_seed(task):
    selector = category_selector(task["Category"], task["Product_Type"])
    candidates = [seed for seed in load_mentor_rows() if mentor_matches_task(seed["Category"], selector)]
    exact = [seed for seed in candidates if clean_text(seed["Category"]).lower() == selector.lower()]
    pool = exact or candidates
    candidates = sorted(pool, key=lambda seed: (design_count_value(seed), seed["_row_index"]))
    for seed in candidates:
        count = design_count_value(seed)
        if count >= LIFECYCLE_LIMIT:
            print(f"[AUDIT] DNA Exceeded Lifecycle ({count}/100). Skipping. Category={seed['Category']} Row={seed['_row_index']}")
            continue
        return seed
    print(f"[AUDIT] No available DNA for Category={task['Category']} Selector={selector}. Skipping.")
    return None


def load_generated_categories():
    path = root_path(GENERATED_LOG)
    if not path.exists():
        return set()
    return {line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()}


def mark_generated(category):
    path = root_path(GENERATED_LOG)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(category + "\n")


def product_type_for(seed_or_category, override=None):
    if override:
        return canonical_product_type(override)
    if isinstance(seed_or_category, dict):
        configured = clean_text(seed_or_category.get("Product_Type"))
        if configured:
            return configured
        text = clean_text(seed_or_category.get("Category")).lower()
    else:
        text = clean_text(seed_or_category).lower()
    if text.startswith("poster-") or "poster" in text:
        return "Poster"
    if "shirt" in text or "t-shirt" in text:
        return "T-shirt"
    if "wall" in text or "canvas" in text:
        return "Acrylic"
    if "acrylic" in text:
        return "Acrylic"
    if "relief" in text or "3d" in text:
        return "Acrylic"
    return "Sticker"


def strip_suffix(prompt):
    prompt = clean_prompt(prompt)
    prompt = re.sub(r"\s--v\s+\S+", "", prompt)
    prompt = re.sub(r"\s--ar\s+\S+", "", prompt)
    prompt = re.sub(r"\s--style\s+\S+", "", prompt)
    prompt = re.sub(r"\s--tile\b", "", prompt)
    prompt = re.sub(r"\s--no\s+.*$", "", prompt)
    return clean_prompt(prompt).rstrip(",")


def product_prompt_tail(product_type):
    return PRODUCT_TEMPLATES[canonical_product_type(product_type)]["suffix"]


def enforce_prompt(raw_prompt, product_type):
    product_type = canonical_product_type(product_type)
    template = PRODUCT_TEMPLATES[product_type]
    prompt = strip_suffix(raw_prompt)
    if product_type != "Sticker":
        prompt = re.sub(r"\b(die[- ]cut|sticker|vinyl decal|white contour border|contour border)\b", "", prompt, flags=re.I)
    tail = template["suffix"]
    if tail.lower() not in prompt.lower():
        prompt = f"{prompt}, {tail}"
    return clean_prompt(f"{prompt}, {template['ar']} {MJ_SUFFIX}")


def seo_hook(title, prompt, material_keywords):
    parts = []
    for value in re.split(r"[,;|/]+", clean_text(material_keywords)):
        if value.strip():
            parts.append(value.strip().lower())
    words = re.findall(r"[A-Za-z][A-Za-z-]{2,}", f"{title} {strip_suffix(prompt)}")
    stop = {"the", "and", "with", "from", "into", "white", "background", "isolated", "style", "raw"}
    for word in words:
        lowered = word.lower()
        if lowered not in stop and lowered not in parts:
            parts.append(lowered)
        if len(parts) >= 15:
            break
    while len(parts) < 10:
        for extra in ("mentor grade", "jade art", "kintsugi", "collectible", "premium decor"):
            if extra not in parts:
                parts.append(extra)
            if len(parts) >= 10:
                break
    return ", ".join(parts[:15])


def extract_json_array(text):
    text = clean_text(text)
    fenced = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.I | re.S)
    if fenced:
        return json.loads(fenced.group(1))
    start = text.find("[")
    end = text.rfind("]")
    if start >= 0 and end > start:
        return json.loads(text[start : end + 1])
    raise ProductLineError("API response did not contain a JSON array")


def build_prompt(seed, product_type, batch_count=BATCH_LIMIT):
    category = clean_text(seed["Category"])
    product_type = canonical_product_type(product_type)
    template = PRODUCT_TEMPLATES[product_type]
    return {
        "mission": "GREY ARCHITECT V15.3 Product_Line mold production",
        "required_count": int(batch_count),
        "sub_category": category,
        "product_type": product_type,
        "style": f"{category} Mentor-Grade",
        "layout": template["layout"],
        "product_ar": template["ar"],
        "product_suffix_keywords": template["suffix"],
        "gold_dna": {
            "title": seed["Title"],
            "prompt": seed["Gold_Prompt_DNA"],
            "materials": seed["Material_Keywords"],
        },
        "rules": [
            f"Generate exactly {int(batch_count)} visually unified but non-duplicate design variants.",
            "Each item must contain Title, MJ_Prompt, SEO_Hook.",
            "Preserve the material logic, mood, and premium visual language from Gold_Prompt_DNA, but DO NOT repeat the same core subject across the batch.",
            "Every batch must cover a subject diversity matrix: creature, botanical object, ritual instrument, architectural relic, talisman/seal, vessel/container, celestial object, weapon/tool, landscape micro-scene, abstract symbol.",
            "Use each primary subject slot at most once per batch. If one item is a phoenix, no other item may be phoenix-like; if one item is a seal, no other item may be seal/medallion/sigil-like.",
            "Do not let the source DNA become a repeated mascot. The source DNA is a material and atmosphere reference, not permission to repeat one object 20 times.",
            "Titles must make the unique focal object obvious in 2 to 5 words.",
            "Adjacent designs must not share the same primary noun or silhouette. Examples: do not output 20 Enso circles, 20 koi, 20 dragons, 20 torii gates, or 20 hourglasses.",
            "Each design must have a visibly different silhouette, focal object, pose/angle, accessory system, and composition, while still belonging to the same Sub_Category aesthetic.",
            "Avoid synonym-only variation. Colorway, lighting, or adjective changes alone are not enough to count as a new design.",
            "Do not preserve the original Midjourney aspect ratio; use product_ar exactly.",
            "MJ_Prompt must follow: [DNA Subject Description], [Product Suffix Keywords], [Product AR] --v 6.1 --style raw --no skin, person, text, watermark",
            "Do not include newline characters.",
            "SEO_Hook must contain 10 to 15 comma-separated keywords.",
            "Return only a JSON array.",
        ],
    }


def stream_claude_objects(seed, product_type, batch_count=BATCH_LIMIT, retries=3, max_seconds=150):
    if not CLAUDE_API_KEY:
        raise ProductLineError("CLAUDE_API_KEY is empty")
    url = BASE_URL.rstrip("/") + "/v1/messages"
    headers = {
        "x-api-key": CLAUDE_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": 8000,
        "temperature": 0.72,
        "stream": True,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": json.dumps(build_prompt(seed, product_type, batch_count), ensure_ascii=False)}],
    }
    last_error = None
    for attempt in range(1, retries + 1):
        started_at = time.monotonic()
        try:
            response = requests.post(url, headers=headers, json=payload, stream=True, timeout=(10, 15))
            if response.status_code >= 400:
                raise ProductLineError(f"{response.status_code} {response.text[:800]}")
            parser = JsonObjectStream()
            for text in iter_anthropic_text(response):
                if time.monotonic() - started_at > max_seconds:
                    response.close()
                    raise TimeoutError(f"Batch stream exceeded {max_seconds}s")
                for item in parser.feed(text):
                    yield item
            return
        except Exception as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(2 * attempt)
    raise ProductLineError(f"Product Line API stream failed after retries: {last_error}")


def fallback_variants(seed):
    category = clean_text(seed["Category"])
    base_prompt = strip_suffix(seed["Gold_Prompt_DNA"])
    themes = [
        "lunar rim lighting",
        "golden dawn glow",
        "obsidian shadow contrast",
        "jade bioluminescent core",
        "kintsugi fracture map",
        "floating relic orbit",
        "vertical talisman layout",
        "sacred geometry frame",
        "crystalline star chart",
        "ink-wash mist trail",
        "museum artifact angle",
        "macro material close-up",
        "halo backlight composition",
        "silent ceremonial pose",
        "celestial gate silhouette",
        "alchemical glow chamber",
        "ancient scholar archive",
        "premium collectible profile",
        "mythic compass symmetry",
        "soft sacred underlighting",
    ]
    return [
        {
            "Title": f"{seed['Title']} {index:02d}",
            "MJ_Prompt": f"{base_prompt}, {theme}, variant {index:02d}",
            "SEO_Hook": seo_hook(seed["Title"], base_prompt, seed["Material_Keywords"]),
        }
        for index, theme in enumerate(themes, 1)
    ]


def main_category(category):
    return clean_text(category).split("-", 1)[0].split("_", 1)[0] or clean_text(category)


def id_prefix(product_type, category):
    return f"{id_product_token(product_type)}-{main_category(category)}"


def next_sequence(product_type, category):
    prefix = id_prefix(product_type, category)
    workbook = load_workbook(root_path(PRODUCTION_FILE), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        id_column = columns["ID"]
        last_number = 0
        width = 4
        for row in sheet.iter_rows(min_row=2, values_only=True):
            value = row[id_column - 1]
            text = str(value or "")
            if not text.startswith(prefix + "-"):
                continue
            match = re.search(r"(\d+)$", text)
            if match:
                last_number = int(match.group(1))
                width = max(width, len(match.group(1)))
        return last_number + 1, width
    finally:
        workbook.close()


def normalize_variants(seed, raw_variants, product_type):
    seen = set()
    variants = []
    for index, item in enumerate(raw_variants, 1):
        if len(variants) >= 20:
            break
        title = clean_text(item.get("Title") or f"{seed['Title']} Variant {index:02d}")
        prompt = enforce_prompt(item.get("MJ_Prompt") or seed["Gold_Prompt_DNA"], product_type)
        hook = clean_text(item.get("SEO_Hook")) or seo_hook(title, prompt, seed["Material_Keywords"])
        hook_parts = [part.strip() for part in hook.split(",") if part.strip()]
        if len(hook_parts) < 10 or len(hook_parts) > 15:
            hook = seo_hook(title, prompt, seed["Material_Keywords"])
        fp = " ".join(sorted(set(re.findall(r"[A-Za-z][A-Za-z-]{3,}", f"{title} {prompt} {hook}".lower()))))
        if fp in seen:
            prompt = enforce_prompt(f"{strip_suffix(prompt)}, unique aesthetic branch {index:02d}", product_type)
            fp = f"{fp}-{index:02d}"
        seen.add(fp)
        variants.append({"Title": title, "MJ_Prompt": prompt, "SEO_Hook": hook})
    if len(variants) != 20:
        raise ProductLineError(f"Expected 20 variants, got {len(variants)}")
    return variants


SIMILARITY_STOPWORDS = {
    "with", "from", "into", "that", "this", "style", "sticker", "design", "white", "background",
    "isolated", "vector", "clean", "edges", "border", "solid", "sharp", "focus", "raw", "skin",
    "person", "text", "watermark", "mentor-grade", "hyper-detailed", "premium", "composition",
    "cinematic", "lighting", "material", "system", "primary", "surface", "relief", "visible",
    "handcrafted", "subtle", "internal", "glow", "crisp", "silhouette", "readability",
    "celestial", "astral", "lunar", "cosmic", "starbound", "starborne", "moonlit", "moonstone",
    "jade", "obsidian", "sapphire", "rainbow", "white", "black", "golden", "silver", "indigo",
    "violet", "emerald", "nebula", "ink-wash", "fragments", "floating", "orbiting", "crafted",
    "carved", "formed", "constructed", "sculpted", "ancient", "sacred", "mythical", "divine",
}


SUBJECT_GROUPS = {
    "phoenix": {"phoenix", "bird", "crane", "eagle", "feather", "wings", "winged"},
    "dragon": {"dragon", "serpent", "wyrm"},
    "koi": {"koi", "fish", "carp"},
    "beast": {"lion", "tiger", "fox", "wolf", "kirin", "qilin", "guardian", "beast"},
    "lotus": {"lotus", "flower", "bloom", "petal", "blossom"},
    "tree": {"tree", "bonsai", "bamboo", "branch", "pine", "willow"},
    "instrument": {"guqin", "bell", "chime", "flute", "drum", "singing", "bowl", "instrument"},
    "vessel": {"vessel", "cauldron", "urn", "chalice", "bowl", "jar", "teapot", "incense", "burner"},
    "gate": {"gate", "torii", "portal", "doorway", "archway", "shrine"},
    "pagoda": {"pagoda", "temple", "tower", "lantern", "pavilion", "bridge"},
    "seal": {"seal", "sigil", "medallion", "emblem", "crest", "talisman", "amulet"},
    "globe": {"globe", "orb", "sphere", "planet", "astrolabe", "compass"},
    "scroll": {"scroll", "manuscript", "tablet", "book", "sutra", "script"},
    "weapon": {"sword", "blade", "dagger", "spear", "staff", "wand", "vajra"},
    "landscape": {"mountain", "waterfall", "river", "island", "garden", "landscape", "pond"},
    "abstract": {"enso", "mandala", "geometry", "knot", "spiral", "circle", "constellation"},
}


def diversity_tokens(title, prompt):
    text = strip_suffix(f"{title} {prompt}").lower()
    words = re.findall(r"[a-z][a-z-]{3,}", text)
    return {
        word
        for word in words
        if word not in SIMILARITY_STOPWORDS and not word.startswith("variant")
    }


def subject_key(title, prompt):
    text = strip_suffix(f"{title} {prompt}").lower()
    words = set(re.findall(r"[a-z][a-z-]{2,}", text))
    for key, aliases in SUBJECT_GROUPS.items():
        if words & aliases:
            return key
    title_words = [
        word
        for word in re.findall(r"[a-z][a-z-]{3,}", str(title or "").lower())
        if word not in SIMILARITY_STOPWORDS
    ]
    return title_words[-1] if title_words else ""


def similarity_score(tokens_a, tokens_b):
    if not tokens_a or not tokens_b:
        return 0.0
    return len(tokens_a & tokens_b) / max(1, len(tokens_a | tokens_b))


def too_similar_to_saved(title, prompt, saved_variants, threshold=0.50):
    current = diversity_tokens(title, prompt)
    current_subject = subject_key(title, prompt)
    for saved in saved_variants:
        saved_subject = saved.get("_subject_key", "")
        if current_subject and saved_subject and current_subject == saved_subject:
            return True, 1.0, saved["Title"]
        score = similarity_score(current, saved["_diversity_tokens"])
        if score >= threshold:
            return True, score, saved["Title"]
    return False, 0.0, ""


def normalize_variant(seed, item, product_type, index=1):
    title = clean_text(item.get("Title") or f"{seed['Title']} Variant {index:02d}")
    prompt = enforce_prompt(item.get("MJ_Prompt") or seed["Gold_Prompt_DNA"], product_type)
    hook = clean_text(item.get("SEO_Hook")) or seo_hook(title, prompt, seed["Material_Keywords"])
    hook_parts = [part.strip() for part in hook.split(",") if part.strip()]
    if len(hook_parts) < 10 or len(hook_parts) > 15:
        hook = seo_hook(title, prompt, seed["Material_Keywords"])
    return {"Title": title, "MJ_Prompt": prompt, "SEO_Hook": hook}


def clone_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    target_cell.number_format = "@"
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)


def last_filled_row(sheet, id_column):
    for row_index in range(sheet.max_row, 1, -1):
        if clean_text(sheet.cell(row=row_index, column=id_column).value):
            return row_index
    return 1


def id_product_token(product_type):
    return clean_text(product_type).replace(" ", "-")


def build_output_rows(seed, variants, product_type):
    dna_category = clean_text(seed["Category"])
    category = main_category(dna_category)
    sequence, width = next_sequence(product_type, dna_category)
    timestamp = excel_timestamp()
    prefix = id_prefix(product_type, dna_category)
    rows = []
    for offset, variant in enumerate(variants):
        rows.append({
            "ID": f"{prefix}-{sequence + offset:0{width}d}",
            "Timestamp": timestamp,
            "Category": category,
            "Product_Type": product_type,
            "Style": f"{category} Mentor-Grade",
            "Title": clean_text(variant["Title"]),
            "MJ_Prompt": clean_prompt(variant["MJ_Prompt"]),
            "SEO_Hook": clean_text(variant["SEO_Hook"]),
            "Status": "Ready_for_production",
        })
    return rows


def build_output_row(seed, variant, product_type):
    return build_output_rows(seed, [variant], product_type)[0]


def save_with_openpyxl(output_rows):
    path = root_path(PRODUCTION_FILE)
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        missing = [column for column in PRODUCTION_COLUMNS if column not in columns]
        if missing:
            raise ProductLineError(f"Production_Line.xlsx missing columns: {', '.join(missing)}")
        last_row = last_filled_row(sheet, columns["ID"])
        template_row = max(last_row, 2)
        for row_data in output_rows:
            row_index = last_row + 1
            for column in PRODUCTION_COLUMNS:
                cell = sheet.cell(row=row_index, column=columns[column])
                clone_style(sheet.cell(row=template_row, column=columns[column]), cell)
                cell.value = row_data[column]
                if column == "Timestamp":
                    cell.number_format = "m/d/yyyy, h:mm:ss AM/PM"
                else:
                    cell.number_format = "@"
            last_row = row_index
        workbook.save(path)
    finally:
        workbook.close()


def save_with_excel_com(output_rows):
    import win32com.client

    excel_app = None
    workbook = None
    try:
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.DisplayAlerts = False
        workbook = excel_app.Workbooks.Open(str(root_path(PRODUCTION_FILE)))
        sheet = workbook.Worksheets(1)
        columns = {}
        column = 1
        while True:
            value = sheet.Cells(1, column).Value
            if value in (None, ""):
                break
            columns[str(value)] = column
            column += 1
        missing = [name for name in PRODUCTION_COLUMNS if name not in columns]
        if missing:
            raise ProductLineError(f"Production_Line.xlsx missing columns: {', '.join(missing)}")
        xl_up = -4162
        last_row = sheet.Cells(sheet.Rows.Count, columns["ID"]).End(xl_up).Row
        template_row = last_row
        for row_data in output_rows:
            target_row = last_row + 1
            sheet.Range(sheet.Cells(template_row, 1), sheet.Cells(template_row, len(columns))).Copy()
            sheet.Range(sheet.Cells(target_row, 1), sheet.Cells(target_row, len(columns))).PasteSpecial(-4122)
            for name in PRODUCTION_COLUMNS:
                cell = sheet.Cells(target_row, columns[name])
                if name == "Timestamp":
                    cell.NumberFormat = "m/d/yyyy, h:mm:ss AM/PM"
                else:
                    cell.NumberFormat = "@"
                cell.Value = row_data[name]
            last_row = target_row
        excel_app.CutCopyMode = False
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=True)
        if excel_app is not None:
            excel_app.Quit()


def append_rows(output_rows):
    try:
        save_with_openpyxl(output_rows)
    except PermissionError:
        save_with_excel_com(output_rows)


def append_row(row):
    append_rows([row])


def increment_design_count(seed, amount=20):
    workbook = load_workbook(root_path(MENTOR_FILE))
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        if "Design_Count" not in columns:
            sheet.cell(row=1, column=sheet.max_column + 1).value = "Design_Count"
            columns = header_map(sheet)
        row_index = int(seed["_row_index"])
        current = sheet.cell(row=row_index, column=columns["Design_Count"]).value or 0
        try:
            current = int(current)
        except (TypeError, ValueError):
            current = 0
        sheet.cell(row=row_index, column=columns["Design_Count"]).value = current + amount
        if "Timestamp" in columns:
            stamp_cell = sheet.cell(row=row_index, column=columns["Timestamp"])
            stamp_cell.value = datetime.now()
            stamp_cell.number_format = "m/d/yyyy, h:mm:ss AM/PM"
        workbook.save(root_path(MENTOR_FILE))
    finally:
        workbook.close()


def increment_design_count_by_row(seed, amount=1):
    increment_design_count(seed, amount)


def process_seed_v135_legacy(seed, product_type="Sticker"):
    saved = 0
    seen = set()
    output_rows = []
    for item in stream_claude_objects(seed, product_type):
        variant = normalize_variant(seed, item, product_type, saved + 1)
        fp = f"{variant['Title']}|{variant['MJ_Prompt']}|{variant['SEO_Hook']}"
        if fp in seen:
            continue
        seen.add(fp)
        row = build_output_row(seed, variant, product_type)
        row["MJ_Prompt"] = clean_prompt(row["MJ_Prompt"])
        append_row(row)
        increment_design_count_by_row(seed, 1)
        saved += 1
        output_rows.append(row)
        print(f"[Dify-Mode] ID: {row['ID']} 写入成功 | 实时保存已完成")
        if saved >= 20:
            break
    if saved != 20:
        raise ProductLineError(f"Expected 20 variants, got {saved}")
    print(f"[PRODUCT_LINE] Generated 20 rows: {seed['Category']}")
    return output_rows


def process_seed(seed, product_type="Sticker", batch_count=BATCH_LIMIT, max_seconds=150, on_saved=None):
    product_type = canonical_product_type(product_type)
    batch_count = min(int(batch_count), BATCH_LIMIT)
    saved = 0
    seen = set()
    output_rows = []
    started_at = time.monotonic()
    for item in stream_claude_objects(seed, product_type, batch_count=batch_count, max_seconds=max_seconds):
        if time.monotonic() - started_at > max_seconds:
            print(f"[AUDIT] Batch exceeded {max_seconds}s after saved={saved}. Breaking for debug-safe resume.")
            break
        variant = normalize_variant(seed, item, product_type, saved + 1)
        too_close, score, near_title = too_similar_to_saved(variant["Title"], variant["MJ_Prompt"], output_rows)
        if too_close:
            print(f"[DIVERSITY] Rejected near-duplicate score={score:.2f}: {variant['Title']} ~ {near_title}")
            continue
        fp = f"{variant['Title']}|{variant['MJ_Prompt']}|{variant['SEO_Hook']}"
        if fp in seen:
            continue
        seen.add(fp)
        row = build_output_row(seed, variant, product_type)
        row["MJ_Prompt"] = clean_prompt(row["MJ_Prompt"])
        row["_diversity_tokens"] = diversity_tokens(row["Title"], row["MJ_Prompt"])
        row["_subject_key"] = subject_key(row["Title"], row["MJ_Prompt"])
        append_row(row)
        increment_design_count_by_row(seed, 1)
        saved += 1
        output_rows.append(row)
        if on_saved:
            on_saved(row, saved)
        print(f"[Dify-Mode] ID: {row['ID']} 写入成功 | 实时保存已完成")
        if saved >= batch_count:
            break
    if saved == 0:
        raise ProductLineError(f"Expected {batch_count} variants, got 0")
    if saved != batch_count:
        print(f"[AUDIT] Partial batch saved={saved}/{batch_count}. Pending_design will keep the remaining demand.")
    print(f"[PRODUCT_LINE] Generated {batch_count} rows: {seed['Category']} -> {product_type}")
    return output_rows


def run_legacy_seed_mode(limit=1, product_type="Sticker"):
    os.chdir(ROOT_DIR)
    validate_schema()
    seeds = []
    for seed in load_mentor_rows():
        try:
            design_count = int(seed.get("Design_Count") or 0)
        except (TypeError, ValueError):
            design_count = 0
        if design_count < 100:
            seeds.append(seed)
    if limit is not None:
        seeds = seeds[:limit]
    if not seeds:
        print("[PRODUCT_LINE] No new Gold DNA seeds.")
        return 0
    processed = 0
    for seed in seeds:
        process_seed(seed, product_type=product_type)
        processed += 1
    return processed


def run_logic(limit=None, product_type=None, max_batches=None, max_seconds=150):
    os.chdir(ROOT_DIR)
    validate_schema()
    tasks = read_pending_design()
    if not tasks:
        print("[PRODUCT_LINE] Pending_design.txt is empty.")
        return 0

    total_saved = 0
    batches_done = 0
    task_index = 0
    while task_index < len(tasks):
        task = tasks[task_index]
        try:
            task["Product_Type"] = canonical_product_type(task["Product_Type"])
        except ProductLineError as exc:
            print(str(exc))
            tasks.pop(task_index)
            write_pending_design(tasks)
            continue

        seed = select_mentor_seed(task)
        if seed is None:
            tasks.pop(task_index)
            write_pending_design(tasks)
            continue

        current_count = design_count_value(seed)
        remaining_life = max(0, LIFECYCLE_LIMIT - current_count)
        if remaining_life <= 0:
            print(f"[AUDIT] DNA Exceeded Lifecycle ({current_count}/100). Skipping.")
            tasks.pop(task_index)
            write_pending_design(tasks)
            continue

        requested = int(task["Count"])
        allowed_for_task = min(requested, remaining_life)
        if allowed_for_task < requested:
            print(f"[AUDIT] Demand clipped: requested={requested}, available_lifecycle={remaining_life}, Category={seed['Category']}")
            task["Count"] = allowed_for_task
            requested = allowed_for_task
            write_pending_design(tasks)

        batch_count = min(requested, BATCH_LIMIT)
        print(
            f"[BATCH] Category={task['Category']} DNA={seed['Category']} Product={task['Product_Type']} "
            f"Batch={batch_count} RemainingTask={requested} Design_Count={current_count}/100"
        )
        task["Count"] = int(task["Count"])

        def checkpoint_saved(_row, _saved_in_batch):
            task["Count"] = max(0, int(task["Count"]) - 1)
            if task["Count"] <= 0:
                tasks.pop(task_index)
            write_pending_design(tasks)

        rows = process_seed(
            seed,
            product_type=task["Product_Type"],
            batch_count=batch_count,
            max_seconds=max_seconds,
            on_saved=checkpoint_saved,
        )
        saved = len(rows)
        total_saved += saved
        batches_done += 1
        if limit is not None and total_saved >= int(limit):
            break
        if max_batches is not None and batches_done >= int(max_batches):
            break

    print(f"[PRODUCT_LINE] V15.3 completed. Batches={batches_done} Saved={total_saved}")
    return total_saved


def self_test():
    os.chdir(ROOT_DIR)
    validate_schema()
    original_root = ROOT_DIR
    original_cwd = Path.cwd()
    with tempfile.TemporaryDirectory(prefix="product_line_v135_") as temp_dir:
        temp_root = Path(temp_dir)
        (temp_root / DATABASE_DIR).mkdir()
        shutil.copy2(root_path(MENTOR_FILE), temp_root / MENTOR_FILE)
        shutil.copy2(root_path(PRODUCTION_FILE), temp_root / PRODUCTION_FILE)
        globals()["ROOT_DIR"] = temp_root
        try:
            seed = load_mentor_rows()[-1]
            next_id, _ = next_sequence("Sticker", seed["Category"])
            prompt = enforce_prompt(f"{seed['Gold_Prompt_DNA']}\nself test branch\r", product_type_for(seed, "Sticker"))
            output = build_output_rows(seed, [{
                "Title": "Self Test DNA",
                "MJ_Prompt": prompt,
                "SEO_Hook": seo_hook("Self Test DNA", prompt, seed["Material_Keywords"]),
            }], "Sticker")
            if int(output[0]["ID"].split("-")[-1]) != next_id:
                raise ProductLineError("ID step validation failed")
            if "\n" in output[0]["MJ_Prompt"] or "\r" in output[0]["MJ_Prompt"]:
                raise ProductLineError("Prompt text cleaning failed")
            if output[0]["Status"] != "Ready_for_production":
                raise ProductLineError("Initial status validation failed")
        finally:
            globals()["ROOT_DIR"] = original_root
            os.chdir(original_cwd)
    print("PRODUCT_LINE_SELF_TEST_OK")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--max-batches", type=int, default=None)
    parser.add_argument("--max-seconds", type=int, default=150)
    parser.add_argument("--product-type", default="Sticker")
    args = parser.parse_args()
    if args.self_test:
        self_test()
    else:
        run_logic(limit=args.limit, product_type=args.product_type, max_batches=args.max_batches, max_seconds=args.max_seconds)


if __name__ == "__main__":
    main()
