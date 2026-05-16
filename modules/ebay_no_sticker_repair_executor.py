from __future__ import annotations

import argparse
import csv
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import Config
from modules.ebay_quiet_jade_pivot import clean, fit_title, request_with_retry, subject_from_title

DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
LEGACY_EXPERIMENT_CSV = DATABASE / "eBay_NoSticker_Repair_Experiment_20260511.csv"
CURRENT_REPAIR_PLAN_CSV = DATABASE / "eBay_API_Repair_Plan.csv"
OUT_PLAN = DATABASE / "eBay_NoSticker_Repair_Apply_Plan.csv"
SYNC_LOG = DATABASE / "eBay_NoSticker_Repair_Sync_Log.csv"
NY = ZoneInfo("America/New_York")

PUBLISH_TITLE_DESC_PRICE = {
    "title": True,
    "description": True,
    "images": False,
    "variants": True,
    "tags": False,
    "keyFeatures": False,
    "shipping_template": False,
}


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def money(value: Any) -> float:
    try:
        return float(re.sub(r"[^0-9.]", "", clean(value)) or "0")
    except ValueError:
        return 0.0


def money_text(value: float) -> str:
    return f"${value:.2f}"


def load_experiment(group: str, limit: int) -> list[dict[str, str]]:
    source = CURRENT_REPAIR_PLAN_CSV if CURRENT_REPAIR_PLAN_CSV.exists() else LEGACY_EXPERIMENT_CSV
    if not source.exists():
        raise FileNotFoundError(f"Missing experiment CSV: {source}")
    rows = list(csv.DictReader(source.open("r", encoding="utf-8-sig", newline="")))
    rows = [
        row for row in rows
        if clean(row.get("Experiment_Group")) == group
        and clean(row.get("Product_Type")) != "Sticker"
        and clean(row.get("Priority")) != "HOLD_GALLERY_FIRST"
    ]
    return rows[:limit] if limit else rows


def row_map(ws, cols: dict[str, int]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx in range(2, ws.max_row + 1):
        item_id = clean(ws.cell(idx, cols["eBay_Item_ID"]).value) if "eBay_Item_ID" in cols else ""
        if item_id:
            status = clean(ws.cell(idx, cols["Status"]).value) if "Status" in cols else ""
            # Duplicate item IDs can exist on experimental clone rows. Prefer the live published row.
            if item_id not in out or status.startswith("Printify_Published"):
                out[item_id] = idx
    return out


def description(product_type: str, title: str, source: dict[str, str]) -> str:
    if product_type == "Acrylic":
        includes = "One 5x7 vertical acrylic photo block desk display."
        placement = "bookshelves, office credenzas, meditation corners, and quiet luxury apartment desks"
        material = "Acrylic block produced on demand through Printify, selected for depth, gloss, and refractive color."
    else:
        includes = "One 12x18 premium matte vertical poster."
        placement = "reading nooks, deep-work offices, dorm rooms, gallery walls, and calm apartment corners"
        material = "Premium matte poster produced on demand through Printify."
    action_note = clean(source.get("Repair_Actions")).replace("|", "; ")
    return (
        f"<h2>{title}</h2>"
        f"<p>A Quiet Jade / OpenClaw Design Studio piece built for {placement}. "
        "The direction favors smoky jade color, cinematic shadow, and collectible room atmosphere over generic wall decor.</p>"
        "<ul>"
        f"<li><strong>Includes:</strong> {includes}</li>"
        f"<li><strong>Material:</strong> {material}</li>"
        "<li><strong>Visual note:</strong> the main image shows the product customers receive. "
        "Additional gallery images are supporting concept/detail views for mood, scale, and surface detail.</li>"
        "</ul>"
        f"<p><strong>Repair experiment:</strong> {action_note}</p>"
    )


def build_title(local_data: dict[str, Any], source: dict[str, str]) -> str:
    product_type = clean(local_data.get("Product_Type")) or clean(source.get("Product_Type"))
    subject = subject_from_title(clean(local_data.get("Title") or source.get("Title")), product_type, clean(local_data.get("Category")))
    if product_type == "Acrylic":
        return fit_title(["Quiet Luxury Desk Object", subject, "5x7 Acrylic Block Shelf Decor"])
    return fit_title(["Deep Work Wall Art", subject, "12x18 Matte Poster Apartment Decor"])


def generate_plan(group: str, limit: int, apply_local: bool, include_free_ship_price: bool = False) -> list[dict[str, str]]:
    selected = load_experiment(group, 0)
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {clean(header): idx + 1 for idx, header in enumerate(headers) if header}
    required = ["ID", "Title", "Description", "Price", "Product_Type", "Category", "Printify_Product_ID", "eBay_Item_ID"]
    missing = [name for name in required if name not in cols]
    if missing:
        wb.close()
        raise RuntimeError(f"Missing eBay workbook columns: {missing}")
    idx_by_item = row_map(ws, cols)
    timestamp = now_text()
    out: list[dict[str, str]] = []
    valid = 0
    for source in selected:
        if limit and valid >= limit:
            break
        item_id = clean(source.get("Item_ID"))
        row_idx = idx_by_item.get(item_id)
        if not row_idx:
            out.append({"Item_ID": item_id, "Result": "SKIP_LOCAL_ROW_NOT_FOUND"})
            continue
        local = {name: ws.cell(row_idx, cols[name]).value for name in required}
        product_type = clean(local.get("Product_Type"))
        new_title = build_title(local, source)
        old_price = money(local.get("Price"))
        free_ship_price = money(source.get("Suggested_FreeShip_Price"))
        # Only use the free-shipping price if the write path also changes shipping.
        # Otherwise this would raise buyer total and hurt conversion.
        new_price = money_text(free_ship_price if include_free_ship_price and free_ship_price else old_price)
        new_desc = description(product_type, new_title, source)
        result = "PLAN_ONLY"
        if apply_local:
            ws.cell(row_idx, cols["Title"]).value = new_title
            ws.cell(row_idx, cols["Description"]).value = new_desc
            ws.cell(row_idx, cols["Price"]).value = new_price
            result = "LOCAL_APPLIED"
        out.append(
            {
                "Timestamp": timestamp,
                "Experiment_Group": clean(source.get("Experiment_Group")),
                "ID": clean(local.get("ID")),
                "Product_Type": product_type,
                "Printify_Product_ID": clean(local.get("Printify_Product_ID")),
                "eBay_Item_ID": item_id,
                "Old_Price": clean(local.get("Price")),
                "New_Price": new_price,
                "Old_Title": clean(local.get("Title")),
                "New_Title": new_title,
                "New_Title_Length": str(len(new_title)),
                "Repair_Actions": clean(source.get("Repair_Actions")),
                "Result": result,
            }
        )
        valid += 1
    if apply_local:
        wb.save(EBAY_BOOK)
    wb.close()
    fields = [
        "Timestamp",
        "Experiment_Group",
        "ID",
        "Product_Type",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Old_Price",
        "New_Price",
        "Old_Title",
        "New_Title",
        "New_Title_Length",
        "Repair_Actions",
        "Result",
    ]
    with OUT_PLAN.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in out:
            writer.writerow({name: row.get(name, "") for name in fields})
    print(f"[EBAY-NOSTICKER-REPAIR] planned={len(out)} apply_local={apply_local} csv={OUT_PLAN}")
    return out


def local_metadata(ids: set[str]) -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {clean(header): idx for idx, header in enumerate(headers) if header}
    out = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        local_id = clean(values[cols["ID"]]) if "ID" in cols else ""
        if local_id not in ids:
            continue
        out[local_id] = {
            "Title": clean(values[cols["Title"]]),
            "Description": clean(values[cols["Description"]]),
            "Price": clean(values[cols["Price"]]),
            "Printify_Product_ID": clean(values[cols["Printify_Product_ID"]]),
            "Product_Type": clean(values[cols["Product_Type"]]),
            "eBay_Item_ID": clean(values[cols["eBay_Item_ID"]]) if "eBay_Item_ID" in cols else "",
        }
    wb.close()
    return out


def sync_printify(limit: int, dry_run: bool, sleep_min: float, sleep_max: float) -> int:
    if not OUT_PLAN.exists():
        raise FileNotFoundError(f"Missing apply plan: {OUT_PLAN}")
    rows = list(csv.DictReader(OUT_PLAN.open("r", encoding="utf-8-sig", newline="")))
    already_synced: set[str] = set()
    if SYNC_LOG.exists() and not dry_run:
        with SYNC_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if clean(row.get("Result")) == "OK":
                    already_synced.add(clean(row.get("ID")))
    rows = [
        row for row in rows
        if clean(row.get("Result")) == "LOCAL_APPLIED"
        and (dry_run or clean(row.get("ID")) not in already_synced)
    ][:limit]
    ids = {clean(row.get("ID")) for row in rows}
    metadata = local_metadata(ids)
    base = Config.Printify_API_URL.rstrip("/")
    done = 0
    exists = SYNC_LOG.exists()
    with SYNC_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        fields = ["Timestamp", "ID", "Printify_Product_ID", "HTTP_Get", "HTTP_Update", "HTTP_Publish", "Result", "Error"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        if not exists:
            writer.writeheader()
        for row in rows:
            local_id = clean(row.get("ID"))
            meta = metadata.get(local_id) or {}
            product_id = meta.get("Printify_Product_ID") or clean(row.get("Printify_Product_ID"))
            get_status = update_status = publish_status = ""
            result = "DRY_RUN" if dry_run else "CHECK"
            error = ""
            try:
                if dry_run:
                    print(f"[EBAY-NOSTICKER-DRY] {local_id} product={product_id} price={meta.get('Price')} title={meta.get('Title')[:72]}")
                else:
                    get_resp = request_with_retry("GET", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json")
                    get_status = str(get_resp.status_code)
                    get_resp.raise_for_status()
                    product = get_resp.json()
                    target_price = int(round(money(meta.get("Price")) * 100))
                    variants = [
                        {
                            "id": variant["id"],
                            "price": target_price if variant.get("is_enabled") else variant.get("price"),
                            "is_enabled": bool(variant.get("is_enabled")),
                        }
                        for variant in product.get("variants") or []
                    ]
                    payload = {"title": meta["Title"], "description": meta["Description"], "variants": variants}
                    update_resp = request_with_retry("PUT", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json", payload=payload)
                    update_status = str(update_resp.status_code)
                    update_resp.raise_for_status()
                    publish_resp = request_with_retry(
                        "POST",
                        f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
                        payload=PUBLISH_TITLE_DESC_PRICE,
                    )
                    publish_status = str(publish_resp.status_code)
                    publish_resp.raise_for_status()
                    result = "OK"
                    done += 1
                    print(f"[EBAY-NOSTICKER-SYNC] {local_id} get={get_status} update={update_status} publish={publish_status}")
            except Exception as exc:  # noqa: BLE001
                result = "FAILED"
                error = f"{type(exc).__name__}: {exc}"[:500]
                print(f"[EBAY-NOSTICKER-FAIL] {local_id}: {error}")
            writer.writerow(
                {
                    "Timestamp": now_text(),
                    "ID": local_id,
                    "Printify_Product_ID": product_id,
                    "HTTP_Get": get_status,
                    "HTTP_Update": update_status,
                    "HTTP_Publish": publish_status,
                    "Result": result,
                    "Error": error,
                }
            )
            time.sleep(random.uniform(sleep_min, max(sleep_min, sleep_max)))
    print(f"[EBAY-NOSTICKER-SYNC-DONE] rows={len(rows)} synced={done} dry_run={dry_run}")
    return done


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply no-sticker eBay repair experiment through local workbook and Printify sync.")
    parser.add_argument("--group", default="TREATMENT_FREE_SHIP_BRAND_COPY")
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument("--apply-local", action="store_true")
    parser.add_argument("--sync-printify", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--include-free-ship-price", action="store_true")
    parser.add_argument("--sleep-min", type=float, default=6.0)
    parser.add_argument("--sleep-max", type=float, default=14.0)
    args = parser.parse_args()
    if args.apply_local or not args.sync_printify:
        generate_plan(
            args.group,
            args.limit,
            apply_local=args.apply_local and not args.dry_run,
            include_free_ship_price=args.include_free_ship_price,
        )
    if args.sync_printify:
        sync_printify(args.limit, dry_run=args.dry_run, sleep_min=args.sleep_min, sleep_max=args.sleep_max)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
