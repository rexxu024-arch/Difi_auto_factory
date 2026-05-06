import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
COPY_CSV = DATABASE_DIR / "Listing_Copy_Optimization.csv"

SAFE_LOCAL_STATUSES = {
    "Ready_for_Printify",
    "Printify_UI_Mockups4",
    "Printify_UI_Mockups5",
    "Printify_UI_Mockups8",
    "Printify_MockupsPending",
    "Printify_UI_Failed",
}


def _load_candidates():
    if not COPY_CSV.exists():
        raise FileNotFoundError(f"Missing copy candidate file: {COPY_CSV}")
    with COPY_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return {row["ID"]: row for row in csv.DictReader(handle)}


def apply_candidates(dry_run=False):
    candidates = _load_candidates()
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    changed_rows = 0
    changed_cells = 0
    skipped_published = 0
    for row_idx in range(2, ws.max_row + 1):
        item_id = str(ws.cell(row_idx, cols["ID"]).value or "").strip()
        status = str(ws.cell(row_idx, cols["Status"]).value or "").strip()
        candidate = candidates.get(item_id)
        if not item_id or not candidate:
            continue
        if status.startswith("Printify_Published"):
            skipped_published += 1
            continue
        if status not in SAFE_LOCAL_STATUSES:
            continue
        row_changed = False
        updates = {
            "Title": candidate.get("Proposed_eBay_Title"),
            "Description": candidate.get("Proposed_eBay_Description"),
        }
        for header, value in updates.items():
            if not value or header not in cols:
                continue
            cell = ws.cell(row_idx, cols[header])
            if cell.value != value:
                if not dry_run:
                    cell.value = value
                changed_cells += 1
                row_changed = True
        if row_changed:
            changed_rows += 1
            print(f"[COPY-APPLY] {item_id} status={status}")
    if not dry_run:
        wb.save(EBAY_BOOK)
    wb.close()
    print(
        f"[COPY-APPLY] changed_rows={changed_rows} changed_cells={changed_cells} "
        f"skipped_published={skipped_published} dry_run={dry_run}"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    apply_candidates(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
