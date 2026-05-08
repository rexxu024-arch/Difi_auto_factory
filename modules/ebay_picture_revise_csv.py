"""Build eBay Seller Hub Reports revise CSVs for buyer-gallery image repair.

The file is intentionally generated locally only. Uploading it to eBay Reports
is a separate, auditable step after a one-listing test.
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.ebay_online_cover_audit import ahash, hamming, load_image


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"
LIVE_GALLERY_AUDIT = DATABASE_DIR / "eBay_Live_Gallery_Duplicate_Audit.csv"
GALLERY_REPAIR_QUEUE = DATABASE_DIR / "Printify_Gallery_Repair_Queue.csv"
OUT_DIR = DATABASE_DIR / "eBay_Picture_Revise"


def clean(value) -> str:
    return str(value or "").strip()


def fetch_product(product_id: str) -> dict:
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers={"Authorization": f"Bearer {Config.Printify_API_KEY}"},
        timeout=90,
    )
    response.raise_for_status()
    return response.json()


def image_hash_from_url(url: str) -> str:
    response = requests.get(url, timeout=90)
    response.raise_for_status()
    with Image.open(io.BytesIO(response.content)) as image:
        image.load()
        return ahash(image.convert("RGB"))


def workbook_rows() -> dict[str, dict[str, str]]:
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[cols["ID"]]:
            continue
        local_id = clean(values[cols["ID"]])
        rows[local_id] = {name: clean(values[idx]) for name, idx in cols.items() if idx < len(values)}
    workbook.close()
    return rows


def fix_ids(limit: int, ids: set[str] | None) -> list[str]:
    if ids:
        return sorted(ids)
    rows = []
    with FIX_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if row.get("Status") == "PENDING_FIX":
                rows.append(clean(row.get("ID")))
            if limit and len(rows) >= limit:
                break
    return rows


def printify_urls_for_row(row: dict[str, str]) -> tuple[list[str], str]:
    product_id = row.get("Printify_Product_ID", "")
    cover_path = Path(row.get("Cover_Path", ""))
    product = fetch_product(product_id)
    images = [image for image in product.get("images") or [] if image.get("is_selected_for_publishing") is not False]
    if not cover_path.exists():
        raise RuntimeError(f"missing local cover: {cover_path}")
    cover_hash = ahash(load_image(cover_path))
    scored = []
    for image in images:
        url = clean(image.get("src"))
        if not url:
            continue
        distance = hamming(cover_hash, image_hash_from_url(url))
        scored.append((distance, url))
    if not scored:
        raise RuntimeError("no selected Printify image URLs")
    scored.sort(key=lambda item: item[0])
    cover_distance, cover_url = scored[0]
    if cover_distance > 5:
        raise RuntimeError(f"no Printify image URL matches local cover closely; best distance={cover_distance}")
    other_urls = [url for distance, url in scored[1:]]
    return [cover_url, *other_urls], f"cover_distance={cover_distance} urls={len(scored)}"


def selected_printify_urls_for_row(row: dict[str, str]) -> tuple[list[str], str]:
    product_id = row.get("Printify_Product_ID", "")
    product = fetch_product(product_id)
    urls = []
    for image in product.get("images") or []:
        if image.get("is_selected_for_publishing") is False:
            continue
        url = clean(image.get("src"))
        if url and url not in urls:
            urls.append(url)
    if len(urls) < 3:
        raise RuntimeError(f"selected Printify image URL count too low: {len(urls)}")
    return urls, f"selected_unique_urls={len(urls)}"


def ids_from_live_gallery(limit: int) -> list[str]:
    if not LIVE_GALLERY_AUDIT.exists():
        return []
    rows = []
    seen = set()
    with LIVE_GALLERY_AUDIT.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            item_id = clean(row.get("ID"))
            if row.get("Result") != "CHECK_LIVE_DUPLICATE" or not item_id or item_id in seen:
                continue
            rows.append(item_id)
            seen.add(item_id)
            if limit and len(rows) >= limit:
                break
    return rows


def ids_from_gallery_repair(limit: int) -> list[str]:
    if not GALLERY_REPAIR_QUEUE.exists():
        return []
    rows = []
    with GALLERY_REPAIR_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            item_id = clean(row.get("ID"))
            if not item_id:
                continue
            rows.append(item_id)
            if limit and len(rows) >= limit:
                break
    return rows


def run(
    limit: int = 1,
    ids: set[str] | None = None,
    source: str = "cover-match",
    from_live_gallery: bool = False,
    from_gallery_repair: bool = False,
) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows_by_id = workbook_rows()
    if ids:
        selected_ids = sorted(ids)
    elif from_live_gallery:
        selected_ids = ids_from_live_gallery(limit=limit)
    elif from_gallery_repair:
        selected_ids = ids_from_gallery_repair(limit=limit)
    else:
        selected_ids = fix_ids(limit=limit, ids=None)
    suffix = source.replace("-", "_")
    out_path = OUT_DIR / f"ebay_picture_revise_{suffix}_{len(selected_ids)}.csv"
    log_path = OUT_DIR / f"ebay_picture_revise_{suffix}_{len(selected_ids)}_log.csv"
    fieldnames = ["Action", "Item number", "Custom label (SKU)", "Item photo URL"]
    log_fields = ["ID", "eBay_Item_ID", "Printify_Product_ID", "Result", "Note"]
    out_rows = []
    log_rows = []
    for local_id in selected_ids:
        row = rows_by_id.get(local_id)
        if not row:
            log_rows.append({"ID": local_id, "Result": "ERROR", "Note": "missing workbook row"})
            continue
        try:
            if source == "printify-selected":
                urls, note = selected_printify_urls_for_row(row)
            else:
                urls, note = printify_urls_for_row(row)
            out_rows.append(
                {
                    "Action": "Revise",
                    "Item number": row.get("eBay_Item_ID", ""),
                    "Custom label (SKU)": row.get("Printify_Product_ID", "") or row.get("SKU", "") or local_id,
                    "Item photo URL": "|".join(urls),
                }
            )
            log_rows.append(
                {
                    "ID": local_id,
                    "eBay_Item_ID": row.get("eBay_Item_ID", ""),
                    "Printify_Product_ID": row.get("Printify_Product_ID", ""),
                    "Result": "READY",
                    "Note": note,
                }
            )
        except Exception as exc:  # noqa: BLE001
            log_rows.append(
                {
                    "ID": local_id,
                    "eBay_Item_ID": row.get("eBay_Item_ID", "") if row else "",
                    "Printify_Product_ID": row.get("Printify_Product_ID", "") if row else "",
                    "Result": "ERROR",
                    "Note": str(exc)[:500],
                }
            )
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)
    with log_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=log_fields)
        writer.writeheader()
        writer.writerows(log_rows)
    print(f"[EBAY-PICTURE-REVISE] rows={len(out_rows)} csv={out_path} log={log_path}")
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=1)
    parser.add_argument("--ids", default="", help="Comma-separated local IDs to include.")
    parser.add_argument("--source", choices=["cover-match", "printify-selected"], default="cover-match")
    parser.add_argument("--from-live-gallery-audit", action="store_true")
    parser.add_argument("--from-gallery-repair", action="store_true")
    args = parser.parse_args()
    ids = {part.strip() for part in args.ids.split(",") if part.strip()} or None
    run(
        limit=args.limit,
        ids=ids,
        source=args.source,
        from_live_gallery=args.from_live_gallery_audit,
        from_gallery_repair=args.from_gallery_repair,
    )


if __name__ == "__main__":
    main()
