from __future__ import annotations

import argparse
import csv
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.ebay_token_manager import get_access_token


DATABASE = PROJECT_ROOT / "Database"
REPORTS = PROJECT_ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
COVER_FIX_QUEUE = DATABASE / "eBay_Online_Cover_Fix_Queue.csv"
LIVE_GALLERY_AUDIT = DATABASE / "eBay_Live_Gallery_Duplicate_Audit.csv"
OUT_DIR = DATABASE / "eBay_Picture_Revise"
NY = ZoneInfo("America/New_York")
TRADING_ENDPOINT = "https://api.ebay.com/ws/api.dll"
TRADING_VERSION = "1209"


def clean(value: object) -> str:
    return str(value or "").strip()


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def trading_headers(call_name: str) -> dict[str, str]:
    return {
        "Content-Type": "text/xml;charset=UTF-8",
        "X-EBAY-API-COMPATIBILITY-LEVEL": TRADING_VERSION,
        "X-EBAY-API-CALL-NAME": call_name,
        "X-EBAY-API-SITEID": "0",
        "X-EBAY-API-IAF-TOKEN": get_access_token(),
    }


def trading_call(call_name: str, xml_body: str) -> ET.Element:
    response = requests.post(
        TRADING_ENDPOINT,
        data=xml_body.encode("utf-8"),
        headers=trading_headers(call_name),
        timeout=60,
    )
    response.raise_for_status()
    return ET.fromstring(response.text)


def ns() -> dict[str, str]:
    return {"e": "urn:ebay:apis:eBLBaseComponents"}


def get_item(item_id: str) -> dict[str, Any]:
    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <Version>{TRADING_VERSION}</Version>
  <ItemID>{item_id}</ItemID>
  <DetailLevel>ReturnAll</DetailLevel>
  <IncludeItemSpecifics>false</IncludeItemSpecifics>
</GetItemRequest>"""
    root = trading_call("GetItem", xml)
    ack = root.findtext("e:Ack", default="", namespaces=ns())
    errors = [clean(node.findtext("e:LongMessage", default="", namespaces=ns())) for node in root.findall("e:Errors", ns())]
    item = root.find("e:Item", ns())
    urls = [
        clean(node.text)
        for node in root.findall(".//e:Item/e:PictureDetails/e:PictureURL", ns())
        if clean(node.text)
    ]
    return {
        "ack": ack,
        "errors": " | ".join(error for error in errors if error),
        "title": clean(item.findtext("e:Title", default="", namespaces=ns())) if item is not None else "",
        "sku": clean(item.findtext("e:SKU", default="", namespaces=ns())) if item is not None else "",
        "picture_source": clean(
            item.findtext("e:PictureDetails/e:PictureSource", default="", namespaces=ns())
        )
        if item is not None
        else "",
        "picture_urls": urls,
        "listing_duration": clean(item.findtext("e:ListingDuration", default="", namespaces=ns())) if item is not None else "",
        "quantity": clean(item.findtext("e:Quantity", default="", namespaces=ns())) if item is not None else "",
        "quantity_sold": clean(item.findtext("e:SellingStatus/e:QuantitySold", default="", namespaces=ns()))
        if item is not None
        else "",
    }


def verify_revise_allowed(item_id: str, existing_urls: list[str]) -> tuple[str, str]:
    if not existing_urls:
        return "NO", "no existing picture URL to verify"
    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<ReviseFixedPriceItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <Version>{TRADING_VERSION}</Version>
  <VerifyOnly>true</VerifyOnly>
  <Item>
    <ItemID>{item_id}</ItemID>
    <PictureDetails>
      <PictureURL>{existing_urls[0]}</PictureURL>
    </PictureDetails>
  </Item>
</ReviseFixedPriceItemRequest>"""
    root = trading_call("ReviseFixedPriceItem", xml)
    ack = root.findtext("e:Ack", default="", namespaces=ns())
    errors = []
    for node in root.findall("e:Errors", ns()):
        severity = clean(node.findtext("e:SeverityCode", default="", namespaces=ns()))
        code = clean(node.findtext("e:ErrorCode", default="", namespaces=ns()))
        short = clean(node.findtext("e:ShortMessage", default="", namespaces=ns()))
        long = clean(node.findtext("e:LongMessage", default="", namespaces=ns()))
        errors.append(f"{severity}:{code}:{short}:{long}"[:350])
    allowed = "YES" if ack in {"Success", "Warning"} else "NO"
    return allowed, " | ".join(errors)[:900]


def workbook_rows() -> dict[str, dict[str, str]]:
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    rows: dict[str, dict[str, str]] = {}
    try:
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or not values[cols["ID"]]:
                continue
            row_id = clean(values[cols["ID"]])
            rows[row_id] = {name: clean(values[idx]) for name, idx in cols.items() if idx < len(values)}
    finally:
        workbook.close()
    return rows


def queue_ids(limit: int, ids: set[str] | None) -> list[str]:
    if ids:
        return sorted(ids)
    if not COVER_FIX_QUEUE.exists():
        return []
    out: list[str] = []
    with COVER_FIX_QUEUE.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if clean(row.get("Status")) != "PENDING_FIX":
                continue
            row_id = clean(row.get("ID"))
            if row_id:
                out.append(row_id)
            if limit and len(out) >= limit:
                break
    return out


def live_gallery_check_ids(limit: int) -> list[str]:
    if not LIVE_GALLERY_AUDIT.exists():
        return []
    out: list[str] = []
    seen = set()
    with LIVE_GALLERY_AUDIT.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            result = clean(row.get("Result"))
            row_id = clean(row.get("ID"))
            if not row_id or row_id in seen:
                continue
            if result.startswith("CHECK"):
                out.append(row_id)
                seen.add(row_id)
            if limit and len(out) >= limit:
                break
    return out


def u_paths_from_u1(u1_path: str) -> list[Path]:
    if not u1_path:
        return []
    first = Path(u1_path)
    paths: list[Path] = []
    for index in range(1, 5):
        candidate = Path(str(first).replace("_U1_", f"_U{index}_"))
        if candidate.exists():
            paths.append(candidate)
    if not paths and first.exists():
        paths.append(first)
    return paths


def local_picture_paths(row: dict[str, str]) -> tuple[list[Path], str]:
    product_type = clean(row.get("Product_Type"))
    cover = Path(clean(row.get("Cover_Path")))
    u_paths = u_paths_from_u1(clean(row.get("Gallery_U1_Path")))
    paths: list[Path] = []
    if cover.exists():
        paths.append(cover)
    if product_type == "Sticker":
        paths.extend(path for path in u_paths if path.exists())
    elif not paths:
        production = Path(clean(row.get("Production_Design_Path")))
        if production.exists():
            paths.append(production)
    deduped: list[Path] = []
    seen = set()
    for path in paths:
        key = str(path.resolve()).lower()
        if key not in seen:
            deduped.append(path)
            seen.add(key)
    note = f"local_paths={len(deduped)} product_type={product_type}"
    return deduped, note


def csv_lookup(path: Path, key_field: str = "ID") -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {clean(row.get(key_field)): row for row in csv.DictReader(handle) if clean(row.get(key_field))}


def candidate_item_id(row_id: str, row: dict[str, str], live_lookup: dict[str, dict[str, str]], queue_lookup: dict[str, dict[str, str]]) -> str:
    return (
        clean(row.get("eBay_Item_ID"))
        or clean(live_lookup.get(row_id, {}).get("eBay_Item_ID"))
        or clean(queue_lookup.get(row_id, {}).get("eBay_Item_ID"))
    )


def unique_live_urls(urls: list[str]) -> list[str]:
    deduped: list[str] = []
    seen = set()
    for url in urls:
        # Strip volatile query parameters but keep the canonical eBay image URL.
        key = url.split("?", 1)[0]
        if key not in seen:
            deduped.append(url)
            seen.add(key)
    return deduped


def planned_reused_live_urls(product_type: str, item: dict[str, Any], local_paths: list[Path]) -> list[str]:
    urls = unique_live_urls(list(item.get("picture_urls") or []))
    if product_type == "Sticker":
        # For stickers, local Cover + U1-U4 is the safest source of truth.
        return []
    if not local_paths:
        return urls
    # Keep buyer-context mockup pictures after the local actual design image.
    # Drop the first live image because the new uploaded local cover should own slot 1.
    return urls[1:12]


def build_revise_xml(item_id: str, local_picture_count: int, reused_live_urls: list[str]) -> str:
    url_lines = [f"      <PictureURL>{{EPS_URL_{index:02d}}}</PictureURL>" for index in range(1, local_picture_count + 1)]
    url_lines.extend(f"      <PictureURL>{url}</PictureURL>" for url in reused_live_urls)
    urls = "\n".join(url_lines)
    return f"""<?xml version="1.0" encoding="utf-8"?>
<ReviseFixedPriceItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <Version>{TRADING_VERSION}</Version>
  <Item>
    <ItemID>{item_id}</ItemID>
    <PictureDetails>
{urls}
    </PictureDetails>
  </Item>
</ReviseFixedPriceItemRequest>"""


def run(
    limit: int,
    ids: set[str] | None = None,
    from_live_gallery_checks: bool = False,
    verify_revise: bool = False,
) -> tuple[Path, Path]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS.mkdir(parents=True, exist_ok=True)
    rows = workbook_rows()
    live_lookup = csv_lookup(LIVE_GALLERY_AUDIT)
    queue_lookup = csv_lookup(COVER_FIX_QUEUE)
    selected = live_gallery_check_ids(limit=limit) if from_live_gallery_checks and not ids else queue_ids(limit=limit, ids=ids)
    stamp = datetime.now(NY).strftime("%Y%m%d_%H%M%S")
    out_csv = OUT_DIR / f"ebay_trading_picture_repair_dry_run_{stamp}.csv"
    out_md = REPORTS / f"eBay_Trading_Picture_Repair_Dry_Run_{stamp}.md"
    fields = [
        "Timestamp",
        "ID",
        "Product_Type",
        "eBay_Item_ID",
        "Printify_Product_ID",
        "Trading_Ack",
        "Trading_SKU",
        "Live_Picture_Source",
        "Live_Picture_Count",
        "Planned_Local_Picture_Count",
        "Planned_Reused_Live_Picture_Count",
        "Planned_Total_Picture_Count",
        "Planned_Method",
        "Trading_Revise_VerifyOnly_Allowed",
        "Trading_Revise_VerifyOnly_Note",
        "Can_Proceed_After_Review",
        "Local_Picture_Paths",
        "Reused_Live_Picture_URLs",
        "Live_Picture_URLs",
        "Revise_XML_Template",
        "Notes",
    ]
    output: list[dict[str, str]] = []
    for row_id in selected:
        row = rows.get(row_id)
        if not row:
            output.append({"Timestamp": now_text(), "ID": row_id, "Can_Proceed_After_Review": "NO", "Notes": "missing workbook row"})
            continue
        item_id = candidate_item_id(row_id, row, live_lookup, queue_lookup)
        if not item_id:
            output.append({"Timestamp": now_text(), "ID": row_id, "Can_Proceed_After_Review": "NO", "Notes": "missing eBay item id"})
            continue
        try:
            item = get_item(item_id)
            local_paths, local_note = local_picture_paths(row)
            product_type = clean(row.get("Product_Type"))
            reused_urls = planned_reused_live_urls(product_type, item, local_paths)
            planned_total = len(local_paths) + len(reused_urls)
            verify_allowed = "NOT_CHECKED"
            verify_note = ""
            if verify_revise:
                verify_allowed, verify_note = verify_revise_allowed(item_id, item.get("picture_urls") or [])
            can_proceed = bool(planned_total and item.get("ack") == "Success" and verify_allowed != "NO")
            output.append(
                {
                    "Timestamp": now_text(),
                    "ID": row_id,
                    "Product_Type": product_type,
                    "eBay_Item_ID": item_id,
                    "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
                    "Trading_Ack": clean(item.get("ack")),
                    "Trading_SKU": clean(item.get("sku")),
                    "Live_Picture_Source": clean(item.get("picture_source")),
                    "Live_Picture_Count": str(len(item.get("picture_urls") or [])),
                    "Planned_Local_Picture_Count": str(len(local_paths)),
                    "Planned_Reused_Live_Picture_Count": str(len(reused_urls)),
                    "Planned_Total_Picture_Count": str(planned_total),
                    "Planned_Method": "UploadSiteHostedPictures_EPS_then_ReviseFixedPriceItem",
                    "Trading_Revise_VerifyOnly_Allowed": verify_allowed,
                    "Trading_Revise_VerifyOnly_Note": verify_note,
                    "Can_Proceed_After_Review": "YES" if can_proceed else "NO",
                    "Local_Picture_Paths": "|".join(str(path) for path in local_paths),
                    "Reused_Live_Picture_URLs": "|".join(reused_urls),
                    "Live_Picture_URLs": "|".join(item.get("picture_urls") or []),
                    "Revise_XML_Template": build_revise_xml(item_id, len(local_paths), reused_urls) if can_proceed else "",
                    "Notes": local_note if can_proceed else f"{local_note}; trading_errors={item.get('errors','')}",
                }
            )
        except Exception as exc:  # noqa: BLE001
            output.append(
                {
                    "Timestamp": now_text(),
                    "ID": row_id,
                    "Product_Type": clean(row.get("Product_Type")),
                    "eBay_Item_ID": item_id,
                    "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
                    "Can_Proceed_After_Review": "NO",
                    "Notes": f"{type(exc).__name__}: {exc}"[:700],
                }
            )
    with out_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(output)
    ready = sum(1 for row in output if row.get("Can_Proceed_After_Review") == "YES")
    lines = [
        "# eBay Trading Picture Repair Dry Run",
        "",
        f"Generated: {now_text()}",
        f"Rows checked: {len(output)}",
        f"Ready after human/code review: {ready}",
        "",
        "## Method",
        "",
        "- Read only: `GetItem` through Trading API with OAuth IAF token.",
        "- Planned write, not executed: upload local Cover/U assets to EPS with `UploadSiteHostedPictures`, then replace `PictureDetails.PictureURL` with `ReviseFixedPriceItem`.",
        "- Guard: this report does not upload pictures and does not revise live listings.",
        "",
        "## Official references",
        "",
        "- https://developer.ebay.com/devzone/XML/docs/Reference/eBay/UploadSiteHostedPictures.html",
        "- https://developer.ebay.com/api-docs/user-guides/static/trading-user-guide/picture-hosting.html",
        "",
        f"CSV: {out_csv}",
    ]
    out_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[EBAY-TRADING-PICTURE-DRY-RUN] rows={len(output)} ready={ready} csv={out_csv} md={out_md}")
    return out_csv, out_md


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=3)
    parser.add_argument("--ids", default="", help="Comma-separated local IDs.")
    parser.add_argument("--from-live-gallery-checks", action="store_true")
    parser.add_argument("--verify-revise", action="store_true", help="Run ReviseFixedPriceItem VerifyOnly with current picture URL.")
    args = parser.parse_args()
    ids = {part.strip() for part in args.ids.split(",") if part.strip()} or None
    run(
        limit=args.limit,
        ids=ids,
        from_live_gallery_checks=args.from_live_gallery_checks,
        verify_revise=args.verify_revise,
    )


if __name__ == "__main__":
    main()
