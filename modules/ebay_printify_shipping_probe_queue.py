from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
MATRIX_CSV = DATABASE / "eBay_Shipping_Repair_Decision_Matrix.csv"
OUT_CSV = DATABASE / "eBay_Printify_Shipping_Template_Probe_Queue.csv"
OUT_MD = REPORTS / "eBay_Printify_Shipping_Template_Probe_Queue.md"
NY = ZoneInfo("America/New_York")


def clean(value: object) -> str:
    return str(value or "").replace("\r", " ").replace("\n", " ").strip()


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def workbook_by_item() -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {clean(name): idx for idx, name in enumerate(headers) if name}
    out: dict[str, dict[str, str]] = {}
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values:
                continue
            item_id = clean(values[cols.get("eBay_Item_ID", -1)]) if "eBay_Item_ID" in cols else ""
            if not item_id:
                continue
            status = clean(values[cols.get("Status", -1)]) if "Status" in cols else ""
            row = {
                "ID": clean(values[cols.get("ID", -1)]) if "ID" in cols else "",
                "Product_Type": clean(values[cols.get("Product_Type", -1)]) if "Product_Type" in cols else "",
                "Printify_Product_ID": clean(values[cols.get("Printify_Product_ID", -1)]) if "Printify_Product_ID" in cols else "",
                "Title": clean(values[cols.get("Title", -1)]) if "Title" in cols else "",
                "Status": status,
            }
            if item_id not in out or status.startswith("Printify_Published"):
                out[item_id] = row
    finally:
        wb.close()
    return out


def to_int(value: object) -> int:
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def build(limit: int = 5) -> tuple[Path, Path]:
    matrix = read_csv(MATRIX_CSV)
    book = workbook_by_item()
    candidates: list[dict[str, str]] = []
    for row in matrix:
        item_id = clean(row.get("Item_ID"))
        local = book.get(item_id, {})
        product_type = clean(row.get("Product_Type") or local.get("Product_Type"))
        if product_type == "Sticker":
            continue
        if clean(row.get("Decision")) != "SOURCE_REBUILD_OR_PRINTIFY_TEMPLATE_PROBE":
            continue
        if to_int(row.get("Image_Count")) < 4:
            continue
        if not clean(local.get("Printify_Product_ID")):
            continue
        candidates.append(
            {
                "Generated": now_text(),
                "Rank": str(len(candidates) + 1),
                "ID": clean(local.get("ID")),
                "Product_Type": product_type,
                "eBay_Item_ID": item_id,
                "Printify_Product_ID": clean(local.get("Printify_Product_ID")),
                "Views_30_Days": clean(row.get("Views_30_Days")),
                "Current_Price": clean(row.get("Current_Price")),
                "Current_Shipping": clean(row.get("Current_Shipping")),
                "Image_Count": clean(row.get("Image_Count")),
                "Recommended_Probe": "Printify publish shipping_template=true only, one item first, then re-read buyer page.",
                "Write_Status": "DRY_RUN_QUEUE_ONLY",
                "Title": clean(row.get("Title") or local.get("Title")),
            }
        )
        if limit and len(candidates) >= limit:
            break

    fields = [
        "Generated",
        "Rank",
        "ID",
        "Product_Type",
        "eBay_Item_ID",
        "Printify_Product_ID",
        "Views_30_Days",
        "Current_Price",
        "Current_Shipping",
        "Image_Count",
        "Recommended_Probe",
        "Write_Status",
        "Title",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(candidates)

    lines = [
        "# eBay Printify Shipping Template Probe Queue",
        "",
        f"Generated: {now_text()}",
        f"Candidates: {len(candidates)}",
        "",
        "## Rule",
        "",
        "- This is a dry-run queue only.",
        "- Candidate must be non-Sticker, already have at least 4 buyer-facing images, have a Printify product id, and be blocked by paid shipping.",
        "- First live test, when allowed, should publish only `shipping_template=true` for one item, then verify the eBay buyer page before any rollout.",
        "",
        "## Candidates",
        "",
    ]
    for row in candidates:
        lines.append(
            f"- {row['Rank']}. {row['ID']} {row['Product_Type']} item={row['eBay_Item_ID']} "
            f"views={row['Views_30_Days']} shipping=${row['Current_Shipping']} images={row['Image_Count']}"
        )
    lines.extend(["", f"CSV: {OUT_CSV}"])
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[EBAY-SHIPPING-PROBE-QUEUE] candidates={len(candidates)} csv={OUT_CSV} md={OUT_MD}")
    return OUT_CSV, OUT_MD


def main() -> int:
    build()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
