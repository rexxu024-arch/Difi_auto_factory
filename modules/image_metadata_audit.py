import csv
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUTPUT_CSV = DATABASE_DIR / "Image_Metadata_Audit.csv"

HEADERS = [
    "ID",
    "Product_Type",
    "Image_Role",
    "Path",
    "Exists",
    "Format",
    "Size",
    "File_Size_KB",
    "Info_Keys",
    "EXIF_Key_Count",
    "Read_Status",
]


def _clean(value):
    return str(value or "").strip()


def _inspect(path):
    target = Path(path)
    if not path or not target.exists():
        return {
            "Exists": False,
            "Format": "",
            "Size": "",
            "File_Size_KB": "",
            "Info_Keys": "",
            "EXIF_Key_Count": "",
            "Read_Status": "MISSING",
        }
    try:
        with Image.open(target) as image:
            image.load()
            exif = image.getexif()
            info_keys = sorted(str(key) for key in image.info.keys())
            return {
                "Exists": True,
                "Format": image.format or target.suffix.lstrip(".").upper(),
                "Size": f"{image.width}x{image.height}",
                "File_Size_KB": round(target.stat().st_size / 1024, 1),
                "Info_Keys": ", ".join(info_keys),
                "EXIF_Key_Count": len(exif or {}),
                "Read_Status": "OK",
            }
    except Exception as exc:
        return {
            "Exists": True,
            "Format": "",
            "Size": "",
            "File_Size_KB": round(target.stat().st_size / 1024, 1),
            "Info_Keys": "",
            "EXIF_Key_Count": "",
            "Read_Status": f"ERROR: {exc}",
        }


def build():
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index for index, header in enumerate(headers)}
    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[cols["ID"]]:
                continue
            item_id = _clean(row[cols["ID"]])
            product = _clean(row[cols["Product_Type"]])
            roles = [("Production", "Production_Path"), ("Cover", "Cover_Path")]
            roles.extend((f"Gallery_U{index}", f"Gallery_U{index}_Path") for index in range(1, 5))
            for role, column in roles:
                path = _clean(row[cols[column]])
                result = _inspect(path)
                rows.append(
                    {
                        "ID": item_id,
                        "Product_Type": product,
                        "Image_Role": role,
                        "Path": path,
                        **result,
                    }
                )
    finally:
        wb.close()

    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    errors = [row for row in rows if row["Read_Status"] != "OK"]
    metadata = [row for row in rows if row["Info_Keys"] or row["EXIF_Key_Count"]]
    print(f"[IMAGE-META] rows={len(rows)} csv={OUTPUT_CSV}")
    print(f"[IMAGE-META] read_errors={len(errors)} rows_with_metadata={len(metadata)}")
    for row in errors[:20]:
        print(f"[IMAGE-META] {row['ID']} {row['Image_Role']} {row['Read_Status']}")


if __name__ == "__main__":
    build()
