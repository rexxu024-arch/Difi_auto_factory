import os, sys, time, requests, datetime, shutil, re, json, csv
from pathlib import Path
from openpyxl import load_workbook
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
from config import Config
from modules.spec_registry import Registry

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")

PRODUCTION_LINE_PATH = PROJECT_ROOT / "Database" / "Production_Line.xlsx"
DEFECT_LOG_PATH = PROJECT_ROOT / "Database" / "mj_defect_log.csv"
PRODUCTION_HEADERS = ["ID", "Timestamp", "Category", "Product_Type", "Style", "Title", "MJ_Prompt", "SEO_Hook", "Status"]
HARVEST_TIMEOUT_SECONDS = int(os.getenv("MJ_HARVEST_TIMEOUT_SECONDS", "120") or "120")
QUEUE_TIMEOUT_SECONDS = int(os.getenv("MJ_HARVEST_QUEUE_TIMEOUT_SECONDS", "600") or "600")
POLL_INTERVAL_SECONDS = float(os.getenv("MJ_HARVEST_POLL_SECONDS", "5") or "5")
GRID_FALLBACK_SECONDS = int(os.getenv("MJ_HARVEST_GRID_FALLBACK_SECONDS", "120") or "120")
MIN_GRID_DIM = int(os.getenv("MJ_MIN_GRID_DIM", "1024") or "1024")
MIN_UPSCALE_DIM = int(os.getenv("MJ_MIN_UPSCALE_DIM", "1024") or "1024")
ALLOW_GRID_FALLBACK = os.getenv("MJ_HARVEST_ALLOW_GRID_FALLBACK", "0").strip() == "1"
PROMPT_PREFIX = "(subject in center of frame:1.5), (clean white background:1.3), "
OFFICIAL_SUFFIX = "--ar 1:1 --v 6.0 --style raw --stylize 250"
PRODUCT_SUFFIXES = {
    "Sticker": "--ar 1:1 --v 6.0 --style raw --stylize 250",
    "Poster": "--ar 2:3 --v 6.0 --style raw --stylize 250",
    "Acrylic": "--ar 5:7 --v 6.0 --style raw --stylize 250",
    "T-Shirt": "--ar 2:3 --v 6.0 --style raw --stylize 250",
    "Wall Art": "--ar 2:3 --v 6.0 --style raw --stylize 250",
}
Config.TOKEN = getattr(Config, "TOKEN", None) or getattr(Config, "DISCORD_TOKEN", None) or os.getenv("DISCORD_TOKEN")
Config.APP_ID = getattr(Config, "APP_ID", None) or os.getenv("APPLICATION_ID")
Config.MJ_ID = getattr(Config, "MJ_ID", None) or os.getenv("MJ_ID")
Config.MJ_VERSION = getattr(Config, "MJ_VERSION", None) or os.getenv("MJ_VERSION")
Config.SESSION_ID = getattr(Config, "SESSION_ID", None) or os.getenv("SESSION_ID")

def _validate_runtime_config():
    required = {
        "DISCORD_TOKEN": Config.TOKEN,
        "GUILD_ID": Config.GUILD_ID,
        "CHANNEL_ID": Config.CHANNEL_ID,
        "APPLICATION_ID": Config.APP_ID,
        "SESSION_ID": Config.SESSION_ID,
        "MJ_VERSION": Config.MJ_VERSION,
        "MJ_ID": Config.MJ_ID,
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise RuntimeError(
            "Discord runtime config missing: "
            + ", ".join(missing)
            + ". Check root .env/config.py before running harvest."
        )

# --- 1. 核心通讯 (保持 V15.9 稳定性) ---
def _interaction(payload):
    headers = {"Authorization": Config.TOKEN, "Content-Type": "application/json"}
    last_error = ""
    for attempt in range(3):
        try:
            response = requests.post("https://discord.com/api/v9/interactions", json=payload, headers=headers, timeout=10)
            if 200 <= response.status_code < 300:
                return response
            last_error = f"HTTP {response.status_code}: {response.text[:300]}"
            if response.status_code in (401, 403):
                raise RuntimeError(f"Discord authorization failed: {last_error}")
            if response.status_code == 429:
                try:
                    retry_after = float(response.json().get("retry_after", 3))
                except Exception:
                    retry_after = 3
                print(f"[Discord] Rate limited. Sleeping {retry_after:.1f}s before retry.")
                time.sleep(min(max(retry_after, 1), 30))
                continue
            break
        except Exception as exc:
            last_error = repr(exc)
            time.sleep(2 + attempt)
    print(f"[Discord] Interaction failed: {last_error}")
    if "NameResolutionError" in last_error or "ConnectionError" in last_error or "ReadTimeout" in last_error:
        raise RuntimeError(f"Discord transient network failed: {last_error}")
    return None

def _load_production_workbook():
    wb = load_workbook(PRODUCTION_LINE_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(PRODUCTION_HEADERS) + 1)]
    if headers != PRODUCTION_HEADERS:
        wb.close()
        raise RuntimeError(f"Production_Line.xlsx header mismatch: {headers}")
    return wb, ws, {name: idx + 1 for idx, name in enumerate(PRODUCTION_HEADERS)}

def _read_ready_tasks(product_type):
    wb, ws, cols = _load_production_workbook()
    tasks = []
    try:
        for row in range(2, ws.max_row + 1):
            row_status = str(ws.cell(row, cols["Status"]).value or "").strip()
            row_product = str(ws.cell(row, cols["Product_Type"]).value or "").strip()
            if row_status != "Ready_for_production":
                continue
            if row_product.lower() != str(product_type).lower():
                continue
            task = {header: ws.cell(row, cols[header]).value for header in PRODUCTION_HEADERS}
            task["_row"] = row
            tasks.append(task)
    finally:
        wb.close()
    return tasks

def _selected_harvest_ids():
    raw_ids = os.getenv("MJ_HARVEST_IDS", "").strip()
    if not raw_ids:
        return None
    return {item.strip() for item in raw_ids.split(",") if item.strip()}

def _update_status_via_excel_com(t_id, status):
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    created_excel = False
    abs_path = str(PRODUCTION_LINE_PATH)
    try:
        try:
            workbook = win32com.client.GetObject(abs_path)
            excel = workbook.Application
        except Exception:
            excel = win32com.client.DispatchEx("Excel.Application")
            created_excel = True
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(abs_path)

        ws = workbook.Worksheets(1)
        headers = [ws.Cells(1, c).Value for c in range(1, len(PRODUCTION_HEADERS) + 1)]
        if headers != PRODUCTION_HEADERS:
            raise RuntimeError(f"Production_Line.xlsx header mismatch: {headers}")

        max_row = ws.UsedRange.Rows.Count
        for row in range(2, max_row + 1):
            if str(ws.Cells(row, 1).Value or "").strip() == str(t_id).strip():
                ws.Cells(row, 9).Value = status
                workbook.Save()
                return True
        return False
    finally:
        if created_excel and workbook is not None:
            workbook.Close(SaveChanges=True)
        if created_excel and excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()

def _update_product_line_status(t_id, status="Completed"):
    try:
        wb, ws, cols = _load_production_workbook()
        try:
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, cols["ID"]).value or "").strip() == str(t_id).strip():
                    ws.cell(row, cols["Status"]).value = status
                    try:
                        wb.save(PRODUCTION_LINE_PATH)
                    except PermissionError:
                        wb.close()
                        if _update_status_via_excel_com(t_id, status):
                            return
                        print(f"[WARN] ID:{t_id} not found in Production_Line.xlsx; status not updated.")
                        return
                    return
            print(f"[WARN] ID:{t_id} not found in Production_Line.xlsx; status not updated.")
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception as exc:
        print(f"[WARN] Failed to update Production_Line.xlsx status for ID:{t_id}: {exc}")

def _attachment_urls(asset):
    if isinstance(asset, dict):
        candidates = [asset.get("url"), asset.get("proxy_url")]
    else:
        candidates = [asset]
    urls = []
    for url in candidates:
        if url and url not in urls:
            urls.append(url)
    return urls

def _image_size(path):
    try:
        from PIL import Image
        with Image.open(path) as image:
            image.load()
            return image.size
    except Exception:
        return None

def _download_asset(asset, save_dir, filename, min_dim=0, label="asset"):
    full_path = os.path.join(save_dir, filename)
    tmp_path = full_path + ".part"
    for url in _attachment_urls(asset):
        try:
            r = requests.get(url, stream=True, timeout=30)
            if r.status_code != 200:
                continue
            with open(tmp_path, "wb") as f:
                for chunk in r.iter_content(8192):
                    if chunk:
                        f.write(chunk)
            size = _image_size(tmp_path)
            if not size:
                try: os.remove(tmp_path)
                except OSError: pass
                print(f"[QUALITY-REJECT] {filename} is not a readable image.")
                continue
            if min_dim and min(size) < min_dim:
                try: os.remove(tmp_path)
                except OSError: pass
                print(f"[QUALITY-REJECT] {filename} {size[0]}x{size[1]} below {min_dim}px min for {label}.")
                continue
            os.replace(tmp_path, full_path)
            return True
        except Exception as exc:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass
            print(f"[WARN] Download failed for {filename}: {exc}")
    return False

def _prepare_discord_prompt(raw_prompt, tid, product_type="Sticker"):
    prompt = str(raw_prompt or "").replace("\n", " ").replace("\r", " ").strip()
    prompt = re.sub(r"\s+", " ", prompt).strip()
    prompt = re.sub(r"\bID_[A-Za-z0-9_-]+\b", " ", prompt).strip()
    prompt = re.sub(r"\s+--(?:ar|v|version|style|stylize|s|q|quality|chaos|c|no)\b(?:\s+[^-][^ ]*)*", " ", prompt, flags=re.I).strip()
    prompt = re.sub(r"\s+", " ", prompt).strip()
    subject = prompt.split("--", 1)[0].strip()
    if str(product_type or "").strip() != "Sticker":
        subject = re.sub(
            r"\b(white contour border|contour border|die[- ]cut sticker style|die[- ]cut|sticker|vinyl decal|solid white background|isolated on white background|isolated on white)\b",
            " ",
            subject,
            flags=re.I,
        )
        subject = re.sub(r"\s*,\s*,+", ", ", subject)
        subject = re.sub(r"\s+", " ", subject).strip(" ,")
    suffix = PRODUCT_SUFFIXES.get(str(product_type or "Sticker").strip(), OFFICIAL_SUFFIX)
    return f"{subject} ID_{tid} {suffix}".strip()

def _message_buttons(message):
    components = message.get("components") or []
    if not components:
        return []
    return components[0].get("components", []) if isinstance(components[0], dict) else []

def _message_progress_percent(message):
    content = str(message.get("content") or "")
    matches = re.findall(r"\((\d{1,3})%\)|\b(\d{1,3})%", content)
    values = []
    for pair in matches:
        for value in pair:
            if value:
                try:
                    values.append(int(value))
                except ValueError:
                    pass
    return max(values) if values else None

def _is_incomplete_midjourney_message(message):
    content = str(message.get("content") or "")
    percent = _message_progress_percent(message)
    if percent is not None and percent < 100:
        return True
    incomplete_markers = (
        "Waiting to start",
        "Waiting to start...",
        "Job queued",
        "queued",
        "paused",
    )
    return any(marker.lower() in content.lower() for marker in incomplete_markers)

def _has_upscale_buttons(message):
    labels = []
    for button in _message_buttons(message):
        label = str(button.get("label") or button.get("emoji", {}).get("name") or "")
        custom_id = str(button.get("custom_id") or "")
        labels.append(label.upper())
        labels.append(custom_id.upper())
    return all(any(f"U{idx}" in label for label in labels) for idx in range(1, 5))

def _message_reference_id(message):
    ref = message.get("message_reference") or {}
    return str(ref.get("message_id") or "")

def _message_unix_time(message):
    try:
        snowflake = int(message.get("id") or 0)
        return ((snowflake >> 22) + 1420070400000) / 1000
    except Exception:
        return 0

def _is_grid_message(message):
    content = message.get("content", "")
    return (
        bool(message.get("attachments"))
        and "Image #" not in content
        and not _message_reference_id(message)
        and not _is_incomplete_midjourney_message(message)
        and _has_upscale_buttons(message)
    )

def _normalize_match_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()

def _prompt_signature(prompt, length=90):
    clean = str(prompt or "").split("--", 1)[0]
    clean = re.sub(r"\bID_[A-Za-z0-9_-]+\b", " ", clean)
    return _normalize_match_text(clean)[:length]

def _message_matches_task(message, info, tid):
    content = message.get("content", "")
    if f"ID_{tid}" in content:
        return True
    grid_message_id = str(info.get("grid_message_id") or "")
    if grid_message_id and _message_reference_id(message) == grid_message_id:
        return True
    if "Image #" not in content and message.get("attachments"):
        return False
    signature = info.get("prompt_signature") or ""
    return bool(signature and signature in _normalize_match_text(content))

def _trigger_upscales(message):
    btns = _message_buttons(message)
    if not btns:
        return False
    triggered = False
    for button in btns[:4]:
        if _interaction({
            "type": 3,
            "application_id": Config.APP_ID,
            "guild_id": Config.GUILD_ID,
            "channel_id": Config.CHANNEL_ID,
            "message_id": message["id"],
            "session_id": Config.SESSION_ID,
            "data": {"component_type": 2, "custom_id": button["custom_id"]},
        }):
            triggered = True
        time.sleep(2)
    return triggered

def _fetch_recent_messages(limit=100):
    return requests.get(
        f"https://discord.com/api/v9/channels/{Config.CHANNEL_ID}/messages?limit={limit}",
        headers={"Authorization": Config.TOKEN},
        timeout=10,
    ).json()

def _find_existing_task_message(tid, prompt_signature):
    if os.getenv("MJ_HARVEST_REATTACH", "0") != "1":
        return None
    try:
        messages = _fetch_recent_messages(100)
    except Exception:
        return None
    needle = f"ID_{tid}"
    now = time.time()
    for message in messages:
        msg_time = _message_unix_time(message)
        if msg_time and now - msg_time > QUEUE_TIMEOUT_SECONDS:
            continue
        content = message.get("content", "")
        if needle in content:
            return message
    return None

def _task_snapshot(t_id):
    try:
        wb, ws, cols = _load_production_workbook()
        try:
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, cols["ID"]).value or "").strip() == str(t_id).strip():
                    return {header: ws.cell(row, cols[header]).value for header in PRODUCTION_HEADERS}
        finally:
            wb.close()
    except Exception:
        pass
    return {"ID": t_id}

def _log_defect(t_id, status, reason="", task=None):
    task = task or _task_snapshot(t_id)
    DEFECT_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    exists = DEFECT_LOG_PATH.exists()
    with open(DEFECT_LOG_PATH, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "Timestamp", "ID", "Status", "Reason", "Category", "Product_Type",
            "Style", "Title", "MJ_Prompt", "SEO_Hook"
        ])
        if not exists:
            writer.writeheader()
        writer.writerow({
            "Timestamp": datetime.datetime.now().strftime("%-m/%-d/%Y  %-I:%M:%S %p") if os.name != "nt" else datetime.datetime.now().strftime("%#m/%#d/%Y  %#I:%M:%S %p"),
            "ID": task.get("ID", t_id),
            "Status": status,
            "Reason": reason,
            "Category": task.get("Category", ""),
            "Product_Type": task.get("Product_Type", ""),
            "Style": task.get("Style", ""),
            "Title": task.get("Title", ""),
            "MJ_Prompt": task.get("MJ_Prompt", ""),
            "SEO_Hook": task.get("SEO_Hook", ""),
        })

def _safe_complete_folder(review_path, final_path):
    if os.path.exists(final_path):
        raise RuntimeError(f"Final folder already exists: {final_path}")
    os.rename(review_path, final_path)

def _finalize_single_art(info, tid, sel_cat, sel_spec):
    u_files = [os.path.join(info["path"], f"{tid}_U{i}.png") for i in range(1, 5)]
    missing = [path for path in u_files if not os.path.exists(path)]
    if missing:
        _purge_asset(info["path"], tid, "Defeated_Quality", "Missing required U1-U4 assets", info.get("task_obj"))
        print(f"[WARN] [{tid}] Missing required U1-U4 assets.")
        return False
    low_res = []
    for path in u_files:
        size = _image_size(path)
        if not size or min(size) < MIN_UPSCALE_DIM:
            low_res.append(f"{os.path.basename(path)}={size}")
    if low_res:
        _purge_asset(info["path"], tid, "Defeated_Quality", "Upscale below quality floor: " + "; ".join(low_res), info.get("task_obj"))
        print(f"[WARN] [{tid}] Upscale below quality floor: {'; '.join(low_res)}.")
        return False
    info["task_obj"]["Status"] = "Completed"
    save_dual_metadata(info["path"], info["task_obj"])
    final_path = os.path.join("Output", sel_cat, sel_spec, f"MASTER_{tid}_Ready_for_Steaming")
    try:
        _safe_complete_folder(info["path"], final_path)
    except Exception as exc:
        print(f"[WARN] [{tid}] Failed to finalize folder: {exc}")
        _purge_asset(info["path"], tid, "Defeated_Prompt", f"Failed to finalize folder: {exc}", info.get("task_obj"))
        return False
    _update_product_line_status(tid)
    print(f"✅ [DONE] {tid} | U1-U4 + metadata harvested.")
    return True

def _finalize_kiss_cut(info, tid):
    grid_file = os.path.join(info["path"], f"{tid}_Grid.png")
    u_files = [os.path.join(info["path"], f"{tid}_U{i}_Grid.png") for i in range(1, 5)]
    missing = [path for path in [grid_file] + u_files if not os.path.exists(path)]
    if missing:
        _purge_asset(info["path"], tid, "Defeated_Quality", "Missing required grid/upscale assets", info.get("task_obj"))
        print(f"[WARN] [{tid}] Missing required grid/upscale assets.")
        return False
    grid_size = _image_size(grid_file)
    if not grid_size or min(grid_size) < MIN_GRID_DIM:
        _purge_asset(info["path"], tid, "Defeated_Quality", f"Grid below quality floor: {grid_size}", info.get("task_obj"))
        print(f"[WARN] [{tid}] Grid below quality floor: {grid_size}.")
        return False
    low_res = []
    for path in u_files:
        size = _image_size(path)
        if not size or min(size) < MIN_UPSCALE_DIM:
            low_res.append(f"{os.path.basename(path)}={size}")
    if low_res:
        _purge_asset(info["path"], tid, "Defeated_Quality", "Upscale below quality floor: " + "; ".join(low_res), info.get("task_obj"))
        print(f"[WARN] [{tid}] Upscale below quality floor: {'; '.join(low_res)}.")
        return False
    info["task_obj"]["Status"] = "Completed"
    save_dual_metadata(info["path"], info["task_obj"])
    final_path = os.path.join("Output", "Sticker", "Kiss-Cut", f"MASTER_{tid}")
    try:
        _safe_complete_folder(info["path"], final_path)
    except Exception as exc:
        print(f"[WARN] [{tid}] Failed to finalize folder: {exc}")
        _purge_asset(info["path"], tid, "Defeated_Prompt", f"Failed to finalize folder: {exc}", info.get("task_obj"))
        return False
    _update_product_line_status(tid)
    print(f"✅ [DONE] {tid} | Kiss-Cut grids + metadata harvested.")
    return True

def _split_grid_as_upscales(info, tid):
    if not ALLOW_GRID_FALLBACK:
        return False
    grid_file = os.path.join(info["path"], f"{tid}_Grid.png")
    if not os.path.exists(grid_file):
        return False
    try:
        from PIL import Image
        image = Image.open(grid_file).convert("RGB")
        width, height = image.size
        boxes = [
            (0, 0, width // 2, height // 2),
            (width // 2, 0, width, height // 2),
            (0, height // 2, width // 2, height),
            (width // 2, height // 2, width, height),
        ]
        for idx, box in enumerate(boxes, 1):
            out = os.path.join(info["path"], f"{tid}_U{idx}_Grid.png")
            if not os.path.exists(out):
                image.crop(box).save(out)
            size = _image_size(out)
            if size and min(size) >= MIN_UPSCALE_DIM:
                info["u_received"].add(str(idx))
            else:
                try: os.remove(out)
                except OSError: pass
                print(f"[QUALITY-REJECT] {tid} grid fallback U{idx} below {MIN_UPSCALE_DIM}px: {size}")
        print(f"[Grid-Fallback] {tid} split grid into U1-U4 files.")
        return True
    except Exception as exc:
        print(f"[WARN] [{tid}] Grid fallback failed: {exc}")
        return False

def _purge_asset(path, t_id, status="Defeated_Prompt", reason="", task=None):
    if path and os.path.exists(path): shutil.rmtree(path)
    _log_defect(t_id, status, reason, task)
    _update_product_line_status(t_id, status)
    print(f"🗑️ [ID:{t_id}] 清理并标记为 {status}")

def save_dual_metadata(save_path, task_a, task_b=None):
    if task_b:
        content = f"MASTER_ID: {task_a.get('ID')}\n[PART A] ID: {task_a.get('ID')}\nMJ_Prompt: {task_a.get('MJ_Prompt')}\n[PART B] ID: {task_b.get('ID')}\nMJ_Prompt: {task_b.get('MJ_Prompt')}\n"
    else:
        content = "".join(f"{header}: {task_a.get(header, '')}\n" for header in PRODUCTION_HEADERS)
    with open(os.path.join(save_path, "metadata.txt"), "w", encoding="utf-8") as f: f.write(content)

# --- 2. 生产引擎 ---
def run_logic():
    print(f"\n{'='*20} V18.3 FULL COVERAGE MODE {'='*20}")
    _validate_runtime_config()
    
    # 模式选择
    for key, obj in Registry.CATALOG.items(): print(f"[{key}] {obj.name}")
    type_key = input("大类: ") or "1"
    selected_prod = Registry.CATALOG.get(type_key, Registry.STICKER)
    for i, spec in enumerate(selected_prod.specs, 1): print(f"[{i}] {spec}")
    spec_idx = int(input("制式: ") or "1") - 1
    sel_cat, sel_spec = selected_prod.name, selected_prod.specs[spec_idx]
    is_kiss_cut = (sel_cat == "Sticker" and "Kiss-Cut" in sel_spec)

    # 任务初始化
    try:
        raw_list = _read_ready_tasks(sel_cat)
        selected_ids = _selected_harvest_ids()
        if selected_ids:
            raw_list = [task for task in raw_list if str(task.get("ID")) in selected_ids]
        style_map = {}
        for t in raw_list:
            s = t.get("Style", "Default"); style_map.setdefault(s, []).append(t)
        
        batch_queue = []
        harvest_limit = int(os.getenv("MJ_HARVEST_LIMIT", "0") or "0")
        for style, tasks in style_map.items():
            while tasks:
                if is_kiss_cut:
                    if tasks:
                        batch_queue.append([tasks.pop(0)])
                    else:
                        # 妥善处理末尾孤儿
                        orphan = tasks.pop(0)
                        _update_product_line_status(orphan['ID'], "Defeated_Orphan")
                        print(f"⚠️ [Orphan] {orphan['ID']} 样式配对失败，已排除。")
                else:
                    batch_queue.append([tasks.pop(0)])
                if harvest_limit and len(batch_queue) >= harvest_limit:
                    break
            if harvest_limit and len(batch_queue) >= harvest_limit:
                break
        print(f"📦 预备批次: {len(batch_queue)}")
    except Exception as exc:
        print(f"[ERROR] Failed to load Production_Line.xlsx tasks: {exc}")
        return

    active_pool = {}
    used_grid_message_ids = set()
    MAX_PARALLEL = int(os.getenv("MJ_HARVEST_MAX_PARALLEL", "3") or "3")

    while batch_queue or active_pool:
        # A. 投放逻辑
        while len(active_pool) < MAX_PARALLEL and batch_queue:
            current_batch = batch_queue.pop(0)
            for task in current_batch:
                tid = str(task['ID'])
                save_path = os.path.join("Output", sel_cat, sel_spec, f"{tid}-Review")
                if not os.path.exists(save_path): os.makedirs(save_path)
                p = task['MJ_Prompt'].strip()
                full_p = _prepare_discord_prompt(p, tid, sel_cat)
                prompt_signature = _prompt_signature(full_p)
                existing = _find_existing_task_message(tid, prompt_signature)
                if existing:
                    active_pool[tid] = {"path": save_path, "u_triggered": False, "u_received": set(), "start_time": _message_unix_time(existing) or time.time(), "queue_seen": time.time(), "task_obj": task, "batch": current_batch, "grid_done": False, "grid_message_id": "", "prompt_signature": prompt_signature}
                    if _is_grid_message(existing) and existing.get("attachments"):
                        if _download_asset(existing["attachments"][0], save_path, f"{tid}_Grid.png", MIN_GRID_DIM, "Grid"):
                            active_pool[tid]["grid_done"] = True
                            active_pool[tid]["grid_message_id"] = str(existing.get("id") or "")
                            active_pool[tid]["grid_time"] = time.time()
                            active_pool[tid]["u_triggered"] = _trigger_upscales(existing)
                    print(f"[Reattach] {tid} existing Discord task found.")
                    time.sleep(2)
                    continue
                
                try:
                    deployed = _interaction({"type": 2, "application_id": Config.APP_ID, "guild_id": Config.GUILD_ID, "channel_id": Config.CHANNEL_ID, "session_id": Config.SESSION_ID, "data": {"version": Config.MJ_VERSION, "id": Config.MJ_ID, "name": "imagine", "type": 1, "options": [{"type": 3, "name": "prompt", "value": full_p}]}})
                except RuntimeError as exc:
                    print(f"[Network-Hold] {tid}: {exc}")
                    batch_queue.insert(0, [task])
                    if os.path.exists(save_path):
                        shutil.rmtree(save_path)
                    time.sleep(30)
                    break
                if deployed:
                    active_pool[tid] = {"path": save_path, "u_triggered": False, "u_received": set(), "start_time": time.time(), "queue_seen": 0, "task_obj": task, "batch": current_batch, "grid_done": False, "grid_message_id": "", "prompt_signature": prompt_signature}
                    print(f"🎬 [{tid}] Deployed.")
                else:
                    _purge_asset(save_path, tid, "Defeated_Prompt", "Discord imagine interaction failed", task)
                time.sleep(5)

        # B. 监听与物理收割
        if active_pool:
            time.sleep(POLL_INTERVAL_SECONDS)
            try:
                msgs = requests.get(f"https://discord.com/api/v9/channels/{Config.CHANNEL_ID}/messages?limit=100", headers={"Authorization": Config.TOKEN}, timeout=10).json()
            except: continue

            now = time.time()
            msgs_chrono = sorted(msgs, key=lambda item: int(item.get("id") or 0))
            claimed_grid_ids = {
                str(item.get("grid_message_id"))
                for item in active_pool.values()
                if item.get("grid_message_id")
            }
            claimed_grid_ids.update(used_grid_message_ids)
            for m in msgs_chrono:
                mid = str(m.get("id") or "")
                if mid in claimed_grid_ids or not _is_grid_message(m):
                    continue
                msg_time = _message_unix_time(m)
                for tid, info in sorted(active_pool.items(), key=lambda item: item[1]["start_time"]):
                    if info["grid_done"]:
                        continue
                    if msg_time and msg_time < info["start_time"] - 10:
                        continue
                    if _message_matches_task(m, info, tid):
                        if _is_incomplete_midjourney_message(m):
                            continue
                        if _download_asset(m["attachments"][0], info["path"], f"{tid}_Grid.png", MIN_GRID_DIM, "Grid"):
                            info["grid_done"] = True
                            info["grid_message_id"] = mid
                            info["grid_time"] = now
                            claimed_grid_ids.add(mid)
                            used_grid_message_ids.add(mid)
                            if not info["u_triggered"]:
                                info["u_triggered"] = _trigger_upscales(m)
                            print(f"[Grid-Bind] {tid} <= {mid}")
                        break

            for tid in list(active_pool.keys()):
                # 核心修复：检查 ID 是否还在池子中
                if tid not in active_pool: continue
                
                info = active_pool[tid]
                for queued_message in msgs:
                    queued_content = queued_message.get("content", "")
                    if _message_matches_task(queued_message, info, tid) and "Waiting to start" in queued_content:
                        info["queue_seen"] = now
                        break
                timeout_limit = QUEUE_TIMEOUT_SECONDS if info.get("queue_seen") else HARVEST_TIMEOUT_SECONDS
                if now - info["start_time"] > timeout_limit:
                    _purge_asset(info["path"], tid, "Defeated_Timeout", f"No complete harvest within {timeout_limit}s", info.get("task_obj")); del active_pool[tid]; continue

                for m in msgs:
                    content = m.get("content", "")
                    ref_matches_grid = str(info.get("grid_message_id") or "") and _message_reference_id(m) == str(info.get("grid_message_id") or "")
                    if not ref_matches_grid and _message_unix_time(m) and _message_unix_time(m) < info["start_time"] - 10:
                        continue
                    if _message_matches_task(m, info, tid):
                        if _is_incomplete_midjourney_message(m):
                            continue
                        # Grid 捕获
                        if "Image #" not in content and m.get("attachments") and not info["grid_done"]:
                            if _download_asset(m["attachments"][0], info["path"], f"{tid}_Grid.png", MIN_GRID_DIM, "Grid"):
                                info["grid_done"] = True
                                info["grid_message_id"] = str(m.get("id") or "")
                                info["grid_time"] = now
                                used_grid_message_ids.add(str(m.get("id") or ""))
                        if info["grid_done"] and not info["u_triggered"] and str(m.get("id") or "") == str(info.get("grid_message_id") or ""):
                            info["u_triggered"] = _trigger_upscales(m)
                        # U图 捕获
                        if "Image #" in content and m.get("attachments"):
                            u_idx = content.split("Image #")[-1][0]
                            if u_idx in {"1", "2", "3", "4"} and u_idx not in info["u_received"]:
                                u_name = f"{tid}_U{u_idx}_Grid.png" if is_kiss_cut else f"{tid}_U{u_idx}.png"
                                min_dim = MIN_UPSCALE_DIM
                                if _download_asset(m["attachments"][0], info["path"], u_name, min_dim, f"U{u_idx}"):
                                    info["u_received"].add(u_idx)

                if is_kiss_cut and info["grid_done"] and not info["u_received"]:
                    grid_time = info.get("grid_time") or info["start_time"]
                    if now - grid_time >= GRID_FALLBACK_SECONDS:
                        _split_grid_as_upscales(info, tid)

                # 判定完成
                success = info["grid_done"] and len(info["u_received"]) >= 4
                if success:
                    if is_kiss_cut:
                        _finalize_kiss_cut(info, tid)
                        del active_pool[tid]
                    else:
                        _finalize_single_art(info, tid, sel_cat, sel_spec)
                        del active_pool[tid]

    print("\n🏁 全覆盖流程结束。")

if __name__ == "__main__":
    run_logic()
