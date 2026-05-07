import argparse
import csv
import hashlib
import io
import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.resilient_http import request_with_retry


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
DATABASE_DIR = PROJECT_ROOT / "Database"
DESIGN_MISMATCH_STATUS = "Printify_DesignMismatch"
STABLE_AUDIT_CSV = DATABASE_DIR / "Printify_Production_Design_Audit.csv"
AUDIT_STATUSES = {
    "Printify_UI_Mockups3",
    "Printify_UI_Mockups5",
    "Printify_UI_Mockups4",
    "Printify_UI_Mockups8",
    "Printify_Published_Mockups3",
    "Printify_Published_Mockups5",
    "Printify_Published_Mockups4",
    "Printify_Published_Mockups8",
    "Printify_PrimaryFix_Needed",
    "Printify_BaseStaged_DefaultMockups3",
    "Printify_UI_Failed",
}
MOCKUP_STATUS_RE = re.compile(r"^Printify_(UI|Published)_Mockups\d+$")


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}"}


def _fetch_product(product_id):
    response = request_with_retry(
        "GET",
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        timeout=120,
        attempts=3,
        backoff=2.5,
    )
    response.raise_for_status()
    return response.json()


def _front_print_image(product):
    for print_area in product.get("print_areas") or []:
        for placeholder in print_area.get("placeholders") or []:
            if placeholder.get("position") != "front":
                continue
            images = placeholder.get("images") or []
            if images:
                return images[0]
    return None


def _sha256_bytes(data):
    return hashlib.sha256(data).hexdigest()


def _remote_bytes(url):
    response = request_with_retry("GET", url, timeout=120, attempts=3, backoff=2.5)
    response.raise_for_status()
    return response.content


def _image_size_from_bytes(data):
    with Image.open(io.BytesIO(data)) as image:
        image.load()
        return image.size


def _ahash_from_bytes(data):
    with Image.open(io.BytesIO(data)) as image:
        image = _white_composite(image).convert("L").resize((16, 16), Image.Resampling.LANCZOS)
        pixels = list(image.tobytes())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if pixel > avg else "0" for pixel in pixels)


def _white_composite(image):
    image = image.convert("RGBA")
    background = Image.new("RGBA", image.size, "white")
    background.alpha_composite(image)
    return background.convert("RGB")


def _distance(left, right):
    return sum(a != b for a, b in zip(left, right))


def product_design_report(product_id, local_path):
    local_path = Path(local_path)
    if not local_path.exists():
        raise FileNotFoundError(f"Missing local Production_Design: {local_path}")
    product = _fetch_product(product_id)
    image = _front_print_image(product)
    if not image or not image.get("src"):
        raise RuntimeError(f"Printify product {product_id} has no front print-area image")
    local_bytes = local_path.read_bytes()
    remote = _remote_bytes(image["src"])
    local_sha = _sha256_bytes(local_bytes)
    remote_sha = _sha256_bytes(remote)
    local_hash = _ahash_from_bytes(local_bytes)
    remote_hash = _ahash_from_bytes(remote)
    return {
        "product_id": product_id,
        "local_path": str(local_path),
        "remote_src": image["src"],
        "remote_image_id": image.get("id") or image.get("imageId") or "",
        "exact_sha_match": local_sha == remote_sha,
        "local_sha256": local_sha,
        "remote_sha256": remote_sha,
        "local_size": _image_size_from_bytes(local_bytes),
        "remote_size": _image_size_from_bytes(remote),
        "remote_declared_size": (image.get("width"), image.get("height")),
        "ahash_distance": _distance(local_hash, remote_hash),
    }


def _visual_match(report):
    return bool(
        report["exact_sha_match"]
        or (
            report["ahash_distance"] <= 5
            and (
                report["local_size"] == report["remote_size"]
                or tuple(report["local_size"]) == tuple(report.get("remote_declared_size") or ())
            )
        )
    )


def assert_product_design_matches(product_id, local_path, require_exact=False):
    report = product_design_report(product_id, local_path)
    report["visual_match"] = _visual_match(report)
    if require_exact and not report["exact_sha_match"]:
        raise RuntimeError(
            "Production design byte mismatch: "
            f"sha local={report['local_sha256'][:12]} remote={report['remote_sha256'][:12]} "
            f"size local={report['local_size']} remote={report['remote_size']} declared={report['remote_declared_size']} "
            f"ahash_distance={report['ahash_distance']}"
        )
    if not report["visual_match"]:
        raise RuntimeError(
            "Production design mismatch: "
            f"sha local={report['local_sha256'][:12]} remote={report['remote_sha256'][:12]} "
            f"size local={report['local_size']} remote={report['remote_size']} "
            f"ahash_distance={report['ahash_distance']}"
        )
    return report


def _product_type_text(value):
    text = str(value or "").strip()
    return text


def audit_workbook(limit=0, mark=False, product_type=None, ids=None, sleep_seconds=0.0, stable_output=True):
    ids = {str(item).strip() for item in (ids or []) if str(item).strip()}
    product_type = str(product_type or "").strip()
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    if "Status" not in cols or "Printify_Product_ID" not in cols or "Production_Path" not in cols:
        wb.close()
        raise RuntimeError("eBay listing workbook is missing required audit columns")
    out_path = STABLE_AUDIT_CSV if stable_output else DATABASE_DIR / f"printify_design_audit_{time.strftime('%Y%m%d_%H%M%S')}.csv"
    checked = 0
    mismatches = 0
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "ID",
                "Product_Type",
                "Status",
                "Printify_Product_ID",
                "exact_sha_match",
                "visual_match",
                "local_size",
                "remote_size",
                "remote_declared_size",
                "ahash_distance",
                "local_sha256",
                "remote_sha256",
                "remote_image_id",
                "error",
            ],
        )
        writer.writeheader()
        for row in range(2, ws.max_row + 1):
            item_id = ws.cell(row, cols["ID"]).value
            row_product_type = ws.cell(row, cols["Product_Type"]).value if "Product_Type" in cols else ""
            status = ws.cell(row, cols["Status"]).value
            product_id = ws.cell(row, cols["Printify_Product_ID"]).value
            production_path = ws.cell(row, cols["Production_Path"]).value
            if not item_id or not product_id or not production_path:
                continue
            if ids and str(item_id) not in ids:
                continue
            if product_type and _product_type_text(row_product_type).lower() != product_type.lower():
                continue
            status_text = str(status or "")
            if (
                status_text not in AUDIT_STATUSES
                and status_text != DESIGN_MISMATCH_STATUS
                and not MOCKUP_STATUS_RE.match(status_text)
            ):
                continue
            record = {
                "ID": item_id,
                "Product_Type": row_product_type,
                "Status": status,
                "Printify_Product_ID": product_id,
                "error": "",
            }
            try:
                report = product_design_report(str(product_id), production_path)
                report["visual_match"] = _visual_match(report)
                record.update(report)
                if report["visual_match"]:
                    print(
                        f"[DESIGN-OK] {item_id} visual_match=True "
                        f"exact_sha={report['exact_sha_match']} size={report['local_size']} "
                        f"remote={report['remote_size']} declared={report['remote_declared_size']} "
                        f"ahash_distance={report['ahash_distance']}"
                    )
                else:
                    mismatches += 1
                    if mark:
                        ws.cell(row, cols["Status"]).value = DESIGN_MISMATCH_STATUS
                    print(
                        f"[DESIGN-MISMATCH] {item_id} exact_sha=False "
                        f"local={report['local_size']} remote={report['remote_size']} "
                        f"ahash_distance={report['ahash_distance']}"
                    )
            except Exception as exc:
                mismatches += 1
                record["error"] = str(exc)
                if mark:
                    ws.cell(row, cols["Status"]).value = DESIGN_MISMATCH_STATUS
                print(f"[DESIGN-ERROR] {item_id}: {exc}")
            writer.writerow({key: record.get(key, "") for key in writer.fieldnames})
            checked += 1
            if sleep_seconds > 0:
                time.sleep(sleep_seconds)
            if limit and checked >= limit:
                break
    if mark:
        wb.save(EBAY_BOOK)
    wb.close()
    print(f"[DONE] design audit checked={checked} mismatches={mismatches} report={out_path}")
    return mismatches


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--mark", action="store_true")
    parser.add_argument("--product-type", default="")
    parser.add_argument("--ids", default="", help="Comma-separated workbook IDs.")
    parser.add_argument("--sleep-seconds", type=float, default=0.0)
    parser.add_argument("--timestamped-output", action="store_true")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()]
    audit_workbook(
        limit=args.limit,
        mark=args.mark,
        product_type=args.product_type,
        ids=ids,
        sleep_seconds=args.sleep_seconds,
        stable_output=not args.timestamped_output,
    )
