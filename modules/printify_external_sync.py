import argparse
import csv
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config

EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
LOG_PATH = PROJECT_ROOT / "Database" / "printify_external_sync.csv"


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}", "Content-Type": "application/json"}


def _fetch_product(product_id):
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        timeout=90,
    )
    response.raise_for_status()
    return response.json()


def _ensure_column(sheet, headers, name):
    if name not in headers:
        sheet.cell(1, sheet.max_column + 1).value = name
        headers.append(name)
    return headers.index(name) + 1


def sync(limit=0, sleep_seconds=3, dry_run=False):
    workbook = load_workbook(EBAY_BOOK)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    ebay_col = _ensure_column(sheet, headers, "eBay_Item_ID")
    external_url_col = _ensure_column(sheet, headers, "eBay_Item_URL")
    external_type_col = _ensure_column(sheet, headers, "External_Type")
    external_sync_col = _ensure_column(sheet, headers, "External_Sync_Timestamp")
    cols = {header: index + 1 for index, header in enumerate(headers)}
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row_idx, cols["Status"]).value or "")
        product_id = str(sheet.cell(row_idx, cols["Printify_Product_ID"]).value or "").strip()
        if not status.startswith("Printify_Published") or not product_id:
            continue
        if str(sheet.cell(row_idx, ebay_col).value or "").strip():
            continue
        rows.append((row_idx, sheet.cell(row_idx, cols["ID"]).value, product_id))
        if limit and len(rows) >= limit:
            break
    log_exists = LOG_PATH.exists()
    with LOG_PATH.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["Timestamp", "ID", "Printify_Product_ID", "eBay_Item_ID", "eBay_Item_URL", "Status", "Error"],
        )
        if not log_exists:
            writer.writeheader()
        updated = 0
        for row_idx, item_id, product_id in rows:
            try:
                product = _fetch_product(product_id)
                external = product.get("external") or {}
                ebay_id = str(external.get("id") or "").strip()
                handle_url = str(external.get("handle") or "").strip()
                external_type = str(external.get("type") or "").strip()
                if not dry_run and ebay_id:
                    sheet.cell(row_idx, ebay_col).value = ebay_id
                    sheet.cell(row_idx, external_url_col).value = handle_url
                    sheet.cell(row_idx, external_type_col).value = external_type
                    sheet.cell(row_idx, external_sync_col).value = datetime.now()
                    workbook.save(EBAY_BOOK)
                updated += 1 if ebay_id else 0
                writer.writerow(
                    {
                        "Timestamp": datetime.now().isoformat(timespec="seconds"),
                        "ID": item_id,
                        "Printify_Product_ID": product_id,
                        "eBay_Item_ID": ebay_id,
                        "eBay_Item_URL": handle_url,
                        "Status": "OK" if ebay_id else "MISSING_EXTERNAL_ID",
                        "Error": "",
                    }
                )
                print(f"[EXTERNAL-SYNC] {item_id} product={product_id} ebay={ebay_id or 'MISSING'}")
            except Exception as exc:
                writer.writerow(
                    {
                        "Timestamp": datetime.now().isoformat(timespec="seconds"),
                        "ID": item_id,
                        "Printify_Product_ID": product_id,
                        "eBay_Item_ID": "",
                        "eBay_Item_URL": "",
                        "Status": "ERROR",
                        "Error": str(exc)[:500],
                    }
                )
                print(f"[EXTERNAL-SYNC-FAIL] {item_id}: {exc}")
            time.sleep(max(0, sleep_seconds))
    workbook.close()
    print(f"[DONE] candidates={len(rows)} updated={updated} dry_run={dry_run}")
    return updated


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--sleep-seconds", type=float, default=3)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    sync(limit=args.limit, sleep_seconds=args.sleep_seconds, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
