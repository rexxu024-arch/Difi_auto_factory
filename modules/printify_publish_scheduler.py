import argparse
import random
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules import ebay_ads_standard
from modules.risk_guard import assert_allowed


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
PUBLISHABLE_PREFIXES = ("Printify_UI_Mockups",)
RETRYABLE_EXTERNAL_PENDING_PREFIXES = ("Printify_PublishExternalPending_Mockups",)
PUBLISHED_PREFIXES = ("Printify_Published", "Printify_Published_Mockups")
PUBLISH_BODY = {
    "title": True,
    "description": True,
    "images": True,
    "variants": True,
    "tags": True,
    "keyFeatures": True,
    "shipping_template": True,
}


def _clean_text(value):
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def _headers():
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def _product_type(value):
    text = str(value or "").strip().lower()
    if text.startswith("poster"):
        return "Poster"
    if text.startswith("acry"):
        return "Acrylic"
    if text.startswith("stick"):
        return "Sticker"
    return "Other"


def _publish_suffix(status):
    text = str(status or "")
    if "Mockups" in text:
        return text.split("Mockups", 1)[1]
    return ""


def _fetch_product(product_id):
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers={"Authorization": _headers()["Authorization"]},
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def _selected_count(product):
    return sum(1 for image in product.get("images") or [] if image.get("is_selected_for_publishing") is not False)


def _selected_images(product):
    return [
        image
        for image in product.get("images") or []
        if image.get("is_selected_for_publishing") is not False
    ]


def _image_payload(image, *, selected=None, default=None):
    payload = {
        key: value
        for key, value in image.items()
        if key in {
            "id",
            "mockup_id",
            "src",
            "variant_ids",
            "position",
            "is_default",
            "is_selected_for_publishing",
            "order",
        }
    }
    if selected is not None:
        payload["is_selected_for_publishing"] = selected
    if default is not None:
        payload["is_default"] = default
    return payload


def _put_product(product_id, product):
    response = requests.put(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        json=product,
        timeout=180,
    )
    response.raise_for_status()
    return response.json() if response.content else _fetch_product(product_id)


def _put_product_partial(product_id, payload):
    response = requests.put(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        json=payload,
        timeout=180,
    )
    response.raise_for_status()
    return _fetch_product(product_id)


def _ensure_clean_clone_public_metadata(product_id, product, row, *, dry_run=False):
    """Prevent eBay clean clones from inheriting neutral Printify titles.

    For normal products, Printify creation may use concise production metadata.
    Clean-clone eBay experiments are different: their whole purpose is to test
    buyer-facing title/description/brand/shipping repairs. If they publish with
    the neutral production title, the readback experiment is invalid and likely
    recreates the low-traffic problem.
    """
    item_id = _clean_text(row.get("ID"))
    if "CLEAN_CLONE" not in item_id:
        return product, "not_clean_clone"
    local_title = _clean_text(row.get("Title"))
    local_description = _clean_text(row.get("Description"))
    product_title = _clean_text(product.get("title"))
    product_description = _clean_text(product.get("description"))
    needs_title = bool(local_title) and product_title != local_title
    needs_description = bool(local_description) and product_description != local_description
    free_shipping = bool((product.get("sales_channel_properties") or {}).get("free_shipping"))
    if not (needs_title or needs_description or not free_shipping):
        return product, "clean_clone_public_metadata_ready"
    if dry_run:
        pieces = []
        if needs_title:
            pieces.append("title")
        if needs_description:
            pieces.append("description")
        if not free_shipping:
            pieces.append("free_shipping")
        return product, "clean_clone_metadata_would_sync:" + ",".join(pieces)
    payload = {
        "sales_channel_properties": {"free_shipping": True},
    }
    if local_title:
        payload["title"] = local_title
    if local_description:
        payload["description"] = local_description
    refreshed = _put_product_partial(product_id, payload)
    return refreshed, "clean_clone_public_metadata_synced"


def _image_kind(image):
    src = str(image.get("src") or "")
    if "images.printify.com/mockup" in src:
        return "official"
    if "pfy-prod-products-mockup-media" in src:
        return "custom"
    return "other"


def _camera_label(image):
    src = str(image.get("src") or "")
    if "camera_label=" not in src:
        return ""
    return src.split("camera_label=", 1)[1].split("&", 1)[0]


def _image_key(image):
    return str(image.get("mockup_id") or image.get("id") or image.get("src") or "")


def _set_default_outside_gallery(images, default_image, gallery_images):
    default_key = _image_key(default_image)
    gallery_counts = Counter(_image_key(image) for image in gallery_images)
    repaired = []
    order = 0
    for image in images:
        item = dict(image)
        key = _image_key(image)
        is_default = key == default_key
        is_selected = gallery_counts.get(key, 0) > 0 and not is_default
        item["is_default"] = is_default
        item["is_selected_for_publishing"] = is_selected
        if is_selected:
            gallery_counts[key] -= 1
            item["order"] = order
            order += 1
        else:
            item["order"] = None
        repaired.append(item)
    return repaired


def _ensure_non_default_gallery(product_id, product, product_type):
    """Keep the official default product photo out of the extra gallery.

    Rex's current standard is to keep the official product-context/default mockup
    so buyers understand the physical product. eBay, however, renders that
    default as the primary photo and then repeats it if it is also selected for
    publishing. For Poster/Acrylic we leave it as default but remove it from the
    selected gallery, then publish only non-default official mockups.
    """
    if product_type not in {"Poster", "Acrylic"}:
        return product, "not_needed"
    images = product.get("images") or []
    official = [image for image in images if _image_kind(image) == "official"]
    if len(official) < 3:
        return product, "not_enough_official_mockups"
    preferred = {
        "Poster": ["front", "close-up", "context-1", "context-2"],
        "Acrylic": ["front", "back", "side-1", "side-2", "context-1"],
    }.get(product_type, [])
    picked = []
    for label in preferred:
        for image in official:
            if image not in picked and _camera_label(image) == label:
                picked.append(image)
                break
    for image in official:
        if len(picked) >= 4:
            break
        if image not in picked:
            picked.append(image)
    default_image = next((image for image in official if image.get("is_default")), None) or picked[0]
    gallery = [image for image in picked if _image_key(image) != _image_key(default_image)]
    if len(gallery) < 2:
        return product, "not_enough_non_default_mockups"
    repaired_images = _set_default_outside_gallery(images, default_image, gallery[:3])
    before = [
        (image.get("is_default"), image.get("is_selected_for_publishing"), image.get("order"))
        for image in images
    ]
    after = [
        (image.get("is_default"), image.get("is_selected_for_publishing"), image.get("order"))
        for image in repaired_images
    ]
    if before == after:
        return product, "default_already_outside_gallery"
    # Printify currently ignores/re-expands image selection flags for some
    # live products when written through the product PUT endpoint. The safe
    # path is UI source repair/rebuild, then this scheduler can publish.
    return product, f"ui_source_repair_required_default_selected selected={len(gallery[:3])}"


def _ensure_column(ws, cols, name):
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def _sync_external_id_for_row(ws, cols, row_idx, item_id, product_id, attempts=5, delay=12):
    ebay_col = _ensure_column(ws, cols, "eBay_Item_ID")
    url_col = _ensure_column(ws, cols, "eBay_Item_URL")
    type_col = _ensure_column(ws, cols, "External_Type")
    sync_col = _ensure_column(ws, cols, "External_Sync_Timestamp")
    existing = str(ws.cell(row_idx, ebay_col).value or "").strip()
    if existing:
        return existing, "existing"
    for attempt in range(1, attempts + 1):
        product = _fetch_product(product_id)
        external = product.get("external") or {}
        ebay_id = str(external.get("id") or "").strip()
        if ebay_id:
            ws.cell(row_idx, ebay_col).value = ebay_id
            ws.cell(row_idx, url_col).value = str(external.get("handle") or "").strip()
            ws.cell(row_idx, type_col).value = str(external.get("type") or "").strip()
            ws.cell(row_idx, sync_col).value = datetime.now()
            return ebay_id, f"synced_attempt_{attempt}"
        if attempt < attempts:
            time.sleep(delay)
    return "", "missing_external_id"


def _preflight(row):
    product_id = str(row.get("Printify_Product_ID") or "").strip()
    if not product_id:
        return False, "missing Printify_Product_ID"
    product = _fetch_product(product_id)
    if not product.get("print_areas"):
        return False, "missing print_areas"
    selected_images = _selected_images(product)
    selected = len(selected_images)
    selected_srcs = [str(image.get("src") or "") for image in selected_images]
    if len(set(selected_srcs)) != len(selected_srcs):
        return False, f"selected gallery contains duplicate image URLs: selected={selected}, unique={len(set(selected_srcs))}"
    defaults = [image for image in selected_images if image.get("is_default")]
    default_count = sum(1 for image in product.get("images") or [] if image.get("is_default"))
    product_type = _product_type(row.get("Product_Type"))
    enabled_variant_ids = {
        int(variant.get("id"))
        for variant in product.get("variants") or []
        if variant.get("is_enabled")
    }
    disabled_selected = []
    for image in selected_images:
        variant_ids = {int(variant_id) for variant_id in image.get("variant_ids") or []}
        if variant_ids and not (variant_ids & enabled_variant_ids):
            disabled_selected.append(image)
    if disabled_selected:
        return False, (
            "selected gallery includes mockups for disabled variants; "
            f"disabled_selected={len(disabled_selected)}. Rebuild product/source mockups before publish"
        )
    if product_type == "Sticker" and selected < 4:
        return False, (
            f"selected mockups={selected}, expected >=4 distinct buyer-facing images; "
            "3 official sticker mockups create repeated eBay picture slots"
        )
    if product_type == "Sticker":
        official_gallery = [
            image for image in selected_images
            if "images.printify.com/mockup" in str(image.get("src") or "")
        ]
        custom_gallery = [
            image for image in selected_images
            if "pfy-prod-products-mockup-media" in str(image.get("src") or "")
        ]
        if custom_gallery and not (len(custom_gallery) == 1 and len(official_gallery) >= 3 and selected >= 4):
            return False, (
                f"sticker custom gallery images selected={len(custom_gallery)}; "
                "only 1 custom Cover plus at least 3 official mockups is allowed"
            )
    if product_type == "Poster" and selected < 3:
        return False, f"selected mockups={selected}, expected >=3 non-default gallery mockups plus default product image"
    if product_type == "Poster" and any("pfy-prod-products-mockup-media" in src for src in selected_srcs):
        return False, "poster custom gallery images selected; use official product mockups only"
    if product_type == "Acrylic" and selected < 3:
        return False, f"selected mockups={selected}, expected >=3 non-default gallery mockups plus default product image"
    if product_type == "Acrylic" and any("pfy-prod-products-mockup-media" in src for src in selected_srcs):
        return False, "acrylic custom gallery images selected; use official product mockups only"
    if product_type in {"Poster", "Acrylic"}:
        if defaults:
            return False, "default product image is selected in gallery; would duplicate primary image on eBay"
        official_selected = [
            src for src in selected_srcs
            if "images.printify.com/mockup" in src
        ]
        if len(official_selected) != selected or len(set(selected_srcs)) != selected:
            return False, "gallery includes non-official or duplicate image URLs"
    if product_type in {"Poster", "Acrylic"} and default_count < 1:
        return False, "default image count=0, expected one product default before publish"
    if product_type == "Sticker" and len(defaults) < 1:
        return False, "default image count=0, expected at least 1 before publish"
    return True, f"selected mockups={selected}, selected_defaults={len(defaults)}, product_defaults={default_count}"


def _publish(product_id):
    last_error = None
    for attempt in range(1, 4):
        try:
            response = requests.post(
                f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
                headers=_headers(),
                json=PUBLISH_BODY,
                timeout=180,
            )
            if response.status_code in {200, 201, 202, 204}:
                return response.status_code
            response.raise_for_status()
        except Exception as exc:
            last_error = exc
            if attempt < 3:
                time.sleep(8 * attempt)
    raise last_error


def _load_publishable(limit, product_cycle, ids=None, retry_pending=False):
    wanted_ids = {str(item).strip() for item in (ids or []) if str(item).strip()}
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    if "Publish_Timestamp" not in cols:
        ws.cell(1, ws.max_column + 1).value = "Publish_Timestamp"
        cols["Publish_Timestamp"] = ws.max_column
    buckets = {product_type: [] for product_type in product_cycle}
    for row_idx in range(2, ws.max_row + 1):
        row_id = str(ws.cell(row_idx, cols["ID"]).value or "").strip()
        if wanted_ids and row_id not in wanted_ids:
            continue
        status = str(ws.cell(row_idx, cols["Status"]).value or "")
        allowed_prefixes = PUBLISHABLE_PREFIXES + (RETRYABLE_EXTERNAL_PENDING_PREFIXES if retry_pending else ())
        if status.startswith(PUBLISHED_PREFIXES) or not status.startswith(allowed_prefixes):
            continue
        product_type = _product_type(ws.cell(row_idx, cols["Product_Type"]).value)
        if product_type not in buckets:
            continue
        row = {header: ws.cell(row_idx, cols[header]).value for header in headers if header in cols}
        row["_row_idx"] = row_idx
        buckets[product_type].append(row)

    selected = []
    # Load extra candidates because preflight guards may hold early rows. The
    # caller stops after `limit` successful publishes; guarded skips should not
    # make the daily drip look complete.
    scan_limit = max(limit, limit * 4)
    while len(selected) < scan_limit and any(buckets.values()):
        for product_type in product_cycle:
            if buckets[product_type] and len(selected) < scan_limit:
                selected.append(buckets[product_type].pop(0))
    return wb, ws, cols, selected


def _mark_publish_hold(ws, cols, row_idx, note):
    status_col = cols.get("Status")
    if status_col:
        ws.cell(row_idx, status_col).value = "Printify_PublishHold_MockupGuard"
    hold_col = _ensure_column(ws, cols, "Publish_Hold_Note")
    hold_ts_col = _ensure_column(ws, cols, "Publish_Hold_Timestamp")
    ws.cell(row_idx, hold_col).value = note
    ws.cell(row_idx, hold_ts_col).value = datetime.now()


def run(limit=8, min_delay=90, max_delay=240, product_cycle=None, dry_run=False, ids=None, retry_pending=False):
    if not dry_run:
        assert_allowed("ebay", "paid_publish")
    product_cycle = product_cycle or ["Poster", "Acrylic", "Sticker"]
    wb, ws, cols, rows = _load_publishable(limit, product_cycle, ids=ids, retry_pending=retry_pending)
    done = 0
    try:
        for row in rows:
            item_id = row["ID"]
            product_id = str(row.get("Printify_Product_ID") or "").strip()
            row_idx = row["_row_idx"]
            try:
                product_type = _product_type(row.get("Product_Type"))
                product = _fetch_product(product_id)
                product, metadata_note = _ensure_clean_clone_public_metadata(
                    product_id,
                    product,
                    row,
                    dry_run=dry_run,
                )
                if metadata_note not in {"not_clean_clone", "clean_clone_public_metadata_ready"}:
                    print(f"[METADATA-GATE] {item_id} {metadata_note}", flush=True)
                if dry_run:
                    if product_type in {"Poster", "Acrylic"}:
                        selected_defaults = sum(
                            1
                            for image in _selected_images(product)
                            if image.get("is_default")
                        )
                        if selected_defaults:
                            print(f"[GALLERY-DRY] {item_id} preserve_official_default", flush=True)
                else:
                    product, gallery_note = _ensure_non_default_gallery(product_id, product, product_type)
                    if gallery_note not in {"not_needed", "preserve_official_default"}:
                        print(f"[GALLERY-SANITIZE] {item_id} {gallery_note}", flush=True)
                ok, note = _preflight(row)
                if not ok:
                    print(f"[PUBLISH-SKIP] {item_id}: {note}")
                    if not dry_run:
                        _mark_publish_hold(ws, cols, row_idx, note)
                        wb.save(EBAY_BOOK)
                    continue
                if dry_run:
                    print(f"[PUBLISH-DRY] {item_id} product={product_id} {note}")
                    continue
                code = _publish(product_id)
                suffix = _publish_suffix(row.get("Status"))
                ws.cell(row_idx, cols["Publish_Timestamp"]).value = datetime.now()
                ebay_id, external_note = _sync_external_id_for_row(ws, cols, row_idx, item_id, product_id)
                if ebay_id:
                    ws.cell(row_idx, cols["Status"]).value = f"Printify_Published_Mockups{suffix}" if suffix else "Printify_Published"
                else:
                    ws.cell(row_idx, cols["Status"]).value = (
                        f"Printify_PublishExternalPending_Mockups{suffix}" if suffix else "Printify_PublishExternalPending"
                    )
                if ebay_id:
                    try:
                        ads_ok = ebay_ads_standard.enroll_listing(item_id, ebay_id)
                        ads_note = "ads_enrolled" if ads_ok else "ads_queued"
                    except Exception as ads_exc:
                        ads_note = f"ads_failed:{ads_exc}"
                else:
                    ads_note = "ads_waiting_for_external_id"
                done += 1 if ebay_id else 0
                wb.save(EBAY_BOOK)
                print(
                    f"[PUBLISH-OK] {item_id} product={product_id} http={code} {note} "
                    f"external={external_note} ebay={ebay_id or 'MISSING'} {ads_note}"
                )
                if done >= limit:
                    break
                if done < limit:
                    delay = random.randint(min_delay, max_delay)
                    print(f"[PUBLISH-SLEEP] {delay}s")
                    time.sleep(delay)
            except Exception as exc:
                print(f"[PUBLISH-FAIL] {item_id}: {exc}")
                continue
    finally:
        wb.close()
    print(f"[DONE] publish attempted={len(rows)} external_confirmed={done}")
    return done


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=8)
    parser.add_argument("--min-delay", type=int, default=90)
    parser.add_argument("--max-delay", type=int, default=240)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--cycle", default="Poster,Acrylic,Sticker")
    parser.add_argument("--ids", default="", help="Comma-separated listing IDs to publish exactly.")
    parser.add_argument("--retry-pending", action="store_true", help="Explicitly retry Printify_PublishExternalPending rows.")
    args = parser.parse_args()
    cycle = [part.strip() for part in args.cycle.split(",") if part.strip()]
    ids = [part.strip() for part in args.ids.split(",") if part.strip()]
    run(
        limit=args.limit,
        min_delay=args.min_delay,
        max_delay=args.max_delay,
        product_cycle=cycle,
        dry_run=args.dry_run,
        ids=ids,
        retry_pending=args.retry_pending,
    )


if __name__ == "__main__":
    main()
