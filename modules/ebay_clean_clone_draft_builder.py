from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


DATABASE = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
PLAN = DATABASE / "eBay_Clean_Clone_Experiment_Plan.csv"
LOG = DATABASE / "eBay_Clean_Clone_Draft_Log.csv"
NY = ZoneInfo("America/New_York")


CLEAR_FIELDS = {
    "Printify_Product_ID",
    "eBay_Item_ID",
    "eBay_Item_URL",
    "External_Type",
    "External_Sync_Timestamp",
    "Publish_Timestamp",
    "Metadata_Sync_Status",
}


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def read_plan() -> list[dict[str, str]]:
    if not PLAN.exists():
        raise FileNotFoundError(PLAN)
    with PLAN.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def append_log(rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = ["Timestamp", "Clone_ID", "Old_Item_ID", "Source_ID", "Status", "Detail"]
    exists = LOG.exists()
    with LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def clone_id(plan_row: dict[str, str], source_id: str) -> str:
    group = clean(plan_row.get("Experiment_Group")) or "CLEAN_CLONE"
    product_type = clean(plan_row.get("Product_Type")) or "Product"
    return f"{product_type}-{group}"


def build_description(plan_row: dict[str, str], source_description: str) -> str:
    title = clean(plan_row.get("Clone_Title_Target"))
    product_type = clean(plan_row.get("Product_Type"))
    brand = clean(plan_row.get("Brand_Target")) or "OpenClaw Design Studio"
    strategy = clean(plan_row.get("Clone_Strategy"))
    image_note = (
        "The main image shows the actual product customers receive. Additional gallery images "
        "are concept, detail, or collection-reference views and are not extra products or selectable variations."
    )
    if product_type == "Acrylic":
        product_note = "One 5x7 acrylic photo block, made on demand through Printify. Price is set to absorb Printify production and shipping cost."
        scene = "quiet luxury desks, collector shelves, reading nooks, and deep-work apartment corners"
    elif product_type == "Poster":
        product_note = "One matte wall poster, made on demand through Printify. Price is set to absorb Printify production and shipping cost."
        scene = "study walls, book rooms, dorm decor, gallery corners, and apartment art walls"
    else:
        product_note = "One physical product made on demand through Printify. Price is set to absorb Printify production and shipping cost."
        scene = "giftable desk setups and collector interiors"
    return (
        f"<h2>{title}</h2>"
        f"<p>{brand} clone-test edition built for higher-trust buyer presentation: clean free-shipping pricing, stronger brand surface, and clearer gallery expectations.</p>"
        f"<p><strong>Best for:</strong> {scene}.</p>"
        f"<ul><li><strong>Includes:</strong> {product_note}</li>"
        f"<li><strong>Experiment lane:</strong> {strategy}</li>"
        f"<li><strong>Gallery note:</strong> {image_note}</li></ul>"
        f"<p><small>Source concept retained for visual continuity. Original note: {clean(source_description)[:260]}</small></p>"
    )


def build(limit: int = 12, dry_run: bool = False) -> list[str]:
    plan_rows = [row for row in read_plan() if clean(row.get("Status")) == "READY_FOR_CLEAN_CLONE_DRAFT"]
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    required = {"ID", "SKU", "Title", "Description", "Price", "Status", "Timestamp", "eBay_Item_ID"}
    missing = sorted(required - set(cols))
    if missing:
        raise RuntimeError(f"Missing workbook columns: {missing}")
    existing_ids = {clean(ws.cell(row, cols["ID"]).value) for row in range(2, ws.max_row + 1)}
    source_by_ebay = {
        clean(ws.cell(row, cols["eBay_Item_ID"]).value): row
        for row in range(2, ws.max_row + 1)
        if clean(ws.cell(row, cols["eBay_Item_ID"]).value)
    }
    created: list[str] = []
    logs: list[dict[str, str]] = []
    try:
        for plan_row in plan_rows:
            if limit and len(created) >= limit:
                break
            old_item = clean(plan_row.get("Old_Item_ID"))
            source_idx = source_by_ebay.get(old_item)
            if not source_idx:
                print(f"[EBAY-CLEAN-CLONE-SKIP] source eBay item missing: {old_item}")
                continue
            source_id = clean(ws.cell(source_idx, cols["ID"]).value)
            new_id = clone_id(plan_row, source_id)
            if new_id in existing_ids:
                print(f"[EBAY-CLEAN-CLONE-SKIP] exists {new_id}")
                continue
            if dry_run:
                print(f"[EBAY-CLEAN-CLONE-DRY] {source_id}/{old_item} -> {new_id}")
                created.append(new_id)
                continue
            new_idx = ws.max_row + 1
            for header in headers:
                value = ws.cell(source_idx, cols[header]).value
                if header == "ID":
                    value = new_id
                elif header == "SKU":
                    value = new_id
                elif header == "Title":
                    value = clean(plan_row.get("Clone_Title_Target"))
                elif header == "Description":
                    value = build_description(plan_row, ws.cell(source_idx, cols[header]).value)
                elif header == "Price":
                    value = f"${float(clean(plan_row.get('Clone_Price_Free_Shipping_USD')) or 0):.2f}"
                elif header == "Status":
                    value = "Ready_for_Printify"
                elif header == "Timestamp":
                    value = now_text()
                elif header == "Quiet_Jade_Target":
                    value = clean(plan_row.get("Clone_Strategy"))
                elif header == "Quiet_Jade_Price_Decision":
                    value = "FREE_SHIPPING_PRICE_INCLUDES_PRINTIFY_SHIPPING"
                elif header in CLEAR_FIELDS:
                    value = ""
                ws.cell(new_idx, cols[header]).value = value
            existing_ids.add(new_id)
            created.append(new_id)
            logs.append(
                {
                    "Timestamp": now_text(),
                    "Clone_ID": new_id,
                    "Old_Item_ID": old_item,
                    "Source_ID": source_id,
                    "Status": "LOCAL_READY_FOR_PRINTIFY",
                    "Detail": "Local clean-clone draft row created; no marketplace write performed.",
                }
            )
            print(f"[EBAY-CLEAN-CLONE] {source_id}/{old_item} -> {new_id}")
        if created and not dry_run:
            wb.save(EBAY_BOOK)
            append_log(logs)
    finally:
        wb.close()
    print(f"[EBAY-CLEAN-CLONE-DONE] created={len(created)}")
    return created


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=12)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    build(limit=args.limit, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
