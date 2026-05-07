"""Compare live eBay listing main images against local Cover_Mockup and U1-U4 files.

This is intentionally read-only. It opens one dedicated Edge CDP tab, navigates item
pages one by one, extracts the visible main gallery image, downloads that image,
and compares it to local cover/U1-U4 assets.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import os
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
import websockets
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
COVER_QA_CSV = DATABASE_DIR / "eBay_Cover_QA.csv"
OUT_CSV = DATABASE_DIR / "eBay_Online_Cover_Audit.csv"
OUT_DIR = DATABASE_DIR / "eBay_Online_Cover_Audit"
IMAGE_DIR = OUT_DIR / "images"
DEFAULT_CDP_PORT = int(os.getenv("OPENCLAW_EBAY_CDP_PORT") or os.getenv("OPENCLAW_CDP_PORT") or "9223")


def now_text() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def http_json(url: str, method: str = "GET") -> dict[str, Any]:
    request = urllib.request.Request(url, method=method)
    with urllib.request.urlopen(request, timeout=10) as response:
        return json.load(response)


async def send(ws, state: dict[str, int], method: str, params: dict | None = None) -> dict:
    state["seq"] += 1
    msg = {"id": state["seq"], "method": method}
    if params is not None:
        msg["params"] = params
    await ws.send(json.dumps(msg))
    while True:
        data = json.loads(await ws.recv())
        if data.get("id") == msg["id"]:
            return data


async def open_tab(cdp_port: int, url: str) -> dict:
    encoded = urllib.parse.quote(url, safe=":/?&=%")
    return http_json(f"http://127.0.0.1:{cdp_port}/json/new?{encoded}", method="PUT")


def close_tab(cdp_port: int, tab_id: str) -> None:
    try:
        http_json(f"http://127.0.0.1:{cdp_port}/json/close/{tab_id}")
    except Exception:
        pass


EXTRACT_JS = r"""
(() => {
  const imgs = [...document.images].map((img, idx) => {
    const r = img.getBoundingClientRect();
    const src = img.currentSrc || img.src || img.getAttribute('data-zoom-src') || img.getAttribute('data-src') || '';
    return {
      idx,
      src,
      alt: img.alt || '',
      naturalWidth: img.naturalWidth || 0,
      naturalHeight: img.naturalHeight || 0,
      x: Math.round(r.x),
      y: Math.round(r.y),
      width: Math.round(r.width),
      height: Math.round(r.height),
      visible: !!(r.width && r.height && r.bottom > 0 && r.right > 0 && r.top < innerHeight * 1.25),
      area: Math.round(r.width * r.height)
    };
  }).filter(x => x.src && !x.src.startsWith('data:'));
  const candidates = imgs
    .filter(x => x.visible && x.area > 20000 && /ebayimg\.com/.test(x.src))
    .sort((a,b) => b.area - a.area);
  return {
    url: location.href,
    title: document.title,
    text: document.body ? document.body.innerText.slice(0, 1200) : '',
    main: candidates[0] || null,
    images: candidates.slice(0, 12),
  };
})()
"""


async def extract_main_image(item_id: str, cdp_port: int = DEFAULT_CDP_PORT, wait_seconds: float = 5.0) -> dict[str, Any]:
    url = f"https://www.ebay.com/itm/{item_id}"
    tab = await open_tab(cdp_port, url)
    tab_id = tab.get("id", "")
    try:
        async with websockets.connect(tab["webSocketDebuggerUrl"], max_size=30_000_000) as ws:
            state = {"seq": 0}
            await send(ws, state, "Page.enable")
            await send(ws, state, "Runtime.enable")
            await send(ws, state, "Page.navigate", {"url": url})
            for _ in range(18):
                await asyncio.sleep(1)
                ready = await send(ws, state, "Runtime.evaluate", {"expression": "document.readyState", "returnByValue": True})
                if ready.get("result", {}).get("result", {}).get("value") == "complete":
                    break
            await asyncio.sleep(wait_seconds)
            result = await send(ws, state, "Runtime.evaluate", {"expression": EXTRACT_JS, "returnByValue": True})
            return result.get("result", {}).get("result", {}).get("value") or {}
    finally:
        if tab_id:
            close_tab(cdp_port, tab_id)


def white_composite(image: Image.Image) -> Image.Image:
    rgba = image.convert("RGBA")
    bg = Image.new("RGBA", rgba.size, "white")
    bg.alpha_composite(rgba)
    return bg.convert("RGB")


def ahash(image: Image.Image, size: int = 24) -> str:
    image = white_composite(image).convert("L").resize((size, size), Image.Resampling.LANCZOS)
    pixels = list(image.tobytes())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if pixel > avg else "0" for pixel in pixels)


def hamming(left: str, right: str) -> int:
    return sum(a != b for a, b in zip(left, right))


def load_image(path: Path) -> Image.Image:
    with Image.open(path) as im:
        im.load()
        return im.convert("RGB")


def download_image(url: str, out_path: Path) -> Image.Image:
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    response.raise_for_status()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(response.content)
    with Image.open(io.BytesIO(response.content)) as im:
        im.load()
        return im.convert("RGB")


def normalize_ebay_image_url(src: str) -> str:
    # Prefer a larger eBay CDN image when the URL exposes a size token.
    return src.replace("/s-l64.", "/s-l1600.").replace("/s-l140.", "/s-l1600.").replace("/s-l500.", "/s-l1600.")


def load_targets(
    ids: list[str] | None = None,
    limit: int = 0,
    source_mode: str = "cover_qa",
    product_types: set[str] | None = None,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source = COVER_QA_CSV if source_mode == "cover_qa" and COVER_QA_CSV.exists() else None
    if source:
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if ids and row.get("ID") not in ids:
                    continue
                if product_types and str(row.get("Product_Type") or "").strip() not in product_types:
                    continue
                rows.append(row)
                if limit and len(rows) >= limit:
                    return rows
    if rows:
        return rows
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    for values in ws.iter_rows(min_row=2, values_only=True):
        local_id = str(values[cols["ID"]] or "")
        if ids and local_id not in ids:
            continue
        if not values[cols.get("eBay_Item_ID", -1)]:
            continue
        product_type = str(values[cols.get("Product_Type")] or "")
        if product_types and product_type not in product_types:
            continue
        status = str(values[cols.get("Status")] or "")
        if not status.startswith("Printify_Published"):
            continue
        rows.append(
            {
                "ID": local_id,
                "Product_Type": product_type,
                "Title": str(values[cols.get("Title")] or ""),
                "eBay_Item_ID": str(values[cols.get("eBay_Item_ID")] or ""),
                "Printify_Product_ID": str(values[cols.get("Printify_Product_ID")] or ""),
                "Cover_Path": str(values[cols.get("Cover_Path")] or ""),
                "Gallery_U1_Path": str(values[cols.get("Gallery_U1_Path")] or ""),
            }
        )
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


def u_paths_from_u1(u1_path: str) -> list[tuple[str, Path]]:
    if not u1_path:
        return []
    first = Path(u1_path)
    paths = []
    for index in range(1, 5):
        candidate = Path(str(first).replace("_U1_", f"_U{index}_"))
        if candidate.exists():
            paths.append((f"U{index}", candidate))
    if not paths and first.exists():
        paths.append(("U1", first))
    return paths


def classify(distance_cover: int | None, best_u_label: str, best_u_distance: int | None) -> tuple[str, str]:
    if distance_cover is None:
        return "ERROR", "missing cover comparison"
    if best_u_distance is None:
        return ("LIKELY_COVER" if distance_cover <= 150 else "UNKNOWN", "no U comparison")
    delta = best_u_distance - distance_cover
    if best_u_distance < distance_cover and -delta >= 18:
        return "LIKELY_SINGLE_U_MISMATCH", f"{best_u_label} closer by {-delta}"
    if distance_cover <= best_u_distance and delta >= 18:
        return "LIKELY_COVER", f"cover closer than {best_u_label} by {delta}"
    return "AMBIGUOUS", f"close distances cover={distance_cover} {best_u_label}={best_u_distance}"


def printify_official_only(product_id: str) -> tuple[bool, str]:
    if not product_id:
        return False, "missing Printify product id"
    try:
        response = requests.get(
            f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
            headers={"Authorization": f"Bearer {Config.Printify_API_KEY}"},
            timeout=60,
        )
        response.raise_for_status()
        product = response.json()
        selected = [
            image
            for image in product.get("images") or []
            if image.get("is_selected_for_publishing") is not False
        ]
        custom = [
            image
            for image in selected
            if "pfy-prod-products-mockup-media" in str(image.get("src") or "")
        ]
        official = [
            image
            for image in selected
            if "images.printify.com/mockup" in str(image.get("src") or "")
        ]
        if selected and official and not custom:
            return True, f"Printify selected official-only mockups ({len(official)}/{len(selected)})"
        return False, f"Printify selected official={len(official)} custom={len(custom)} total={len(selected)}"
    except Exception as exc:  # noqa: BLE001
        return False, f"Printify official-only check failed: {str(exc)[:160]}"


def make_contact_sheet(records: list[dict[str, Any]]) -> Path:
    if not records:
        return OUT_DIR / "Online_Cover_Audit_Contact_Sheet.jpg"
    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 24) if Path("C:/Windows/Fonts/arial.ttf").exists() else ImageFont.load_default()
    w, h = 360, 360
    label_h = 90
    canvas = Image.new("RGB", (w * 3, (h + label_h) * len(records)), "white")
    draw = ImageDraw.Draw(canvas)
    for row_idx, record in enumerate(records):
        y0 = row_idx * (h + label_h)
        paths = [record.get("Cover_Path"), record.get("Online_Image_Path"), record.get("Best_U_Path") or record.get("Gallery_U1_Path")]
        labels = ["LOCAL COVER", f"ONLINE {record.get('Result')}", f"BEST LOCAL {record.get('Best_U_Label') or 'U'}"]
        for col, (path, label) in enumerate(zip(paths, labels)):
            x0 = col * w
            draw.rectangle([x0, y0, x0 + w - 1, y0 + h + label_h - 1], outline=(190, 190, 190))
            if path and Path(path).exists():
                im = load_image(Path(path))
                im.thumbnail((w - 20, h - 20), Image.Resampling.LANCZOS)
                canvas.paste(im, (x0 + (w - im.width) // 2, y0 + 10 + (h - 20 - im.height) // 2))
            draw.text((x0 + 10, y0 + h + 8), label[:38], fill=(30, 30, 30), font=font)
            draw.text((x0 + 10, y0 + h + 40), str(record.get("ID", ""))[:38], fill=(80, 80, 80), font=font)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "Online_Cover_Audit_Contact_Sheet.jpg"
    canvas.save(out, "JPEG", quality=90, optimize=True)
    return out


def append_rows(records: list[dict[str, Any]]) -> None:
    headers = [
        "Timestamp",
        "ID",
        "eBay_Item_ID",
        "Printify_Product_ID",
        "Title",
        "Online_URL",
        "Online_Image_URL",
        "Online_Image_Path",
        "Cover_Path",
        "Gallery_U1_Path",
        "Best_U_Label",
        "Best_U_Path",
        "Distance_To_Cover",
        "Best_U_Distance",
        "Result",
        "Note",
        "Error",
    ]
    exists = OUT_CSV.exists()
    if exists:
        try:
            with OUT_CSV.open("r", encoding="utf-8-sig", newline="") as existing_handle:
                existing_header = (existing_handle.readline() or "").strip().split(",")
            if existing_header != headers:
                backup = OUT_CSV.with_name(f"{OUT_CSV.stem}.backup_schema_{datetime.now():%Y%m%d_%H%M%S}{OUT_CSV.suffix}")
                OUT_CSV.replace(backup)
                print(f"[ONLINE-COVER] migrated old audit schema to {backup}")
                exists = False
        except Exception:
            exists = False
    with OUT_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not exists:
            writer.writeheader()
        for record in records:
            writer.writerow({key: record.get(key, "") for key in headers})


async def audit_async(targets: list[dict[str, str]], cdp_port: int, wait_seconds: float) -> list[dict[str, Any]]:
    records = []
    for target in targets:
        record: dict[str, Any] = {
            "Timestamp": now_text(),
            "ID": target.get("ID", ""),
            "eBay_Item_ID": target.get("eBay_Item_ID", ""),
            "Printify_Product_ID": target.get("Printify_Product_ID", ""),
            "Title": target.get("Title", ""),
            "Cover_Path": target.get("Cover_Path", ""),
            "Gallery_U1_Path": target.get("Gallery_U1_Path", ""),
            "Online_URL": f"https://www.ebay.com/itm/{target.get('eBay_Item_ID', '')}",
            "Error": "",
        }
        try:
            page = await extract_main_image(str(target.get("eBay_Item_ID", "")), cdp_port=cdp_port, wait_seconds=wait_seconds)
            text = f"{page.get('title','')} {page.get('text','')} {page.get('url','')}"
            if "Access Denied" in text or "Service Unavailable" in text or "Zero size object" in text:
                raise RuntimeError("eBay page returned access/server error")
            main = page.get("main")
            if not main or not main.get("src"):
                raise RuntimeError("no visible eBay main image extracted")
            image_url = normalize_ebay_image_url(str(main["src"]))
            record["Online_Image_URL"] = image_url
            online_path = IMAGE_DIR / f"{record['ID']}_online.jpg"
            online = download_image(image_url, online_path)
            record["Online_Image_Path"] = str(online_path)
            online_hash = ahash(online)
            cover_path = Path(record["Cover_Path"])
            distance_cover = hamming(online_hash, ahash(load_image(cover_path))) if cover_path.exists() else None
            u_distances: list[tuple[str, Path, int]] = []
            for label, path in u_paths_from_u1(record["Gallery_U1_Path"]):
                u_distances.append((label, path, hamming(online_hash, ahash(load_image(path)))))
            best_u_label = ""
            best_u_path = ""
            best_u_distance = None
            if u_distances:
                best_u_label, best_u_path_obj, best_u_distance = sorted(u_distances, key=lambda item: item[2])[0]
                best_u_path = str(best_u_path_obj)
            official_only, official_note = printify_official_only(str(target.get("Printify_Product_ID", "")))
            if official_only:
                result, note = "LIKELY_COVER_OFFICIAL", official_note
            else:
                result, note = classify(distance_cover, best_u_label, best_u_distance)
            record["Distance_To_Cover"] = distance_cover if distance_cover is not None else ""
            record["Best_U_Label"] = best_u_label
            record["Best_U_Path"] = best_u_path
            record["Best_U_Distance"] = best_u_distance if best_u_distance is not None else ""
            record["Result"] = result
            record["Note"] = note
            print(
                f"[ONLINE-COVER] {record['ID']} {result} "
                f"cover={record['Distance_To_Cover']} best_u={best_u_label}:{record['Best_U_Distance']} {note}"
            )
        except Exception as exc:  # noqa: BLE001
            record["Result"] = "ERROR"
            record["Error"] = str(exc)[:500]
            print(f"[ONLINE-COVER-FAIL] {record['ID']}: {exc}")
        records.append(record)
        time.sleep(1)
    return records


def run(
    ids: list[str] | None = None,
    limit: int = 3,
    cdp_port: int = DEFAULT_CDP_PORT,
    wait_seconds: float = 5.0,
    source_mode: str = "cover_qa",
    product_types: set[str] | None = None,
) -> list[dict[str, Any]]:
    targets = load_targets(ids=ids, limit=limit, source_mode=source_mode, product_types=product_types)
    records = asyncio.run(audit_async(targets, cdp_port=cdp_port, wait_seconds=wait_seconds))
    append_rows(records)
    contact = make_contact_sheet(records)
    print(f"[ONLINE-COVER-DONE] checked={len(records)} csv={OUT_CSV} contact={contact}")
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ids", default="", help="Comma-separated local IDs, e.g. Sticker-Zen-0025,Sticker-Zen-0045")
    parser.add_argument("--limit", type=int, default=3)
    parser.add_argument("--cdp-port", type=int, default=DEFAULT_CDP_PORT)
    parser.add_argument("--wait-seconds", type=float, default=5.0)
    parser.add_argument("--source", choices=["cover_qa", "workbook"], default="cover_qa")
    parser.add_argument("--product-types", default="", help="Comma-separated Product_Type filter, e.g. Sticker,Poster")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()] or None
    product_types = {part.strip() for part in args.product_types.split(",") if part.strip()} or None
    run(
        ids=ids,
        limit=args.limit,
        cdp_port=args.cdp_port,
        wait_seconds=args.wait_seconds,
        source_mode=args.source,
        product_types=product_types,
    )


if __name__ == "__main__":
    main()
