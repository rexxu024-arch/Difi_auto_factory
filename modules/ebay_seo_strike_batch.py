from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.ebay_quiet_jade_pivot import (
    IMAGE_NOTE,
    api_headers,
    clean,
    fit_title,
    money,
    money_text,
    request_with_retry,
    subject_from_title,
)


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
PLAN_CSV = DATABASE_DIR / "eBay_SEO_Strike_Batch.csv"
ROLLBACK_CSV = DATABASE_DIR / "eBay_SEO_Strike_Rollback.csv"
SYNC_LOG = DATABASE_DIR / "eBay_SEO_Strike_Sync_Log.csv"
STATE_JSON = DATABASE_DIR / "eBay_SEO_Strike_State.json"
NY = ZoneInfo("America/New_York")


@dataclass
class StrikeRow:
    row_idx: int
    local_id: str
    product_type: str
    category: str
    status: str
    printify_product_id: str
    old_title: str
    old_description: str
    old_price: str
    group: str
    new_title: str
    new_description: str
    new_price: str


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def ensure_column(ws, cols: dict[str, int], name: str) -> int:
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def group_for(local_id: str, product_type: str) -> str:
    seed = sum(ord(ch) for ch in local_id)
    if product_type == "Acrylic":
        return "A_QUIET_LUXURY_DESK_OBJECT" if seed % 2 else "B_GOTHIC_COLLECTOR_SHELF"
    return "A_DEEP_WORK_STICKER_SET" if seed % 2 else "B_BOOK_NOOK_LAPTOP_DECALS"


def strike_title(row: dict[str, Any], group: str) -> str:
    product_type = clean(row.get("Product_Type"))
    category = clean(row.get("Category"))
    subject = subject_from_title(clean(row.get("Title")), product_type, category)
    if product_type == "Acrylic":
        if group.startswith("A_"):
            return fit_title(["Quiet Luxury Desk Relic", subject, "5x7 Acrylic Block Shelf Decor"])
        return fit_title(["Smoky Jade Gothic Object", subject, "5x7 Acrylic Display Gift"])
    if group.startswith("A_"):
        return fit_title(["Deep Work Sticker Set", subject, "4pc 6x6 Vinyl Laptop Journal"])
    return fit_title(["Book Nook Laptop Decals", subject, "4pc 6x6 Kiss-Cut Sticker Set"])


def strike_price(row: dict[str, Any]) -> str:
    product_type = clean(row.get("Product_Type"))
    current = money(row.get("Price"))
    if product_type == "Acrylic":
        return money_text(current if current and current >= 79.99 else 89.99)
    if product_type == "Poster":
        return money_text(current if current and current >= 29.99 else 34.99)
    return money_text(current if current and current >= 9.99 else 11.99)


def strike_description(row: dict[str, Any], new_title: str, group: str) -> str:
    product_type = clean(row.get("Product_Type"))
    local_id = clean(row.get("ID"))
    category = clean(row.get("Category")) or "OpenClaw"
    if product_type == "Acrylic":
        includes = "One 5x7 vertical acrylic photo block, produced on demand through Printify."
        use_case = "quiet luxury apartment shelves, gothic study desks, collector corners, and deep-work rooms"
        texture = "smoky jade glow, acrylic depth, refractive light, and gallery-object presence"
    else:
        includes = "One physical 4pc 6x6 kiss-cut vinyl sticker sheet, produced on demand through Printify."
        use_case = "laptops, journals, planners, reading notebooks, book nooks, and focused desk setups"
        texture = "smoky jade accents, quiet desk mood, sharp die-cut presentation, and giftable detail"
    return (
        f"<h2>{new_title}</h2>"
        f"<p>This is part of a fast A/B traffic test built around higher-intent buyer scenes: {use_case}. "
        f"The style leans into {texture}, avoiding generic mass-market wording.</p>"
        "<ul>"
        f"<li><strong>Includes:</strong> {includes}</li>"
        f"<li><strong>Visual Lane:</strong> {group.replace('_', ' ').title()}</li>"
        f"<li><strong>Style:</strong> {category} with Quiet Jade / Deep Work positioning.</li>"
        "</ul>"
        f"<p><strong>Image Note:</strong> {IMAGE_NOTE}</p>"
        f"<p><small>Reference SKU: {local_id}</small></p>"
    )


def load_ready_candidates(limit: int) -> tuple[Any, Any, dict[str, int], list[StrikeRow]]:
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        data = {header: ws.cell(row_idx, cols[header]).value for header in headers if header in cols}
        data["_row_idx"] = row_idx
        status = clean(data.get("Status"))
        product_id = clean(data.get("Printify_Product_ID"))
        ebay_id = clean(data.get("eBay_Item_ID"))
        product_type = clean(data.get("Product_Type"))
        if not status.startswith("Printify_UI_Mockups") or not product_id or ebay_id:
            continue
        if product_type not in {"Acrylic", "Sticker", "Poster"}:
            continue
        rows.append(data)

    def priority(data: dict[str, Any]) -> tuple[int, str]:
        product_type = clean(data.get("Product_Type"))
        rank = {"Acrylic": 0, "Poster": 1, "Sticker": 2}.get(product_type, 9)
        return rank, clean(data.get("ID"))

    out = []
    for data in sorted(rows, key=priority)[:limit]:
        product_type = clean(data.get("Product_Type"))
        group = group_for(clean(data.get("ID")), product_type)
        new_title = strike_title(data, group)
        out.append(
            StrikeRow(
                row_idx=int(data["_row_idx"]),
                local_id=clean(data.get("ID")),
                product_type=product_type,
                category=clean(data.get("Category")),
                status=clean(data.get("Status")),
                printify_product_id=clean(data.get("Printify_Product_ID")),
                old_title=clean(data.get("Title")),
                old_description=clean(data.get("Description")),
                old_price=clean(data.get("Price")),
                group=group,
                new_title=new_title,
                new_description=strike_description(data, new_title, group),
                new_price=strike_price(data),
            )
        )
    return wb, ws, cols, out


def prepare(limit: int = 10, apply_local: bool = False) -> list[StrikeRow]:
    timestamp = now_text()
    wb, ws, cols, selected = load_ready_candidates(limit)
    if apply_local:
        backup = EBAY_BOOK.with_name(f"eBay_listing.backup_seo_strike_{datetime.now(NY):%Y%m%d_%H%M%S}.xlsx")
        shutil.copy2(EBAY_BOOK, backup)
    else:
        backup = None

    title_fail = [row.local_id for row in selected if not (75 <= len(row.new_title) <= 79)]
    if title_fail:
        wb.close()
        raise RuntimeError(f"Title length failures: {title_fail}")

    strike_col = ensure_column(ws, cols, "SEO_Strike_Timestamp")
    group_col = ensure_column(ws, cols, "SEO_Strike_Group")
    sync_col = ensure_column(ws, cols, "Metadata_Sync_Status")
    if apply_local:
        for row in selected:
            ws.cell(row.row_idx, cols["Title"]).value = row.new_title
            ws.cell(row.row_idx, cols["Description"]).value = row.new_description
            ws.cell(row.row_idx, cols["Price"]).value = row.new_price
            ws.cell(row.row_idx, strike_col).value = timestamp
            ws.cell(row.row_idx, group_col).value = row.group
            ws.cell(row.row_idx, sync_col).value = "SEO_STRIKE_PENDING_PRINTIFY_UPDATE"
        wb.save(EBAY_BOOK)
    wb.close()

    plan_rows = [
        {
            "Timestamp": timestamp,
            "ID": row.local_id,
            "Product_Type": row.product_type,
            "Status": row.status,
            "Group": row.group,
            "Printify_Product_ID": row.printify_product_id,
            "Old_Price": row.old_price,
            "New_Price": row.new_price,
            "Old_Title_Length": len(row.old_title),
            "New_Title_Length": len(row.new_title),
            "Old_Title": row.old_title,
            "New_Title": row.new_title,
            "Apply_Status": "LOCAL_APPLIED" if apply_local else "DRY_PLAN",
        }
        for row in selected
    ]
    rollback_rows = [
        {
            "Timestamp": timestamp,
            "ID": row.local_id,
            "Printify_Product_ID": row.printify_product_id,
            "Old_Title": row.old_title,
            "Old_Description": row.old_description,
            "Old_Price": row.old_price,
            "New_Title": row.new_title,
            "New_Description": row.new_description,
            "New_Price": row.new_price,
        }
        for row in selected
    ]
    write_csv(PLAN_CSV, plan_rows, list(plan_rows[0].keys()) if plan_rows else ["Timestamp"])
    write_csv(ROLLBACK_CSV, rollback_rows, list(rollback_rows[0].keys()) if rollback_rows else ["Timestamp"])
    STATE_JSON.write_text(
        json.dumps(
            {
                "timestamp": timestamp,
                "selected": len(selected),
                "apply_local": apply_local,
                "backup": str(backup) if backup else "",
                "ids": [row.local_id for row in selected],
                "counts": {ptype: sum(1 for row in selected if row.product_type == ptype) for ptype in ["Acrylic", "Poster", "Sticker"]},
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"[SEO-STRIKE] selected={len(selected)} apply_local={apply_local} ids={','.join(row.local_id for row in selected)}")
    return selected


def _load_local_rows(ids: set[str]) -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        local_id = clean(row[cols["ID"]])
        if local_id not in ids:
            continue
        out[local_id] = {
            "Title": clean(row[cols["Title"]]),
            "Description": clean(row[cols["Description"]]),
            "Price": clean(row[cols["Price"]]),
            "Product_Type": clean(row[cols["Product_Type"]]),
            "Printify_Product_ID": clean(row[cols["Printify_Product_ID"]]),
        }
    wb.close()
    return out


def _update_sync_status(ids_done: set[str], ids_failed: set[str]) -> None:
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    sync_col = ensure_column(ws, cols, "Metadata_Sync_Status")
    for row_idx in range(2, ws.max_row + 1):
        local_id = clean(ws.cell(row_idx, cols["ID"]).value)
        if local_id in ids_done:
            ws.cell(row_idx, sync_col).value = "SEO_STRIKE_SYNCED_PRINTIFY_DRAFT"
        elif local_id in ids_failed:
            ws.cell(row_idx, sync_col).value = "SEO_STRIKE_SYNC_FAILED"
    wb.save(EBAY_BOOK)
    wb.close()


def sync_printify(limit: int = 10, dry_run: bool = False, sleep_min: float = 1.0, sleep_max: float = 3.0) -> int:
    plan = list(csv.DictReader(PLAN_CSV.open("r", encoding="utf-8-sig", newline=""))) if PLAN_CSV.exists() else []
    plan = [row for row in plan if clean(row.get("Apply_Status")) == "LOCAL_APPLIED"][:limit]
    ids = {clean(row.get("ID")) for row in plan}
    local = _load_local_rows(ids)
    base = Config.Printify_API_URL.rstrip("/")
    done: set[str] = set()
    failed: set[str] = set()
    exists = SYNC_LOG.exists()
    with SYNC_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        fields = ["Timestamp", "ID", "Printify_Product_ID", "HTTP_Get", "HTTP_Update", "Result", "Error"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        if not exists:
            writer.writeheader()
        for row in plan:
            local_id = clean(row.get("ID"))
            meta = local.get(local_id) or {}
            product_id = meta.get("Printify_Product_ID") or clean(row.get("Printify_Product_ID"))
            get_status = update_status = ""
            result = "CHECK"
            error = ""
            try:
                if dry_run:
                    print(f"[SEO-STRIKE-DRY] {local_id} product={product_id} title={meta.get('Title','')[:75]}")
                    continue
                get_resp = request_with_retry("GET", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json")
                get_status = get_resp.status_code
                get_resp.raise_for_status()
                product = get_resp.json()
                target_price = int(round((money(meta.get("Price")) or 0) * 100))
                variants = []
                for variant in product.get("variants") or []:
                    variants.append(
                        {
                            "id": variant["id"],
                            "price": target_price if variant.get("is_enabled") else variant.get("price"),
                            "is_enabled": bool(variant.get("is_enabled")),
                        }
                    )
                payload = {"title": meta["Title"], "description": meta["Description"], "variants": variants}
                update_resp = request_with_retry("PUT", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json", payload=payload)
                update_status = update_resp.status_code
                update_resp.raise_for_status()
                done.add(local_id)
                result = "OK"
                print(f"[SEO-STRIKE-SYNC] {local_id} get={get_status} update={update_status} result=OK")
            except Exception as exc:  # noqa: BLE001
                failed.add(local_id)
                error = f"{type(exc).__name__}: {exc}"[:500]
                print(f"[SEO-STRIKE-FAIL] {local_id}: {error}")
            writer.writerow(
                {
                    "Timestamp": now_text(),
                    "ID": local_id,
                    "Printify_Product_ID": product_id,
                    "HTTP_Get": get_status,
                    "HTTP_Update": update_status,
                    "Result": result,
                    "Error": error,
                }
            )
            time.sleep(random.uniform(sleep_min, sleep_max))
    if not dry_run:
        _update_sync_status(done, failed)
    print(f"[SEO-STRIKE-DONE] synced={len(done)} failed={len(failed)} dry_run={dry_run}")
    return len(done)


def main() -> None:
    parser = argparse.ArgumentParser(description="SEO strike batch for ready eBay/Printify drafts.")
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--prepare", action="store_true")
    parser.add_argument("--apply-local", action="store_true")
    parser.add_argument("--sync-printify", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sleep-min", type=float, default=1.0)
    parser.add_argument("--sleep-max", type=float, default=3.0)
    args = parser.parse_args()
    if args.prepare or args.apply_local:
        prepare(limit=args.limit, apply_local=args.apply_local and not args.dry_run)
    if args.sync_printify:
        sync_printify(limit=args.limit, dry_run=args.dry_run, sleep_min=args.sleep_min, sleep_max=max(args.sleep_min, args.sleep_max))


if __name__ == "__main__":
    main()
