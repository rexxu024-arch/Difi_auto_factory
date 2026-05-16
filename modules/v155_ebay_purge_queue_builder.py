from __future__ import annotations

import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
BOOK = DATABASE / "eBay_listing.xlsx"
CANDIDATES = DATABASE / "V155_eBay_Purge_Candidates.csv"
QUEUE = DATABASE / "V155_eBay_Purge_Execution_Queue.csv"
REPORT = REPORTS / "V155_eBay_Purge_Execution_Queue_Report.md"
NY = ZoneInfo("America/New_York")


def now_text() -> str:
    return datetime.now(NY).isoformat(timespec="seconds")


def clean(value) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def workbook_index() -> dict[str, dict[str, str]]:
    wb = load_workbook(BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {value: idx for idx, value in enumerate(headers) if value}
    index: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = clean(row[cols["ID"]]) if "ID" in cols else ""
        if not item_id:
            continue
        index[item_id] = {
            "Printify_Product_ID": clean(row[cols["Printify_Product_ID"]]) if "Printify_Product_ID" in cols else "",
            "eBay_Item_ID": clean(row[cols["eBay_Item_ID"]]) if "eBay_Item_ID" in cols else "",
            "Product_Type": clean(row[cols["Product_Type"]]) if "Product_Type" in cols else "",
            "Status": clean(row[cols["Status"]]) if "Status" in cols else "",
            "Price": clean(row[cols["Price"]]) if "Price" in cols else "",
            "Title": clean(row[cols["Title"]]) if "Title" in cols else "",
        }
    wb.close()
    return index


def safe_price(value: str) -> float:
    try:
        return float(clean(value).replace("$", "").replace(",", ""))
    except ValueError:
        return 999999.0


def build_queue(limit: int | None = None, overwrite: bool = False) -> tuple[list[dict[str, str]], dict[str, int]]:
    candidates = read_csv(CANDIDATES)
    existing = read_csv(QUEUE) if QUEUE.exists() and not overwrite else []
    existing_keys = {clean(row.get("Old_eBay_Item_ID")) for row in existing if clean(row.get("Old_eBay_Item_ID"))}
    idx = workbook_index()
    rows: list[dict[str, str]] = list(existing)
    stats = {
        "candidate_rows": len(candidates),
        "added": 0,
        "skipped_existing": 0,
        "skipped_not_sticker": 0,
        "skipped_price": 0,
        "skipped_missing_ebay_id": 0,
    }
    for cand in candidates:
        item_id = clean(cand.get("ID"))
        ebay_id = clean(cand.get("eBay_Item_ID"))
        product_type = clean(cand.get("Product_Type"))
        price = clean(cand.get("Price"))
        if not ebay_id:
            stats["skipped_missing_ebay_id"] += 1
            continue
        if ebay_id in existing_keys:
            stats["skipped_existing"] += 1
            continue
        if product_type.lower() != "sticker":
            stats["skipped_not_sticker"] += 1
            continue
        if safe_price(price) >= 15:
            stats["skipped_price"] += 1
            continue
        local = idx.get(item_id, {})
        rows.append(
            {
                "Timestamp": now_text(),
                "Old_ID": item_id,
                "Old_eBay_Item_ID": ebay_id,
                "Old_Printify_Product_ID": local.get("Printify_Product_ID", ""),
                "Replacement_ID": "",
                "Replacement_eBay_Item_ID": "",
                "Reason": clean(cand.get("Reason")) or "V15.5 low-value sticker purge candidate.",
                "Status": "WAIT_SAFE_END_LISTING_PATH",
                "Retire_Attempted_At": "",
                "Retire_Result": "Queued from V155_eBay_Purge_Candidates; requires Seller Hub active item confirmation before ending.",
            }
        )
        existing_keys.add(ebay_id)
        stats["added"] += 1
        if limit is not None and stats["added"] >= limit:
            break
    return rows, stats


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Convert V15.5 eBay purge candidates into a safe Seller Hub retire queue.")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows, stats = build_queue(args.limit, args.overwrite)
    if args.dry_run:
        print({"dry_run": True, **stats, "queue_rows_after": len(rows)})
        return
    if QUEUE.exists() and not args.overwrite:
        backup = QUEUE.with_name(f"V155_eBay_Purge_Execution_Queue.backup_{datetime.now(NY):%Y%m%d_%H%M%S}.csv")
        shutil.copy2(QUEUE, backup)
    fieldnames = [
        "Timestamp",
        "Old_ID",
        "Old_eBay_Item_ID",
        "Old_Printify_Product_ID",
        "Replacement_ID",
        "Replacement_eBay_Item_ID",
        "Reason",
        "Status",
        "Retire_Attempted_At",
        "Retire_Result",
    ]
    write_csv(QUEUE, rows, fieldnames)
    REPORTS.mkdir(exist_ok=True)
    REPORT.write_text(
        "\n".join(
            [
                "# V15.5 eBay Purge Execution Queue",
                "",
                f"Timestamp: {now_text()}",
                "",
                f"- Candidate rows: {stats['candidate_rows']}",
                f"- Added to execution queue: {stats['added']}",
                f"- Queue rows now: {len(rows)}",
                f"- Skipped existing: {stats['skipped_existing']}",
                f"- Skipped not sticker: {stats['skipped_not_sticker']}",
                f"- Skipped price >= 15: {stats['skipped_price']}",
                f"- Skipped missing eBay id: {stats['skipped_missing_ebay_id']}",
                "",
                "Execution rule: every row remains `WAIT_SAFE_END_LISTING_PATH` until Seller Hub active search confirms the item id.",
                "Queue: `Database/V155_eBay_Purge_Execution_Queue.csv`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    print({"dry_run": False, **stats, "queue_rows_after": len(rows), "queue": str(QUEUE)})


if __name__ == "__main__":
    main()
