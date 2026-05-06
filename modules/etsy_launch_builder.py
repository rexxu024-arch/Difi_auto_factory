import argparse
import csv
import hashlib
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUTPUT_XLSX = DATABASE_DIR / "Etsy_launch_plan.xlsx"
OUTPUT_CSV = DATABASE_DIR / "Etsy_launch_plan.csv"
BRAND_SHELL = DATABASE_DIR / "Etsy_brand_shell.md"

TARGET_MIX = {
    "Poster": 14,
    "Acrylic": 10,
    "Sticker": 6,
}

PRODUCT_SECTIONS = {
    "Poster": "Scholar Wall Art",
    "Acrylic": "Acrylic Desk Objects",
    "Sticker": "Sticker Sheets",
}

PRODUCT_MATERIALS = {
    "Poster": ["premium matte paper", "archival wall art print", "museum-inspired artwork"],
    "Acrylic": ["acrylic photo block", "refractive display art", "gallery desk decor"],
    "Sticker": ["vinyl sticker sheet", "kiss-cut finish", "water-resistant vinyl"],
}

PRODUCT_TAGS = {
    "Poster": [
        "dark academia art",
        "zen wall art",
        "study room art",
        "wabi sabi decor",
        "jade decor",
        "library poster",
        "scholar gift",
        "mystical print",
        "wall art print",
        "dorm room decor",
        "book lover gift",
        "moody wall decor",
        "art print",
    ],
    "Acrylic": [
        "acrylic block",
        "desk decor",
        "dark academia",
        "zen decor",
        "jade decor",
        "collector gift",
        "shelf decor",
        "library decor",
        "wabi sabi decor",
        "mystical gift",
        "study desk art",
        "acrylic art",
        "unique gift",
    ],
    "Sticker": [
        "sticker sheet",
        "kiss cut sticker",
        "vinyl sticker",
        "dark academia",
        "zen sticker",
        "journal sticker",
        "laptop sticker",
        "book lover gift",
        "study decor",
        "jade aesthetic",
        "mystical sticker",
        "scrapbook decal",
        "unique sticker",
    ],
}

HEADERS = [
    "ID",
    "Product_Type",
    "Category",
    "Source_Status",
    "Printify_Product_ID",
    "Etsy_Title",
    "Etsy_Description",
    "Etsy_Tags",
    "Etsy_Materials",
    "Etsy_Section",
    "Price",
    "Production_Path",
    "Cover_Path",
    "Gallery_U1_Path",
    "Gallery_U2_Path",
    "Gallery_U3_Path",
    "Gallery_U4_Path",
    "Launch_Status",
    "Selection_Rationale",
    "Image_Note",
    "Created_Timestamp",
]


def _now():
    try:
        return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")
    except Exception:
        # Windows project venv may not have tzdata. Use local timestamp with
        # explicit Eastern label; current project operating timezone is New York.
        return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %z")


def _clean(value):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text.replace("\n", " ").replace("\r", " ")


def _ascii(value):
    return re.sub(r"[^\x00-\x7F]+", " ", _clean(value)).strip()


def _row_dicts(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {headers[index]: row[index] for index in range(len(headers))}
        if data.get("ID"):
            rows.append(data)
    workbook.close()
    return rows


def _eligible(row):
    status = _clean(row.get("Status"))
    product_type = _clean(row.get("Product_Type"))
    if product_type not in TARGET_MIX:
        return False
    if not (status.startswith("Printify_UI_Mockups") or status.startswith("Printify_Published_Mockups")):
        return False
    production = Path(_clean(row.get("Production_Path")))
    cover = Path(_clean(row.get("Cover_Path")))
    return production.exists() and cover.exists()


def _subject(title, product_type, category):
    text = _ascii(title)
    remove = [
        "4pc",
        "6x6",
        "12x18",
        "5x7",
        "Kiss-Cut",
        "Kiss Cut",
        "Sticker",
        "Sticker Sheet",
        "Vinyl",
        "Laptop",
        "Journal",
        "Gift",
        "Matte Poster",
        "Poster",
        "Wall Decor",
        "Wall Art",
        "Acrylic Photo Block",
        "Acrylic Block",
        "Acrylic",
        "Desk Display",
        "Shelf Decor",
        category,
    ]
    for term in remove:
        if term:
            text = re.sub(rf"\b{re.escape(term)}\b", " ", text, flags=re.I)
    text = re.sub(r"\b(Dark|Zen|Aesthetic|Academia|Vintage|Gothic)\b", " ", text, flags=re.I)
    text = re.sub(r"[^A-Za-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    words = text.split()
    if len(words) > 7:
        text = " ".join(words[:7])
    return text or f"{category} Art"


def _profile(row):
    blob = " ".join(
        _clean(row.get(key))
        for key in ("Title", "Description", "DNA Profile", "Category")
    ).lower()
    if "academia" in blob or "gothic" in blob or "library" in blob:
        return "Dark Academia"
    if "zen" in blob or "wabi" in blob or "lotus" in blob or "jade" in blob:
        return "Zen Aesthetic"
    if "alchemy" in blob or "alchemical" in blob:
        return "Alchemical"
    return "OpenClaw"


def _etsy_title(row):
    product_type = _clean(row.get("Product_Type"))
    category = _clean(row.get("Category"))
    subject = _subject(row.get("Title"), product_type, category)
    profile = _profile(row)
    if product_type == "Poster":
        title = f"{subject} {profile} Wall Art Print, Zen Study Poster, Wabi Sabi Room Decor"
    elif product_type == "Acrylic":
        title = f"{subject} {profile} Acrylic Block, Jade Desk Decor, Collector Shelf Gift"
    else:
        title = f"{subject} {profile} Sticker Sheet, Kiss Cut Vinyl, Journal Laptop Gift"
    title = re.sub(r"\s+", " ", title).strip(" ,")
    return title[:140].rstrip(" ,")


def _tags(row):
    product_type = _clean(row.get("Product_Type"))
    tags = list(PRODUCT_TAGS[product_type])
    blob = " ".join(_clean(row.get(key)).lower() for key in ("Title", "DNA Profile", "Description"))
    contextual = []
    if "kintsugi" in blob:
        contextual.append("kintsugi art")
    if "jade" in blob:
        contextual.append("jade art")
    if "moon" in blob or "celestial" in blob:
        contextual.append("celestial decor")
    if "library" in blob:
        contextual.append("library gift")
    if "lotus" in blob:
        contextual.append("lotus decor")
    merged = contextual + tags
    final = []
    seen = set()
    for tag in merged:
        tag = tag.lower().strip()
        if len(tag) > 20 or tag in seen:
            continue
        seen.add(tag)
        final.append(tag)
        if len(final) == 13:
            break
    return final


def _description(row):
    product_type = _clean(row.get("Product_Type"))
    category = _clean(row.get("Category"))
    subject = _subject(row.get("Title"), product_type, category)
    profile = _profile(row)
    if product_type == "Poster":
        product_line = "premium matte wall art print"
        use = "study rooms, reading corners, dorm walls, home offices, and dark academia gallery walls"
        includes = "One physical 12x18 vertical matte poster."
    elif product_type == "Acrylic":
        product_line = "5x7 acrylic display block"
        use = "desks, bookshelves, bedside tables, library nooks, and collector displays"
        includes = "One physical 5x7 acrylic photo block."
    else:
        product_line = "6x6 kiss-cut vinyl sticker sheet"
        use = "journals, laptops, notebooks, scrapbooks, water bottles, and study desk decor"
        includes = "One physical 6x6 kiss-cut sticker sheet with coordinated designs."
    tags = ", ".join(_tags(row)[:8])
    return (
        f"{subject} is part of a Quiet Relic Studio {profile} collection built around jade textures, "
        f"quiet ritual objects, scholar-room atmosphere, and collectible visual detail.\n\n"
        f"Product: {product_line}\n"
        f"Includes: {includes}\n"
        f"Best for: {use}\n"
        f"Style keywords: {tags}\n\n"
        "Image note: the main product image represents the item customers receive. "
        "Additional images are included as concept, detail, or collection-reference views and do not represent extra products or selectable variations.\n\n"
        "Production note: this item is made through a print-on-demand production partner after purchase. "
        "Minor color differences may occur between screen previews and the finished physical item."
    )


def _score(row):
    product_type = _clean(row.get("Product_Type"))
    status = _clean(row.get("Status"))
    blob = " ".join(_clean(row.get(key)).lower() for key in ("Title", "Description", "DNA Profile", "Category"))
    score = {"Acrylic": 35, "Poster": 32, "Sticker": 15}.get(product_type, 0)
    if status.startswith("Printify_Published"):
        score += 8
    for term in ("jade", "kintsugi", "dark academia", "zen", "wabi", "library", "celestial", "alchemy", "lotus"):
        if term in blob:
            score += 4
    digest = hashlib.sha1(_clean(row.get("ID")).encode("utf-8")).hexdigest()
    score += int(digest[:2], 16) / 1000
    return score


def _select(rows, limit):
    eligible = [row for row in rows if _eligible(row)]
    buckets = {product_type: [] for product_type in TARGET_MIX}
    for row in eligible:
        buckets[_clean(row.get("Product_Type"))].append(row)
    for bucket in buckets.values():
        bucket.sort(key=_score, reverse=True)
    selected = []
    for product_type, count in TARGET_MIX.items():
        selected.extend(buckets[product_type][:count])
        buckets[product_type] = buckets[product_type][count:]
    if len(selected) < limit:
        remainder = [row for bucket in buckets.values() for row in bucket]
        remainder.sort(key=_score, reverse=True)
        selected.extend(remainder[: limit - len(selected)])
    return selected[:limit]


def _plan_row(row):
    product_type = _clean(row.get("Product_Type"))
    status = _clean(row.get("Status"))
    image_note = (
        "Main image represents the physical item received; additional images are concept/detail reference views."
    )
    return {
        "ID": _clean(row.get("ID")),
        "Product_Type": product_type,
        "Category": _clean(row.get("Category")),
        "Source_Status": status,
        "Printify_Product_ID": _clean(row.get("Printify_Product_ID")),
        "Etsy_Title": _etsy_title(row),
        "Etsy_Description": _description(row),
        "Etsy_Tags": ", ".join(_tags(row)),
        "Etsy_Materials": ", ".join(PRODUCT_MATERIALS[product_type]),
        "Etsy_Section": PRODUCT_SECTIONS[product_type],
        "Price": _clean(row.get("Price")),
        "Production_Path": _clean(row.get("Production_Path")),
        "Cover_Path": _clean(row.get("Cover_Path")),
        "Gallery_U1_Path": _clean(row.get("Gallery_U1_Path")),
        "Gallery_U2_Path": _clean(row.get("Gallery_U2_Path")),
        "Gallery_U3_Path": _clean(row.get("Gallery_U3_Path")),
        "Gallery_U4_Path": _clean(row.get("Gallery_U4_Path")),
        "Launch_Status": "Draft_Prepared_Not_Published",
        "Selection_Rationale": f"Selected for Etsy Phase 1 mix; score={_score(row):.3f}; no listing fee charged.",
        "Image_Note": image_note,
        "Created_Timestamp": _now(),
    }


def _write_csv(rows):
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Etsy Launch Plan"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 22,
        "B": 14,
        "C": 16,
        "D": 26,
        "F": 58,
        "G": 90,
        "H": 50,
        "I": 44,
        "J": 24,
        "K": 12,
        "R": 28,
        "S": 60,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(OUTPUT_XLSX)
    workbook.close()


def _write_brand_shell(rows):
    counts = {}
    for row in rows:
        counts[row["Product_Type"]] = counts.get(row["Product_Type"], 0) + 1
    top_tags = []
    seen = set()
    for row in rows:
        for tag in row["Etsy_Tags"].split(", "):
            if tag not in seen:
                seen.add(tag)
                top_tags.append(tag)
            if len(top_tags) >= 24:
                break
        if len(top_tags) >= 24:
            break
    BRAND_SHELL.write_text(
        "\n".join(
            [
                "# Etsy Brand Shell Draft",
                "",
                f"Generated: {_now()}",
                "",
                "## Positioning",
                "",
                "Quiet Relic Studio is a small-batch art object shop for dark academia rooms, zen study corners, jade-inspired decor, and collectible ritual-detail visuals.",
                "",
                "## Shop Announcement",
                "",
                "Small-batch wall art, acrylic display objects, and sticker sheets inspired by jade textures, scholar rooms, quiet ritual objects, and wabi-sabi detail. Each piece is produced on demand and prepared as part of a curated visual collection.",
                "",
                "## About",
                "",
                "Quiet Relic Studio creates collectible visual objects for readers, students, desk dwellers, and quiet-room aesthetes. The collection blends dark academia, zen minimalism, jade mineral textures, kintsugi repair, and surreal object design. Products are made through print-on-demand partners so each order is produced after purchase rather than mass stocked.",
                "",
                "## FAQ",
                "",
                "**What will I receive?**",
                "You will receive the physical item named in the listing title and shown in the main product image.",
                "",
                "**Are the other images included as separate products?**",
                "No. Additional images are concept, detail, or collection-reference views to help show the design language. They are not extra products or selectable variations unless a listing explicitly says so.",
                "",
                "**When is it made?**",
                "Each item is made after purchase through a print-on-demand production partner.",
                "",
                "**Will colors match my screen exactly?**",
                "Small color differences can happen between screen previews and physical print or acrylic production.",
                "",
                "## Shipping / Production Note",
                "",
                "Production begins after purchase. Shipping time depends on the production partner and destination. Please check the listing estimate before ordering. If there is a production issue, message the shop with photos so it can be reviewed.",
                "",
                "## Suggested Sections",
                "",
                "- Scholar Wall Art",
                "- Acrylic Desk Objects",
                "- Sticker Sheets",
                "- Zen Study Decor",
                "- Dark Academia Gifts",
                "",
                "## Phase 1 Launch Mix",
                "",
                *[f"- {product_type}: {count}" for product_type, count in sorted(counts.items())],
                "",
                "## Tag Universe",
                "",
                ", ".join(top_tags),
                "",
            ]
        ),
        encoding="utf-8",
    )


def build(limit=30):
    rows = _row_dicts(EBAY_BOOK)
    selected = _select(rows, limit)
    plan_rows = [_plan_row(row) for row in selected]
    _write_csv(plan_rows)
    _write_xlsx(plan_rows)
    _write_brand_shell(plan_rows)
    print(f"[ETSY-LAUNCH] selected={len(plan_rows)} xlsx={OUTPUT_XLSX}")
    counts = {}
    for row in plan_rows:
        counts[row["Product_Type"]] = counts.get(row["Product_Type"], 0) + 1
    print(f"[ETSY-LAUNCH] mix={counts}")
    return plan_rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=30)
    args = parser.parse_args()
    build(args.limit)


if __name__ == "__main__":
    main()
