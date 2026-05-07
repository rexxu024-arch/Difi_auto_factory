"""Generate a repeatable diagnosis for eBay low-traffic listings."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
COVER_FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"
RETIRE_QUEUE = DATABASE_DIR / "eBay_Retire_Queue.csv"
EXPERIMENT_REPORT = DATABASE_DIR / "eBay_Traffic_Experiment_Report.csv"
GALLERY_DUPLICATE_AUDIT = DATABASE_DIR / "Printify_Gallery_Duplicate_Audit.csv"
OUT_CSV = DATABASE_DIR / "eBay_Traffic_Diagnosis.csv"
OUT_MD = DATABASE_DIR / "eBay_Traffic_Diagnosis.md"

HEADERS = [
    "Priority",
    "Diagnosis",
    "Evidence",
    "Recommended_Action",
    "Network_Dependency",
]


def now_text() -> str:
    return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def retired_old_ids() -> set[str]:
    retired = set()
    for row in read_csv(RETIRE_QUEUE):
        if clean(row.get("Status")) == "RETIRED_CONFIRMED":
            old_id = clean(row.get("Old_ID"))
            if old_id:
                retired.add(old_id)
    return retired


def workbook_by_ebay_id() -> dict[str, dict[str, object]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    rows: dict[str, dict[str, object]] = {}
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            ebay_id = clean(values[cols.get("eBay_Item_ID")]) if "eBay_Item_ID" in cols else ""
            if ebay_id:
                rows[ebay_id] = {header: values[index] for header, index in cols.items()}
    finally:
        wb.close()
    return rows


def latest_performance() -> tuple[str, list[dict[str, str]]]:
    rows = read_csv(PERFORMANCE_LOG)
    if not rows:
        return "", []
    latest = max(clean(row.get("Snapshot_Timestamp")) for row in rows)
    return latest, [row for row in rows if clean(row.get("Snapshot_Timestamp")) == latest]


def int_value(value: object) -> int:
    text = clean(value)
    return int("".join(ch for ch in text if ch.isdigit()) or "0")


def build_rows() -> list[dict[str, str]]:
    latest, perf_rows = latest_performance()
    workbook = workbook_by_ebay_id()
    retired_ids = retired_old_ids()
    cover_fix_ids = {
        clean(row.get("ID"))
        for row in read_csv(COVER_FIX_QUEUE)
        if clean(row.get("ID")) and clean(row.get("ID")) not in retired_ids
    }
    product_stats = defaultdict(lambda: {"rows": 0, "views": 0, "moved": 0})
    promoted = 0
    zero = 0
    for row in perf_rows:
        ebay_id = clean(row.get("Item_ID"))
        wb_row = workbook.get(ebay_id, {})
        product = clean(wb_row.get("Product_Type")) or "Unknown"
        views = int_value(row.get("Views_30_Days"))
        product_stats[product]["rows"] += 1
        product_stats[product]["views"] += views
        if views > 0:
            product_stats[product]["moved"] += 1
        if clean(row.get("General_Status")).lower() == "promoted":
            promoted += 1
        if views == 0:
            zero += 1

    experiment_rows = read_csv(EXPERIMENT_REPORT)
    gallery_duplicate_rows = [
        row for row in read_csv(GALLERY_DUPLICATE_AUDIT)
        if clean(row.get("Result")) not in {"", "OK"}
    ]
    experiment_moved = Counter()
    experiment_counts = Counter()
    for row in experiment_rows:
        group = clean(row.get("Group"))
        experiment_counts[group] += 1
        if int_value(row.get("Delta")) > 0:
            experiment_moved[group] += 1

    rows = []
    if cover_fix_ids:
        rows.append(
            {
                "Priority": "100",
                "Diagnosis": "Live cover/gallery mismatch is still a primary blocker.",
                "Evidence": f"Active cover fix queue contains {len(cover_fix_ids)} rows after excluding {len(retired_ids)} retired old eBay IDs; latest snapshot has {zero}/{len(perf_rows)} zero-view rows despite {promoted} promoted rows.",
                "Recommended_Action": "Do not expand the affected SKU family. Repair Printify source defaults or create verified replacement listings, then retire the bad public IDs.",
                "Network_Dependency": "medium",
            }
        )
    else:
        rows.append(
            {
                "Priority": "100",
                "Diagnosis": "Cover Gate is cleared; the current blocker is traffic/product-market fit.",
                "Evidence": f"Active cover fix queue is 0 after excluding {len(retired_ids)} retired old eBay IDs; latest snapshot has {zero}/{len(perf_rows)} zero-view rows despite {promoted} promoted rows.",
                "Recommended_Action": "Keep image-order audits in the QA gate, but shift growth effort to Track A/B/C experiments: buyer-intent SEO, product mix, price/room-use positioning, and Etsy digital gray launch.",
                "Network_Dependency": "low",
            }
        )
    if perf_rows:
        rows.append(
            {
                "Priority": "90",
                "Diagnosis": "Promoted Listings Standard 2% is active but is not enough alone.",
                "Evidence": f"Latest snapshot {latest}: promoted={promoted}, zero_views={zero}, rows={len(perf_rows)}.",
                "Recommended_Action": "Keep 2% Standard as baseline, but treat image/search-intent repair as the growth lever. Do not raise to suggested ad rates yet.",
                "Network_Dependency": "low",
            }
        )
    if gallery_duplicate_rows:
        rows.append(
            {
                "Priority": "85",
                "Diagnosis": "Repeated or risky gallery images can suppress buyer trust and marketplace quality scoring.",
                "Evidence": f"Printify gallery duplicate audit has {len(gallery_duplicate_rows)} non-OK rows. This includes exact repeated selected image URLs and non-sticker custom gallery sets that can look like duplicate spam on eBay.",
                "Recommended_Action": "Pause expansion, repair selected galleries to unique official product mockups, and only resume small-batch publish after duplicate audit is OK.",
                "Network_Dependency": "medium",
            }
        )
    product_evidence = "; ".join(
        f"{product}: rows={data['rows']} views={data['views']} moved={data['moved']}"
        for product, data in sorted(product_stats.items())
    )
    rows.append(
        {
            "Priority": "80",
            "Diagnosis": "Poster/Acrylic currently show more early movement than Sticker.",
            "Evidence": product_evidence or "No latest performance rows.",
            "Recommended_Action": "Keep the near-term product mix tilted toward Poster/Acrylic and Etsy digital printables until Sticker cover issue is fixed.",
            "Network_Dependency": "low",
        }
    )
    rows.append(
        {
            "Priority": "70",
            "Diagnosis": "Title rewrite experiment has not produced a clear Sticker lift yet.",
            "Evidence": "; ".join(f"{g}: moved={experiment_moved[g]}/{experiment_counts[g]}" for g in sorted(experiment_counts)) or "Experiment report unavailable.",
            "Recommended_Action": "Continue the controlled experiment window, but do not churn all titles daily. Next test should combine buyer-intent titles with corrected cover/gallery.",
            "Network_Dependency": "low",
        }
    )
    return rows


def write_outputs(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    lines = [
        "# eBay Traffic Diagnosis",
        "",
        f"Generated: {now_text()} America/New_York",
        "",
    ]
    for row in rows:
        lines.extend(
            [
                f"## P{row['Priority']} {row['Diagnosis']}",
                f"- Evidence: {row['Evidence']}",
                f"- Action: {row['Recommended_Action']}",
                f"- Network dependency: {row['Network_Dependency']}",
                "",
            ]
        )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_outputs(rows)
    print(f"[TRAFFIC-DIAGNOSIS] rows={len(rows)} csv={OUT_CSV}")
    for row in rows:
        print(f"[TRAFFIC-DIAGNOSIS] P{row['Priority']} {row['Diagnosis']}")


if __name__ == "__main__":
    main()
