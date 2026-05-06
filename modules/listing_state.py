import argparse
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"

STABLE_PREFIXES = (
    "Printify_UI_Mockups",
    "Printify_Published_Mockups",
    "Printify_Published",
)


def load_rows():
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        cols = {name: idx for idx, name in enumerate(headers)}
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[cols["ID"]]:
                rows.append(row)
        return headers, cols, rows
    finally:
        workbook.close()


def is_stable(status):
    return str(status or "").startswith(STABLE_PREFIXES)


def print_summary():
    _, cols, rows = load_rows()
    print(f"rows={len(rows)}")
    print("status=" + repr(Counter(row[cols["Status"]] for row in rows)))
    for product_type in ("Sticker", "Poster", "Acrylic"):
        stable = [
            row for row in rows
            if row[cols["Product_Type"]] == product_type and is_stable(row[cols["Status"]])
        ]
        published = [
            row for row in stable
            if str(row[cols["Status"]]).startswith("Printify_Published")
        ]
        ready = [
            row[cols["ID"]] for row in rows
            if row[cols["Product_Type"]] == product_type and row[cols["Status"]] == "Ready_for_Printify"
        ]
        failed = [
            (row[cols["ID"]], row[cols["Status"]], row[cols["Printify_Product_ID"]])
            for row in rows
            if row[cols["Product_Type"]] == product_type
            and (
                "Failed" in str(row[cols["Status"]])
                or "Hold" in str(row[cols["Status"]])
                or "DefaultMockups" in str(row[cols["Status"]])
            )
        ]
        print(
            f"{product_type}: stable={len(stable)} "
            f"published={len(published)} ready={len(ready)}"
        )
        if ready:
            print(f"  ready_sample={ready[:30]}")
        if failed:
            print(f"  failed_sample={failed[:30]}")


def main():
    parser = argparse.ArgumentParser()
    parser.parse_args()
    print_summary()


if __name__ == "__main__":
    main()
