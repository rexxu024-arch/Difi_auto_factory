from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.ebay_ads_standard import _headers

import requests

DATABASE = PROJECT_ROOT / "Database"
REPORTS = PROJECT_ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
OUT_CSV = DATABASE / "eBay_Ad_Recommendation_Probe.csv"
OUT_MD = REPORTS / "eBay_Ad_Recommendation_Probe.md"

FIELDS = [
    "Timestamp",
    "ID",
    "SKU",
    "Product_Type",
    "Status",
    "eBay_Item_ID",
    "HTTP_Status",
    "Recommendation_Returned",
    "Promote_With_Ad",
    "Bid_Item",
    "Bid_Trending",
    "Message",
    "Diagnosis",
    "Raw",
    "Error",
]


def clean(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def load_recent_non_sticker(limit: int = 10) -> list[dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        data = {str(k): clean(v) for k, v in zip(headers, raw)}
        if not data.get("eBay_Item_ID"):
            continue
        if not data.get("Status", "").startswith("Printify_Published"):
            continue
        title = data.get("Title", "").lower()
        if "sticker" in title:
            continue
        rows.append(data)
    wb.close()
    rows.sort(key=lambda item: item.get("Publish_Timestamp") or item.get("Timestamp") or "", reverse=True)
    return rows[:limit]


def parse_recommendations(rows: list[dict[str, str]], response) -> list[dict[str, Any]]:
    timestamp = now_text()
    by_listing: dict[str, dict[str, Any]] = {}
    http_status = response.status_code
    error_text = ""
    try:
        body = response.json()
    except Exception:
        body = {"raw": response.text[:1800]}
    if http_status >= 400:
        error_text = json.dumps(body, ensure_ascii=False)[:1800]
    for rec in body.get("listingRecommendations") or []:
        by_listing[clean(rec.get("listingId"))] = rec
    out: list[dict[str, Any]] = []
    for row in rows:
        listing_id = row["eBay_Item_ID"]
        rec = by_listing.get(listing_id, {})
        marketing = rec.get("marketing") or {}
        ad = ((marketing.get("ad")) or {})
        bids = {clean(item.get("basis")): clean(item.get("value")) for item in ad.get("bidPercentages") or []}
        diagnosis = ""
        if http_status == 204:
            diagnosis = "NO_RECOMMENDED_LISTINGS_ACCOUNT_WIDE"
        elif rec and not marketing:
            diagnosis = "NO_MARKETING_CONTAINER: per eBay docs this usually means not eligible, already in campaign, or invalid"
        elif ad:
            diagnosis = "AD_RECOMMENDATION_AVAILABLE"
        elif http_status >= 400:
            diagnosis = "API_ERROR"
        out.append(
            {
                "Timestamp": timestamp,
                "ID": row.get("ID", ""),
                "SKU": row.get("SKU", ""),
                "Product_Type": row.get("Product_Type", ""),
                "Status": row.get("Status", ""),
                "eBay_Item_ID": listing_id,
                "HTTP_Status": http_status,
                "Recommendation_Returned": "YES" if rec else "NO",
                "Promote_With_Ad": clean(ad.get("promoteWithAd") or rec.get("promotedWithAd")),
                "Bid_Item": bids.get("ITEM", ""),
                "Bid_Trending": bids.get("TRENDING", ""),
                "Message": clean(marketing.get("message")),
                "Diagnosis": diagnosis,
                "Raw": json.dumps(rec, ensure_ascii=False)[:1800] if rec else "",
                "Error": error_text if http_status >= 400 else "",
            }
        )
    return out


def append_rows(rows: list[dict[str, Any]]) -> None:
    exists = OUT_CSV.exists()
    with OUT_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def write_report(rows: list[dict[str, Any]]) -> None:
    REPORTS.mkdir(exist_ok=True)
    returned = [r for r in rows if r.get("Recommendation_Returned") == "YES"]
    recommended = [r for r in rows if r.get("Promote_With_Ad") == "RECOMMENDED"]
    no_rec = [r for r in rows if r.get("Recommendation_Returned") != "YES"]
    ad_info = [r for r in rows if r.get("Diagnosis") == "AD_RECOMMENDATION_AVAILABLE"]
    no_marketing = [r for r in rows if str(r.get("Diagnosis", "")).startswith("NO_MARKETING_CONTAINER")]
    lines = [
        "# eBay Ad Recommendation Probe",
        "",
        f"Generated: {now_text()}",
        f"Rows checked: {len(rows)}",
        "",
        "## Summary",
        "",
        f"- Recommendation returned: {len(returned)}",
        f"- AD recommendation payload available: {len(ad_info)}",
        f"- RECOMMENDED: {len(recommended)}",
        f"- ListingId only / no marketing container: {len(no_marketing)}",
        f"- No recommendation returned: {len(no_rec)}",
        "",
        "## Meaning",
        "",
        "- This is read-only and uses eBay's Recommendation API for Promoted Listings diagnostics.",
        "- Official eBay docs say the marketing container is omitted when a listing is not eligible for Promoted Listings, already in a campaign, or invalid.",
        "- If an account-wide request returns 204/empty, eBay currently has no listings it recommends for Standard/General ads.",
        "- Use this before ad-write retries so we do not keep hammering Marketing API with ineligible IDs.",
        "",
        "CSV: `Database/eBay_Ad_Recommendation_Probe.csv`",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(limit: int) -> list[dict[str, Any]]:
    rows = load_recent_non_sticker(limit=limit)
    listing_ids = [row["eBay_Item_ID"] for row in rows]
    if not listing_ids:
        return []
    headers = _headers()
    url = f"{Config.EBAY_API_BASE_URL.rstrip()}/sell/recommendation/v1/find?filter=recommendationTypes:{{AD}}&limit={limit}"
    response = requests.post(url, headers=headers, json={"listingIds": listing_ids}, timeout=90)
    out = parse_recommendations(rows, response)
    append_rows(out)
    write_report(out)
    print(
        f"[EBAY-AD-RECO] rows={len(out)} http={response.status_code} returned={sum(1 for r in out if r['Recommendation_Returned']=='YES')}"
    )
    return out


def run_account_wide(limit: int) -> dict[str, Any]:
    headers = _headers()
    url = f"{Config.EBAY_API_BASE_URL.rstrip()}/sell/recommendation/v1/find?filter=recommendationTypes:{{AD}}&limit={limit}&offset=0"
    response = requests.post(url, headers=headers, timeout=90)
    try:
        body = response.json() if response.text else {}
    except Exception:
        body = {"raw": response.text[:1800]}
    REPORTS.mkdir(exist_ok=True)
    out = {
        "timestamp": now_text(),
        "http_status": response.status_code,
        "total": body.get("total", 0) if isinstance(body, dict) else 0,
        "returned": len((body.get("listingRecommendations") or [])) if isinstance(body, dict) else 0,
        "body": body,
    }
    (REPORTS / "eBay_Ad_Recommendation_AccountWide.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        f"[EBAY-AD-RECO-ACCOUNT] http={response.status_code} returned={out['returned']} total={out['total']}"
    )
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--account-wide", action="store_true")
    args = parser.parse_args()
    if args.account_wide:
        run_account_wide(limit=args.limit)
    else:
        run(limit=args.limit)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
