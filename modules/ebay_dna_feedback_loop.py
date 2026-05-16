from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE / "Performance_Log.csv"
OUT_CSV = DATABASE / "eBay_DNA_Feedback_Loop.csv"
LOW_PERFORMER_CSV = DATABASE / "eBay_Low_Performer_Candidates.csv"
OUT_MD = REPORTS / "eBay_DNA_Feedback_Loop.md"
NY = ZoneInfo("America/New_York")


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def to_int(value: Any) -> int:
    try:
        return int(re.sub(r"[^0-9]", "", clean(value)) or "0")
    except ValueError:
        return 0


def parse_timestamp(value: Any) -> datetime | None:
    text = clean(value)
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    formats = [
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S %z",
        "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=NY)
            return parsed.astimezone(NY)
        except ValueError:
            continue
    return None


def age_days(row: dict[str, Any]) -> int:
    parsed = parse_timestamp(row.get("Timestamp"))
    if not parsed:
        return 0
    return max(0, (datetime.now(NY) - parsed).days)


def load_latest_performance() -> tuple[str, dict[str, dict[str, Any]]]:
    if not PERFORMANCE_LOG.exists():
        return "", {}
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return "", {}
    latest = max(clean(row.get("Snapshot_Timestamp")) for row in rows)
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        if clean(row.get("Snapshot_Timestamp")) != latest:
            continue
        item_id = clean(row.get("Item_ID"))
        if not item_id:
            continue
        out[item_id] = row
    return latest, out


def load_ebay_rows() -> list[dict[str, Any]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index for index, header in enumerate(headers) if header}
    rows: list[dict[str, Any]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not values[cols.get("ID", 0)]:
                continue
            rows.append({header: values[index] if index < len(values) else "" for header, index in cols.items()})
    finally:
        wb.close()
    return rows


def title_tokens(title: str) -> list[str]:
    stop = {
        "the",
        "and",
        "for",
        "with",
        "gift",
        "decor",
        "poster",
        "acrylic",
        "photo",
        "block",
        "sticker",
        "vinyl",
        "kiss",
        "cut",
        "matte",
        "wall",
        "art",
        "set",
        "4pc",
        "5x7",
        "12x18",
    }
    words = re.findall(r"[a-zA-Z][a-zA-Z0-9]+", title.lower())
    return [word for word in words if word not in stop and len(word) > 2]


def decision(row: dict[str, Any], views: int, live_age_days: int) -> tuple[str, str]:
    product_type = clean(row.get("Product_Type"))
    status = clean(row.get("Status"))
    if "Retired" in status:
        return "RETIRED_IGNORE", "Already retired/replaced."
    if "Published" not in status:
        return "NOT_LIVE_IGNORE", "Not live; evaluate only after publish."
    if live_age_days >= 14 and views <= 2:
        return "RETIRE_OR_REPLACE_CANDIDATE", "Two-week low signal: 0-2 views. Do not scale; replace with different DNA/product angle unless visually strategic."
    if live_age_days >= 14 and views <= 10:
        return "PAUSE_REWORK_CANDIDATE", "Two-week weak signal: 3-10 views. Pause scaling; rework title/gallery/price or test a different product type."
    if views >= 11:
        return "KEEP_AND_VARIATE", "Reached 11+ views; preserve DNA and generate variants."
    if views >= 3:
        return "WATCH_72H", "Some signal; keep for another read cycle before changing."
    if product_type in {"Poster", "Acrylic"} and views >= 1:
        return "WATCH_PREMIUM", "Premium product got some movement; avoid premature churn."
    return "LOW_SIGNAL_REVIEW", "Low/zero view; inspect title, first image, gallery uniqueness, category, and price."


def build() -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    latest, performance = load_latest_performance()
    rows = []
    token_scores: dict[str, list[int]] = defaultdict(list)
    product_scores: dict[str, list[int]] = defaultdict(list)
    decision_counts = Counter()
    for row in load_ebay_rows():
        ebay_id = clean(row.get("eBay_Item_ID"))
        perf = performance.get(ebay_id, {})
        views = to_int(perf.get("Views_30_Days"))
        live_age = age_days(row)
        action, reason = decision(row, views, live_age)
        decision_counts[action] += 1
        product_type = clean(row.get("Product_Type"))
        product_scores[product_type].append(views)
        title = clean(row.get("Title"))
        for token in title_tokens(title):
            token_scores[token].append(views)
        rows.append(
            {
                "Snapshot_Timestamp": latest,
                "ID": clean(row.get("ID")),
                "Product_Type": product_type,
                "Category": clean(row.get("Category")),
                "Status": clean(row.get("Status")),
                "eBay_Item_ID": ebay_id,
                "Views_30_Days": views,
                "Age_Days": live_age,
                "Decision": action,
                "Reason": reason,
                "Title": title,
            }
        )
    summary = {
        "decision_counts": dict(decision_counts),
        "product_median_views": {
            key: median(values) if values else 0 for key, values in sorted(product_scores.items())
        },
        "top_tokens": sorted(
            (
                {
                    "token": token,
                    "count": len(values),
                    "avg_views": sum(values) / len(values),
                    "max_views": max(values),
                }
                for token, values in token_scores.items()
                if len(values) >= 2
            ),
            key=lambda item: (item["avg_views"], item["max_views"], item["count"]),
            reverse=True,
        )[:20],
        "bottom_tokens": sorted(
            (
                {
                    "token": token,
                    "count": len(values),
                    "avg_views": sum(values) / len(values),
                    "max_views": max(values),
                }
                for token, values in token_scores.items()
                if len(values) >= 3
            ),
            key=lambda item: (item["avg_views"], item["max_views"], item["count"]),
        )[:20],
    }
    return latest, rows, summary


def write(rows: list[dict[str, Any]], summary: dict[str, Any], latest: str) -> None:
    DATABASE.mkdir(exist_ok=True)
    REPORTS.mkdir(exist_ok=True)
    fields = [
        "Snapshot_Timestamp",
        "ID",
        "Product_Type",
        "Category",
        "Status",
        "eBay_Item_ID",
        "Views_30_Days",
        "Age_Days",
        "Decision",
        "Reason",
        "Title",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    low_rows = [
        row for row in rows
        if row["Decision"] in {"RETIRE_OR_REPLACE_CANDIDATE", "PAUSE_REWORK_CANDIDATE"}
    ]
    with LOW_PERFORMER_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(low_rows)

    lines = [
        "# eBay DNA Feedback Loop",
        "",
        f"Generated: {now_text()}",
        f"Latest Seller Hub snapshot: {latest or 'missing'}",
        "",
        "## Decision Counts",
        "",
    ]
    for key, value in sorted(summary["decision_counts"].items()):
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Product Median Views", ""])
    for key, value in summary["product_median_views"].items():
        lines.append(f"- {key or 'Unknown'}: {value}")
    lines.extend(["", "## Tokens To Preserve / Test Variations", ""])
    for item in summary["top_tokens"]:
        lines.append(
            f"- {item['token']}: count={item['count']} avg_views={item['avg_views']:.2f} max={item['max_views']}"
        )
    lines.extend(["", "## Tokens To De-emphasize If Repeatedly Low", ""])
    for item in summary["bottom_tokens"]:
        lines.append(
            f"- {item['token']}: count={item['count']} avg_views={item['avg_views']:.2f} max={item['max_views']}"
        )
    lines.extend(
        [
            "",
            "## Operating Rule",
            "",
            "- 0-2 views after two weeks: idle or replace; inspect cover/gallery/category/price only if the visual is strategically important.",
            "- 3-10 views after two weeks: pause scaling, rework first image/title/price, or move the DNA to a different product type.",
            "- 11+ views or any order: preserve DNA, generate variants, and prioritize controlled restock/new listing tests.",
            f"- Low-performer candidate CSV: {LOW_PERFORMER_CSV}",
        ]
    )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.parse_args()
    latest, rows, summary = build()
    write(rows, summary, latest)
    print(f"[EBAY-DNA] rows={len(rows)} csv={OUT_CSV} md={OUT_MD}")
    for key, value in sorted(summary["decision_counts"].items()):
        print(f"[EBAY-DNA] {key}={value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
