from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.ebay_ads_standard import _api, resolve_campaign_id

DATABASE = PROJECT_ROOT / "Database"
REPORTS = PROJECT_ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
PENDING_ADS = DATABASE / "ebay_ads_pending_2pct.csv"
OUT_CSV = DATABASE / "eBay_Marketing_Reference_Probe.csv"
OUT_MD = REPORTS / "eBay_Marketing_Reference_Probe.md"

FIELDS = [
    "Timestamp",
    "ID",
    "SKU",
    "Product_Type",
    "Status",
    "eBay_Item_ID",
    "Campaign_ID",
    "Probe_Type",
    "Probe_Reference",
    "HTTP_Status",
    "Ad_Count",
    "Listing_IDs",
    "Result",
    "Error",
]


def clean(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def load_pending_ids(limit: int = 0) -> list[str]:
    if not PENDING_ADS.exists():
        return []
    ids: list[str] = []
    with PENDING_ADS.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            item_id = clean(row.get("ID"))
            status = clean(row.get("Status"))
            if item_id and status.startswith("PENDING"):
                ids.append(item_id)
            if limit and len(ids) >= limit:
                break
    return ids


def load_listing_rows(limit: int = 0, pending_only: bool = True) -> list[dict[str, str]]:
    wanted = set(load_pending_ids(limit=0)) if pending_only else set()
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not clean(raw[0]):
            continue
        data = {str(k): clean(v) for k, v in zip(headers, raw)}
        if pending_only and data.get("ID") not in wanted:
            continue
        if not data.get("eBay_Item_ID"):
            continue
        title = data.get("Title", "").lower()
        if "sticker" in title:
            continue
        rows.append(data)
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


def response_summary(response) -> tuple[int, int, str, str]:
    status = int(response.status_code)
    try:
        data = response.json()
    except Exception:
        body = clean(response.text)[:1000]
        return status, 0, "", body
    ads = data.get("ads") or []
    listing_ids = ",".join(clean(ad.get("listingId")) for ad in ads if clean(ad.get("listingId")))
    if status >= 400:
        return status, len(ads), listing_ids, json.dumps(data, ensure_ascii=False)[:1500]
    return status, len(ads), listing_ids, json.dumps(data, ensure_ascii=False)[:1500]


def safe_response_summary(callable_request) -> tuple[str, int, str, str, str]:
    """Return a probe row summary without letting one network fault stop the run."""
    try:
        response = callable_request()
        status, ad_count, listing_ids, result = response_summary(response)
        return str(status), ad_count, listing_ids, result, "" if status < 400 else result
    except Exception as exc:
        message = f"{type(exc).__name__}: {exc}"
        return "EXCEPTION", 0, "", message[:1500], message[:1500]


def get_ads_by_listing_ids(campaign_id: str, listing_ids: list[str]):
    joined = ",".join(listing_ids)
    return _api(
        "GET",
        f"/sell/marketing/v1/ad_campaign/{campaign_id}/ad?listing_ids={quote(joined)}&limit={max(len(listing_ids), 1)}",
    )


def get_ads_by_inventory_reference(campaign_id: str, reference_id: str, reference_type: str):
    return _api(
        "GET",
        f"/sell/marketing/v1/ad_campaign/{campaign_id}/get_ads_by_inventory_reference"
        f"?inventory_reference_id={quote(reference_id)}&inventory_reference_type={reference_type}",
    )


def append_rows(rows: list[dict[str, Any]]) -> None:
    DATABASE.mkdir(exist_ok=True)
    exists = OUT_CSV.exists()
    with OUT_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def write_report(campaign_id: str, rows: list[dict[str, Any]], listing_rows: list[dict[str, str]]) -> None:
    REPORTS.mkdir(exist_ok=True)
    ok = [r for r in rows if str(r.get("HTTP_Status")).startswith("2") and int(r.get("Ad_Count") or 0) > 0]
    valid_empty = [r for r in rows if str(r.get("HTTP_Status")).startswith("2") and int(r.get("Ad_Count") or 0) == 0]
    errors = [r for r in rows if not str(r.get("HTTP_Status")).startswith("2")]
    lines = [
        "# eBay Marketing Reference Probe",
        "",
        f"Generated: {now_text()}",
        f"Campaign ID: `{campaign_id}`",
        f"Listings checked: {len(listing_rows)}",
        "",
        "## Summary",
        "",
        f"- Ads found: {len(ok)}",
        f"- Valid reads but no ads: {len(valid_empty)}",
        f"- API errors: {len(errors)}",
        "",
        "## Interpretation",
        "",
        "- This probe is read-only. It does not create, update, pause, or delete any eBay ads.",
        "- Official eBay Marketing API supports reading ads by listing id and by inventory reference for listings managed by Inventory API.",
        "- If listing-id reads and inventory-reference reads both return no ads or invalid references, the next safe path is Seller Hub/manual Standard Ads for a tiny control group, or eBay support/API clarification before scaling.",
        "",
        "CSV: `Database/eBay_Marketing_Reference_Probe.csv`",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(limit: int, pending_only: bool) -> list[dict[str, Any]]:
    campaign_id = resolve_campaign_id()
    if not campaign_id:
        raise RuntimeError("No usable eBay campaign id found for Marketing API reference probe.")
    listing_rows = load_listing_rows(limit=limit, pending_only=pending_only)
    out: list[dict[str, Any]] = []
    timestamp = now_text()
    if listing_rows:
        status, ad_count, listing_ids, result, error = safe_response_summary(
            lambda: get_ads_by_listing_ids(campaign_id, [r["eBay_Item_ID"] for r in listing_rows])
        )
        out.append(
            {
                "Timestamp": timestamp,
                "ID": "BATCH",
                "SKU": "",
                "Product_Type": "",
                "Status": "",
                "eBay_Item_ID": ",".join(r["eBay_Item_ID"] for r in listing_rows),
                "Campaign_ID": campaign_id,
                "Probe_Type": "LISTING_IDS_BATCH",
                "Probe_Reference": ",".join(r["eBay_Item_ID"] for r in listing_rows),
                "HTTP_Status": status,
                "Ad_Count": ad_count,
                "Listing_IDs": listing_ids,
                "Result": result,
                "Error": error,
            }
        )
    for row in listing_rows:
        probes: list[tuple[str, str]] = []
        for value in [row.get("SKU"), row.get("ID"), row.get("eBay_Item_ID"), row.get("Printify_Product_ID")]:
            ref = clean(value)
            if ref and ref not in [p[0] for p in probes]:
                probes.append((ref, "INVENTORY_ITEM"))
                probes.append((ref, "INVENTORY_ITEM_GROUP"))
        for reference_id, reference_type in probes:
            status, ad_count, listing_ids, result, error = safe_response_summary(
                lambda reference_id=reference_id, reference_type=reference_type: get_ads_by_inventory_reference(
                    campaign_id, reference_id, reference_type
                )
            )
            out.append(
                {
                    "Timestamp": timestamp,
                    "ID": row.get("ID", ""),
                    "SKU": row.get("SKU", ""),
                    "Product_Type": row.get("Product_Type", ""),
                    "Status": row.get("Status", ""),
                    "eBay_Item_ID": row.get("eBay_Item_ID", ""),
                    "Campaign_ID": campaign_id,
                    "Probe_Type": reference_type,
                    "Probe_Reference": reference_id,
                    "HTTP_Status": status,
                    "Ad_Count": ad_count,
                    "Listing_IDs": listing_ids,
                    "Result": result,
                    "Error": error,
                }
            )
    append_rows(out)
    write_report(campaign_id, out, listing_rows)
    print(f"[EBAY-MARKETING-PROBE] listings={len(listing_rows)} rows={len(out)} campaign={campaign_id}")
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument("--all", action="store_true", help="Probe all workbook rows instead of pending ad rows.")
    args = parser.parse_args()
    run(limit=args.limit, pending_only=not args.all)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
