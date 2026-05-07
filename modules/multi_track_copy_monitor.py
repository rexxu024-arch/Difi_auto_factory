from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE = PROJECT_ROOT / "Database"
REVIEW = PROJECT_ROOT / "Review_Packets"
BATCH = DATABASE / "Multi_Track_Copy_Batch.csv"
SYNC_LOG = DATABASE / "Multi_Track_Copy_Sync_Log.csv"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
PERFORMANCE = DATABASE / "Performance_Log.csv"
OUT_CSV = DATABASE / "Multi_Track_Copy_Performance_Monitor.csv"
OUT_MD = REVIEW / f"MULTI_TRACK_COPY_MONITOR_{datetime.now():%Y%m%d}.md"
NY = ZoneInfo("America/New_York")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def as_int(value: object) -> int | None:
    text = re.sub(r"[^0-9-]", "", clean(value))
    if text in {"", "-"}:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def latest_performance() -> dict[str, dict[str, str]]:
    by_item: dict[str, dict[str, str]] = {}
    for row in read_csv(PERFORMANCE):
        item_id = clean(row.get("Item_ID"))
        if item_id:
            by_item[item_id] = row
    return by_item


def workbook_by_id() -> dict[str, dict[str, str]]:
    if not EBAY_BOOK.exists():
        return {}
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers) if header}
    result: dict[str, dict[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        item_id = clean(values[cols["ID"]]) if "ID" in cols and cols["ID"] < len(values) else ""
        if not item_id:
            continue
        row = {}
        for key in [
            "ID",
            "Product_Type",
            "Title",
            "eBay_Item_ID",
            "Multi_Track_Track",
            "Multi_Track_Primary_Intent",
            "Multi_Track_Mockup_Mood",
            "Metadata_Sync_Status",
        ]:
            if key in cols and cols[key] < len(values):
                row[key] = clean(values[cols[key]])
        result[item_id] = row
    wb.close()
    return result


def experiment_rows() -> list[dict[str, str]]:
    """Monitor all synced copy experiments, not just the most recent batch file."""
    by_id: dict[str, dict[str, str]] = {}
    for row in read_csv(BATCH):
        item_id = clean(row.get("ID"))
        if item_id:
            by_id[item_id] = dict(row)
    for row in read_csv(SYNC_LOG):
        if clean(row.get("Result")) != "OK":
            continue
        item_id = clean(row.get("ID"))
        if not item_id:
            continue
        by_id.setdefault(
            item_id,
            {
                "ID": item_id,
                "Product_Type": clean(row.get("Product_Type")),
                "eBay_Item_ID": clean(row.get("eBay_Item_ID")),
                "Apply_Status": "SYNCED_PRINTIFY",
                "Sync_Result": "OK",
                "New_Title": "",
            },
        )
    local = workbook_by_id()
    merged: list[dict[str, str]] = []
    for item_id in sorted(by_id):
        row = dict(by_id[item_id])
        meta = local.get(item_id, {})
        row["Product_Type"] = row.get("Product_Type") or meta.get("Product_Type", "")
        row["eBay_Item_ID"] = row.get("eBay_Item_ID") or meta.get("eBay_Item_ID", "")
        row["Track"] = row.get("Track") or meta.get("Multi_Track_Track", "")
        row["Primary_Intent"] = row.get("Primary_Intent") or meta.get("Multi_Track_Primary_Intent", "")
        row["Mockup_Mood"] = row.get("Mockup_Mood") or meta.get("Multi_Track_Mockup_Mood", "")
        row["New_Title"] = row.get("New_Title") or meta.get("Title", "")
        row["Sync_Result"] = row.get("Sync_Result") or ("OK" if meta.get("Metadata_Sync_Status") == "MULTI_TRACK_SYNCED_PRINTIFY_PUBLISH" else "")
        row["Apply_Status"] = row.get("Apply_Status") or meta.get("Metadata_Sync_Status", "")
        merged.append(row)
    return merged


def build_rows() -> list[dict[str, str]]:
    performance = latest_performance()
    rows = []
    for batch in experiment_rows():
        item_id = clean(batch.get("eBay_Item_ID"))
        perf = performance.get(item_id, {})
        latest_views = as_int(perf.get("Views_30_Days"))
        rows.append(
            {
                "Timestamp": datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z"),
                "ID": clean(batch.get("ID")),
                "Product_Type": clean(batch.get("Product_Type")),
                "Track": clean(batch.get("Track")),
                "Primary_Intent": clean(batch.get("Primary_Intent")),
                "Mockup_Mood": clean(batch.get("Mockup_Mood")),
                "eBay_Item_ID": item_id,
                "Apply_Status": clean(batch.get("Apply_Status")),
                "Sync_Result": clean(batch.get("Sync_Result")),
                "New_Title": clean(batch.get("New_Title")),
                "Latest_Views_30d": "" if latest_views is None else str(latest_views),
                "Latest_General_Status": clean(perf.get("General_Status")),
                "Latest_Priority_Status": clean(perf.get("Priority_Status")),
                "Read_Status": clean(perf.get("Read_Status")),
                "Monitor_Action": action(latest_views, batch, perf),
            }
        )
    return rows


def action(latest_views: int | None, batch: dict[str, str], perf: dict[str, str]) -> str:
    if clean(batch.get("Sync_Result")) != "OK":
        return "RECONCILE_SYNC_BEFORE_JUDGING_TRAFFIC"
    if not perf:
        return "WAIT_FOR_NEXT_SELLER_HUB_READ"
    if latest_views is None:
        return "READBACK_MISSING_VIEWS"
    if latest_views == 0:
        return "ZERO_VIEW_WAIT_48H_THEN_DOWNGRADE_OR_TRACK_C"
    if latest_views <= 2:
        return "NONZERO_TRAFFIC_MONITOR_CLICK_SIGNAL"
    return "HAS_TRAFFIC_CONSIDER_VARIATION"


def write_outputs(rows: list[dict[str, str]]) -> None:
    headers = [
        "Timestamp",
        "ID",
        "Product_Type",
        "Track",
        "Primary_Intent",
        "Mockup_Mood",
        "eBay_Item_ID",
        "Apply_Status",
        "Sync_Result",
        "New_Title",
        "Latest_Views_30d",
        "Latest_General_Status",
        "Latest_Priority_Status",
        "Read_Status",
        "Monitor_Action",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    counts = defaultdict(int)
    for row in rows:
        counts[row["Monitor_Action"]] += 1
    lines = [
        "# Multi-Track Copy Monitor",
        "",
        f"Generated: {datetime.now(NY).strftime('%Y-%m-%d %H:%M:%S %z')}",
        "",
        "## Action Counts",
        "",
    ]
    for key, value in sorted(counts.items()):
        lines.append(f"- `{key}`: {value}")
    lines.extend(["", "## Rows", "", "| ID | Intent | Views | Action |", "|---|---|---:|---|"])
    for row in rows:
        lines.append(f"| {row['ID']} | {row['Primary_Intent']} | {row['Latest_Views_30d'] or 'n/a'} | {row['Monitor_Action']} |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_outputs(rows)
    print(f"[MULTI-MONITOR] rows={len(rows)} csv={OUT_CSV} report={OUT_MD}")


if __name__ == "__main__":
    main()
