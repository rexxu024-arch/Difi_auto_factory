"""Audit selected Printify gallery images for repeated buyer-facing photos.

Cover correctness alone is not enough. eBay can display repeated gallery
thumbnails, especially for Poster/Acrylic official mockups. This gate checks
that a published product has enough visually distinct selected images before we
trust it as a polished listing.
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
import time
from collections import Counter
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
RETIRE_QUEUE = DATABASE_DIR / "eBay_Retire_Queue.csv"
OUT_CSV = DATABASE_DIR / "Printify_Gallery_Duplicate_Audit.csv"

PRODUCT_TYPES = {"Sticker", "Poster", "Acrylic"}


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def retired_old_ids() -> set[str]:
    retired = set()
    for row in read_csv(RETIRE_QUEUE):
        if clean(row.get("Status")) == "RETIRED_CONFIRMED":
            old_id = clean(row.get("Old_ID"))
            if old_id:
                retired.add(old_id)
    return retired


def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}"}


def fetch_product(product_id: str) -> dict:
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=headers(),
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def ahash_from_url(url: str) -> str:
    response = requests.get(url, timeout=90)
    response.raise_for_status()
    image = Image.open(io.BytesIO(response.content)).convert("L").resize((16, 16), Image.Resampling.LANCZOS)
    pixels = list(image.getdata())
    average = sum(pixels) / len(pixels)
    return "".join("1" if pixel > average else "0" for pixel in pixels)


def hamming(left: str, right: str) -> int:
    return sum(a != b for a, b in zip(left, right))


def selected_images(product: dict) -> list[dict]:
    return [
        image
        for image in product.get("images") or []
        if image.get("is_selected_for_publishing") is not False and clean(image.get("src"))
    ]


def image_role(src: str) -> str:
    if "images.printify.com/mockup" in src:
        query = parse_qs(urlparse(src).query)
        label = (query.get("camera_label") or ["official"])[0]
        return f"official:{label}"
    if "pfy-prod-products-mockup-media" in src:
        return "custom"
    return "other"


def expected_unique(product_type: str, selected_count: int) -> int:
    product = clean(product_type)
    if product == "Sticker":
        return min(3, selected_count)
    if product in {"Poster", "Acrylic"}:
        return min(3, selected_count)
    return min(2, selected_count)


def workbook_rows(limit: int = 0, ids: set[str] | None = None) -> list[dict[str, str]]:
    retired_ids = retired_old_ids()
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers_row = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers_row)}
    rows: list[dict[str, str]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not values[cols["ID"]]:
                continue
            item_id = clean(values[cols["ID"]])
            if ids and item_id not in ids:
                continue
            product_type = clean(values[cols.get("Product_Type")])
            if product_type not in PRODUCT_TYPES:
                continue
            product_id = clean(values[cols.get("Printify_Product_ID")])
            ebay_id = clean(values[cols.get("eBay_Item_ID")])
            status = clean(values[cols.get("Status")])
            if not product_id or not ebay_id:
                continue
            if ebay_id in retired_ids:
                continue
            if status.startswith("Retired"):
                continue
            rows.append(
                {
                    "ID": item_id,
                    "Product_Type": product_type,
                    "Status": status,
                    "Printify_Product_ID": product_id,
                    "eBay_Item_ID": ebay_id,
                    "Title": clean(values[cols.get("Title")]) if "Title" in cols else "",
                }
            )
            if limit and len(rows) >= limit:
                break
    finally:
        wb.close()
    return rows


def classify_hashes(hashes: list[str], near_threshold: int) -> tuple[int, int, list[str]]:
    groups: list[list[int]] = []
    group_labels: list[str] = []
    for idx, value in enumerate(hashes):
        placed = False
        for group_idx, group in enumerate(groups):
            if any(hamming(value, hashes[other]) <= near_threshold for other in group):
                group.append(idx)
                placed = True
                break
        if not placed:
            groups.append([idx])
    duplicate_pairs = 0
    notes = []
    for group in groups:
        if len(group) > 1:
            duplicate_pairs += len(group) - 1
            notes.append("near_duplicate_indexes=" + "|".join(str(index) for index in group))
    return len(groups), duplicate_pairs, notes


def audit_row(row: dict[str, str], near_threshold: int = 6, deep_hash: bool = False) -> dict[str, str]:
    record = dict(row)
    try:
        product = fetch_product(row["Printify_Product_ID"])
        images = selected_images(product)
        roles = []
        for image in images:
            src = clean(image.get("src"))
            roles.append(image_role(src))
        srcs = [clean(image.get("src")) for image in images]
        exact_counts = Counter(srcs)
        exact_duplicates = sum(count - 1 for count in exact_counts.values() if count > 1)
        role_counts = Counter(roles)
        unique_count = len(set(srcs))
        near_duplicates = 0
        notes = []
        if deep_hash:
            hashes = [ahash_from_url(src) for src in srcs]
            hash_exact_counts = Counter(hashes)
            exact_duplicates = max(exact_duplicates, sum(count - 1 for count in hash_exact_counts.values() if count > 1))
            unique_count, near_duplicates, notes = classify_hashes(hashes, near_threshold=near_threshold)
        expected = expected_unique(row["Product_Type"], len(images))
        result = "OK"
        if len(images) < expected:
            result = "CHECK_TOO_FEW_IMAGES"
        elif row["Product_Type"] in {"Poster", "Acrylic"} and role_counts.get("custom", 0) >= 3:
            result = "CHECK_CUSTOM_GALLERY_REPEATS_RISK"
        elif exact_duplicates:
            result = "CHECK_EXACT_DUPLICATE"
        elif unique_count < expected:
            result = "CHECK_NEAR_DUPLICATE"
        record.update(
            {
                "Selected_Count": str(len(images)),
                "Unique_Visual_Count": str(unique_count),
                "Expected_Unique_Count": str(expected),
                "Exact_Duplicate_Count": str(exact_duplicates),
                "Near_Duplicate_Count": str(near_duplicates),
                "Roles": "|".join(roles),
                "Result": result,
                "Notes": "; ".join(notes + [f"{key}={value}" for key, value in sorted(role_counts.items())]),
                "Error": "",
            }
        )
    except Exception as exc:  # noqa: BLE001
        record.update(
            {
                "Selected_Count": "",
                "Unique_Visual_Count": "",
                "Expected_Unique_Count": "",
                "Exact_Duplicate_Count": "",
                "Near_Duplicate_Count": "",
                "Roles": "",
                "Result": "ERROR",
                "Notes": "",
                "Error": str(exc)[:500],
            }
        )
    return record


def run(limit: int = 0, ids: set[str] | None = None, sleep_seconds: float = 0.5, near_threshold: int = 6, deep_hash: bool = False) -> list[dict[str, str]]:
    records = []
    for row in workbook_rows(limit=limit, ids=ids):
        record = audit_row(row, near_threshold=near_threshold, deep_hash=deep_hash)
        records.append(record)
        print(
            f"[GALLERY-DUP-AUDIT] {record['ID']} result={record['Result']} "
            f"selected={record.get('Selected_Count')} unique={record.get('Unique_Visual_Count')} "
            f"exact_dup={record.get('Exact_Duplicate_Count')} near_dup={record.get('Near_Duplicate_Count')}"
            ,
            flush=True,
        )
        time.sleep(max(0.0, sleep_seconds))

    fieldnames = [
        "ID",
        "Product_Type",
        "Status",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Title",
        "Selected_Count",
        "Unique_Visual_Count",
        "Expected_Unique_Count",
        "Exact_Duplicate_Count",
        "Near_Duplicate_Count",
        "Roles",
        "Result",
        "Notes",
        "Error",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    checks = sum(1 for record in records if record.get("Result") != "OK")
    print(f"[GALLERY-DUP-AUDIT-DONE] rows={len(records)} checks={checks} csv={OUT_CSV}")
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--ids", default="", help="Comma-separated listing IDs.")
    parser.add_argument("--sleep-seconds", type=float, default=0.5)
    parser.add_argument("--near-threshold", type=int, default=6)
    parser.add_argument("--deep-hash", action="store_true")
    args = parser.parse_args()
    ids = {part.strip() for part in args.ids.split(",") if part.strip()} or None
    run(limit=args.limit, ids=ids, sleep_seconds=args.sleep_seconds, near_threshold=args.near_threshold, deep_hash=args.deep_hash)


if __name__ == "__main__":
    main()
