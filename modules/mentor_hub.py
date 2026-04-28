import json
import os
import re
import sys
import time
import argparse
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from config import BASE_URL, CLAUDE_API_KEY, DEEPSEEK_API_KEY
from modules.streaming_json import JsonObjectStream, iter_anthropic_text

DATABASE_DIR = Path("Database")
PENDING_FILE = DATABASE_DIR / "pending_tasks.txt"
MENTOR_FILE = DATABASE_DIR / "Mentor_Hub.xlsx"
RUN_AUDIT_LOG = DATABASE_DIR / "mentor_run_audit.log"

MENTOR_COLUMNS = [
    "Category",
    "Layout",
    "Title",
    "Gold_Prompt_DNA",
    "Material_Keywords",
    "Timestamp",
    "Design_Count",
    "Performance",
]

NEW_SPECIES = {
    "Celestial_Gate": "Moonstone and Imperial Jade, torii gate silhouette, ink-wash nebula, floating talisman fragments",
    "Alchemical_Vessel": "Corroded Iron, cracked glass, Kintsugi repair seams, toxic jade bioluminescence",
    "Astrological_Globe": "Obsidian rings, Starlight Jade sphere, crystalline star charts, floating stardust nodes",
    "Silent_Guqin": "Petrified Wood, White Jade tuning pegs, vibrating strings, floating jade petals",
}

CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-5")
SYSTEM_PROMPT = 'Strictly output raw JSON array. Start directly with "[". No preamble, no markdown, no explanations.'


class MentorHubError(RuntimeError):
    pass


def root_path(relative_path):
    return ROOT_DIR / relative_path


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_mj_suffix(prompt):
    prompt = clean_text(prompt)
    prompt = re.sub(r"\s--v\s+\S+", "", prompt)
    prompt = re.sub(r"\s--ar\s+\S+", "", prompt)
    prompt = re.sub(r"\s--style\s+\S+", "", prompt)
    prompt = re.sub(r"\s--tile\b", "", prompt)
    prompt = re.sub(r"\s--no\s+.*$", "", prompt)
    return prompt.rstrip(", ")


def enrich_gold_prompt(prompt, category, title, materials):
    material_terms = [part.strip() for part in clean_text(materials).split(",") if part.strip()]
    material_phrase = ", ".join(material_terms[:5]) or "Translucent Imperial Jade, Kintsugi Gold, ethereal luminous material"
    base = strip_mj_suffix(prompt)
    if not base or base == ", isolated on white background":
        base = f"{title}, {category} artifact concept built from {material_phrase}"
    detail = (
        f"{base}, hyper-detailed mentor-grade object design, primary material system: {material_phrase}, "
        "micro-engraved surface relief, visible handcrafted edge bevels, layered translucent depth, "
        "subtle internal bioluminescent glow, refined kintsugi-like vein logic, floating accent fragments arranged with intentional negative space, "
        "cinematic rim lighting plus soft museum-grade overhead fill, crisp silhouette readability, centered premium product composition, "
        "clean contour separation, high-end collectible artifact aesthetic, isolated on white background"
    )
    suffix = "--v 6.1 --ar 9:16 --style raw --no skin, person, blurry edges, text, watermark"
    return f"{clean_text(detail)} {suffix}"


def excel_timestamp():
    return datetime.now()


def audit_event(message):
    path = root_path(RUN_AUDIT_LOG)
    path.parent.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
    with path.open("a", encoding="utf-8") as handle:
        handle.write(f"{stamp} | {message}\n")


def read_pending_tasks():
    path = root_path(PENDING_FILE)
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("", encoding="utf-8")
        return []
    raw = path.read_text(encoding="utf-8-sig").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            return [parsed]
    except json.JSONDecodeError:
        pass
    return [line.strip() for line in raw.splitlines() if line.strip()]


def task_key(task):
    if isinstance(task, dict):
        return clean_text(task.get("Sub_Category") or task.get("Category") or task.get("Title") or task)
    return clean_text(task)


def remove_completed_task(task):
    path = root_path(PENDING_FILE)
    raw = path.read_text(encoding="utf-8-sig") if path.exists() else ""
    try:
        parsed = json.loads(raw) if raw.strip() else []
        if isinstance(parsed, list):
            target = json.dumps(task, ensure_ascii=False, sort_keys=True)
            kept = []
            removed = False
            for item in parsed:
                if not removed and json.dumps(item, ensure_ascii=False, sort_keys=True) == target:
                    removed = True
                    continue
                kept.append(item)
            temp_path = path.with_suffix(".txt.tmp")
            temp_path.write_text(json.dumps(kept, ensure_ascii=False, indent=2) + ("\n" if kept else ""), encoding="utf-8")
            os.replace(temp_path, path)
            return
    except json.JSONDecodeError:
        pass

    target = task_key(task)
    kept = []
    removed = False
    for line in raw.splitlines():
        if not removed and line.strip() == target:
            removed = True
            continue
        kept.append(line)
    temp_path = path.with_suffix(".txt.tmp")
    temp_path.write_text("\n".join(kept) + ("\n" if kept else ""), encoding="utf-8")
    os.replace(temp_path, path)


def ensure_mentor_schema():
    path = root_path(MENTOR_FILE)
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        changed = False
        for column_index in range(len(headers), 0, -1):
            if headers[column_index - 1] in {"Product_Type", "Logic_Protocol", "Note"}:
                sheet.delete_cols(column_index)
                changed = True
        headers = [cell.value for cell in sheet[1]]
        for column in MENTOR_COLUMNS:
            if column not in headers:
                sheet.cell(row=1, column=len(headers) + 1).value = column
                headers.append(column)
                changed = True
        if changed:
            workbook.save(path)
    except PermissionError:
        ensure_mentor_schema_com()
    finally:
        workbook.close()


def ensure_mentor_schema_com():
    import win32com.client

    workbook = win32com.client.GetObject(str(root_path(MENTOR_FILE)))
    sheet = workbook.Worksheets(1)
    headers = []
    column = 1
    while True:
        value = sheet.Cells(1, column).Value
        if value in (None, ""):
            break
        headers.append(str(value))
        column += 1
    for name in MENTOR_COLUMNS:
        if name not in headers:
            sheet.Cells(1, column).Value = name
            headers.append(name)
            column += 1
    for column_index in range(len(headers), 0, -1):
        if headers[column_index - 1] in {"Product_Type", "Logic_Protocol", "Note"}:
            sheet.Columns(column_index).Delete()
    workbook.Save()


def header_map(sheet):
    return {cell.value: index + 1 for index, cell in enumerate(sheet[1]) if cell.value}


def extract_json_array(text):
    text = clean_text(text)
    fenced = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.I | re.S)
    if fenced:
        return json.loads(fenced.group(1))
    start = text.find("[")
    end = text.rfind("]")
    if start >= 0 and end > start:
        return json.loads(text[start : end + 1])
    raise MentorHubError("API response did not contain a JSON array")


def build_seed_prompt(task):
    seed = task if isinstance(task, dict) else {"Seed": task_key(task)}
    return {
        "mission": "GREY ARCHITECT V13.5 Mentor_Hub Gold DNA cultivation",
        "seed": seed,
        "required_count": 20,
        "rules": [
            "Use the seed Sub_Category as the exact Category for every row.",
            "Create exactly 20 representative Gold DNA rows under the same Category.",
            "Prefer integrating one of Celestial_Gate, Alchemical_Vessel, Astrological_Globe, Silent_Guqin when compatible.",
            "Return only JSON array with 20 objects.",
            "Object keys: Category, Layout, Title, Gold_Prompt_DNA, Material_Keywords.",
            "Gold_Prompt_DNA must preserve isolated on white background and include material, lighting, composition.",
            "Each Gold_Prompt_DNA should be 420 to 650 characters before parameters: dense like a production prompt, not essay-length.",
            "Do not use markdown, commentary, escaped newlines, or multiline strings.",
            "Each row must be visually distinct: colorway, lighting style, object pose, composition, or material emphasis.",
        ],
        "new_species_knowledge": NEW_SPECIES,
    }


def stream_claude_objects(task, retries=3):
    if not CLAUDE_API_KEY:
        raise MentorHubError("CLAUDE_API_KEY is empty")
    url = BASE_URL.rstrip("/") + "/v1/messages"
    headers = {
        "x-api-key": CLAUDE_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": 7000,
        "temperature": 0.65,
        "stream": True,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": json.dumps(build_seed_prompt(task), ensure_ascii=False)}],
    }
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            response = requests.post(url, headers=headers, json=payload, stream=True, timeout=(10, 15))
            if response.status_code >= 400:
                raise MentorHubError(f"{response.status_code} {response.text[:800]}")
            parser = JsonObjectStream()
            for text in iter_anthropic_text(response):
                for item in parser.feed(text):
                    yield item
            return
        except Exception as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(2 * attempt)
    raise MentorHubError(f"Mentor API stream failed after retries: {last_error}")


def fallback_gold_dna(task):
    label = task_key(task)
    species_name, species_logic = next(
        ((name, logic) for name, logic in NEW_SPECIES.items() if name.lower() in label.lower()),
        ("Celestial_Gate", NEW_SPECIES["Celestial_Gate"]),
    )
    sub_category = label if "-" in label or "_" in label else f"Zen-{species_name}"
    title = clean_text(task.get("Title")) if isinstance(task, dict) and task.get("Title") else sub_category.replace("_", " ")
    gold_prompt = clean_text(task.get("Gold_Prompt_DNA")) if isinstance(task, dict) and task.get("Gold_Prompt_DNA") else (
        f"{title}, {species_logic}, luminous jade core, cinematic rim lighting, centered relic composition, "
        "isolated on white background --v 6.1 --ar 3:4 --style raw --tile"
    )
    protocol = clean_text(task.get("Logic_Protocol")) if isinstance(task, dict) and task.get("Logic_Protocol") else species_logic
    themes = [
        "pale jade dawn light",
        "deep emerald rim glow",
        "moonstone blue underlight",
        "antique gold dust halo",
        "obsidian shadow contrast",
        "white porcelain negative space",
        "amber sacred backlight",
        "ink-wash mist silhouette",
        "crystalline star flare",
        "aged bronze micro-detail",
        "floating talisman orbit",
        "vertical museum composition",
        "macro carved material focus",
        "ritual seal symmetry",
        "soft celestial bloom",
        "fractured kintsugi geometry",
        "quiet scholar archive light",
        "mythic relic front view",
        "diagonal motion aura",
        "minimal luxury product pose",
    ]
    return [{
        "Category": sub_category,
        "Layout": "Isolated",
        "Title": f"{title} Variant {index:02d}",
        "Gold_Prompt_DNA": f"{gold_prompt}, {themes[index - 1]}",
        "Material_Keywords": protocol,
        "Timestamp": excel_timestamp(),
        "Design_Count": 0,
        "Performance": "",
    } for index in range(1, 21)]


def normalize_gold_row(item, task):
    label = task_key(task)
    category = label
    if "_" not in category and "-" not in category:
        category = f"{category}-Relic_Instrument"
    logic = clean_text(item.get("Material_Keywords") or (task.get("Logic_Protocol") if isinstance(task, dict) else ""))
    prompt = clean_text(item.get("Gold_Prompt_DNA"))
    if "isolated on white" not in prompt.lower():
        prompt = f"{prompt}, isolated on white background"
    prompt = enrich_gold_prompt(prompt, category, clean_text(item.get("Title") or category.replace("_", " ")), clean_text(item.get("Material_Keywords") or logic))
    return {
        "Category": category,
        "Layout": clean_text(item.get("Layout") or "Isolated"),
        "Title": clean_text(item.get("Title") or category.replace("_", " ")),
        "Gold_Prompt_DNA": prompt,
        "Material_Keywords": clean_text(item.get("Material_Keywords") or logic),
        "Timestamp": excel_timestamp(),
        "Design_Count": int(item.get("Design_Count") or 0),
        "Performance": clean_text(item.get("Performance") or ""),
    }


def append_mentor_rows(rows):
    path = root_path(MENTOR_FILE)
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        for row_data in rows:
            row_index = sheet.max_row + 1
            for column in MENTOR_COLUMNS:
                cell = sheet.cell(row=row_index, column=columns[column])
                cell.value = row_data.get(column)
                if column == "Timestamp":
                    cell.number_format = "m/d/yyyy, h:mm:ss AM/PM"
        workbook.save(path)
    except PermissionError:
        append_mentor_rows_com(rows)
    finally:
        workbook.close()


def append_mentor_row(row):
    append_mentor_rows([row])


def append_mentor_rows_com(rows):
    import win32com.client

    workbook = win32com.client.GetObject(str(root_path(MENTOR_FILE)))
    sheet = workbook.Worksheets(1)
    columns = {}
    column = 1
    while True:
        value = sheet.Cells(1, column).Value
        if value in (None, ""):
            break
        columns[str(value)] = column
        column += 1
    xl_up = -4162
    last_row = sheet.Cells(sheet.Rows.Count, 1).End(xl_up).Row
    for row_data in rows:
        last_row += 1
        for name in MENTOR_COLUMNS:
            cell = sheet.Cells(last_row, columns[name])
            cell.NumberFormat = "m/d/yyyy, h:mm:ss AM/PM" if name == "Timestamp" else "@"
            cell.Value = row_data.get(name)
    workbook.Save()


def existing_category_state(category):
    workbook = load_workbook(root_path(MENTOR_FILE), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        columns = header_map(sheet)
        count = 0
        fingerprints = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[columns["Category"] - 1] != category:
                continue
            count += 1
            title = row[columns["Title"] - 1]
            prompt = row[columns["Gold_Prompt_DNA"] - 1]
            fingerprints.add(f"{category}|{title}|{prompt}")
        return count, fingerprints
    finally:
        workbook.close()


def run_logic(limit=1, max_seconds=120):
    os.chdir(ROOT_DIR)
    ensure_mentor_schema()
    tasks = read_pending_tasks()
    if limit is not None:
        tasks = tasks[:limit]
    if not tasks:
        print("[MENTOR_HUB] No pending tasks.")
        return 0
    processed = 0
    for task in tasks:
        try:
            started_at = time.monotonic()
            category = task_key(task)
            audit_event(f"START {category}")
            existing_count, seen = existing_category_state(category)
            if existing_count >= 20:
                remove_completed_task(task)
                processed += 1
                print(f"[MENTOR_HUB] Existing Gold DNA group complete: {category}")
                continue
            saved = existing_count
            for item in stream_claude_objects(task):
                if time.monotonic() - started_at > max_seconds:
                    print(f"[MENTOR_HUB] Seed time budget exceeded: {category} | saved {saved}/20 | keep pending for resume")
                    break
                row = normalize_gold_row(item, task)
                fp = f"{row['Category']}|{row['Title']}|{row['Gold_Prompt_DNA']}"
                if fp in seen:
                    continue
                seen.add(fp)
                append_mentor_row(row)
                saved += 1
                print(f"[Dify-Mode] Category: {row['Category']} | Gold DNA {saved}/20 写入成功 | 实时保存已完成")
                if saved >= 20:
                    break
            if saved != 20:
                print(f"[MENTOR_HUB] Incomplete Gold DNA group: {category} | saved {saved}/20 | not removing pending")
                audit_event(f"INCOMPLETE {category} saved={saved}/20 elapsed={time.monotonic() - started_at:.1f}s")
                continue
            remove_completed_task(task)
            processed += 1
            audit_event(f"DONE {category} saved=20/20 elapsed={time.monotonic() - started_at:.1f}s")
            print(f"[MENTOR_HUB] Stored Gold DNA group: {task_key(task)}")
        except Exception as exc:
            print(f"[MENTOR_HUB] Failed: {task_key(task)} | {exc}")
            raise
    return processed


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=1)
    parser.add_argument("--max-seconds", type=int, default=120)
    args = parser.parse_args()
    run_logic(limit=args.limit, max_seconds=args.max_seconds)
