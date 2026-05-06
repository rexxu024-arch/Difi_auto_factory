import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
ETSY_PLAN = DATABASE_DIR / "Etsy_launch_plan.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
OUTPUT_CSV = DATABASE_DIR / "Unified_Listing_Registry.csv"
OUTPUT_XLSX = DATABASE_DIR / "Unified_Listing_Registry.xlsx"

HEADERS = [
    "ID",
    "Product_Type",
    "Category",
    "Local_Status",
    "Printify_Product_ID",
    "eBay_Item_ID",
    "eBay_Item_URL",
    "eBay_Title",
    "eBay_Price",
    "Latest_eBay_Views_30_Days",
    "Latest_eBay_General_Status",
    "Latest_eBay_Priority_Status",
    "Etsy_Planned",
    "Etsy_Title",
    "Etsy_Launch_Status",
    "Production_Path",
    "Cover_Path",
    "Gallery_Ready",
    "Image_Note_Ready",
    "Action_Bucket",
]


def _clean(value):
    return str(value or "").strip()


def _rows_from_xlsx(path):
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[index]: row[index] for index in range(len(headers))}
            if data.get("ID"):
                rows.append(data)
    finally:
        wb.close()
    return rows


def _latest_performance_by_item():
    latest = {}
    if not PERFORMANCE_LOG.exists():
        return latest
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            item_id = _clean(row.get("Item_ID"))
            ts = _clean(row.get("Snapshot_Timestamp"))
            if not item_id:
                continue
            if item_id not in latest or ts >= _clean(latest[item_id].get("Snapshot_Timestamp")):
                latest[item_id] = row
    return latest


def _etsy_by_id():
    return {row["ID"]: row for row in _rows_from_xlsx(ETSY_PLAN)}


def _gallery_ready(row):
    if _clean(row.get("Product_Type")) == "Sticker":
        return all(_clean(row.get(f"Gallery_U{index}_Path")) for index in range(1, 5))
    return all("Gallery_U" in _clean(row.get(f"Gallery_U{index}_Path")) for index in range(1, 5))


def _bucket(row, perf, etsy):
    status = _clean(row.get("Status"))
    views = _clean(perf.get("Views_30_Days")) if perf else ""
    try:
        view_count = int(views)
    except ValueError:
        view_count = None
    if not _gallery_ready(row):
        return "Fix_Gallery_First"
    if status == "Ready_for_Printify":
        return "Ready_For_Printify_When_Network_OK"
    if status.startswith("Printify_UI_Mockups"):
        return "Stable_Draft_Publish_When_Scheduled"
    if status.startswith("Printify_Published") and view_count == 0:
        return "Published_Zero_View_Copy_Ad_Review"
    if status.startswith("Printify_Published") and view_count and view_count > 0:
        return "Published_Has_View_Monitor"
    if etsy:
        return "Etsy_Draft_Prepared"
    return "Hold"


def build_rows():
    perf_by_item = _latest_performance_by_item()
    etsy = _etsy_by_id()
    output = []
    for row in _rows_from_xlsx(EBAY_BOOK):
        item_id = _clean(row.get("eBay_Item_ID"))
        perf = perf_by_item.get(item_id, {})
        etsy_row = etsy.get(row["ID"], {})
        desc = _clean(row.get("Description")).lower()
        output.append(
            {
                "ID": _clean(row.get("ID")),
                "Product_Type": _clean(row.get("Product_Type")),
                "Category": _clean(row.get("Category")),
                "Local_Status": _clean(row.get("Status")),
                "Printify_Product_ID": _clean(row.get("Printify_Product_ID")),
                "eBay_Item_ID": item_id,
                "eBay_Item_URL": _clean(row.get("eBay_Item_URL")),
                "eBay_Title": _clean(row.get("Title")),
                "eBay_Price": _clean(row.get("Price")),
                "Latest_eBay_Views_30_Days": _clean(perf.get("Views_30_Days")),
                "Latest_eBay_General_Status": _clean(perf.get("General_Status")),
                "Latest_eBay_Priority_Status": _clean(perf.get("Priority_Status")),
                "Etsy_Planned": bool(etsy_row),
                "Etsy_Title": _clean(etsy_row.get("Etsy_Title")),
                "Etsy_Launch_Status": _clean(etsy_row.get("Launch_Status")),
                "Production_Path": _clean(row.get("Production_Path")),
                "Cover_Path": _clean(row.get("Cover_Path")),
                "Gallery_Ready": _gallery_ready(row),
                "Image_Note_Ready": "main image shows the actual product customers receive" in desc,
                "Action_Bucket": _bucket(row, perf, etsy_row),
            }
        )
    return output


def write_outputs(rows):
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Unified Registry"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column, width in {
        "A": 24,
        "B": 14,
        "D": 28,
        "E": 26,
        "F": 18,
        "H": 62,
        "N": 62,
        "T": 34,
    }.items():
        ws.column_dimensions[column].width = width
    wb.save(OUTPUT_XLSX)
    wb.close()


def main():
    rows = build_rows()
    write_outputs(rows)
    buckets = defaultdict(int)
    for row in rows:
        buckets[row["Action_Bucket"]] += 1
    print(f"[REGISTRY] rows={len(rows)} csv={OUTPUT_CSV}")
    for key in sorted(buckets):
        print(f"[REGISTRY] {key}={buckets[key]}")


if __name__ == "__main__":
    main()
