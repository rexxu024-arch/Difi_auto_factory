from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path
from statistics import median
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


DATABASE = PROJECT_ROOT / "Database"
REPORTS = PROJECT_ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
OUT_CSV = DATABASE / "eBay_Ad_Rate_Experiment_Plan.csv"
OUT_MD = REPORTS / "eBay_Ad_Rate_Experiment_Plan.md"
OUT_SHORTLIST = DATABASE / "eBay_Ad_Rate_Execution_Shortlist_NoSticker.csv"
SHIPPING_MATRIX_CSV = DATABASE / "eBay_Shipping_Repair_Decision_Matrix.csv"
NY = ZoneInfo("America/New_York")


PRODUCT_COSTS = {
    # Current enabled Printify variants, read back from live products.
    # Shipping is US first-item shipping from the existing Printify guardrail.
    "Sticker": {"production": 2.32, "shipping": 4.29, "floor_margin": 0.75},
    "Poster": {"production": 9.34, "shipping": 5.99, "floor_margin": 4.00},
    "Acrylic": {"production": 35.43, "shipping": 15.99, "floor_margin": 10.00},
}

TARGET_PROFIT = {
    # Stickers are review / traffic builders: keep them safely above water,
    # but do not force premium-profit pricing that would kill conversion.
    "Sticker": 0.75,
    "Poster": 8.00,
    "Acrylic": 15.00,
}

PRICE_CEILING = {
    "Sticker": 11.99,
    "Poster": 44.99,
    "Acrylic": 99.99,
}

PRICE_SWEET_SPOT = {
    # eBay sticker packs are a commodity lane.  Rex's sticker listings should
    # be priced as review/traffic builders, not premium-margin products.
    "Sticker": 9.99,
    "Poster": 36.99,
    "Acrylic": 94.99,
}

STICKER_TRAFFIC_PRICE_LADDER = {
    0.0: 8.99,
    2.0: 9.99,
    4.0: 9.99,
    6.0: 10.99,
    8.0: 10.99,
    10.0: 10.99,
    12.0: 10.99,
}

INTENT_TERMS = {
    "quiet luxury": 3,
    "smoky jade": 3,
    "deep work": 3,
    "reading nook": 3,
    "collector": 2,
    "shelf": 2,
    "desk": 2,
    "study": 2,
    "gothic": 2,
    "grimdark": 2,
    "premium": 2,
    "acrylic": 2,
    "poster": 2,
    "wall art": 2,
    "apartment": 2,
    "meditation": 2,
    "laptop": 1,
    "journal": 1,
    "gift": 1,
}

LOW_QUALITY_TITLE_PATTERNS = [
    re.compile(r"\b(.{4,24})\b(?:\s+\1\b){2,}", re.IGNORECASE),
]

# Keep 2% as the existing baseline, then test a controlled ladder.  A listing
# should only be assigned to one live ad-rate lane at a time, otherwise the
# attribution data becomes useless.
AD_RATES = [2.0, 4.0, 6.0, 8.0, 10.0, 12.0]
EBAY_FVF_RATE = 0.136
EBAY_FIXED_FEE = 0.40


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def money(value: object) -> float:
    raw = re.sub(r"[^0-9.]", "", clean(value))
    if not raw:
        return 0.0
    try:
        return float(raw)
    except ValueError:
        return 0.0


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def load_rows() -> list[dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers) if name}
    rows: list[dict[str, str]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or not values[cols["ID"]]:
                continue
            product_type = clean(values[cols.get("Product_Type")])
            ebay_id = clean(values[cols.get("eBay_Item_ID")])
            if product_type not in PRODUCT_COSTS or not ebay_id:
                continue
            rows.append(
                {
                    "ID": clean(values[cols["ID"]]),
                    "Product_Type": product_type,
                    "Category": clean(values[cols.get("Category")]),
                    "Title": clean(values[cols.get("Title")]),
                    "Price": clean(values[cols.get("Price")]),
                    "Status": clean(values[cols.get("Status")]),
                    "eBay_Item_ID": ebay_id,
                    "Metadata_Sync_Status": clean(values[cols.get("Metadata_Sync_Status")]) if "Metadata_Sync_Status" in cols else "",
                }
            )
    finally:
        wb.close()
    return rows


def load_registry() -> dict[str, dict[str, str]]:
    path = DATABASE / "Unified_Listing_Registry.csv"
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {clean(row.get("ID")): {key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)}


def load_shipping_matrix() -> dict[str, dict[str, str]]:
    if not SHIPPING_MATRIX_CSV.exists():
        return {}
    with SHIPPING_MATRIX_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return {clean(row.get("Item_ID")): {key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)}


def estimate(price: float, product_type: str, ad_rate: float) -> dict[str, float]:
    cost = PRODUCT_COSTS[product_type]
    production = cost["production"]
    shipping = cost["shipping"]
    ebay_fee = price * EBAY_FVF_RATE + EBAY_FIXED_FEE
    ad_fee = price * (ad_rate / 100.0)
    profit = price - production - shipping - ebay_fee - ad_fee
    return {
        "production": production,
        "shipping": shipping,
        "ebay_fee": ebay_fee,
        "ad_fee": ad_fee,
        "profit": profit,
        "margin_pct": (profit / price * 100.0) if price else 0.0,
    }


def price_for_profit(product_type: str, ad_rate: float, target_profit: float) -> float:
    cost = PRODUCT_COSTS[product_type]
    denominator = 1.0 - EBAY_FVF_RATE - (ad_rate / 100.0)
    if denominator <= 0:
        return 9999.0
    return (cost["production"] + cost["shipping"] + EBAY_FIXED_FEE + target_profit) / denominator


def charm_price(value: float) -> float:
    """Round up into buyer-friendly .99 pricing."""
    if value <= 0:
        return 0.0
    dollars = int(value)
    candidate = dollars + 0.99
    if candidate + 1e-9 < value:
        candidate = dollars + 1.99
    return round(candidate, 2)


def normalized_free_shipping_price(row: dict[str, str]) -> float:
    """Estimate the source price that should be set in Printify.

    Older listings often had item price plus buyer-paid shipping.  The current
    strategy is free shipping, so we lift underpriced legacy rows to at least
    a product-specific free-shipping floor before doing ad math.
    """
    current = money(row["Price"])
    product_type = row["Product_Type"]
    if not current:
        return 0.0
    if product_type == "Sticker" and current < 9.0:
        return PRICE_SWEET_SPOT["Sticker"]
    if product_type == "Poster" and current < 36.0:
        return 36.99
    if product_type == "Acrylic" and current < 94.0:
        return 94.99
    return current


def recommended_source_price(row: dict[str, str], ad_rate: float) -> tuple[float, str]:
    product_type = row["Product_Type"]
    listed_price = money(row["Price"])
    current_source = normalized_free_shipping_price(row)
    target = TARGET_PROFIT[product_type]
    needed = charm_price(price_for_profit(product_type, ad_rate, target))
    ceiling = PRICE_CEILING[product_type]
    if needed > ceiling:
        floor_needed = charm_price(price_for_profit(product_type, ad_rate, PRODUCT_COSTS[product_type]["floor_margin"]))
        if floor_needed <= ceiling:
            needed = floor_needed
            action = "RAISE_TO_FLOOR_MARGIN_PRICE"
        else:
            return current_source, "RATE_PRICE_NOT_VIABLE"
    else:
        action = "RAISE_TO_TARGET_PROFIT_PRICE"

    if product_type == "Sticker":
        sweet_spot = STICKER_TRAFFIC_PRICE_LADDER.get(float(ad_rate), PRICE_SWEET_SPOT["Sticker"])
        sweet_est = estimate(sweet_spot, product_type, ad_rate)
        if sweet_est["profit"] >= PRODUCT_COSTS[product_type]["floor_margin"]:
            if listed_price > sweet_spot + 0.50:
                return sweet_spot, "LOWER_STICKER_TRAFFIC_PRICE"
            if listed_price <= sweet_spot + 0.50:
                return max(current_source, sweet_spot), "KEEP_STICKER_TRAFFIC_PRICE"

    recommended = max(current_source, needed)
    if listed_price + 0.01 < recommended:
        return recommended, "UPDATE_PRINTIFY_SOURCE_PRICE"
    if current_source + 0.01 >= needed:
        return current_source, "KEEP_CURRENT_SOURCE_PRICE"
    return needed, action


def title_quality_issue(title: str) -> str:
    lowered = title.lower()
    for pattern in LOW_QUALITY_TITLE_PATTERNS:
        if pattern.search(lowered):
            return "repeated_phrase"
    tokens = re.findall(r"[a-z0-9]+", lowered)
    if len(tokens) >= 6:
        unique_ratio = len(set(tokens)) / len(tokens)
        if unique_ratio < 0.55:
            return "low_unique_word_ratio"
    return ""


def effectiveness_score(row: dict[str, str], registry: dict[str, dict[str, str]]) -> tuple[int, str]:
    reg = registry.get(row["ID"], {})
    title = row["Title"] or reg.get("eBay_Title", "")
    lowered = title.lower()
    score = 0
    reasons: list[str] = []

    for term, weight in INTENT_TERMS.items():
        if term in lowered:
            score += weight
            reasons.append(term)

    product_type = row["Product_Type"]
    if product_type == "Acrylic":
        score += 4
        reasons.append("premium-margin-product")
    elif product_type == "Poster":
        score += 3
        reasons.append("mid-margin-product")
    elif product_type == "Sticker":
        score += 1
        reasons.append("traffic-builder")

    if reg.get("Gallery_Ready", "").lower() == "true":
        score += 2
        reasons.append("gallery-ready")
    if reg.get("Production_Path"):
        score += 1
        reasons.append("production-path")
    views = int(float(reg.get("Latest_eBay_Views_30_Days") or 0))
    if views > 0:
        score += min(5, views)
        reasons.append(f"views={views}")

    issue = title_quality_issue(title)
    if issue:
        score -= 6
        reasons.append(f"title-issue={issue}")
    if not reg:
        reasons.append("registry-missing")

    return score, "; ".join(reasons)


def max_effective_rate(row: dict[str, str], score: int) -> float:
    if row.get("Title_Quality_Issue"):
        return 0.0
    product_type = row["Product_Type"]
    if product_type == "Sticker":
        return 4.0 if score >= 1 else 2.0
    if product_type == "Poster":
        if score >= 10:
            return 8.0
        if score >= 6:
            return 6.0
        return 4.0
    if product_type == "Acrylic":
        if score >= 12:
            return 12.0
        if score >= 9:
            return 10.0
        if score >= 6:
            return 8.0
        return 6.0
    return 2.0


def choose_lane(row: dict[str, str], ad_rate: float, profit: float, margin_pct: float, effective_cap: float) -> str:
    product_type = row["Product_Type"]
    floor = PRODUCT_COSTS[product_type]["floor_margin"]
    if profit < floor:
        return "DO_NOT_USE_RATE"
    if ad_rate > effective_cap:
        return "RATE_EFFECTIVENESS_HOLD"
    if product_type == "Sticker":
        if ad_rate <= 2.0:
            return "STICKER_REVIEW_BUILDER_CONTROL"
        if ad_rate <= 4.0:
            return "STICKER_TRAFFIC_TEST_SMALL"
        return "STICKER_RATE_TOO_HIGH_FOR_LOW_TICKET"
    if product_type == "Poster":
        if ad_rate <= 4.0:
            return "POSTER_MID_AD_TEST"
        if ad_rate <= 8.0 and margin_pct >= 24:
            return "POSTER_AGGRESSIVE_TEST"
        return "POSTER_HOLD_RATE"
    if product_type == "Acrylic":
        if ad_rate <= 6.0:
            return "ACRYLIC_PREMIUM_AD_TEST"
        if ad_rate <= 12.0 and profit >= PRODUCT_COSTS[product_type]["floor_margin"] + 5:
            return "ACRYLIC_AGGRESSIVE_PREMIUM_TEST"
        return "ACRYLIC_HOLD_RATE"
    return "HOLD"


def target_lane_for_row(row: dict[str, str], lane_counts: dict[str, int], limit_per_lane: int) -> tuple[float, str] | None:
    """Pick one ad-rate assignment per listing.

    The ladder intentionally puts more aggressive rates on higher-margin
    products first. Sticker listings are no longer selected for new ad tests:
    Rex froze sticker expansion after the lane proved too commoditized on eBay.
    """
    product_type = row["Product_Type"]
    if product_type == "Sticker":
        return None
    if row.get("Shipping_Ad_Gate", "").startswith("BLOCK_AD_UNTIL"):
        return None
    if row.get("Shipping_Repair_Decision") in {"HOLD_REBUILD_GALLERY_FIRST", "SOURCE_REBUILD_OR_PRINTIFY_TEMPLATE_PROBE", "HOLD_API_READ_FAILED", "UNAUDITED_SHIPPING"}:
        return None
    effective_cap = float(row.get("Max_Effective_Ad_Rate_Pct") or 2.0)
    if product_type == "Acrylic":
        ladder = [12.0, 10.0, 8.0, 6.0, 4.0, 2.0]
    elif product_type == "Poster":
        ladder = [8.0, 6.0, 4.0, 2.0]
    else:
        ladder = [4.0, 2.0]
    for ad_rate in ladder:
        source_price, price_action = recommended_source_price(row, ad_rate)
        if price_action == "RATE_PRICE_NOT_VIABLE":
            continue
        est = estimate(source_price, product_type, ad_rate)
        lane = choose_lane(row, ad_rate, est["profit"], est["margin_pct"], effective_cap)
        if lane == "DO_NOT_USE_RATE" or lane.endswith("HOLD_RATE") or lane.endswith("TOO_HIGH_FOR_LOW_TICKET"):
            continue
        if lane == "RATE_EFFECTIVENESS_HOLD":
            continue
        if lane_counts.get(lane, 0) >= limit_per_lane:
            continue
        lane_counts[lane] = lane_counts.get(lane, 0) + 1
        return ad_rate, lane
    return None


def build(limit_per_lane: int = 8) -> list[dict[str, str]]:
    rows = load_rows()
    registry = load_registry()
    shipping_matrix = load_shipping_matrix()
    output: list[dict[str, str]] = []
    selected_counts: dict[str, int] = {}
    for row in rows:
        current_price = money(row["Price"])
        if not current_price:
            continue
        score, reasons = effectiveness_score(row, registry)
        title_issue = title_quality_issue(row["Title"] or registry.get(row["ID"], {}).get("eBay_Title", ""))
        row["Effectiveness_Score"] = str(score)
        row["Effectiveness_Reason"] = reasons
        row["Title_Quality_Issue"] = title_issue
        row["Max_Effective_Ad_Rate_Pct"] = f"{max_effective_rate(row, score):.1f}"
        shipping_gate = shipping_matrix.get(row["eBay_Item_ID"], {})
        row["Shipping_Repair_Decision"] = clean(shipping_gate.get("Decision")) or "UNAUDITED_SHIPPING"
        row["Shipping_Ad_Gate"] = clean(shipping_gate.get("Ad_Gate")) or "BLOCK_AD_UNTIL_SHIPPING_AUDITED"
        assignment = target_lane_for_row(row, selected_counts, limit_per_lane)
        for ad_rate in AD_RATES:
            source_price, price_action = recommended_source_price(row, ad_rate)
            est = estimate(source_price, row["Product_Type"], ad_rate)
            lane = choose_lane(row, ad_rate, est["profit"], est["margin_pct"], float(row["Max_Effective_Ad_Rate_Pct"]))
            if price_action == "RATE_PRICE_NOT_VIABLE":
                lane = "RATE_PRICE_NOT_VIABLE"
            selected = "Yes" if assignment == (ad_rate, lane) else "No"
            output.append(
                {
                    "Timestamp": now_text(),
                    "Selected_For_Test": selected,
                    "Lane": lane,
                    "ID": row["ID"],
                    "Product_Type": row["Product_Type"],
                    "Category": row["Category"],
                    "eBay_Item_ID": row["eBay_Item_ID"],
                    "Current_Listed_Price_USD": f"{current_price:.2f}",
                    "Recommended_Printify_Source_Price_USD": f"{source_price:.2f}",
                    "Price_Action": price_action,
                    "Ad_Rate_Pct": f"{ad_rate:.1f}",
                    "Max_Effective_Ad_Rate_Pct": row["Max_Effective_Ad_Rate_Pct"],
                    "Effectiveness_Score": row["Effectiveness_Score"],
                    "Effectiveness_Reason": row["Effectiveness_Reason"],
                    "Title_Quality_Issue": row["Title_Quality_Issue"],
                    "Shipping_Repair_Decision": row["Shipping_Repair_Decision"],
                    "Shipping_Ad_Gate": row["Shipping_Ad_Gate"],
                    "Production_USD": f"{est['production']:.2f}",
                    "Shipping_USD": f"{est['shipping']:.2f}",
                    "Estimated_eBay_Fee_USD": f"{est['ebay_fee']:.2f}",
                    "Estimated_Ad_Fee_USD": f"{est['ad_fee']:.2f}",
                    "Estimated_Profit_USD": f"{est['profit']:.2f}",
                    "Estimated_Margin_Pct": f"{est['margin_pct']:.1f}",
                    "Title": row["Title"],
                }
            )
    return output


def break_even_price(product_type: str, ad_rate: float, target_profit: float = 0.0) -> float:
    return charm_price(price_for_profit(product_type, ad_rate, target_profit))


def sticker_price_ladder_lines() -> list[str]:
    lines = [
        "## Sticker Break-Even Guardrail",
        "",
        "Sticker is treated as a traffic/review product, but never as a loss leader.",
        "",
        "| Ad rate | Break-even price | Floor-margin price | Recommended source price | Est. profit at recommendation |",
        "|---:|---:|---:|---:|---:|",
    ]
    for rate in AD_RATES:
        rec = STICKER_TRAFFIC_PRICE_LADDER.get(rate, PRICE_SWEET_SPOT["Sticker"])
        est = estimate(rec, "Sticker", rate)
        lines.append(
            f"| {rate:.0f}% | ${break_even_price('Sticker', rate, 0):.2f} | "
            f"${break_even_price('Sticker', rate, PRODUCT_COSTS['Sticker']['floor_margin']):.2f} | "
            f"${rec:.2f} | ${est['profit']:.2f} |"
        )
    lines.extend(["", ""])
    return lines


def write_outputs(rows: list[dict[str, str]]) -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    headers = [
        "Timestamp",
        "Selected_For_Test",
        "Lane",
        "ID",
        "Product_Type",
        "Category",
        "eBay_Item_ID",
        "Current_Listed_Price_USD",
        "Recommended_Printify_Source_Price_USD",
        "Price_Action",
        "Ad_Rate_Pct",
        "Max_Effective_Ad_Rate_Pct",
        "Effectiveness_Score",
        "Effectiveness_Reason",
        "Title_Quality_Issue",
        "Shipping_Repair_Decision",
        "Shipping_Ad_Gate",
        "Production_USD",
        "Shipping_USD",
        "Estimated_eBay_Fee_USD",
        "Estimated_Ad_Fee_USD",
        "Estimated_Profit_USD",
        "Estimated_Margin_Pct",
        "Title",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    selected = [row for row in rows if row["Selected_For_Test"] == "Yes"]
    with OUT_SHORTLIST.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows([row for row in selected if row.get("Product_Type") != "Sticker"])
    by_lane: dict[str, list[float]] = {}
    for row in selected:
        by_lane.setdefault(row["Lane"], []).append(float(row["Estimated_Profit_USD"]))
    lines = [
        "# eBay Ad Rate Experiment Plan",
        "",
        f"Generated: {now_text()}",
        "",
        "## Conservative Cost Assumptions",
        "",
        f"- eBay fee estimate: {EBAY_FVF_RATE * 100:.1f}% + ${EBAY_FIXED_FEE:.2f}.",
        "- General/Standard ad fee is modeled as a sale-attributed percentage of item price.",
        "- eBay official help says General strategy is charged only when the promoted item sells after an ad click; Priority/PPC remains excluded.",
        "- Product price is treated as free-shipping price, so Printify shipping is included as seller cost.",
        "- Legacy listings that appear to be item-price-plus-shipping are lifted to a Printify source price that absorbs shipping before ad math.",
        "- If an ad lane would leave too little profit, the plan recommends a Printify source price change first; it does not accept loss-making ad exposure.",
        "- Higher ad rates are only assigned when the listing passes an effectiveness gate: buyer-intent keywords, product margin tier, gallery readiness, and clean title quality.",
        "- Listings with obvious repeated-title or low-unique-word issues are held for copy repair before ad budget is increased.",
        "- Listings whose buyer-facing shipping/gallery is still blocked by the shipping repair matrix are excluded from selected ad lanes.",
        "- Sticker listings remain in the cost model for historical analysis, but are frozen out of new selected test lanes after the eBay sticker lane proved too commoditized.",
        "- Rates that fall below each product's floor margin are excluded.",
        "- Each listing is assigned to at most one experimental ad-rate lane.",
        "",
        *sticker_price_ladder_lines(),
        "## Selected Test Lanes",
        "",
    ]
    for lane, profits in sorted(by_lane.items()):
        lines.append(f"- {lane}: {len(profits)} listings, median estimated profit ${median(profits):.2f}")
    lines.extend(["", "## Selected Rows", ""])
    for row in selected:
        lines.append(
            f"- {row['ID']} {row['Product_Type']} rate={row['Ad_Rate_Pct']}% "
            f"source_price=${row['Recommended_Printify_Source_Price_USD']} "
            f"price_action={row['Price_Action']} profit=${row['Estimated_Profit_USD']} "
            f"margin={row['Estimated_Margin_Pct']}% "
            f"lane={row['Lane']}"
        )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[EBAY-AD-PLAN] rows={len(rows)} selected={len(selected)} csv={OUT_CSV}")
    for lane, profits in sorted(by_lane.items()):
        print(f"[EBAY-AD-PLAN] {lane} count={len(profits)} median_profit={median(profits):.2f}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit-per-lane", type=int, default=8)
    args = parser.parse_args()
    write_outputs(build(limit_per_lane=args.limit_per_lane))


if __name__ == "__main__":
    main()
