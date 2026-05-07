import argparse
import csv
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.network_guard import report as network_report
from modules.resilient_http import request_with_retry


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
SYNC_LOG = PROJECT_ROOT / "Database" / "printify_external_sync.csv"
DECISION_CSV = PROJECT_ROOT / "Database" / "Self_Healing_Decisions.csv"
RUN_LOG = PROJECT_ROOT / "Database" / "Self_Healing_Daemon_Log.csv"
PRINTIFY_STATUS_URL = "https://printify.com/network-fulfillment-status/"

DECISION_FIELDS = [
    "Timestamp",
    "ID",
    "Product_Type",
    "Status",
    "Printify_Product_ID",
    "eBay_Item_ID",
    "Network_Mode",
    "Network_Reason",
    "Printify_Status",
    "Error_Rate_30",
    "External_Missing_Age_Minutes",
    "Attempts",
    "Decision",
    "Reason",
    "Execute",
]


@dataclass
class PlatformPulse:
    network_mode: str
    network_reason: str
    printify_status: str
    printify_status_reason: str
    error_rate_30: float


@dataclass
class ExternalCandidate:
    row_idx: int
    item_id: str
    product_type: str
    status: str
    printify_product_id: str
    ebay_item_id: str
    first_seen: datetime | None
    attempts: int


def _now():
    return datetime.now()


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}", "Content-Type": "application/json"}


def _iso(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return str(value)


def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %I:%M:%S %p"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            pass
    return None


def _ensure_column(sheet, headers, name):
    if name not in headers:
        sheet.cell(1, sheet.max_column + 1).value = name
        headers.append(name)
    return headers.index(name) + 1


def _append_csv(path, rows, fieldnames):
    exists = path.exists()
    with path.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def fetch_printify_status(timeout=15):
    try:
        response = requests.get(PRINTIFY_STATUS_URL, timeout=timeout)
        text = response.text.lower()
        if response.status_code >= 500:
            return "unknown", f"http_{response.status_code}"
        degraded_markers = [
            "delayed",
            "delay",
            "disruption",
            "incident",
            "degraded",
            "maintenance",
            "issue",
        ]
        steady_markers = ["all systems go", "steady", "normal"]
        # The status page can include explanatory copy such as "delays" even
        # when the headline is healthy, so explicit healthy headline wins.
        if any(marker in text for marker in steady_markers):
            return "steady", "status_page_steady_marker"
        if any(marker in text for marker in degraded_markers):
            return "degraded_or_delayed", "status_page_degraded_marker"
        return "unknown", "status_page_readable_no_known_marker"
    except Exception as exc:
        return "unknown", f"status_page_error={type(exc).__name__}: {exc}"


def external_sync_error_rate(limit=30):
    if not SYNC_LOG.exists():
        return 0.0
    rows = []
    try:
        with SYNC_LOG.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)[-limit:]
    except Exception:
        return 1.0
    if not rows:
        return 0.0
    bad = 0
    for row in rows:
        status = str(row.get("Status") or "").upper()
        if status in {"ERROR", "MISSING_EXTERNAL_ID"}:
            bad += 1
    return bad / len(rows)


def external_missing_history():
    history = {}
    if not SYNC_LOG.exists():
        return history
    try:
        with SYNC_LOG.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                item_id = str(row.get("ID") or "").strip()
                status = str(row.get("Status") or "").upper()
                if not item_id or status not in {"MISSING_EXTERNAL_ID", "ERROR"}:
                    continue
                ts = _parse_datetime(row.get("Timestamp"))
                current = history.setdefault(item_id, {"first_seen": ts, "attempts": 0})
                current["attempts"] += 1
                if ts and (not current.get("first_seen") or ts < current["first_seen"]):
                    current["first_seen"] = ts
    except Exception:
        return history
    return history


def build_pulse():
    try:
        network_payload = network_report(count=4)
        strategy = network_payload.get("strategy") or {}
        network_mode = strategy.get("mode", "unknown")
        network_reason = strategy.get("reason", "")
    except Exception as exc:
        network_mode = "unknown"
        network_reason = f"network_guard_error={type(exc).__name__}: {exc}"
    printify_status, printify_reason = fetch_printify_status()
    return PlatformPulse(
        network_mode=network_mode,
        network_reason=network_reason,
        printify_status=printify_status,
        printify_status_reason=printify_reason,
        error_rate_30=external_sync_error_rate(),
    )


def _candidate_rows(sheet, headers, limit=0, ids=None, missing_history=None):
    ids = {str(item).strip() for item in (ids or []) if str(item).strip()}
    missing_history = missing_history or {}
    cols = {header: index + 1 for index, header in enumerate(headers)}
    candidates = []
    for row_idx in range(2, sheet.max_row + 1):
        item_id = str(sheet.cell(row_idx, cols.get("ID", 1)).value or "").strip()
        if ids and item_id not in ids:
            continue
        status = str(sheet.cell(row_idx, cols.get("Status", 1)).value or "").strip()
        product_id = str(sheet.cell(row_idx, cols.get("Printify_Product_ID", 1)).value or "").strip()
        ebay_id = str(sheet.cell(row_idx, cols.get("eBay_Item_ID", 1)).value or "").strip()
        external_pending = "PublishExternalPending" in status or status == "PublishExternalMissing"
        published_missing = status.startswith("Printify_Published") and product_id and not ebay_id
        if not product_id or ebay_id or not (external_pending or published_missing):
            continue
        first_seen = _parse_datetime(
            sheet.cell(row_idx, cols.get("External_Missing_First_Seen", 1)).value
            if "External_Missing_First_Seen" in cols
            else None
        )
        hist = missing_history.get(item_id) or {}
        if not first_seen:
            first_seen = hist.get("first_seen")
        attempts = int(hist.get("attempts") or 0)
        if "External_Missing_Attempts" in cols:
            try:
                attempts = max(attempts, int(sheet.cell(row_idx, cols["External_Missing_Attempts"]).value or 0))
            except ValueError:
                pass
        candidates.append(
            ExternalCandidate(
                row_idx=row_idx,
                item_id=item_id,
                product_type=str(sheet.cell(row_idx, cols.get("Product_Type", 1)).value or ""),
                status=status,
                printify_product_id=product_id,
                ebay_item_id=ebay_id,
                first_seen=first_seen,
                attempts=attempts,
            )
        )
        if limit and len(candidates) >= limit:
            break
    return candidates


def fetch_printify_product(product_id):
    url = f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json"
    response = request_with_retry("GET", url, headers=_headers(), timeout=90, attempts=3, backoff=3, jitter=1.5)
    response.raise_for_status()
    return response.json()


def decide(candidate, pulse, product_payload=None, api_error=None):
    now = _now()
    first_seen = candidate.first_seen or now
    age_minutes = max(0.0, (now - first_seen).total_seconds() / 60.0)

    if pulse.network_mode == "pause":
        return "DELAY_POLL", f"network pause: {pulse.network_reason}", age_minutes
    if pulse.printify_status == "degraded_or_delayed":
        return "DELAY_POLL", f"Printify status degraded: {pulse.printify_status_reason}", age_minutes
    if pulse.error_rate_30 >= 0.45:
        return "DELAY_POLL", f"recent external sync error rate too high: {pulse.error_rate_30:.2f}", age_minutes
    if api_error:
        if candidate.attempts >= 5 and age_minutes >= 24 * 60 and pulse.network_mode == "full_throughput":
            return "QUARANTINE_LOSS_ASSET", f"API still failing after {candidate.attempts} attempts and {age_minutes:.0f} minutes: {api_error}", age_minutes
        return "DELAY_POLL", f"transient API failure: {api_error}", age_minutes

    external = (product_payload or {}).get("external") or {}
    ebay_id = str(external.get("id") or "").strip()
    handle = str(external.get("handle") or "").strip()
    if ebay_id:
        return "FORCE_ASSOCIATE", f"Printify product external.id={ebay_id} handle={handle}", age_minutes

    locked = bool((product_payload or {}).get("is_locked"))
    if age_minutes < 45:
        return "DELAY_POLL", f"recent publish; wait for external propagation age={age_minutes:.0f}m locked={locked}", age_minutes
    if candidate.attempts < 3:
        return "DELAY_POLL", f"insufficient retries attempts={candidate.attempts} age={age_minutes:.0f}m locked={locked}", age_minutes
    if age_minutes >= 24 * 60 and candidate.attempts >= 5 and pulse.network_mode == "full_throughput":
        return "QUARANTINE_LOSS_ASSET", f"healthy network/status but no external after {age_minutes:.0f}m and {candidate.attempts} attempts", age_minutes
    return "DELAY_POLL", f"still within delayed propagation envelope age={age_minutes:.0f}m attempts={candidate.attempts}", age_minutes


def apply_decision(sheet, headers, candidate, decision, reason, product_payload, execute=False):
    cols = {header: index + 1 for index, header in enumerate(headers)}
    now = _now()
    external = (product_payload or {}).get("external") or {}
    ebay_id = str(external.get("id") or "").strip()
    handle = str(external.get("handle") or "").strip()
    external_type = str(external.get("type") or "").strip()

    missing_first_col = _ensure_column(sheet, headers, "External_Missing_First_Seen")
    attempts_col = _ensure_column(sheet, headers, "External_Missing_Attempts")
    decision_col = _ensure_column(sheet, headers, "Self_Healing_Decision")
    reason_col = _ensure_column(sheet, headers, "Self_Healing_Reason")
    ts_col = _ensure_column(sheet, headers, "Self_Healing_Timestamp")
    ebay_col = _ensure_column(sheet, headers, "eBay_Item_ID")
    url_col = _ensure_column(sheet, headers, "eBay_Item_URL")
    type_col = _ensure_column(sheet, headers, "External_Type")
    sync_col = _ensure_column(sheet, headers, "External_Sync_Timestamp")

    if not sheet.cell(candidate.row_idx, missing_first_col).value:
        sheet.cell(candidate.row_idx, missing_first_col).value = now
    sheet.cell(candidate.row_idx, attempts_col).value = candidate.attempts + 1
    sheet.cell(candidate.row_idx, decision_col).value = decision
    sheet.cell(candidate.row_idx, reason_col).value = reason[:500]
    sheet.cell(candidate.row_idx, ts_col).value = now

    if execute and decision == "FORCE_ASSOCIATE" and ebay_id:
        sheet.cell(candidate.row_idx, ebay_col).value = ebay_id
        sheet.cell(candidate.row_idx, url_col).value = handle
        sheet.cell(candidate.row_idx, type_col).value = external_type
        sheet.cell(candidate.row_idx, sync_col).value = now
        if "Status" in cols and "PublishExternalPending" in candidate.status:
            match = re.search(r"Mockups(\\d+)", candidate.status)
            suffix = match.group(1) if match else ""
            sheet.cell(candidate.row_idx, cols["Status"]).value = (
                f"Printify_Published_Mockups{suffix}" if suffix else "Printify_Published"
            )
    elif execute and decision == "QUARANTINE_LOSS_ASSET":
        if "Status" in cols:
            sheet.cell(candidate.row_idx, cols["Status"]).value = "Quarantined_ExternalMissing"


def run(limit=0, execute=False, ids=None, sleep_seconds=2.0):
    pulse = build_pulse()
    workbook = load_workbook(EBAY_BOOK)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    # Ensure daemon metadata columns exist before row scanning.
    for name in [
        "External_Missing_First_Seen",
        "External_Missing_Attempts",
        "Self_Healing_Decision",
        "Self_Healing_Reason",
        "Self_Healing_Timestamp",
        "eBay_Item_ID",
        "eBay_Item_URL",
        "External_Type",
        "External_Sync_Timestamp",
    ]:
        _ensure_column(sheet, headers, name)
    candidates = _candidate_rows(sheet, headers, limit=limit, ids=ids, missing_history=external_missing_history())
    decisions = []
    for candidate in candidates:
        payload = None
        api_error = ""
        try:
            payload = fetch_printify_product(candidate.printify_product_id)
        except Exception as exc:
            api_error = f"{type(exc).__name__}: {exc}"
        decision, reason, age_minutes = decide(candidate, pulse, payload, api_error)
        apply_decision(sheet, headers, candidate, decision, reason, payload, execute=execute)
        row = {
            "Timestamp": _now().isoformat(timespec="seconds"),
            "ID": candidate.item_id,
            "Product_Type": candidate.product_type,
            "Status": candidate.status,
            "Printify_Product_ID": candidate.printify_product_id,
            "eBay_Item_ID": str(((payload or {}).get("external") or {}).get("id") or ""),
            "Network_Mode": pulse.network_mode,
            "Network_Reason": pulse.network_reason,
            "Printify_Status": f"{pulse.printify_status}:{pulse.printify_status_reason}",
            "Error_Rate_30": f"{pulse.error_rate_30:.3f}",
            "External_Missing_Age_Minutes": f"{age_minutes:.1f}",
            "Attempts": str(candidate.attempts + 1),
            "Decision": decision,
            "Reason": reason,
            "Execute": str(bool(execute)),
        }
        decisions.append(row)
        print(f"[SELF-HEAL] {candidate.item_id} decision={decision} reason={reason}")
        time.sleep(max(0.0, sleep_seconds))
    if execute:
        workbook.save(EBAY_BOOK)
    workbook.close()
    _append_csv(DECISION_CSV, decisions, DECISION_FIELDS)
    _append_csv(
        RUN_LOG,
        [
            {
                "Timestamp": _now().isoformat(timespec="seconds"),
                "Candidates": len(candidates),
                "Execute": str(bool(execute)),
                "Network_Mode": pulse.network_mode,
                "Network_Reason": pulse.network_reason,
                "Printify_Status": pulse.printify_status,
                "Printify_Status_Reason": pulse.printify_status_reason,
                "Error_Rate_30": f"{pulse.error_rate_30:.3f}",
            }
        ],
        [
            "Timestamp",
            "Candidates",
            "Execute",
            "Network_Mode",
            "Network_Reason",
            "Printify_Status",
            "Printify_Status_Reason",
            "Error_Rate_30",
        ],
    )
    return {"pulse": asdict(pulse), "decisions": decisions}


def main():
    parser = argparse.ArgumentParser(description="Self-healing daemon for Printify/eBay external sync gaps.")
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--execute", action="store_true", help="Apply workbook association/quarantine changes.")
    parser.add_argument("--ids", default="", help="Comma-separated workbook IDs to evaluate.")
    parser.add_argument("--sleep-seconds", type=float, default=2.0)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()]
    result = run(limit=args.limit, execute=args.execute, ids=ids, sleep_seconds=args.sleep_seconds)
    if args.json:
        print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
