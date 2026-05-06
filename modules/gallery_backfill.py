import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.art_asset_builder import PRODUCT_SPECS, process_folder


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"


def _product_type(value):
    value = str(value or "").strip().lower()
    if value.startswith("acry"):
        return "Acrylic"
    if value.startswith("poster"):
        return "Poster"
    return ""


def _folder_for(item_id, product_type):
    spec = PRODUCT_SPECS[product_type]["spec"]
    return PROJECT_ROOT / "Output" / product_type / spec / f"MASTER_{item_id}_Ready_for_Steaming"


def ensure_gallery_assets(product_type=None, limit=0):
    products = [product_type] if product_type else ["Poster", "Acrylic"]
    built = 0
    checked = 0
    notes = []
    for product in products:
        spec = PRODUCT_SPECS[product]["spec"]
        root = PROJECT_ROOT / "Output" / product / spec
        folders = sorted(root.glob("MASTER_*_Ready_for_Steaming"))
        for folder in folders:
            if limit and checked >= limit:
                return checked, built, notes
            checked += 1
            missing = [
                folder / f"Gallery_U{index}.png"
                for index in range(1, 5)
                if not (folder / f"Gallery_U{index}.png").exists()
            ]
            if not missing:
                continue
            ok, note = process_folder(folder, product, force=False)
            notes.append(("OK" if ok else "HOLD") + " | " + note)
            if ok:
                built += 1
                print(f"[GALLERY] {folder.name} derived from production image")
            else:
                print(f"[GALLERY-HOLD] {note}")
    return checked, built, notes


def update_listing_gallery_paths(limit=0):
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    required = ["ID", "Product_Type", *[f"Gallery_U{index}_Path" for index in range(1, 5)]]
    missing = [header for header in required if header not in cols]
    if missing:
        wb.close()
        raise RuntimeError(f"eBay listing workbook missing columns: {missing}")

    changed = 0
    checked = 0
    for row in range(2, ws.max_row + 1):
        item_id = str(ws.cell(row, cols["ID"]).value or "").strip()
        product = _product_type(ws.cell(row, cols["Product_Type"]).value)
        if not item_id or product not in {"Poster", "Acrylic"}:
            continue
        checked += 1
        if limit and checked > limit:
            break
        folder = _folder_for(item_id, product)
        if not folder.exists():
            continue
        for index in range(1, 5):
            gallery = folder / f"Gallery_U{index}.png"
            if not gallery.exists():
                continue
            col = cols[f"Gallery_U{index}_Path"]
            new_value = str(gallery.resolve())
            if ws.cell(row, col).value != new_value:
                ws.cell(row, col).value = new_value
                changed += 1
    wb.save(EBAY_BOOK)
    wb.close()
    print(f"[GALLERY-PATHS] checked={checked} cells_changed={changed}")
    return checked, changed


def audit_listing_gallery_paths():
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index for index, header in enumerate(headers)}
    bad = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        product = _product_type(row[cols["Product_Type"]])
        if product not in {"Poster", "Acrylic"}:
            continue
        total += 1
        paths = [str(row[cols[f"Gallery_U{index}_Path"]] or "") for index in range(1, 5)]
        if any("_U" in path and "Gallery_U" not in path for path in paths):
            bad.append(str(row[cols["ID"]]))
    wb.close()
    print(f"[GALLERY-AUDIT] non_sticker_rows={total} raw_u_gallery_rows={len(bad)}")
    if bad:
        print("[GALLERY-AUDIT] sample=" + ", ".join(bad[:20]))
    return total, bad


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--product-type", choices=["Poster", "Acrylic"], default=None)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--audit-only", action="store_true")
    args = parser.parse_args()
    if not args.audit_only:
        checked, built, _ = ensure_gallery_assets(args.product_type, args.limit)
        print(f"[GALLERY-ASSETS] checked={checked} rebuilt={built}")
        update_listing_gallery_paths(args.limit)
    audit_listing_gallery_paths()


if __name__ == "__main__":
    main()
