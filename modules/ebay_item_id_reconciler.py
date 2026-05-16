from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATABASE = ROOT / "Database"
BOOK = DATABASE / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE / "Performance_Log.csv"
OUT_LOG = DATABASE / "eBay_Item_ID_Reconcile_Log.csv"

FIELDS = [
    "Timestamp",
    "Mode",
    "ID",
    "Product_Type",
    "Title",
    "Old_eBay_Item_ID",
    "New_eBay_Item_ID",
    "Result",
]


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def latest_performance_map() -> tuple[str, dict[str, str]]:
    if not PERFORMANCE_LOG.exists():
        return "", {}
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return "", {}
    latest = max(str(row.get("Snapshot_Timestamp") or "") for row in rows)
    title_to_id: dict[str, str] = {}
    for row in rows:
        if str(row.get("Snapshot_Timestamp") or "") != latest:
            continue
        title = norm(row.get("Title"))
        item_id = str(row.get("Item_ID") or "").strip()
        if not title or not item_id:
            continue
        lower_title = title.lower()
        if "toothbrush" in lower_title or "dumbbell" in lower_title:
            continue
        title_to_id.setdefault(title, item_id)
    return latest, title_to_id


def write_log(rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    exists = OUT_LOG.exists()
    with OUT_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def run(limit: int = 0, execute: bool = False, product_types: set[str] | None = None) -> int:
    snapshot, title_to_id = latest_performance_map()
    if not title_to_id:
        raise RuntimeError("No latest Performance_Log title/id map available.")
    wb = load_workbook(BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    if "eBay_Item_ID" not in cols:
        ws.cell(1, ws.max_column + 1).value = "eBay_Item_ID"
        cols["eBay_Item_ID"] = ws.max_column
    logs: list[dict[str, str]] = []
    changed = 0
    try:
        for row_idx in range(2, ws.max_row + 1):
            item_id = str(ws.cell(row_idx, cols["ID"]).value or "").strip()
            title = str(ws.cell(row_idx, cols["Title"]).value or "").strip()
            product_type = str(ws.cell(row_idx, cols.get("Product_Type", 0)).value or "").strip()
            if not item_id or not title:
                continue
            if product_types and product_type not in product_types:
                continue
            new_ebay_id = title_to_id.get(norm(title))
            if not new_ebay_id:
                continue
            old_ebay_id = str(ws.cell(row_idx, cols["eBay_Item_ID"]).value or "").strip()
            if old_ebay_id == new_ebay_id:
                continue
            logs.append(
                {
                    "Timestamp": datetime.now().isoformat(timespec="seconds"),
                    "Mode": "EXECUTE" if execute else "DRY_RUN",
                    "ID": item_id,
                    "Product_Type": product_type,
                    "Title": title,
                    "Old_eBay_Item_ID": old_ebay_id,
                    "New_eBay_Item_ID": new_ebay_id,
                    "Result": f"matched latest Performance_Log snapshot {snapshot}",
                }
            )
            if execute:
                ws.cell(row_idx, cols["eBay_Item_ID"]).value = new_ebay_id
            changed += 1
            if limit and changed >= limit:
                break
        if execute and changed:
            wb.save(BOOK)
    finally:
        wb.close()
    write_log(logs)
    print(f"[EBAY-ID-RECONCILE] snapshot={snapshot} changes={changed} execute={execute} log={OUT_LOG}")
    return changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--product-types", default="Poster,Acrylic")
    args = parser.parse_args()
    product_types = {item.strip() for item in args.product_types.split(",") if item.strip()}
    run(limit=args.limit, execute=args.execute, product_types=product_types)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
