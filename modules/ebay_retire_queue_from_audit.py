"""Queue old eBay listings for retirement after replacement live audit passes."""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DATABASE_DIR = PROJECT_ROOT / "Database"
REPLACEMENT_QUEUE = DATABASE_DIR / "eBay_Cover_Replacement_Queue.csv"
ONLINE_AUDIT = DATABASE_DIR / "eBay_Online_Cover_Audit.csv"
RETIRE_QUEUE = DATABASE_DIR / "eBay_Retire_Queue.csv"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"

PASS_RESULTS = {"LIKELY_COVER", "LIKELY_COVER_OFFICIAL"}

RETIRE_HEADERS = [
    "Timestamp",
    "Old_ID",
    "Old_eBay_Item_ID",
    "Old_Printify_Product_ID",
    "Replacement_ID",
    "Replacement_eBay_Item_ID",
    "Reason",
    "Status",
    "Retire_Attempted_At",
    "Retire_Result",
]


def now_text() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def workbook_ids() -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cols = {name: idx for idx, name in enumerate(headers) if name}
        data: dict[str, dict[str, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[cols["ID"]]:
                continue
            row_id = str(row[cols["ID"]] or "").strip()
            data[row_id] = {
                "Printify_Product_ID": str(row[cols.get("Printify_Product_ID", -1)] or "").strip() if "Printify_Product_ID" in cols else "",
                "eBay_Item_ID": str(row[cols.get("eBay_Item_ID", -1)] or "").strip() if "eBay_Item_ID" in cols else "",
                "Status": str(row[cols.get("Status", -1)] or "").strip() if "Status" in cols else "",
            }
        return data
    finally:
        wb.close()


def latest_audit_by_id() -> dict[str, dict[str, str]]:
    latest: dict[str, dict[str, str]] = {}
    for row in read_csv(ONLINE_AUDIT):
        row_id = str(row.get("ID") or "").strip()
        if row_id:
            latest[row_id] = row
    return latest


def build(limit: int = 0, dry_run: bool = False) -> int:
    replacement_rows = read_csv(REPLACEMENT_QUEUE)
    retire_rows = read_csv(RETIRE_QUEUE)
    existing_old_ids = {str(row.get("Old_ID") or "").strip() for row in retire_rows}
    existing_replacements = {str(row.get("Replacement_ID") or "").strip() for row in retire_rows}
    workbook = workbook_ids()
    audits = latest_audit_by_id()

    added = 0
    for repl in replacement_rows:
        old_id = str(repl.get("ID") or "").strip()
        replacement_id = str(repl.get("Replacement_SKU") or "").strip()
        status = str(repl.get("Replacement_Status") or "").strip()
        if status not in {"READY_TO_REPLACE_VERIFIED", "REPLACEMENT_PUBLISHED_LIVE_PASS"}:
            continue
        if old_id in existing_old_ids or replacement_id in existing_replacements:
            continue
        audit = audits.get(replacement_id) or {}
        result = str(audit.get("Result") or "").strip()
        if result not in PASS_RESULTS:
            continue
        replacement_ebay = str(audit.get("eBay_Item_ID") or "").strip() or workbook.get(replacement_id, {}).get("eBay_Item_ID", "")
        if not replacement_ebay:
            continue
        old_ebay = str(repl.get("Old_eBay_Item_ID") or "").strip()
        old_printify = str(repl.get("Old_Printify_Product_ID") or "").strip()
        new_row = {
            "Timestamp": now_text(),
            "Old_ID": old_id,
            "Old_eBay_Item_ID": old_ebay,
            "Old_Printify_Product_ID": old_printify,
            "Replacement_ID": replacement_id,
            "Replacement_eBay_Item_ID": replacement_ebay,
            "Reason": f"Cover-only replacement passed live buyer-page audit ({result}); old listing had U/detail main-image risk.",
            "Status": "WAIT_SAFE_END_LISTING_PATH",
            "Retire_Attempted_At": "",
            "Retire_Result": "",
        }
        retire_rows.append(new_row)
        existing_old_ids.add(old_id)
        existing_replacements.add(replacement_id)
        added += 1
        print(f"[RETIRE-QUEUE] {old_id} -> {replacement_id} replacement_ebay={replacement_ebay}")
        if limit and added >= limit:
            break

    if added and not dry_run:
        write_csv(RETIRE_QUEUE, retire_rows, RETIRE_HEADERS)
    print(f"[RETIRE-QUEUE-DONE] added={added} dry_run={dry_run}")
    return added


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    build(limit=args.limit, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
