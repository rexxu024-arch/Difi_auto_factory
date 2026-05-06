import csv
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUTPUT_CSV = DATABASE_DIR / "Local_Listing_QA.csv"

EXPECTED = {
    "Sticker": {"production_min": 1500, "cover_min": 1000},
    "Poster": {"production_size": (3600, 5400), "cover_ratio": 2 / 3},
    "Acrylic": {"production_size": (1538, 2138), "cover_ratio": 5 / 7},
}

HEADERS = [
    "ID",
    "Product_Type",
    "Status",
    "Issue_Count",
    "Issues",
    "Production_Exists",
    "Production_Size",
    "Cover_Exists",
    "Cover_Size",
    "Gallery_All_Exist",
    "Gallery_Paths_Derived",
    "Title_Length",
    "Title_Length_OK",
    "Image_Note_Ready",
]


def _clean(value):
    return str(value or "").strip()


def _image_size(path):
    if not path or not Path(path).exists():
        return None, "missing"
    try:
        with Image.open(path) as image:
            image.load()
            return image.size, ""
    except Exception as exc:
        return None, f"corrupt:{exc}"


def _ratio_ok(size, expected, tolerance=0.035):
    if not size:
        return False
    width, height = size
    if not width or not height:
        return False
    return abs((width / height) - expected) <= tolerance


def _qa_row(row, cols):
    item_id = _clean(row[cols["ID"]])
    product = _clean(row[cols["Product_Type"]]) or "Sticker"
    status = _clean(row[cols["Status"]])
    title = _clean(row[cols["Title"]])
    desc = _clean(row[cols["Description"]]).lower()
    production = _clean(row[cols["Production_Path"]])
    cover = _clean(row[cols["Cover_Path"]])
    gallery_paths = [_clean(row[cols[f"Gallery_U{index}_Path"]]) for index in range(1, 5)]
    issues = []

    production_size, production_error = _image_size(production)
    cover_size, cover_error = _image_size(cover)
    if not production_size:
        issues.append("missing_production" if production_error == "missing" else "corrupt_production")
    if not cover_size:
        issues.append("missing_cover" if cover_error == "missing" else "corrupt_cover")

    if product in {"Poster", "Acrylic"}:
        expected_size = EXPECTED[product]["production_size"]
        if production_size and production_size != expected_size:
            issues.append(f"production_size_{production_size}_expected_{expected_size}")
        if cover_size and not _ratio_ok(cover_size, EXPECTED[product]["cover_ratio"]):
            issues.append(f"cover_ratio_{cover_size}")
        if any("Gallery_U" not in path for path in gallery_paths):
            issues.append("non_sticker_gallery_not_derived")
    else:
        if production_size and min(production_size) < EXPECTED["Sticker"]["production_min"]:
            issues.append(f"sticker_production_low_{production_size}")
        if cover_size and min(cover_size) < EXPECTED["Sticker"]["cover_min"]:
            issues.append(f"sticker_cover_low_{cover_size}")

    gallery_exist = all(path and Path(path).exists() for path in gallery_paths)
    if not gallery_exist:
        issues.append("missing_gallery")

    title_ok = 75 <= len(title) <= 79 if status.startswith("Printify_Published") else True
    if not title_ok:
        issues.append(f"title_length_{len(title)}")
    image_note = "main image shows the actual product customers receive" in desc
    if product in {"Poster", "Acrylic"} and not image_note:
        issues.append("missing_image_note")

    return {
        "ID": item_id,
        "Product_Type": product,
        "Status": status,
        "Issue_Count": len(issues),
        "Issues": "; ".join(issues),
        "Production_Exists": bool(production_size),
        "Production_Size": f"{production_size[0]}x{production_size[1]}" if production_size else "",
        "Cover_Exists": bool(cover_size),
        "Cover_Size": f"{cover_size[0]}x{cover_size[1]}" if cover_size else "",
        "Gallery_All_Exist": gallery_exist,
        "Gallery_Paths_Derived": all("Gallery_U" in path for path in gallery_paths) if product in {"Poster", "Acrylic"} else True,
        "Title_Length": len(title),
        "Title_Length_OK": title_ok,
        "Image_Note_Ready": image_note,
    }


def build():
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index for index, header in enumerate(headers)}
    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[cols["ID"]]:
                rows.append(_qa_row(row, cols))
    finally:
        wb.close()

    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[LOCAL-QA] rows={len(rows)} csv={OUTPUT_CSV}")
    issue_rows = [row for row in rows if row["Issue_Count"]]
    print(f"[LOCAL-QA] issue_rows={len(issue_rows)}")
    for row in issue_rows[:20]:
        print(f"[LOCAL-QA] {row['ID']} {row['Issues']}")
    return rows


if __name__ == "__main__":
    build()
