import argparse
import csv
import math
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageOps

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
OUTPUT_ROOT = PROJECT_ROOT / "Output" / "Digital" / "Printable_Wall_Art"
INDEX_PATH = PROJECT_ROOT / "Database" / "Digital_Printable_Pack_Index.csv"

PRINT_SPECS = [
    ("2x3", (3600, 5400), "12x18, 16x24, 20x30"),
    ("3x4", (3600, 4800), "9x12, 12x16, 18x24"),
    ("4x5", (3600, 4500), "8x10, 12x15, 16x20"),
    ("5x7", (3600, 5040), "5x7, 10x14"),
    ("11x14", (3300, 4200), "11x14"),
]

MAX_ZIP_MB = 19.0


def _slug(value):
    text = re.sub(r"[^A-Za-z0-9]+", "-", str(value or "")).strip("-").lower()
    return text[:80] or "printable-wall-art"


def _clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _fit_cover(image, size):
    src_w, src_h = image.size
    dst_w, dst_h = size
    scale = max(dst_w / src_w, dst_h / src_h)
    resized = image.resize((math.ceil(src_w * scale), math.ceil(src_h * scale)), Image.Resampling.LANCZOS)
    left = max(0, (resized.width - dst_w) // 2)
    top = max(0, (resized.height - dst_h) // 2)
    return resized.crop((left, top, left + dst_w, top + dst_h))


def _polish(image):
    image = ImageOps.exif_transpose(image).convert("RGB")
    image = ImageEnhance.Sharpness(image).enhance(1.04)
    image = ImageEnhance.Contrast(image).enhance(1.02)
    return image


def _load_candidates(limit=10, ids=None):
    wanted = {item.strip() for item in (ids or []) if item.strip()}
    existing = set()
    if INDEX_PATH.exists() and not wanted:
        with INDEX_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                item_id = _clean(row.get("ID"))
                if item_id:
                    existing.add(item_id)
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {header: index for index, header in enumerate(headers)}
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        item_id = str(row[cols["ID"]]).strip()
        if wanted and item_id not in wanted:
            continue
        if not wanted and item_id in existing:
            continue
        product_type = str(row[cols.get("Product_Type")] or "").strip()
        status = str(row[cols.get("Status")] or "").strip()
        production_path = Path(str(row[cols.get("Production_Path")] or ""))
        if product_type != "Poster":
            continue
        if not production_path.exists():
            continue
        if not (status.startswith("Printify_Published") or status.startswith("Printify_UI_Mockups")):
            continue
        rows.append(
            {
                "ID": item_id,
                "Title": _clean(row[cols.get("Title")] or item_id),
                "Description": _clean(row[cols.get("Description")] or ""),
                "Production_Path": production_path,
                "Status": status,
            }
        )
        if limit and len(rows) >= limit:
            break
    workbook.close()
    return rows


def _write_readme(folder, row):
    text = f"""Digital Printable Wall Art Pack

Title:
{row['Title']}

Files included:
- 2x3 ratio JPG for 12x18, 16x24, 20x30
- 3x4 ratio JPG for 9x12, 12x16, 18x24
- 4x5 ratio JPG for 8x10, 12x15, 16x20
- 5x7 ratio JPG for 5x7, 10x14
- 11x14 ratio JPG

Production note:
This is a digital download. No physical item is shipped.

AI disclosure:
This artwork is an original AI-assisted design curated, edited, formatted, and prepared by Quiet Relic Studio / OpenClaw.

Use:
Personal use only. Do not resell, redistribute, or upload the files as a competing digital product.
"""
    path = folder / "README_License_and_Printing_Note.txt"
    path.write_text(text, encoding="utf-8")
    return path


def _zip_folder(folder, zip_path, quality):
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=7) as archive:
        for path in sorted(folder.glob("*.jpg")):
            archive.write(path, path.name)
        readme = folder / "README_License_and_Printing_Note.txt"
        if readme.exists():
            archive.write(readme, readme.name)
    return zip_path.stat().st_size / (1024 * 1024)


def _build_one(row, force=False):
    folder = OUTPUT_ROOT / f"{row['ID']}_{_slug(row['Title'])}"
    folder.mkdir(parents=True, exist_ok=True)
    zip_path = folder.with_suffix(".zip")
    if zip_path.exists() and not force:
        return folder, zip_path, zip_path.stat().st_size / (1024 * 1024), "exists"

    with Image.open(row["Production_Path"]) as source:
        source = _polish(source)
        quality = 92
        while quality >= 82:
            for ratio, size, _label in PRINT_SPECS:
                image = _fit_cover(source, size)
                out = folder / f"{row['ID']}_{ratio}_printable_wall_art.jpg"
                image.save(out, "JPEG", quality=quality, optimize=True, progressive=True, dpi=(300, 300))
            _write_readme(folder, row)
            zip_mb = _zip_folder(folder, zip_path, quality)
            if zip_mb <= MAX_ZIP_MB:
                return folder, zip_path, zip_mb, f"quality_{quality}"
            quality -= 4
    return folder, zip_path, zip_path.stat().st_size / (1024 * 1024), "oversize_review"


def _append_index(records):
    exists = INDEX_PATH.exists()
    with INDEX_PATH.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Timestamp",
                "ID",
                "Title",
                "Source_Path",
                "Pack_Folder",
                "Zip_Path",
                "Zip_MB",
                "Build_Status",
                "Suggested_Etsy_Price",
                "Listing_Status",
            ],
        )
        if not exists:
            writer.writeheader()
        writer.writerows(records)


def build(limit=10, ids=None, force=False):
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    rows = _load_candidates(limit=limit, ids=ids)
    done = 0
    for row in rows:
        try:
            folder, zip_path, zip_mb, status = _build_one(row, force=force)
            record = {
                "Timestamp": datetime.now().isoformat(timespec="seconds"),
                "ID": row["ID"],
                "Title": row["Title"],
                "Source_Path": str(row["Production_Path"]),
                "Pack_Folder": str(folder),
                "Zip_Path": str(zip_path),
                "Zip_MB": f"{zip_mb:.2f}",
                "Build_Status": status,
                "Suggested_Etsy_Price": "$6.99",
                "Listing_Status": "LOCAL_READY_NOT_PUBLISHED",
            }
            print(f"[DIGITAL-PACK] {row['ID']} zip={zip_mb:.2f}MB {status}")
        except Exception as exc:
            record = {
                "Timestamp": datetime.now().isoformat(timespec="seconds"),
                "ID": row["ID"],
                "Title": row["Title"],
                "Source_Path": str(row["Production_Path"]),
                "Pack_Folder": "",
                "Zip_Path": "",
                "Zip_MB": "",
                "Build_Status": f"ERROR_{type(exc).__name__}: {_clean(exc)}"[:120],
                "Suggested_Etsy_Price": "$6.99",
                "Listing_Status": "HOLD_SOURCE_IMAGE_REVIEW",
            }
            print(f"[DIGITAL-PACK-HOLD] {row['ID']} {record['Build_Status']}")
        _append_index([record])
        done += 1
    print(f"[DONE] digital printable pack records={done}")
    return done


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--ids", default="", help="Comma-separated Poster IDs.")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()]
    build(limit=args.limit, ids=ids, force=args.force)


if __name__ == "__main__":
    main()
