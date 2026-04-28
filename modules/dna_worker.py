import argparse
import json
import os
import random
import re
import shutil
import sys
import tempfile
import time
from copy import copy
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from config import CLAUDE_API_KEY, BASE_URL

MENTOR_FILE = Path("Database") / "Mentor_Hub.xlsx"
PRODUCTION_FILE = Path("Database") / "Production_Line.xlsx"
PENDING_FILE = Path("Database") / "pending_tasks.txt"
PROCESSED_LOG = Path("Database") / "processed_dna.log"

MENTOR_COLUMNS = ["Category", "Layout", "Title", "Gold_Prompt_DNA", "Material_Keywords"]
PRODUCTION_COLUMNS = [
    "ID",
    "Timestamp",
    "Category",
    "Product_Type",
    "Style",
    "Title",
    "MJ_Prompt",
    "SEO_Hook",
    "Status",
]

MJ_SUFFIX = "--v 6.1 --ar 9:16 --style raw --no skin, person, blurry edges, text, watermark"
CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-5")
STOP_WORDS = {
    "the",
    "and",
    "with",
    "from",
    "into",
    "white",
    "background",
    "isolated",
    "style",
    "raw",
    "skin",
    "person",
    "text",
    "watermark",
    "blurry",
    "edges",
    "hyper",
    "detailed",
}


class DnaWorkerError(RuntimeError):
    pass


def root_path(relative_path):
    return ROOT_DIR / relative_path


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def task_label(task):
    if isinstance(task, dict):
        return normalize_text(
            task.get("Sub_Category")
            or task.get("Category")
            or task.get("Title")
            or json.dumps(task, ensure_ascii=False)
        )
    return normalize_text(task)


def slug_family(task):
    token = re.split(r"[-_\s]+", task_label(task))[0] or "Zen"
    return re.sub(r"[^A-Za-z0-9]+", "", token) or "Zen"


def excel_timestamp():
    epoch = datetime(1899, 12, 30)
    return (datetime.now() - epoch).total_seconds() / 86400


def load_headers(path, required):
    workbook = load_workbook(root_path(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
    finally:
        workbook.close()
    missing = [column for column in required if column not in headers]
    if missing:
        raise DnaWorkerError(f"{path} missing columns: {', '.join(missing)}")
    return headers


def validate_schema():
    load_headers(MENTOR_FILE, MENTOR_COLUMNS)
    load_headers(PRODUCTION_FILE, PRODUCTION_COLUMNS)


def read_pending_tasks():
    path = root_path(PENDING_FILE)
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("", encoding="utf-8")
        return []
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            return [parsed]
    except json.JSONDecodeError:
        pass
    return [line.strip() for line in raw.splitlines() if line.strip()]


def remove_completed_task(completed_task):
    path = root_path(PENDING_FILE)
    raw = path.read_text(encoding="utf-8") if path.exists() else ""
    try:
        parsed = json.loads(raw) if raw.strip() else []
        if isinstance(parsed, list):
            removed = False
            kept = []
            completed_key = json.dumps(completed_task, ensure_ascii=False, sort_keys=True)
            for item in parsed:
                item_key = json.dumps(item, ensure_ascii=False, sort_keys=True)
                if not removed and item_key == completed_key:
                    removed = True
                    continue
                kept.append(item)
            temp_path = path.with_suffix(".txt.tmp")
            temp_path.write_text(
                json.dumps(kept, ensure_ascii=False, indent=2) + ("\n" if kept else ""),
                encoding="utf-8",
            )
            os.replace(temp_path, path)
            return
    except json.JSONDecodeError:
        pass
    lines = raw.splitlines()
    removed = False
    kept = []
    completed_label = task_label(completed_task)
    for line in lines:
        if not removed and line.strip() == completed_label:
            removed = True
            continue
        kept.append(line)
    temp_path = path.with_suffix(".txt.tmp")
    temp_path.write_text("\n".join(kept) + ("\n" if kept else ""), encoding="utf-8")
    os.replace(temp_path, path)


def build_row_map(sheet):
    headers = [cell.value for cell in sheet[1]]
    return {name: index + 1 for index, name in enumerate(headers)}


def read_mentor_rows():
    workbook = load_workbook(root_path(MENTOR_FILE), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        columns = build_row_map(sheet)
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item = {name: row[columns[name] - 1] for name in MENTOR_COLUMNS}
            if item.get("Category") and item.get("Gold_Prompt_DNA"):
                rows.append(item)
        return rows
    finally:
        workbook.close()


def match_mentor_protocol(task, mentor_rows):
    seed_protocol = None
    if isinstance(task, dict) and task.get("Gold_Prompt_DNA"):
        material_source = normalize_text(task.get("Logic_Protocol") or task.get("Gold_Prompt_DNA"))
        material_match = re.search(r"Material:\s*([^.;]+)", material_source, re.I)
        seed_protocol = {
            "Category": task_label(task),
            "Layout": "Isolated",
            "Title": normalize_text(task.get("Title") or task_label(task)),
            "Gold_Prompt_DNA": normalize_text(task.get("Gold_Prompt_DNA")),
            "Material_Keywords": normalize_text(material_match.group(1) if material_match else material_source),
        }
    task_key = task_label(task).lower()
    task_family = slug_family(task).lower()
    exact = [row for row in mentor_rows if normalize_text(row["Category"]).lower() == task_key]
    if exact:
        return ([seed_protocol] if seed_protocol else []) + exact[:5]
    contains = [
        row
        for row in mentor_rows
        if task_key in normalize_text(row["Category"]).lower()
        or task_key in normalize_text(row["Title"]).lower()
    ]
    if contains:
        return ([seed_protocol] if seed_protocol else []) + contains[:5]
    family = [row for row in mentor_rows if normalize_text(row["Category"]).lower().startswith(task_family)]
    if family:
        return ([seed_protocol] if seed_protocol else []) + family[:5]
    if seed_protocol:
        return [seed_protocol]
    raise DnaWorkerError(f"No Logic_Protocol match in Mentor_Hub.xlsx for task: {task}")


def strip_mj_suffix(prompt):
    prompt = normalize_text(prompt)
    prompt = re.sub(r"\s--v\s+\S+", "", prompt)
    prompt = re.sub(r"\s--ar\s+\S+", "", prompt)
    prompt = re.sub(r"\s--style\s+\S+", "", prompt)
    prompt = re.sub(r"\s--tile\b", "", prompt)
    prompt = re.sub(r"\s--no\s+.*$", "", prompt)
    return normalize_text(prompt).rstrip(",")


def enforce_prompt(prompt, material_keywords):
    prompt = strip_mj_suffix(prompt)
    lower_prompt = prompt.lower()
    for keyword in split_keywords(material_keywords)[:4]:
        if keyword.lower() not in lower_prompt:
            prompt = f"{prompt}, {keyword}"
    if "isolated on white" not in prompt.lower() and "solid white background" not in prompt.lower():
        prompt = f"{prompt}, isolated on white background"
    return f"{prompt} {MJ_SUFFIX}"


def split_keywords(value):
    parts = re.split(r"[,;|/]+", normalize_text(value))
    return [part.strip() for part in parts if part.strip()]


def seo_keywords(title, prompt, material_keywords):
    seed = split_keywords(material_keywords)
    text = f"{title} {strip_mj_suffix(prompt)}"
    words = [
        word.lower()
        for word in re.findall(r"[A-Za-z][A-Za-z-]{2,}", text)
        if word.lower() not in STOP_WORDS and len(word) > 2
    ]
    result = []
    for item in seed + words + ["sticker", "jade art", "kintsugi", "mythic decor", "vinyl decal"]:
        key = normalize_text(item).lower()
        if key and key not in result:
            result.append(key)
        if len(result) >= 15:
            break
    return ", ".join(result[:15])


def fingerprint(text):
    words = sorted(
        {
            word.lower()
            for word in re.findall(r"[A-Za-z][A-Za-z-]{3,}", normalize_text(text))
            if word.lower() not in STOP_WORDS
        }
    )
    return " ".join(words[:24])


def load_processed_fingerprints():
    seen = set()
    log_path = root_path(PROCESSED_LOG)
    if log_path.exists():
        for line in log_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                seen.add(line.strip())
    production_path = root_path(PRODUCTION_FILE)
    if production_path.exists():
        workbook = load_workbook(production_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            columns = build_row_map(sheet)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                prompt = row[columns["MJ_Prompt"] - 1] if columns.get("MJ_Prompt") else ""
                hook = row[columns["SEO_Hook"] - 1] if columns.get("SEO_Hook") else ""
                if prompt:
                    seen.add(fingerprint(f"{prompt} {hook}"))
        finally:
            workbook.close()
    return seen


def append_processed_log(rows):
    log_path = root_path(PROCESSED_LOG)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        for row in rows:
            handle.write(fingerprint(f"{row['MJ_Prompt']} {row['SEO_Hook']}") + "\n")


def next_sequence():
    workbook = load_workbook(root_path(PRODUCTION_FILE), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        columns = build_row_map(sheet)
        id_column = columns["ID"]
        last_number = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            value = row[id_column - 1]
            match = re.search(r"(\d+)$", str(value or ""))
            if match:
                last_number = int(match.group(1))
        return last_number + 1
    finally:
        workbook.close()


def extract_json_array(text):
    text = normalize_text(text)
    fenced = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.I | re.S)
    if fenced:
        return json.loads(fenced.group(1))
    start = text.find("[")
    end = text.rfind("]")
    if start >= 0 and end > start:
        return json.loads(text[start : end + 1])
    raise ValueError("API response did not contain a JSON array")


class ClaudeClient:
    def __init__(self, api_key, base_url, retries=3, timeout=120):
        self.api_key = api_key
        self.base_url = (base_url or "").rstrip("/")
        self.retries = retries
        self.timeout = timeout

    def generate(self, task, protocols):
        if not self.api_key:
            raise DnaWorkerError("CLAUDE_API_KEY is empty")
        payload = self._payload(task, protocols)
        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        url = f"{self.base_url}/v1/messages"
        last_error = None
        for attempt in range(1, self.retries + 1):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=self.timeout)
                if response.status_code >= 400:
                    raise DnaWorkerError(f"{response.status_code} {response.text[:800]}")
                body = response.json()
                text = "\n".join(
                    block.get("text", "")
                    for block in body.get("content", [])
                    if block.get("type") == "text"
                )
                return extract_json_array(text)
            except Exception as exc:
                last_error = exc
                if attempt < self.retries:
                    time.sleep(2 * attempt)
        raise DnaWorkerError(f"API generation failed after retries: {last_error}")

    def _payload(self, task, protocols):
        references = []
        for row in protocols[:5]:
            references.append(
                {
                    "Category": row["Category"],
                    "Layout": row["Layout"],
                    "Title": row["Title"],
                    "Gold_Prompt_DNA": row["Gold_Prompt_DNA"],
                    "Material_Keywords": row["Material_Keywords"],
                }
            )
        prompt = {
            "task": task,
            "required_count": 20,
            "reference_format": "Academia-Time_Mechanism",
            "rules": [
                "Generate exactly 20 visually unified but distinct DNA variants.",
                "Evolve from Gold_Prompt_DNA, preserving material logic such as Jade and Kintsugi.",
                "Each MJ_Prompt must include material detail, lighting, composition, and isolated on white background.",
                "Return only a JSON array. Each object must contain Title, MJ_Prompt, SEO_Hook.",
                "SEO_Hook must contain 10 to 15 comma-separated keywords.",
            ],
            "mentor_protocols": references,
        }
        return {
            "model": CLAUDE_MODEL,
            "max_tokens": 8000,
            "temperature": 0.72,
            "messages": [{"role": "user", "content": json.dumps(prompt, ensure_ascii=False)}],
        }


class MockClient:
    def __init__(self):
        self.calls = 0

    def generate(self, task, protocols):
        self.calls += 1
        if self.calls == 1:
            raise requests.Timeout("simulated retry failure")
        base = protocols[0]
        materials = split_keywords(base["Material_Keywords"])
        variants = []
        angles = [
            "orbital relic",
            "vertical talisman",
            "floating core",
            "fractured mandala",
            "celestial mechanism",
            "ritual compass",
            "jade shard crown",
            "luminous sigil",
            "ancient seal plate",
            "spiral ascension",
            "relic gateway",
            "halo fragment",
            "sacred geometry shell",
            "moonlit archive",
            "radiant axis",
            "kinetic emblem",
            "imperial charm",
            "crystalized aura",
            "mythic instrument",
            "silent monument",
        ]
        for index, angle in enumerate(angles, 1):
            material = materials[(index - 1) % max(len(materials), 1)] if materials else "Imperial Jade"
            title = f"{task} {angle.title()} {index:02d}"
            prompt = (
                f"{angle} evolved from {base['Title']}, {material}, Kintsugi gold vein inlays, "
                "soft rim lighting, centered vertical composition, clean contour silhouette, "
                "isolated on white background"
            )
            variants.append(
                {
                    "Title": title,
                    "MJ_Prompt": prompt,
                    "SEO_Hook": seo_keywords(title, prompt, base["Material_Keywords"]),
                }
            )
        return variants


class RetryingClient:
    def __init__(self, client, retries=3):
        self.client = client
        self.retries = retries

    def generate(self, task, protocols):
        last_error = None
        for attempt in range(1, self.retries + 1):
            try:
                return self.client.generate(task, protocols)
            except Exception as exc:
                last_error = exc
                if attempt < self.retries:
                    time.sleep(0.1 * attempt)
        raise DnaWorkerError(f"API generation failed after retries: {last_error}")


def normalize_variants(raw_variants, task, protocols, seen):
    rows = []
    reference = protocols[0]
    material_keywords = reference["Material_Keywords"]
    for index, item in enumerate(raw_variants, 1):
        if len(rows) >= 20:
            break
        title = normalize_text(item.get("Title") or f"{task} DNA Variant {index:02d}")
        prompt = enforce_prompt(item.get("MJ_Prompt") or reference["Gold_Prompt_DNA"], material_keywords)
        seo = normalize_text(item.get("SEO_Hook")) or seo_keywords(title, prompt, material_keywords)
        seo_parts = split_keywords(seo)
        if len(seo_parts) < 10 or len(seo_parts) > 15:
            seo = seo_keywords(title, prompt, material_keywords)
        fp = fingerprint(f"{prompt} {seo}")
        if fp in seen:
            prompt = enforce_prompt(f"{strip_mj_suffix(prompt)}, unique variant index {index:02d}", material_keywords)
            fp = fingerprint(f"{prompt} {seo}")
        if fp in seen:
            continue
        seen.add(fp)
        rows.append({"Title": title, "MJ_Prompt": prompt, "SEO_Hook": seo})
    if len(rows) != 20:
        raise DnaWorkerError(f"Expected 20 non-duplicate DNA variants, got {len(rows)}")
    return rows


def append_production_rows(task, variants):
    family = slug_family(task)
    start_number = next_sequence()
    timestamp = excel_timestamp()
    output_rows = []
    for offset, variant in enumerate(variants):
        output_rows.append(
            {
                "ID": f"Sticker-{family}-{start_number + offset:04d}",
                "Timestamp": timestamp,
                "Category": family,
                "Product_Type": "Sticker",
                "Style": "V13.0 DNA Worker",
                "Title": variant["Title"],
                "MJ_Prompt": variant["MJ_Prompt"],
                "SEO_Hook": variant["SEO_Hook"],
                "Status": "Completed",
            }
        )
    save_with_retry(output_rows)
    return output_rows


def clone_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)


def save_with_retry(output_rows, retries=5):
    path = root_path(PRODUCTION_FILE)
    last_error = None
    for attempt in range(1, retries + 1):
        workbook = None
        try:
            workbook = load_workbook(path)
            sheet = workbook.active
            columns = build_row_map(sheet)
            missing = [column for column in PRODUCTION_COLUMNS if column not in columns]
            if missing:
                raise DnaWorkerError(f"Production_Line.xlsx missing columns: {', '.join(missing)}")
            template_row = sheet.max_row
            for row_data in output_rows:
                target_row = sheet.max_row + 1
                for column_name in PRODUCTION_COLUMNS:
                    column_index = columns[column_name]
                    target_cell = sheet.cell(row=target_row, column=column_index)
                    clone_style(sheet.cell(row=template_row, column=column_index), target_cell)
                    target_cell.value = row_data[column_name]
            workbook.save(path)
            return
        except PermissionError as exc:
            last_error = exc
            try:
                save_with_excel_com(output_rows)
                return
            except Exception as com_exc:
                last_error = com_exc
            time.sleep(attempt)
        finally:
            if workbook:
                workbook.close()
    raise DnaWorkerError(f"Excel write failed after retries: {last_error}")


def save_with_excel_com(output_rows):
    import win32com.client

    path = str(root_path(PRODUCTION_FILE))
    workbook = win32com.client.GetObject(path)
    sheet = workbook.Worksheets(1)
    columns = {}
    column_index = 1
    while True:
        value = sheet.Cells(1, column_index).Value
        if value in (None, ""):
            break
        columns[str(value)] = column_index
        column_index += 1
    missing = [column for column in PRODUCTION_COLUMNS if column not in columns]
    if missing:
        raise DnaWorkerError(f"Production_Line.xlsx missing columns: {', '.join(missing)}")

    xl_up = -4162
    id_column = columns["ID"]
    last_row = sheet.Cells(sheet.Rows.Count, id_column).End(xl_up).Row
    template_row = last_row
    for row_data in output_rows:
        target_row = last_row + 1
        source = sheet.Range(sheet.Cells(template_row, 1), sheet.Cells(template_row, len(columns)))
        target = sheet.Range(sheet.Cells(target_row, 1), sheet.Cells(target_row, len(columns)))
        source.Copy()
        target.PasteSpecial(-4122)
        for column_name in PRODUCTION_COLUMNS:
            sheet.Cells(target_row, columns[column_name]).Value = row_data[column_name]
        last_row = target_row
    workbook.Application.CutCopyMode = False
    workbook.Save()


def process_task(task, client, mentor_rows, seen):
    protocols = match_mentor_protocol(task, mentor_rows)
    raw_variants = client.generate(task, protocols)
    variants = normalize_variants(raw_variants, task, protocols, seen)
    rows = append_production_rows(task, variants)
    append_processed_log(rows)
    remove_completed_task(task)
    return rows


def run(client=None, max_tasks=None):
    os.chdir(ROOT_DIR)
    validate_schema()
    tasks = read_pending_tasks()
    if not tasks:
        return 0
    mentor_rows = read_mentor_rows()
    seen = load_processed_fingerprints()
    active_client = client or ClaudeClient(CLAUDE_API_KEY, BASE_URL)
    processed = 0
    for task in tasks:
        process_task(task, active_client, mentor_rows, seen)
        processed += 1
        if max_tasks and processed >= max_tasks:
            break
    return processed


def self_test():
    os.chdir(ROOT_DIR)
    validate_schema()
    old_cwd = Path.cwd()
    with tempfile.TemporaryDirectory(prefix="dna_worker_") as tmp_dir:
        tmp_root = Path(tmp_dir)
        (tmp_root / "Database").mkdir()
        shutil.copy2(root_path(MENTOR_FILE), tmp_root / MENTOR_FILE)
        shutil.copy2(root_path(PRODUCTION_FILE), tmp_root / PRODUCTION_FILE)
        (tmp_root / PENDING_FILE).write_text("Zen-Mythic-Beast-Azure-Dragon\n", encoding="utf-8")
        old_root = globals()["ROOT_DIR"]
        globals()["ROOT_DIR"] = tmp_root
        try:
            processed = run(client=RetryingClient(MockClient()), max_tasks=1)
            if processed != 1:
                raise DnaWorkerError("self-test did not process one task")
            workbook = load_workbook(tmp_root / PRODUCTION_FILE, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                if headers[: len(PRODUCTION_COLUMNS)] != PRODUCTION_COLUMNS:
                    raise DnaWorkerError("self-test production headers changed")
                rows = list(sheet.iter_rows(min_row=sheet.max_row - 19, max_row=sheet.max_row, values_only=True))
                if len(rows) != 20:
                    raise DnaWorkerError("self-test did not append 20 rows")
                status_index = headers.index("Status")
                prompt_index = headers.index("MJ_Prompt")
                if any(row[status_index] != "Completed" for row in rows):
                    raise DnaWorkerError("self-test status mismatch")
                if any(MJ_SUFFIX not in row[prompt_index] for row in rows):
                    raise DnaWorkerError("self-test prompt suffix mismatch")
            finally:
                workbook.close()
            pending = (tmp_root / PENDING_FILE).read_text(encoding="utf-8").strip()
            if pending:
                raise DnaWorkerError("self-test pending queue was not cleared")
        finally:
            globals()["ROOT_DIR"] = old_root
            os.chdir(old_cwd)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--max-tasks", type=int, default=None)
    args = parser.parse_args()
    if args.self_test:
        self_test()
        print("SELF_TEST_OK")
        return
    processed = run(max_tasks=args.max_tasks)
    print(f"DNA_WORKER_PROCESSED={processed}")


if __name__ == "__main__":
    main()
