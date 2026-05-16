from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import Config


DATABASE = ROOT / "Database"
REPORTS = ROOT / "Reports"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
OUT_CSV = DATABASE / "Printify_Sales_Channel_Audit.csv"
OUT_MD = REPORTS / "Printify_Sales_Channel_Audit.md"
NY = ZoneInfo("America/New_York")


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def now_et() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}", "Content-Type": "application/json"}


def load_rows() -> list[dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(raw_headers)}
    rows: list[dict[str, str]] = []
    try:
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {header: values[cols[header]] if cols[header] < len(values) else "" for header in raw_headers}
            ptype = clean(row.get("Product_Type"))
            pid = clean(row.get("Printify_Product_ID"))
            if pid and not ptype.lower().startswith("stick"):
                rows.append(row)
    finally:
        wb.close()
    return rows


def fetch_product(product_id: str) -> tuple[int, dict | None, str]:
    try:
        response = requests.get(
            f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
            headers=headers(),
            timeout=60,
        )
    except Exception as exc:
        return 0, None, f"{type(exc).__name__}: {exc}"
    if not 200 <= response.status_code < 300:
        return response.status_code, None, response.text[:500]
    try:
        return response.status_code, response.json(), ""
    except json.JSONDecodeError as exc:
        return response.status_code, None, f"JSON_ERROR: {exc}"


def state(value) -> str:
    if value is True:
        return "FREE_SHIPPING_TRUE"
    if value is False:
        return "FREE_SHIPPING_FALSE"
    return "FREE_SHIPPING_MISSING"


def run(limit: int = 0) -> int:
    rows = load_rows()
    if limit:
        rows = rows[:limit]
    out: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        product_id = clean(row.get("Printify_Product_ID"))
        status, product, error = fetch_product(product_id)
        props = (product or {}).get("sales_channel_properties") or {}
        external = (product or {}).get("external") or {}
        item = {
            "Timestamp": now_et(),
            "Workbook_ID": clean(row.get("ID")),
            "Product_Type": clean(row.get("Product_Type")),
            "Workbook_Status": clean(row.get("Status")),
            "Printify_Product_ID": product_id,
            "HTTP_Status": str(status),
            "Printify_Title": clean((product or {}).get("title")),
            "External_ID": clean(external.get("id")),
            "External_Type": clean(external.get("type")),
            "Visible": clean((product or {}).get("visible")),
            "Locked": clean((product or {}).get("is_locked")),
            "Free_Shipping_State": state(props.get("free_shipping")),
            "Sales_Channel_Properties": json.dumps(props, ensure_ascii=False, sort_keys=True),
            "Error": error,
        }
        out.append(item)
        print(
            f"[PRINTIFY-SALES-AUDIT] {index}/{len(rows)} {item['Workbook_ID']} "
            f"{item['Free_Shipping_State']} external={item['External_ID']}",
            flush=True,
        )

    fields = [
        "Timestamp",
        "Workbook_ID",
        "Product_Type",
        "Workbook_Status",
        "Printify_Product_ID",
        "HTTP_Status",
        "Printify_Title",
        "External_ID",
        "External_Type",
        "Visible",
        "Locked",
        "Free_Shipping_State",
        "Sales_Channel_Properties",
        "Error",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(out)
    counts = Counter(row["Free_Shipping_State"] for row in out)
    lines = [
        "# Printify Sales Channel Audit",
        "",
        f"Generated: {now_et()}",
        f"Rows: {len(out)}",
        "",
        "## Free Shipping State",
        "",
    ]
    for key, value in counts.most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(
        [
            "",
            "## Use",
            "",
            "- `FREE_SHIPPING_FALSE` products are not suitable for eBay traffic tests until rebuilt or channel settings are repaired.",
            "- Future products should be created with `sales_channel_properties.free_shipping=true` from the source payload.",
            "",
            f"CSV: {OUT_CSV}",
        ]
    )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[PRINTIFY-SALES-AUDIT-DONE] rows={len(out)} csv={OUT_CSV} md={OUT_MD}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()
    return run(limit=args.limit)


if __name__ == "__main__":
    raise SystemExit(main())
