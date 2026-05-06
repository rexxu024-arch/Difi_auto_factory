import argparse
import asyncio
import io
import json
import sys
import time
import urllib.request
from pathlib import Path

import requests
import websockets
from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
CHROME_DEBUG_URL = "http://127.0.0.1:9222"
TARGET_VARIANT_ID = 45754
VARIANT_LABELS = {
    "Sticker": ['6" × 6", White', '6" x 6", White', '6" × 6"'],
    "Poster": ['12″ x 18″ / Matte', '12" x 18" / Matte', '12″ x 18″', '12" x 18"'],
    "Acrylic": ["5'' × 7'' (Vertical)", "5'' x 7'' (Vertical)", "5 × 7", "5x7"],
}


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}"}


def _fetch_product(product_id):
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def _publish_product_images(product_id):
    response = requests.post(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
        headers={**_headers(), "Content-Type": "application/json"},
        json={
            "title": False,
            "description": False,
            "images": True,
            "variants": False,
            "tags": False,
            "keyFeatures": False,
            "shipping_template": False,
        },
        timeout=180,
    )
    response.raise_for_status()
    return response


def _selected_count(product):
    return sum(1 for image in product.get("images") or [] if image.get("is_selected_for_publishing") is not False)


def _default_count(product):
    return sum(
        1
        for image in product.get("images") or []
        if image.get("is_selected_for_publishing") is not False and image.get("is_default")
    )


def _official_mockup_count(product):
    return sum(
        1
        for image in product.get("images") or []
        if image.get("is_selected_for_publishing") is not False and "images.printify.com/mockup" in str(image.get("src") or "")
    )


def _selected_images(product):
    return [image for image in product.get("images") or [] if image.get("is_selected_for_publishing") is not False]


def _ahash(image):
    image = image.convert("L").resize((16, 16), Image.Resampling.LANCZOS)
    pixels = list(image.getdata())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if pixel > avg else "0" for pixel in pixels)


def _distance(a, b):
    return sum(left != right for left, right in zip(a, b))


def _first_selected_matches_cover(product, cover_path, threshold=8):
    selected = _selected_images(product)
    if not selected or not cover_path or not Path(cover_path).exists():
        return False
    url = selected[0].get("src")
    if not url:
        return False
    local_hash = _ahash(Image.open(cover_path))
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    remote_hash = _ahash(Image.open(io.BytesIO(response.content)))
    return _distance(local_hash, remote_hash) <= threshold


def _load_rows(limit=0, ids=None, allow_any_status=False):
    ids = set(ids or [])
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        data = {header: ws.cell(row_idx, cols[header]).value for header in headers}
        data["_row_idx"] = row_idx
        item_id = str(data.get("ID") or "").strip()
        if ids and item_id not in ids:
            continue
        if not allow_any_status and not ids:
            if data.get("Status") not in {"Ready_for_Printify", "Printify_BaseStaged_DefaultMockups3", "Printify_UI_Failed", "Printify_PrimaryFix_Needed"}:
                continue
        if data.get("Printify_Product_ID") and data.get("Status") != "Ready_for_Printify":
            rows.append(data)
        elif data.get("Printify_Product_ID") and data.get("Status") == "Ready_for_Printify":
            rows.append(data)
        if limit and len(rows) >= limit:
            break
    return wb, ws, headers, cols, rows


def _set_status(ws, cols, row_idx, status):
    ws.cell(row_idx, cols["Status"]).value = status
    if "Timestamp" in cols:
        ws.cell(row_idx, cols["Timestamp"]).value = time.strftime("%#m/%#d/%Y  %#I:%M:%S %p")


def _assets(row):
    display_paths = [Path(row["Cover_Path"])]
    display_paths.extend(Path(row[f"Gallery_U{i}_Path"]) for i in range(1, 5))
    paths = display_paths
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("; ".join(missing))
    return [str(path.resolve()) for path in paths]


def _product_type(row):
    value = str(row.get("Product_Type") or "Sticker").strip()
    if value.lower().startswith("poster"):
        return "Poster"
    if value.lower().startswith("acry"):
        return "Acrylic"
    return "Sticker"


def _target_ws(product_id):
    with urllib.request.urlopen(f"{CHROME_DEBUG_URL}/json/list", timeout=10) as response:
        pages = json.load(response)
    target_urls = (f"product-details/{product_id}", f"products/{product_id}")
    for page in pages:
        page_url = page.get("url", "")
        if any(target_url in page_url for target_url in target_urls):
            return page["webSocketDebuggerUrl"]
    url = f"https://printify.com/app/mockup-library/shops/{Config.Printify_SHOP_ID}/products/{product_id}?revealUploads=true"
    tab = next(
        (page for page in pages if "printify.com/app/" in page.get("url", "") and page.get("webSocketDebuggerUrl")),
        None,
    )
    if tab and tab.get("webSocketDebuggerUrl"):
        return tab["webSocketDebuggerUrl"]
    request = urllib.request.Request(f"{CHROME_DEBUG_URL}/json/new", data=url.encode("utf-8"), method="PUT")
    tab = json.loads(urllib.request.urlopen(request, timeout=10).read().decode("utf-8", "ignore"))
    if tab.get("webSocketDebuggerUrl"):
        return tab["webSocketDebuggerUrl"]
    raise RuntimeError(f"Open Printify product page first: {product_id}")


class CdpPage:
    def __init__(self, ws_url):
        self.ws_url = ws_url
        self.seq = 1
        self.sock = None

    async def __aenter__(self):
        self.sock = await websockets.connect(self.ws_url, max_size=20_000_000)
        await self.send("Runtime.enable")
        await self.send("DOM.enable")
        await self.send("Page.enable")
        return self

    async def __aexit__(self, exc_type, exc, tb):
        await self.sock.close()

    async def send(self, method, params=None, timeout=30):
        msg = {"id": self.seq, "method": method}
        if params is not None:
            msg["params"] = params
        await self.sock.send(json.dumps(msg))
        my_id = self.seq
        self.seq += 1
        deadline = time.time() + timeout
        while True:
            remaining = max(1, deadline - time.time())
            data = json.loads(await asyncio.wait_for(self.sock.recv(), timeout=remaining))
            if data.get("id") == my_id:
                return data

    async def eval(self, expression):
        result = await self.send("Runtime.evaluate", {"expression": expression, "returnByValue": True, "awaitPromise": True})
        if "exceptionDetails" in result.get("result", {}):
            details = result["result"]["exceptionDetails"]
            message = details.get("text", "Runtime.evaluate failed")
            exception = details.get("exception") or {}
            description = exception.get("description") or exception.get("value")
            if description:
                message = f"{message}: {description}"
            raise RuntimeError(message)
        return result["result"]["result"].get("value")

    async def navigate(self, url):
        await self.send("Page.navigate", {"url": url})

    async def click(self, x, y):
        await self.send("Input.dispatchMouseEvent", {"type": "mouseMoved", "x": x, "y": y})
        await self.send("Input.dispatchMouseEvent", {"type": "mousePressed", "x": x, "y": y, "button": "left", "clickCount": 1})
        await self.send("Input.dispatchMouseEvent", {"type": "mouseReleased", "x": x, "y": y, "button": "left", "clickCount": 1})

    async def set_file_input(self, files):
        doc = await self.send("DOM.getDocument", {"depth": 1, "pierce": True})
        root = doc["result"]["root"]["nodeId"]
        query = await self.send("DOM.querySelectorAll", {"nodeId": root, "selector": "input[type=file]"})
        node_ids = query["result"].get("nodeIds") or []
        if not node_ids:
            raise RuntimeError("Printify upload file input not found")
        await self.send("DOM.setFileInputFiles", {"nodeId": node_ids[-1], "files": files})
        await self.eval(
            r"""(() => {
                const input = [...document.querySelectorAll('input[type=file]')].slice(-1)[0];
                if (!input) return false;
                input.dispatchEvent(new Event('input', {bubbles:true}));
                input.dispatchEvent(new Event('change', {bubbles:true}));
                return input.files ? input.files.length : 0;
            })()"""
        )


async def _dismiss_connect_dialog(page):
    dismissed = await page.eval(
        r"""(() => {
            const text = (document.body && document.body.innerText) || '';
            if (!/Connect your mockups|Unlinked mockups/.test(text)) return false;
            const buttons = [...document.querySelectorAll('button')]
              .filter(e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length));
            const close = buttons.reverse().find(e => (e.innerText || '').trim() === 'close'
              || (e.getAttribute('aria-label') || '').toLowerCase().includes('close'));
            if (!close) return false;
            close.click();
            return true;
        })()"""
    )
    if dismissed:
        await asyncio.sleep(1)
    return dismissed


async def _open_upload_slot(page):
    opened = await page.eval(
        r"""(() => {
            const visible = e => !!(e.offsetWidth || e.offsetHeight || e.getClientRects().length);
            const buttons = [...document.querySelectorAll('button,[role="button"]')].filter(visible);
            const add = buttons.find(e => /Add image/i.test(e.innerText || e.ariaLabel || ''));
            if (add) {
                add.click();
                return 'ADD_IMAGE';
            }
            const browse = buttons.find(e => /Browse/i.test(e.innerText || e.ariaLabel || ''));
            if (browse) {
                browse.click();
                return 'BROWSE';
            }
            return '';
        })()"""
    )
    if opened:
        await asyncio.sleep(1)
    return opened


async def _set_files_and_wait_for_continue(page, files, wait_checks=30):
    attempts = []
    for attempt in range(2):
        slot = await _open_upload_slot(page)
        await page.set_file_input(files)
        for _ in range(wait_checks):
            upload_ready = await page.eval(
                """((expectedFiles) => {
                    const inputs=[...document.querySelectorAll('input[type=file]')];
                    const file_count=inputs.reduce((n,i)=>Math.max(n, i.files ? i.files.length : 0), 0);
                    const buttons=[...document.querySelectorAll('button')]
                      .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                    const continue_buttons=buttons.filter(e=>(e.innerText||'').trim()==='Continue');
                    return {
                        slot: '',
                        file_count,
                        enabled: continue_buttons.some(e=>!e.disabled),
                        continue_count: continue_buttons.length
                    };
                })(""" + str(len(files)) + """)"""
            )
            upload_ready["slot"] = slot
            attempts.append(upload_ready)
            if upload_ready and upload_ready.get("file_count") == len(files) and upload_ready.get("enabled"):
                return upload_ready
            await asyncio.sleep(1)
    raise RuntimeError(f"Upload files did not enable Continue: {attempts[-4:]}")


async def _upload_mockups_once(product_id, files, keep_default_mockups=False, expected_count=5, publish=False, product_type="Sticker"):
    ws_url = _target_ws(product_id)
    async with CdpPage(ws_url) as page:
        await page.navigate(f"https://printify.com/app/mockup-library/shops/{Config.Printify_SHOP_ID}/products/{product_id}?revealUploads=true")
        for _ in range(30):
            if await page.eval("!!location.href && /\\/auth\\/login/.test(location.href)"):
                raise RuntimeError("Printify login required in Codex browser")
            if await page.eval("!!document.body && /Mockup library/.test(document.body.innerText || '')"):
                break
            await asyncio.sleep(1)
        await _dismiss_connect_dialog(page)
        opened = await page.eval(
            """(() => {
                const buttons=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                const b=buttons.find(e=>/Upload/.test(e.innerText||e.ariaLabel||''));
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not opened:
            raise RuntimeError("Upload button not found")
        for _ in range(15):
            if await page.eval("!!document.body && /Upload mockups/.test(document.body.innerText || '') && document.querySelectorAll('input[type=file]').length"):
                break
            await asyncio.sleep(1)

        await _set_files_and_wait_for_continue(page, files, wait_checks=90)
        if keep_default_mockups:
            await page.eval(
                r"""(() => {
                    const labels=[...document.querySelectorAll('label,div,span')].filter(e=>/Replace previous mockup selection/.test(e.innerText||''));
                    const target=labels[0];
                    if (!target) return false;
                    const row=target.closest('label,div') || target;
                    const input=row.querySelector('input[type=checkbox]') || document.querySelector('input[type=checkbox]:checked:last-of-type');
                    if (input && input.checked) input.click();
                    return true;
                })()"""
            )
        continued = await page.eval(
            """(() => {
                const buttons=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length)
                    && (e.innerText||'').trim()==='Continue' && !e.disabled);
                const b=buttons[buttons.length-1];
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not continued:
            raise RuntimeError("Continue button not found")
        for _ in range(30):
            if await page.eval("!!document.body && /Unlinked mockups|Confirm/.test(document.body.innerText || '')"):
                break
            await asyncio.sleep(1)

        for _ in range(10):
            state = await page.eval("""(() => ({unlinked:(document.body.innerText.match(/Unlinked mockups/g)||[]).length}))()""")
            if not state["unlinked"]:
                break
            await page.eval(
                """(() => {const c=document.querySelector('.side-panel-content'); if(!c)return; const b=[...document.querySelectorAll('[role=combobox]')].find(e=>/Unlinked/.test(e.innerText||'')); if(!b)return; const r=b.getBoundingClientRect(); if(r.y>560)c.scrollTop += 180; if(r.y<180)c.scrollTop -= 120;})()"""
            )
            await asyncio.sleep(0.25)
            box = await page.eval(
                r"""(() => {const b=[...document.querySelectorAll('[role="combobox"]')].filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length)&&/Unlinked/.test(e.innerText||''))[0]; if(!b)return null; const r=b.getBoundingClientRect(); return {x:r.x+r.width/2,y:r.y+r.height/2};})()"""
            )
            if not box:
                raise RuntimeError("Unlinked combobox not found")
            labels = json.dumps(VARIANT_LABELS.get(product_type, VARIANT_LABELS["Sticker"]))
            selected = await page.eval(
                r"""(async (labels) => {
                    const clean = s => (s || '').replace(/\s+/g, ' ').trim().replace(/Ã—/g, '×');
                    const box=[...document.querySelectorAll('[role="combobox"]')]
                      .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length))
                      .find(e=>/Unlinked/.test(e.innerText||''));
                    if(!box) return false;
                    box.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                    box.click();
                    await new Promise(resolve => setTimeout(resolve, 350));
                    const candidates=[...document.querySelectorAll('[role="option"],pfy-dropdown-option-v2')]
                      .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                    const o=candidates.find(e=>labels.some(label=>clean(e.innerText).includes(clean(label)))) || candidates[0];
                    if(!o)return false;
                    o.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                    o.click();
                    return true;
                })(""" + labels + """)"""
            )
            if not selected:
                await page.click(box["x"], min(660, box["y"] + 52))
            await asyncio.sleep(0.8)

        final_state = await page.eval("""(() => ({unlinked:(document.body.innerText.match(/Unlinked mockups/g)||[]).length}))()""")
        if final_state["unlinked"]:
            raise RuntimeError(f"Mockups still unlinked: {final_state['unlinked']}")

        await page.eval("const c=document.querySelector('.side-panel-content'); if(c)c.scrollTop=c.scrollHeight;")
        await asyncio.sleep(0.4)
        confirmed = await page.eval(
            """(() => {
                const b=[...document.querySelectorAll('button')].find(e=>(e.innerText||'').trim()==='Confirm');
                if(!b)return false;
                const r=b.getBoundingClientRect();
                b.click();
                return {ok:true,x:r.x+r.width/2,y:r.y+r.height/2};
            })()"""
        )
        if not confirmed:
            raise RuntimeError("Confirm button not found")
        await asyncio.sleep(8)
        selected = await page.eval(
            r"""(() => {
                const text = (document.body && document.body.innerText) || '';
                return ((text.match(/Selected mockups\s+(\d+)\s+selected/) || text.match(/(\d+)\s+selected/)) || [])[1] || null;
            })()"""
        )
        if selected is not None:
            try:
                selected_int = int(selected)
            except ValueError:
                selected_int = -1
            if selected_int < expected_count:
                raise RuntimeError(f"Unexpected selected mockups count after confirm: {selected}, expected at least {expected_count}")
        primary_count = await page.eval(
            r"""(() => {
                const candidates=[...document.querySelectorAll('button.mockup-container, button')]
                    .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length))
                    .filter(e=>(e.innerText||'').trim()==='check' || e.className.toString().includes('mockup'));
                if (candidates.length) candidates[0].click();
                return candidates.length;
            })()"""
        )
        await asyncio.sleep(1)
        saved = await page.eval(
            """(() => {
                const labels=['Save as draft','Save selection'];
                const b=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length))
                  .find(e=>labels.includes((e.innerText||'').trim()));
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not saved:
            raise RuntimeError("Save button not found")
        await asyncio.sleep(8)
        if publish:
            published = await page.eval(
                """(() => {const b=[...document.querySelectorAll('button')].find(e=>(e.innerText||'').trim()==='Publish'); if(!b)return false; b.click(); return true;})()"""
            )
            if not published:
                _publish_product_images(product_id)
            else:
                await asyncio.sleep(8)


async def _upload_files_to_library(product_id, files, keep_default_mockups=True):
    ws_url = _target_ws(product_id)
    async with CdpPage(ws_url) as page:
        await page.navigate(f"https://printify.com/app/mockup-library/shops/{Config.Printify_SHOP_ID}/products/{product_id}?revealUploads=true")
        for _ in range(30):
            if await page.eval("!!location.href && /\\/auth\\/login/.test(location.href)"):
                raise RuntimeError("Printify login required in Codex browser")
            if await page.eval("!!document.body && /Mockup library/.test(document.body.innerText || '')"):
                break
            await asyncio.sleep(1)
        await _dismiss_connect_dialog(page)
        opened = await page.eval(
            """(() => {
                const buttons=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length));
                const b=buttons.find(e=>/Upload/.test(e.innerText||e.ariaLabel||''));
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not opened:
            raise RuntimeError("Upload button not found")
        for _ in range(15):
            if await page.eval("!!document.body && /Upload mockups/.test(document.body.innerText || '') && document.querySelectorAll('input[type=file]').length"):
                break
            await asyncio.sleep(1)
        await _set_files_and_wait_for_continue(page, files, wait_checks=60)
        if keep_default_mockups:
            await page.eval(
                r"""(() => {
                    const labels=[...document.querySelectorAll('label,div,span')].filter(e=>/Replace previous mockup selection/.test(e.innerText||''));
                    const target=labels[0];
                    if (!target) return false;
                    const row=target.closest('label,div') || target;
                    const input=row.querySelector('input[type=checkbox]') || document.querySelector('input[type=checkbox]:checked:last-of-type');
                    if (input && input.checked) input.click();
                    return true;
                })()"""
            )
        continued = await page.eval(
            """(() => {
                const buttons=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length)
                    && (e.innerText||'').trim()==='Continue' && !e.disabled);
                const b=buttons[buttons.length-1];
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not continued:
            raise RuntimeError("Continue button not found")
        expected_library_count = len(files) + 1 if keep_default_mockups else len(files)
        for _ in range(60):
            ready = await page.eval(
                """((expectedCount) => {
                    const text=(document.body && document.body.innerText) || '';
                    const uploadCount=[...document.querySelectorAll('button.mockup-container, .mockup-container')]
                      .filter(e=>{const r=e.getBoundingClientRect(); return r.width>100 && r.height>100;}).length;
                    return /My Uploads/.test(text) && uploadCount >= expectedCount && !/Upload mockups/.test(text);
                })(""" + str(expected_library_count) + """)"""
            )
            if ready:
                return
            await asyncio.sleep(1)
        raise RuntimeError("Uploaded files did not return to My Uploads library")


async def _select_latest_library_mockups(product_id, add_count, expected_count=5, publish=False):
    ws_url = _target_ws(product_id)
    async with CdpPage(ws_url) as page:
        await page.navigate(f"https://printify.com/app/mockup-library/shops/{Config.Printify_SHOP_ID}/products/{product_id}?revealUploads=true")
        for _ in range(30):
            if await page.eval("!!location.href && /\\/auth\\/login/.test(location.href)"):
                raise RuntimeError("Printify login required in Codex browser")
            ready = await page.eval(
                """((addCount) => {
                    if (!(document.body && /Mockup library/.test(document.body.innerText || '') && /My Uploads/.test(document.body.innerText || ''))) return false;
                    const items=[...document.querySelectorAll('button.mockup-container, .mockup-container')]
                      .filter(e=>{const r=e.getBoundingClientRect(); return r.width>100 && r.height>100 && r.x < 1100;});
                    const unselected=items.filter(e=>{
                        const ctrl=e.querySelector('[data-testid="checkboxWrapper"], [role="checkbox"]');
                        return ctrl && ctrl.getAttribute('aria-checked') !== 'true';
                    }).length;
                    return items.length >= addCount + 1 && unselected >= addCount;
                })(""" + str(add_count) + """)"""
            )
            if ready:
                break
            await asyncio.sleep(1)
        await _dismiss_connect_dialog(page)
        selected = await page.eval(
            """(async (addCount) => {
                [...document.querySelectorAll('button')].filter(b=>(b.innerText||b.ariaLabel||'').trim()==='close').slice(-5).forEach(b=>b.click());
                await new Promise(r=>setTimeout(r, 250));
                const items=[...document.querySelectorAll('button.mockup-container, .mockup-container')]
                  .filter(e=>{const r=e.getBoundingClientRect(); return r.width>100 && r.height>100 && r.x < 1100;});
                let changed=0;
                for (const e of items) {
                    if (changed >= addCount) break;
                    const ctrl=e.querySelector('[data-testid="checkboxWrapper"], [role="checkbox"]');
                    if (ctrl && ctrl.getAttribute('aria-checked') !== 'true') {
                        ctrl.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
                        ctrl.click();
                        changed++;
                        await new Promise(r=>setTimeout(r, 220));
                    }
                }
                await new Promise(r=>setTimeout(r, 700));
                return {
                    changed,
                    selected:items.filter(e=>{
                        const ctrl=e.querySelector('[data-testid="checkboxWrapper"], [role="checkbox"]');
                        return ctrl && ctrl.getAttribute('aria-checked') === 'true';
                    }).length
                };
            })(""" + str(add_count) + """)"""
        )
        if not selected or selected.get("changed") != add_count:
            raise RuntimeError(f"Could not select latest library mockups: {selected}")
        saved = await page.eval(
            """(() => {
                const b=[...document.querySelectorAll('button')]
                  .filter(e=>!!(e.offsetWidth||e.offsetHeight||e.getClientRects().length))
                  .find(e=>['Save selection','Save as draft'].includes((e.innerText||'').trim()) && !e.disabled);
                if(!b)return false;
                b.click();
                return true;
            })()"""
        )
        if not saved:
            raise RuntimeError("Save button not found")
        await asyncio.sleep(10)
        product = _fetch_product(product_id)
        count = _selected_count(product)
        if count < expected_count:
            raise RuntimeError(f"API selected mockup count is {count}, expected at least {expected_count}")
        if publish:
            published = await page.eval(
                """(() => {const b=[...document.querySelectorAll('button')].find(e=>(e.innerText||'').trim()==='Publish'); if(!b)return false; b.click(); return true;})()"""
            )
            if not published:
                _publish_product_images(product_id)
            else:
                await asyncio.sleep(8)


async def _upload_mockups(product_id, files, keep_default_mockups=False, expected_count=5, publish=False, product_type="Sticker"):
    if len(files) == 5:
        cover = [files[0]]
        gallery = files[1:]
        await _upload_mockups_once(
            product_id,
            cover,
            keep_default_mockups=True,
            expected_count=1,
            publish=False,
            product_type=product_type,
        )
        await _upload_files_to_library(product_id, gallery, keep_default_mockups=True)
        await _select_latest_library_mockups(product_id, add_count=len(gallery), expected_count=expected_count, publish=publish)
        return
    await _upload_mockups_once(
        product_id,
        files,
        keep_default_mockups=keep_default_mockups,
        expected_count=expected_count,
        publish=publish,
        product_type=product_type,
    )


def upload_from_open_page(limit=0, expected_count=5, publish=False, ids=None, allow_any_status=False):
    wb, ws, headers, cols, rows = _load_rows(limit, ids=ids, allow_any_status=allow_any_status)
    done = 0
    try:
        for row in rows:
            product_id = str(row.get("Printify_Product_ID") or "").strip()
            if not product_id:
                continue
            item_id = row["ID"]
            try:
                files = _assets(row)
                asyncio.run(_upload_mockups(product_id, files, expected_count=expected_count, publish=publish, product_type=_product_type(row)))
                product = _fetch_product(product_id)
                count = _selected_count(product)
                if count < expected_count:
                    raise RuntimeError(f"API selected mockup count is {count}, expected at least {expected_count}")
                defaults = _default_count(product)
                if defaults < 1:
                    raise RuntimeError("API default mockup count is 0, expected at least 1")
                old_status = str(row.get("Status") or "")
                if publish and old_status.startswith("Printify_Published"):
                    _set_status(ws, cols, row["_row_idx"], f"Printify_Published_Mockups{expected_count}")
                else:
                    _set_status(ws, cols, row["_row_idx"], f"Printify_UI_Mockups{expected_count}")
                wb.save(EBAY_BOOK)
                done += 1
                print(f"[MOCKUP-UI] {item_id} product={product_id} selected_mockups={expected_count}")
            except Exception as exc:
                if ids or allow_any_status:
                    _set_status(ws, cols, row["_row_idx"], "Printify_SourceRepair_Failed")
                else:
                    _set_status(ws, cols, row["_row_idx"], "Printify_UI_Failed")
                wb.save(EBAY_BOOK)
                print(f"[MOCKUP-UI-FAIL] {item_id}: {exc}")
                break
    finally:
        wb.close()
    print(f"[DONE] UI mockup uploads: {done}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--expected-count", type=int, default=5)
    parser.add_argument("--publish", action="store_true")
    parser.add_argument("--ids", default="", help="Comma-separated workbook IDs to repair regardless of current status.")
    parser.add_argument("--allow-any-status", action="store_true")
    args = parser.parse_args()
    ids = [value.strip() for value in args.ids.split(",") if value.strip()]
    upload_from_open_page(
        limit=args.limit,
        expected_count=args.expected_count,
        publish=args.publish,
        ids=ids,
        allow_any_status=args.allow_any_status,
    )
