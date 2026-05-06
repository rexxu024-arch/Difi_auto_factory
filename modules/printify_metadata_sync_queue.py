"""Sync queued title/description changes from local workbook to Printify/eBay.

The queue is produced by modules/ebay_traffic_experiment.py. This script is
deliberately narrow: it updates only title + description and asks Printify to
publish only those fields. It does not touch images, variants, pricing, or
shipping templates.
"""

from __future__ import annotations

import argparse
import csv
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
QUEUE_CSV = DATABASE_DIR / "eBay_Metadata_Sync_Queue.csv"
LOG_CSV = DATABASE_DIR / "eBay_Metadata_Sync_Log.csv"

PUBLISH_TITLE_DESC_ONLY = {
    "title": True,
    "description": True,
    "images": False,
    "variants": False,
    "tags": False,
    "keyFeatures": False,
    "shipping_template": False,
}


def headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def clean(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def load_workbook_rows() -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers_row = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers_row)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        item_id = clean(row[cols["ID"]])
        out[item_id] = {
            "ID": item_id,
            "Title": clean(row[cols["Title"]]),
            "Description": clean(row[cols["Description"]]),
            "Printify_Product_ID": clean(row[cols["Printify_Product_ID"]]),
            "eBay_Item_ID": clean(row[cols.get("eBay_Item_ID")]) if "eBay_Item_ID" in cols else "",
        }
    wb.close()
    return out


def load_queue() -> list[dict[str, str]]:
    if not QUEUE_CSV.exists():
        return []
    with QUEUE_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def save_queue(rows: list[dict[str, str]]) -> None:
    if not rows:
        QUEUE_CSV.unlink(missing_ok=True)
        return
    fieldnames = ["Timestamp", "ID", "Product_Type", "Printify_Product_ID", "eBay_Item_ID", "Action", "Status"]
    with QUEUE_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def append_log(row: dict[str, Any]) -> None:
    fieldnames = [
        "Timestamp",
        "ID",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Action",
        "HTTP_Update",
        "HTTP_Publish",
        "Result",
        "Error",
    ]
    exists = LOG_CSV.exists()
    with LOG_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerow({name: row.get(name, "") for name in fieldnames})


def request_with_retry(method: str, url: str, *, json_payload=None, attempts: int = 3) -> requests.Response:
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            response = requests.request(method, url, headers=headers(), json=json_payload, timeout=120)
            if response.status_code >= 500 and attempt < attempts:
                time.sleep(5 * attempt)
                continue
            return response
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt < attempts:
                time.sleep(5 * attempt)
    raise last_exc or RuntimeError("request failed")


def sync(limit: int = 3, dry_run: bool = False, sleep_seconds: float = 4.0) -> int:
    if not Config.Printify_API_KEY:
        raise RuntimeError("Printify_API_KEY is missing")
    workbook_rows = load_workbook_rows()
    queue = load_queue()
    pending = [row for row in queue if clean(row.get("Status")) == "PENDING"]
    selected = pending[:limit] if limit else pending
    base = Config.Printify_API_URL.rstrip("/")
    done_ids = set()
    done = 0
    for qrow in selected:
        item_id = clean(qrow.get("ID"))
        local = workbook_rows.get(item_id)
        if not local:
            append_log(
                {
                    "Timestamp": datetime.now().isoformat(timespec="seconds"),
                    "ID": item_id,
                    "Action": "SYNC_TITLE_DESCRIPTION",
                    "Result": "SKIP",
                    "Error": "missing local workbook row",
                }
            )
            continue
        product_id = clean(qrow.get("Printify_Product_ID")) or local["Printify_Product_ID"]
        if not product_id:
            append_log(
                {
                    "Timestamp": datetime.now().isoformat(timespec="seconds"),
                    "ID": item_id,
                    "Action": "SYNC_TITLE_DESCRIPTION",
                    "Result": "SKIP",
                    "Error": "missing Printify_Product_ID",
                }
            )
            continue
        if dry_run:
            print(f"[META-DRY] {item_id} product={product_id} title={local['Title'][:70]}")
            continue
        update = request_with_retry(
            "PUT",
            f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
            json_payload={"title": local["Title"], "description": local["Description"]},
        )
        publish = request_with_retry(
            "POST",
            f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
            json_payload=PUBLISH_TITLE_DESC_ONLY,
        )
        ok = update.status_code in {200, 201, 202, 204} and publish.status_code in {200, 201, 202, 204}
        append_log(
            {
                "Timestamp": datetime.now().isoformat(timespec="seconds"),
                "ID": item_id,
                "Printify_Product_ID": product_id,
                "eBay_Item_ID": clean(qrow.get("eBay_Item_ID")) or local["eBay_Item_ID"],
                "Action": "SYNC_TITLE_DESCRIPTION",
                "HTTP_Update": update.status_code,
                "HTTP_Publish": publish.status_code,
                "Result": "OK" if ok else "CHECK",
                "Error": "" if ok else (update.text[:300] + " " + publish.text[:300]),
            }
        )
        print(f"[META-SYNC] {item_id} update={update.status_code} publish={publish.status_code}")
        if ok:
            done_ids.add(item_id)
            done += 1
        time.sleep(max(0.0, sleep_seconds))

    if not dry_run and done_ids:
        for row in queue:
            if clean(row.get("ID")) in done_ids:
                row["Status"] = "SYNCED"
        save_queue(queue)
    print(f"[META-DONE] selected={len(selected)} synced={done} dry_run={dry_run}")
    return done


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=3)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sleep-seconds", type=float, default=4.0)
    args = parser.parse_args()
    sync(limit=args.limit, dry_run=args.dry_run, sleep_seconds=args.sleep_seconds)


if __name__ == "__main__":
    main()
