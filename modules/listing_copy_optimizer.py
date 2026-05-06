import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.edit_for_platforms import (
    _build_local_description,
    _clean_text,
    _ensure_image_note,
    _fit_ebay_title,
    _retitle_duplicate,
)
from modules.etsy_launch_builder import _description as _etsy_description
from modules.etsy_launch_builder import _etsy_title, _tags


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUTPUT_CSV = DATABASE_DIR / "Listing_Copy_Optimization.csv"
OUTPUT_XLSX = DATABASE_DIR / "Listing_Copy_Optimization.xlsx"

HEADERS = [
    "ID",
    "Product_Type",
    "Category",
    "Status",
    "Printify_Product_ID",
    "Existing_Title",
    "Existing_Title_Length",
    "Proposed_eBay_Title",
    "Proposed_eBay_Title_Length",
    "Title_Length_OK",
    "Existing_Description_Has_Image_Note",
    "Proposed_eBay_Description",
    "Etsy_Title",
    "Etsy_Tags",
    "Etsy_Tag_Count",
    "Etsy_Description",
    "Needs_Local_Update",
    "Reason",
]


def _row_dicts():
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows = []
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {headers[index]: row[index] for index in range(len(headers))}
            if data.get("ID"):
                rows.append(data)
    finally:
        workbook.close()
    return rows


def _eligible(row, mode):
    status = _clean_text(row.get("Status"))
    if mode == "published":
        return status.startswith("Printify_Published")
    if mode == "stable":
        return status.startswith("Printify_Published") or status.startswith("Printify_UI_Mockups")
    return True


def _metadata(row):
    return {
        "ID": row.get("ID"),
        "Title": row.get("Title"),
        "Category": row.get("Category"),
        "Product_Type": row.get("Product_Type"),
        "SEO_Hook": row.get("DNA Profile"),
        "MJ_Prompt": row.get("DNA Profile"),
        "DNA Profile": row.get("DNA Profile"),
        "Description": row.get("Description"),
        "Price": row.get("Price"),
        "Status": row.get("Status"),
        "Printify_Product_ID": row.get("Printify_Product_ID"),
    }


def build(mode="published", limit=0):
    rows = [row for row in _row_dicts() if _eligible(row, mode)]
    if limit:
        rows = rows[:limit]
    output = []
    used_titles = set()
    for row in rows:
        meta = _metadata(row)
        item_id = _clean_text(row.get("ID"))
        product_type = _clean_text(row.get("Product_Type")) or "Sticker"
        existing_title = _clean_text(row.get("Title"))
        proposed = _fit_ebay_title(existing_title, row.get("DNA Profile"), product_type)
        if proposed in used_titles:
            proposed = _retitle_duplicate(proposed, item_id, row.get("DNA Profile"), product_type, used_titles)
        used_titles.add(proposed)

        existing_desc = _clean_text(row.get("Description"))
        proposed_desc = _ensure_image_note(existing_desc or _build_local_description(meta))
        tag_list = _tags(row)
        reasons = []
        if existing_title != proposed:
            reasons.append("title")
        if "main image shows the actual product customers receive" not in proposed_desc.lower():
            reasons.append("image_note")
        if len(tag_list) != 13:
            reasons.append("etsy_tags")
        output.append(
            {
                "ID": item_id,
                "Product_Type": product_type,
                "Category": _clean_text(row.get("Category")),
                "Status": _clean_text(row.get("Status")),
                "Printify_Product_ID": _clean_text(row.get("Printify_Product_ID")),
                "Existing_Title": existing_title,
                "Existing_Title_Length": len(existing_title),
                "Proposed_eBay_Title": proposed,
                "Proposed_eBay_Title_Length": len(proposed),
                "Title_Length_OK": 75 <= len(proposed) <= 79,
                "Existing_Description_Has_Image_Note": "main image shows the actual product customers receive" in existing_desc.lower(),
                "Proposed_eBay_Description": proposed_desc,
                "Etsy_Title": _etsy_title(row),
                "Etsy_Tags": ", ".join(tag_list),
                "Etsy_Tag_Count": len(tag_list),
                "Etsy_Description": _etsy_description(row),
                "Needs_Local_Update": bool(reasons),
                "Reason": ", ".join(reasons),
            }
        )
    return output


def write_outputs(rows):
    DATABASE_DIR.mkdir(exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Copy Optimization"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 24,
        "B": 14,
        "C": 14,
        "D": 28,
        "F": 58,
        "H": 58,
        "L": 90,
        "M": 68,
        "N": 52,
        "P": 90,
        "R": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    wb.save(OUTPUT_XLSX)
    wb.close()
    print(f"[COPY] rows={len(rows)} csv={OUTPUT_CSV}")
    print(f"[COPY] xlsx={OUTPUT_XLSX}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["published", "stable", "all"], default="published")
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()
    rows = build(mode=args.mode, limit=args.limit)
    write_outputs(rows)
    bad_titles = [row["ID"] for row in rows if not row["Title_Length_OK"]]
    bad_tags = [row["ID"] for row in rows if row["Etsy_Tag_Count"] != 13]
    print(f"[COPY-AUDIT] bad_title_lengths={len(bad_titles)} bad_tag_counts={len(bad_tags)}")
    if bad_titles:
        print("[COPY-AUDIT] bad_title_sample=" + ", ".join(bad_titles[:20]))
    if bad_tags:
        print("[COPY-AUDIT] bad_tag_sample=" + ", ".join(bad_tags[:20]))


if __name__ == "__main__":
    main()
