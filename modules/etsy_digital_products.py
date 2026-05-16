"""Local Etsy digital product R&D and printable art pack builder.

This module intentionally stays local/low-bandwidth. It reads already-approved
OpenClaw artwork from Database/eBay_listing.xlsx and builds candidate digital
download products without touching Etsy, Printify, or eBay.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageOps
from zoneinfo import ZoneInfo


ROOT = Path(__file__).resolve().parents[1]
DATABASE = ROOT / "Database"
OUTPUT = ROOT / "Output" / "Digital"
SOURCE_BOOK = DATABASE / "eBay_listing.xlsx"
NY_TZ = ZoneInfo("America/New_York")
HOLD_LOG = DATABASE / "Etsy_Digital_Source_Hold.csv"


PRINT_RATIOS = [
    {
        "code": "2x3",
        "label": "2x3 ratio - 12x18, 16x24, 20x30",
        "size": (3600, 5400),
    },
    {
        "code": "3x4",
        "label": "3x4 ratio - 9x12, 12x16, 18x24",
        "size": (3600, 4800),
    },
    {
        "code": "4x5",
        "label": "4x5 ratio - 8x10, 16x20",
        "size": (3600, 4500),
    },
    {
        "code": "11x14",
        "label": "11x14 ratio",
        "size": (3300, 4200),
    },
    {
        "code": "A-Series",
        "label": "ISO A ratio - A5, A4, A3",
        "size": (3508, 4961),
    },
]


PRODUCT_LINES = [
    {
        "priority": 1,
        "line": "Printable Wall Art Set",
        "format": "Instant download",
        "fit": 95,
        "support_load": "Low",
        "why": "Reuses high-resolution poster art, no shipping, no buyer photos, immediate Etsy test.",
        "first_offer": "5 JPG ratios per listing, framed as Zen / Dark Academia gallery decor.",
        "pilot_count": 20,
        "price": "$7.99-$12.99 single, $14.99-$24.99 bundle",
        "risk": "Saturated market; must win with coherent niche and premium visuals.",
    },
    {
        "priority": 2,
        "line": "Personalized Ex Libris / Bookplate",
        "format": "Made-to-order digital",
        "fit": 90,
        "support_load": "Medium-low",
        "why": "Dark academia buyers understand library/book rituals; only needs buyer name/text, not face likeness.",
        "first_offer": "Printable bookplate PDF/JPG with buyer name and 1 revision.",
        "pilot_count": 10,
        "price": "$9.99-$18.99",
        "risk": "Needs a simple personalization queue and revision policy.",
    },
    {
        "priority": 3,
        "line": "Phone/Desktop Wallpaper Pack",
        "format": "Instant download",
        "fit": 78,
        "support_load": "Low",
        "why": "Good use for acrylic-style vertical art and detail crops; easy local resizing.",
        "first_offer": "Phone lock screen, desktop, tablet, and square social wallpaper variants.",
        "pilot_count": 10,
        "price": "$3.99-$7.99",
        "risk": "Lower buyer intent than printable wall art.",
    },
    {
        "priority": 4,
        "line": "Custom Pet Familiar Portrait",
        "format": "Made-to-order digital",
        "fit": 72,
        "support_load": "High",
        "why": "Large market and fits 'familiar' / memorial / fantasy portrait angle, but requires buyer photo QA.",
        "first_offer": "Dark Academia Familiar or Zen Guardian portrait, 24-72h delivery, 1 revision.",
        "pilot_count": 5,
        "price": "$24.99-$49.99",
        "risk": "Revision pressure, likeness disputes, customer-photo privacy, more manual inspection.",
    },
    {
        "priority": 5,
        "line": "Custom Family / Couple Relic Portrait",
        "format": "Made-to-order digital",
        "fit": 58,
        "support_load": "Very high",
        "why": "Can sell, but conflicts with current low-touch factory goal.",
        "first_offer": "Defer until pet workflow and revision policy are proven.",
        "pilot_count": 0,
        "price": "$39.99-$89.99",
        "risk": "Likeness, emotional expectations, privacy, many revisions.",
    },
]


@dataclass
class ListingCandidate:
    row: int
    listing_id: str
    product_type: str
    category: str
    title: str
    source_path: Path
    status: str


def slug(text: str, max_len: int = 68) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-")
    return text[:max_len].strip("-") or "quiet-relic-art"


def clean_one_line(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace("\n", " ").replace("\r", " ")).strip()


def log_source_hold(item: ListingCandidate, reason: str) -> None:
    exists = HOLD_LOG.exists()
    with HOLD_LOG.open("a", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["Timestamp", "ID", "Title", "Source_Path", "Reason"],
        )
        if not exists:
            writer.writeheader()
        writer.writerow(
            {
                "Timestamp": datetime.now(NY_TZ).isoformat(timespec="seconds"),
                "ID": item.listing_id,
                "Title": item.title,
                "Source_Path": str(item.source_path),
                "Reason": clean_one_line(reason)[:240],
            }
        )


def read_candidates() -> list[ListingCandidate]:
    wb = load_workbook(SOURCE_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers)}
    out: list[ListingCandidate] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[cols["ID"]]:
            continue
        product_type = row[cols["Product_Type"]]
        status = str(row[cols["Status"]] or "")
        source = Path(str(row[cols["Production_Path"]] or ""))
        if product_type != "Poster":
            continue
        if "Published" not in status and "Mockups" not in status:
            continue
        if not source.exists():
            continue
        out.append(
            ListingCandidate(
                row=row_idx,
                listing_id=str(row[cols["ID"]]),
                product_type=str(product_type),
                category=str(row[cols["Category"]] or ""),
                title=clean_one_line(row[cols["Title"]]),
                source_path=source,
                status=status,
            )
        )
    wb.close()
    return out


def score_candidate(item: ListingCandidate) -> int:
    title = item.title.lower()
    score = 50
    for word in ("zen", "celestial", "dark academia", "astrolabe", "library", "bonsai", "lotus", "alchemy"):
        if word in title:
            score += 8
    if "poster" in title:
        score += 5
    if item.category in ("Zen", "Academia"):
        score += 8
    try:
        with Image.open(item.source_path) as im:
            w, h = im.size
        if w >= 3000 and h >= 4500:
            score += 15
        if 0.62 <= w / h <= 0.70:
            score += 8
    except Exception as exc:
        log_source_hold(item, f"score_open_failed:{type(exc).__name__}:{exc}")
        score -= 50
    return score


def etsy_title(item: ListingCandidate) -> str:
    base = item.title
    base = re.sub(r"\b12x18\b", "", base, flags=re.IGNORECASE)
    base = re.sub(r"\bposter\b", "Printable Wall Art", base, flags=re.IGNORECASE)
    base = clean_one_line(base)
    suffixes = [
        "Digital Download",
        "Zen Study Decor",
        "Dark Academia Print",
        "Printable Gallery Art",
    ]
    title = f"{base} {suffixes[0]}"
    if len(title) < 75:
        title = f"{base} {suffixes[0]} {suffixes[1]}"
    return title[:140]


def etsy_tags(item: ListingCandidate) -> list[str]:
    tags = [
        "digital download",
        "printable art",
        "wall art print",
        "zen decor",
        "study room decor",
        "dark academia",
        "gallery wall",
        "wabi sabi decor",
        "japanese decor",
        "meditation room",
        "scholar decor",
        "instant download",
        "quiet relic",
    ]
    if item.category.lower() == "academia":
        tags[3] = "library decor"
        tags[8] = "vintage study"
    return tags[:13]


def etsy_description(item: ListingCandidate) -> str:
    tags = ", ".join(etsy_tags(item)[:5])
    return (
        f"{etsy_title(item)}\n\n"
        "Instant digital download. No physical item will be shipped.\n\n"
        "You receive five high-resolution JPG files prepared for common frame sizes: "
        "2:3, 3:4, 4:5, 11x14, and ISO A-series. Print at home, through a local print shop, "
        "or through an online photo lab.\n\n"
        "Style notes: quiet relic atmosphere, premium zen / dark academia visual language, "
        "gallery-wall composition, and contemplative study-room decor.\n\n"
        "AI-assisted disclosure: this artwork was created with AI-assisted tools and then curated, "
        "cropped, formatted, and prepared by Quiet Relic Studio.\n\n"
        "Personal use only. Commercial resale, redistribution, and file sharing are not included.\n\n"
        f"Search themes: {tags}."
    )


def center_crop_resize(im: Image.Image, size: tuple[int, int]) -> Image.Image:
    im = ImageOps.exif_transpose(im).convert("RGB")
    target_w, target_h = size
    source_w, source_h = im.size
    target_ratio = target_w / target_h
    source_ratio = source_w / source_h
    if source_ratio > target_ratio:
        new_w = int(source_h * target_ratio)
        left = (source_w - new_w) // 2
        box = (left, 0, left + new_w, source_h)
    else:
        new_h = int(source_w / target_ratio)
        top = (source_h - new_h) // 2
        box = (0, top, source_w, top + new_h)
    cropped = im.crop(box)
    return cropped.resize(size, Image.Resampling.LANCZOS)


def build_printable_pack(item: ListingCandidate, root: Path) -> dict:
    pack_dir = root / f"{item.listing_id}_{slug(item.title, 48)}"
    pack_dir.mkdir(parents=True, exist_ok=True)
    files = []
    try:
        with Image.open(item.source_path) as source:
            source.load()
            source_size = source.size
            for ratio in PRINT_RATIOS:
                image = center_crop_resize(source, ratio["size"])
                filename = f"{slug(item.listing_id)}_{ratio['code']}_QuietRelicStudio.jpg"
                path = pack_dir / filename
                image.save(path, "JPEG", quality=92, optimize=True, dpi=(300, 300))
                files.append(
                    {
                        "ratio": ratio["code"],
                        "label": ratio["label"],
                        "path": str(path),
                        "size_px": f"{ratio['size'][0]}x{ratio['size'][1]}",
                        "size_mb": round(path.stat().st_size / (1024 * 1024), 2),
                        "etsy_file_limit_ok": path.stat().st_size <= 20 * 1024 * 1024,
                    }
                )
    except Exception as exc:
        log_source_hold(item, f"pack_build_failed:{type(exc).__name__}:{exc}")
        raise
    manifest = {
        "generated_at": datetime.now(NY_TZ).isoformat(timespec="seconds"),
        "source_listing_id": item.listing_id,
        "source_title": item.title,
        "source_path": str(item.source_path),
        "source_size_px": f"{source_size[0]}x{source_size[1]}",
        "etsy_title": etsy_title(item),
        "etsy_description": etsy_description(item),
        "etsy_tags": etsy_tags(item),
        "files": files,
    }
    (pack_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    return manifest


def write_csv(path: Path, rows: Iterable[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_rnd_docs(product_rows: list[dict], candidate_rows: list[dict], pack_rows: list[dict]) -> None:
    write_csv(
        DATABASE / "Etsy_Digital_Product_RnD.csv",
        product_rows,
        ["priority", "line", "format", "fit", "support_load", "why", "first_offer", "pilot_count", "price", "risk"],
    )
    write_csv(
        DATABASE / "Etsy_Digital_Candidates.csv",
        candidate_rows,
        [
            "rank",
            "score",
            "ID",
            "Category",
            "Current_Title",
            "Digital_Title",
            "Tags",
            "Source_Path",
            "Status",
        ],
    )
    write_csv(
        DATABASE / "Etsy_Digital_Pilot_Packs.csv",
        pack_rows,
        ["ID", "Digital_Title", "Pack_Dir", "File_Count", "Max_File_MB", "All_Under_20MB", "Source_Path"],
    )
    upload_rows = []
    for row in pack_rows:
        manifest_path = Path(row["Pack_Dir"]) / "manifest.json"
        if not manifest_path.exists():
            continue
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        files = manifest.get("files") or []
        upload_row = {
            "ID": row["ID"],
            "Listing_Type": "Digital",
            "Product_Line": "Printable Wall Art Set",
            "Price_USD": "9.99",
            "Title": manifest.get("etsy_title", row["Digital_Title"]),
            "Description": manifest.get("etsy_description", ""),
            "Tags": ", ".join(manifest.get("etsy_tags") or []),
            "AI_Disclosure": "AI-assisted artwork, curated and formatted by Quiet Relic Studio.",
            "License": "Personal use only; no resale, redistribution, or commercial file use.",
            "Pack_Dir": row["Pack_Dir"],
            "Max_File_MB": row["Max_File_MB"],
            "All_Under_20MB": row["All_Under_20MB"],
            "Source_Path": row["Source_Path"],
        }
        for idx in range(5):
            upload_row[f"Digital_File_{idx + 1}"] = files[idx]["path"] if idx < len(files) else ""
        upload_rows.append(upload_row)
    write_csv(
        DATABASE / "Etsy_Digital_Upload_Queue.csv",
        upload_rows,
        [
            "ID",
            "Listing_Type",
            "Product_Line",
            "Price_USD",
            "Title",
            "Description",
            "Tags",
            "AI_Disclosure",
            "License",
            "Pack_Dir",
            "Digital_File_1",
            "Digital_File_2",
            "Digital_File_3",
            "Digital_File_4",
            "Digital_File_5",
            "Max_File_MB",
            "All_Under_20MB",
            "Source_Path",
        ],
    )
    md = [
        "# Etsy Digital Product R&D",
        "",
        f"Generated: {datetime.now(NY_TZ):%Y-%m-%d %I:%M:%S %p %Z}",
        "",
        "## Decision",
        "",
        "Start with printable wall art instant downloads. This best matches the current Poster pipeline, avoids Printify production cost, avoids shipping, and does not require customer photos or revisions.",
        "",
        "Do not start with digital family portraits. Keep custom pet portraits as a second-stage experiment after a QA/revision queue exists.",
        "",
        "## Pilot Product Lines",
        "",
    ]
    for row in product_rows:
        md.extend(
            [
                f"### {row['priority']}. {row['line']}",
                f"- Format: {row['format']}",
                f"- Fit score: {row['fit']}/100",
                f"- Support load: {row['support_load']}",
                f"- First offer: {row['first_offer']}",
                f"- Suggested price: {row['price']}",
                f"- Risk: {row['risk']}",
                "",
            ]
        )
    md.extend(
        [
            "## Etsy Listing Guardrails",
            "",
            "- Digital download only; no physical item shipped.",
            "- Upload no more than 5 files per listing unless using a ZIP/link workflow.",
            "- Keep each Etsy-uploaded file under 20MB.",
            "- Include AI-assisted disclosure in description.",
            "- Personal-use license only unless a commercial-license variation is deliberately created.",
            "",
            "## First Pack Output",
            "",
        ]
    )
    for row in pack_rows:
        md.append(f"- {row['ID']}: {row['Pack_Dir']} ({row['File_Count']} files, max {row['Max_File_MB']}MB)")
    md.append("")
    (DATABASE / "Etsy_Digital_Product_RnD.md").write_text("\n".join(md), encoding="utf-8")


def build_contact_sheet(limit: int) -> Path:
    root = OUTPUT / "PrintableWallArt"
    items = []
    for pack in sorted(root.iterdir()) if root.exists() else []:
        if not pack.is_dir():
            continue
        candidates = list(pack.glob("*_2x3_QuietRelicStudio.jpg"))
        if candidates:
            items.append((pack.name, candidates[0]))
    items = items[:limit]
    if not items:
        raise RuntimeError("No printable wall art pack images found for contact sheet.")
    thumb_w, thumb_h = 220, 330
    pad, label_h = 18, 60
    cols = 4
    rows = (len(items) + cols - 1) // cols
    sheet = Image.new("RGB", (cols * (thumb_w + pad) + pad, rows * (thumb_h + label_h + pad) + pad), "white")
    draw = ImageDraw.Draw(sheet)
    for idx, (name, path) in enumerate(items):
        c = idx % cols
        r = idx // cols
        x = pad + c * (thumb_w + pad)
        y = pad + r * (thumb_h + label_h + pad)
        with Image.open(path) as im:
            im = im.convert("RGB")
            im.thumbnail((thumb_w, thumb_h), Image.Resampling.LANCZOS)
            ox = x + (thumb_w - im.width) // 2
            oy = y + (thumb_h - im.height) // 2
            sheet.paste(im, (ox, oy))
        draw.rectangle([x, y, x + thumb_w, y + thumb_h], outline=(180, 180, 180), width=1)
        draw.text((x, y + thumb_h + 8), name[:38], fill=(20, 20, 20))
    out = root / f"Pilot_Contact_Sheet_{limit}.jpg"
    sheet.save(out, "JPEG", quality=92, optimize=True)
    return out


def run(limit: int = 3) -> None:
    candidates = read_candidates()
    scored = sorted(((score_candidate(item), item) for item in candidates), key=lambda x: x[0], reverse=True)
    product_rows = PRODUCT_LINES
    candidate_rows = []
    for rank, (score, item) in enumerate(scored[:30], start=1):
        candidate_rows.append(
            {
                "rank": rank,
                "score": score,
                "ID": item.listing_id,
                "Category": item.category,
                "Current_Title": item.title,
                "Digital_Title": etsy_title(item),
                "Tags": ", ".join(etsy_tags(item)),
                "Source_Path": str(item.source_path),
                "Status": item.status,
            }
        )
    pack_root = OUTPUT / "PrintableWallArt"
    pack_rows = []
    for _, item in scored:
        if len(pack_rows) >= limit:
            break
        try:
            manifest = build_printable_pack(item, pack_root)
        except Exception as exc:
            print(f"[DIGITAL-HOLD] {item.listing_id} {type(exc).__name__}: {clean_one_line(exc)}")
            continue
        max_mb = max(f["size_mb"] for f in manifest["files"])
        pack_rows.append(
            {
                "ID": item.listing_id,
                "Digital_Title": manifest["etsy_title"],
                "Pack_Dir": str(Path(manifest["files"][0]["path"]).parent),
                "File_Count": len(manifest["files"]),
                "Max_File_MB": max_mb,
                "All_Under_20MB": all(f["etsy_file_limit_ok"] for f in manifest["files"]),
                "Source_Path": manifest["source_path"],
            }
        )
    write_rnd_docs(product_rows, candidate_rows, pack_rows)
    contact_sheet = build_contact_sheet(limit)
    print(f"[DIGITAL] product_lines={len(product_rows)} candidates={len(candidate_rows)} packs={len(pack_rows)}")
    for row in pack_rows:
        print(f"[DIGITAL] {row['ID']} files={row['File_Count']} max_mb={row['Max_File_MB']} ok={row['All_Under_20MB']}")
    print(f"[DIGITAL] contact_sheet={contact_sheet}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=3)
    args = parser.parse_args()
    run(limit=args.limit)
