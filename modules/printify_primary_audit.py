import argparse
import io
import re
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config

EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
FIX_STATUS = "Printify_PrimaryFix_Needed"
EXPECTED_MOCKUPS = {
    "Sticker": 4,
    "Poster": 4,
    "Acrylic": 4,
}
MOCKUP_STATUS_RE = re.compile(r"^Printify_(UI|Published)_Mockups\d+$")


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}"}


def _ahash(image):
    image = image.convert("L").resize((16, 16), Image.Resampling.LANCZOS)
    pixels = list(image.getdata())
    avg = sum(pixels) / len(pixels)
    return "".join("1" if pixel > avg else "0" for pixel in pixels)


def _distance(a, b):
    return sum(left != right for left, right in zip(a, b))


def _fetch_product(product_id):
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=_headers(),
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def _remote_hash(url):
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    return _ahash(Image.open(io.BytesIO(response.content)))


def _expected_count(product_type):
    text = str(product_type or "Sticker").strip().lower()
    if text.startswith("acry"):
        return EXPECTED_MOCKUPS["Acrylic"]
    if text.startswith("poster"):
        return EXPECTED_MOCKUPS["Poster"]
    return EXPECTED_MOCKUPS["Sticker"]


def _default_matches_cover(product_id, cover_path, threshold=8, product_type="Sticker"):
    product = _fetch_product(product_id)
    images = product.get("images") or []
    expected = _expected_count(product_type)
    product_text = str(product_type or "").strip().lower()
    if product_text.startswith("acry"):
        selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
        selected_srcs = [str(image.get("src") or "") for image in selected]
        if len(set(selected_srcs)) != len(selected_srcs):
            return False, "official acrylic mockups contain duplicate selected image URLs"
        if any("pfy-prod-products-mockup-media" in src for src in selected_srcs):
            return False, "acrylic custom gallery images are selected; use official product mockups only"
        labels = {
            str(image.get("src") or "").split("camera_label=")[-1].split("&")[0]
            for image in images
            if image.get("is_selected_for_publishing") is not False
        }
        required = {"front", "back", "side-1", "side-2"}
        if not required.issubset(labels):
            return False, f"official acrylic views missing: {sorted(labels)}"
        return True, f"official acrylic front/back/side mockups present; selected image count is {len(images)}"
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    if product_text.startswith("poster"):
        selected_srcs = [str(image.get("src") or "") for image in selected]
        if len(set(selected_srcs)) != len(selected_srcs):
            return False, "official poster mockups contain duplicate selected image URLs"
        if any("pfy-prod-products-mockup-media" in src for src in selected_srcs):
            return False, "poster custom gallery images are selected; use official product mockups only"
        official = [
            image for image in selected
            if "images.printify.com/mockup" in str(image.get("src") or "")
        ]
        if len(selected) < expected or not official:
            return False, f"poster official mockups missing: selected={len(selected)}, official={len(official)}"
        return True, f"official poster mockups present; selected image count is {len(selected)}"
    official = [
        image for image in selected
        if "images.printify.com/mockup" in str(image.get("src") or "")
    ]
    selected_srcs = [str(image.get("src") or "") for image in selected]
    if len(set(selected_srcs)) != len(selected_srcs):
        return False, "sticker official mockups contain duplicate selected image URLs"
    custom_gallery = [
        image for image in selected
        if "pfy-prod-products-mockup-media" in str(image.get("src") or "")
    ]
    if len(official) >= expected and not custom_gallery:
        defaults = [image for image in selected if image.get("is_default")]
        if not defaults:
            return False, "default image count is 0, expected at least 1"
        return True, f"official sticker mockups present; selected={len(selected)} official={len(official)} defaults={len(defaults)}"
    return False, f"sticker gallery unsafe: selected={len(selected)}, official={len(official)}, custom_gallery={len(custom_gallery)}"


def audit_and_mark(limit=0, ids=None):
    ids = {str(item).strip() for item in (ids or []) if str(item).strip()}
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    checked = 0
    fixes = 0
    try:
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row, cols["Status"]).value
            status_text = str(status or "")
            if not (MOCKUP_STATUS_RE.match(status_text) or status_text == FIX_STATUS):
                continue
            product_id = ws.cell(row, cols["Printify_Product_ID"]).value
            cover_path = ws.cell(row, cols["Cover_Path"]).value
            item_id = ws.cell(row, cols["ID"]).value
            if ids and str(item_id) not in ids:
                continue
            if not product_id or not cover_path or not Path(cover_path).exists():
                continue
            checked += 1
            product_type = ws.cell(row, cols.get("Product_Type", cols["ID"])).value if "Product_Type" in cols else "Sticker"
            ok, note = _default_matches_cover(str(product_id), cover_path, product_type=product_type)
            if ok:
                if not status_text.startswith("Printify_Published"):
                    ws.cell(row, cols["Status"]).value = f"Printify_UI_Mockups{_expected_count(product_type)}"
                print(f"[PRIMARY-OK] {item_id} {note}")
            else:
                ws.cell(row, cols["Status"]).value = FIX_STATUS
                fixes += 1
                print(f"[PRIMARY-FIX] {item_id} {note}")
            wb.save(EBAY_BOOK)
            if limit and checked >= limit:
                break
    finally:
        wb.close()
    print(f"[DONE] primary audit checked={checked} fixes={fixes}")
    return fixes


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--ids", default="", help="Comma-separated listing IDs to audit.")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()] or None
    audit_and_mark(limit=args.limit, ids=ids)
