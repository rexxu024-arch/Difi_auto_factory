import argparse
import csv
import json
import os
import sys
from datetime import datetime
from datetime import timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.ebay_token_manager import EbayTokenError, get_access_token

DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
LOG_PATH = DATABASE_DIR / "ebay_ads_standard_2pct.csv"
PENDING_PATH = DATABASE_DIR / "ebay_ads_pending_2pct.csv"
OAUTH_TOKEN_FILE = DATABASE_DIR / ".ebay_oauth_tokens.json"

CAMPAIGN_NAME = "Fixed_2_Percent_Strategy"
DEFAULT_CAMPAIGN_ID = os.getenv("EBAY_AD_CAMPAIGN_ID", "165251921016")
MARKETPLACE_ID = os.getenv("EBAY_MARKETPLACE_ID", "EBAY_US")
AD_RATE = "2.0"
BASE_URL = os.getenv("EBAY_API_BASE_URL", "https://api.ebay.com")
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"

HEADERS = [
    "Timestamp",
    "Action",
    "ID",
    "eBay_Item_ID",
    "Campaign_ID",
    "HTTP_Status",
    "Result",
    "Error",
]
PENDING_HEADERS = [
    "Timestamp",
    "ID",
    "eBay_Item_ID",
    "Campaign_ID",
    "Ad_Rate",
    "Status",
    "Error",
]


def _access_token():
    token = os.getenv("EBAY_ACCESS_TOKEN") or os.getenv("EBAY_OAUTH_TOKEN")
    if not token:
        try:
            token = get_access_token()
        except EbayTokenError:
            token = ""
    if not token and OAUTH_TOKEN_FILE.exists():
        try:
            token = json.loads(OAUTH_TOKEN_FILE.read_text(encoding="utf-8")).get("access_token")
        except Exception:
            token = ""
    if not token:
        raise RuntimeError(
            "Missing EBAY_ACCESS_TOKEN / EBAY_OAUTH_TOKEN. eBay Marketing API cannot be used until OAuth is configured."
        )
    return token


def has_access_token():
    if os.getenv("EBAY_ACCESS_TOKEN") or os.getenv("EBAY_OAUTH_TOKEN"):
        return True
    try:
        return bool(get_access_token())
    except EbayTokenError:
        pass
    if not OAUTH_TOKEN_FILE.exists():
        return False
    try:
        return bool(json.loads(OAUTH_TOKEN_FILE.read_text(encoding="utf-8")).get("access_token"))
    except Exception:
        return False


def _headers():
    return {
        "Authorization": f"Bearer {_access_token()}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "X-EBAY-C-MARKETPLACE-ID": MARKETPLACE_ID,
    }


def _api(method, path, **kwargs):
    response = requests.request(method, f"{BASE_URL}{path}", headers=_headers(), timeout=90, **kwargs)
    return response


def _log(rows):
    exists = LOG_PATH.exists()
    with LOG_PATH.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        if not exists:
            writer.writeheader()
        writer.writerows(rows)


def enqueue_listing(item_id, ebay_item_id, campaign_id=None, status="PENDING_API"):
    ebay_item_id = str(ebay_item_id or "").strip()
    if not ebay_item_id:
        return False
    campaign_id = campaign_id or DEFAULT_CAMPAIGN_ID or CAMPAIGN_NAME
    existing = set()
    if PENDING_PATH.exists():
        with PENDING_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                existing.add((row.get("eBay_Item_ID"), row.get("Campaign_ID")))
    key = (ebay_item_id, campaign_id)
    if key in existing:
        return False
    file_exists = PENDING_PATH.exists()
    with PENDING_PATH.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PENDING_HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerow(
            {
                "Timestamp": datetime.now().isoformat(timespec="seconds"),
                "ID": item_id,
                "eBay_Item_ID": ebay_item_id,
                "Campaign_ID": campaign_id,
                "Ad_Rate": AD_RATE,
                "Status": status,
                "Error": "",
            }
        )
    return True


def _published_listing_ids(limit=0):
    workbook = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    cols = {header: index for index, header in enumerate(headers)}
    if "eBay_Item_ID" not in cols:
        workbook.close()
        return _published_listing_ids_from_performance_log(limit=limit)
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[cols["ID"]]:
            continue
        status = str(row[cols.get("Status")] or "")
        item_id = str(row[cols["eBay_Item_ID"]] or "").strip()
        if status.startswith("Printify_Published") and item_id:
            rows.append(
                {
                    "ID": row[cols["ID"]],
                    "Product_Type": row[cols.get("Product_Type")],
                    "Title": row[cols.get("Title")],
                    "eBay_Item_ID": item_id,
                }
            )
            if limit and len(rows) >= limit:
                break
    workbook.close()
    return rows


def _published_listing_ids_from_performance_log(limit=0):
    if not PERFORMANCE_LOG.exists():
        raise RuntimeError("eBay_Item_ID column is missing and Performance_Log.csv is unavailable.")
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return []
    latest = max(str(row.get("Snapshot_Timestamp") or "") for row in rows)
    out = []
    for row in rows:
        if str(row.get("Snapshot_Timestamp") or "") != latest:
            continue
        title = str(row.get("Title") or "")
        if "toothbrush" in title.lower() or "dumbbell" in title.lower():
            continue
        ebay_id = str(row.get("Item_ID") or "").strip()
        if not ebay_id:
            continue
        out.append(
            {
                "ID": ebay_id,
                "Product_Type": "",
                "Title": title,
                "eBay_Item_ID": ebay_id,
            }
        )
        if limit and len(out) >= limit:
            break
    return out


def find_campaign():
    response = _api("GET", f"/sell/marketing/v1/ad_campaign?campaign_name={CAMPAIGN_NAME}")
    if response.status_code == 404:
        return None
    response.raise_for_status()
    data = response.json()
    campaigns = data.get("campaigns") or data.get("adCampaigns") or []
    for campaign in campaigns:
        if campaign.get("campaignName") == CAMPAIGN_NAME or campaign.get("name") == CAMPAIGN_NAME:
            return campaign
    return campaigns[0] if len(campaigns) == 1 else None


def campaign_exists(campaign_id):
    campaign_id = str(campaign_id or "").strip()
    if not campaign_id:
        return False
    response = _api("GET", f"/sell/marketing/v1/ad_campaign/{campaign_id}")
    if response.status_code == 404:
        return False
    response.raise_for_status()
    return True


def resolve_campaign_id(campaign_id=None):
    campaign_id = str(campaign_id or "").strip()
    if campaign_id and campaign_exists(campaign_id):
        return campaign_id
    campaign = find_campaign()
    if not campaign:
        return ""
    return (
        campaign.get("campaignId")
        or campaign.get("campaign_id")
        or campaign.get("adCampaignId")
        or ""
    )


def create_campaign():
    payload = {
        "campaignName": CAMPAIGN_NAME,
        "startDate": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "channels": ["ON_SITE"],
        "fundingStrategy": {
            "fundingModel": "COST_PER_SALE",
            "adRateStrategy": "FIXED",
            "bidPercentage": AD_RATE,
        },
        "marketplaceId": MARKETPLACE_ID,
    }
    response = _api("POST", "/sell/marketing/v1/ad_campaign", json=payload)
    if response.status_code not in {200, 201, 202, 204}:
        response.raise_for_status()
    location = response.headers.get("Location", "")
    campaign_id = location.rstrip("/").split("/")[-1] if location else ""
    return campaign_id, response.status_code, location


def bulk_create_ads(campaign_id, listings, dry_run=True):
    logs = []
    if dry_run:
        for listing in listings:
            logs.append(
                {
                    "Timestamp": datetime.now().isoformat(timespec="seconds"),
                    "Action": "DRY_RUN_ADD_AD",
                    "ID": listing["ID"],
                    "eBay_Item_ID": listing["eBay_Item_ID"],
                    "Campaign_ID": campaign_id,
                    "HTTP_Status": "",
                    "Result": f"Would add listing at fixed bidPercentage={AD_RATE}",
                    "Error": "",
                }
            )
        _log(logs)
        print(f"[ADS-DRY] campaign={campaign_id or CAMPAIGN_NAME} listings={len(listings)} rate={AD_RATE}")
        return logs
    requests_payload = [
        {"listingId": listing["eBay_Item_ID"], "bidPercentage": AD_RATE}
        for listing in listings
    ]
    response = _api(
        "POST",
        f"/sell/marketing/v1/ad_campaign/{campaign_id}/bulk_create_ads_by_listing_id",
        json={"requests": requests_payload},
    )
    status_code = response.status_code
    if response.status_code not in {200, 201, 202, 207}:
        response.raise_for_status()
    try:
        response_body = response.json()
        responses = response_body.get("responses") or []
    except Exception:
        responses = []
    for listing in listings:
        item_response = next(
            (item for item in responses if str(item.get("listingId") or "") == str(listing["eBay_Item_ID"])),
            {},
        )
        child_status = str(item_response.get("statusCode") or item_response.get("status") or status_code)
        errors = item_response.get("errors") or []
        error_text = "; ".join(str(err.get("message") or err.get("errorId") or err) for err in errors)
        result_text = json.dumps(item_response or {"http_status": status_code}, ensure_ascii=False)[:1500]
        logs.append(
            {
                "Timestamp": datetime.now().isoformat(timespec="seconds"),
                "Action": "ADD_AD_FIXED_2",
                "ID": listing["ID"],
                "eBay_Item_ID": listing["eBay_Item_ID"],
                "Campaign_ID": campaign_id,
                "HTTP_Status": child_status,
                "Result": result_text,
                "Error": error_text,
            }
        )
    _log(logs)
    ok_count = sum(1 for row in logs if str(row["HTTP_Status"]).startswith("2") and not row["Error"])
    print(f"[ADS-RESULT] campaign={campaign_id} listings={len(listings)} ok={ok_count} http={status_code}")
    return logs


def enroll_listing(item_id, ebay_item_id, campaign_id=None, dry_run=False):
    requested_campaign_id = campaign_id or DEFAULT_CAMPAIGN_ID
    campaign_id = requested_campaign_id
    listing = {"ID": item_id, "eBay_Item_ID": str(ebay_item_id).strip()}
    if not listing["eBay_Item_ID"]:
        return False
    if dry_run:
        bulk_create_ads(campaign_id or CAMPAIGN_NAME, [listing], dry_run=True)
        return True
    if not has_access_token():
        enqueue_listing(item_id, listing["eBay_Item_ID"], campaign_id=campaign_id, status="PENDING_OAUTH")
        print(f"[ADS-PENDING] {item_id} ebay={listing['eBay_Item_ID']} reason=missing_oauth")
        return False
    campaign_id = resolve_campaign_id(campaign_id)
    if not campaign_id:
        enqueue_listing(item_id, listing["eBay_Item_ID"], campaign_id=requested_campaign_id or CAMPAIGN_NAME, status="PENDING_CAMPAIGN_ID")
        print(f"[ADS-PENDING] {item_id} ebay={listing['eBay_Item_ID']} reason=missing_or_stale_campaign_id")
        return False
    logs = bulk_create_ads(campaign_id, [listing], dry_run=False)
    ok = any(str(row.get("HTTP_Status") or "").startswith("2") and not row.get("Error") for row in logs)
    if not ok:
        enqueue_listing(
            item_id,
            listing["eBay_Item_ID"],
            campaign_id=campaign_id,
            status="PENDING_API_RETRY",
        )
        return False
    return True


def run(limit=0, dry_run=True, campaign_id=None, create_if_missing=False):
    listings = _published_listing_ids(limit=limit)
    print(f"[ADS] eligible_with_ebay_id={len(listings)} dry_run={dry_run} rate={AD_RATE}")
    if dry_run:
        return bulk_create_ads(campaign_id or CAMPAIGN_NAME, listings, dry_run=True)
    if campaign_id:
        resolved_campaign_id = resolve_campaign_id(campaign_id)
    else:
        resolved_campaign_id = resolve_campaign_id(DEFAULT_CAMPAIGN_ID)
    if not resolved_campaign_id and create_if_missing:
        resolved_campaign_id, status, location = create_campaign()
        print(f"[ADS-CAMPAIGN-CREATED] id={resolved_campaign_id} http={status} location={location}")
    if not resolved_campaign_id:
        raise RuntimeError(
            f"Campaign {CAMPAIGN_NAME} was not found. Create it in Seller Hub or rerun with --create-if-missing after final confirmation."
        )
    return bulk_create_ads(resolved_campaign_id, listings, dry_run=False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--campaign-id", default="")
    parser.add_argument("--create-if-missing", action="store_true")
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args()
    run(
        limit=args.limit,
        dry_run=not args.execute,
        campaign_id=args.campaign_id or None,
        create_if_missing=args.create_if_missing,
    )


if __name__ == "__main__":
    main()
