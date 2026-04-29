import argparse
import base64
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from PIL import Image

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
PRINTIFY_SPECS = {
    "Sticker": {"blueprint_id": 400, "provider_id": 99, "variant_id": 45754, "default_price": 1199, "production_min": 1600, "cover_min": 1600},
    "Poster": {"blueprint_id": 282, "provider_id": 99, "variant_id": 43138, "default_price": 3499, "production_min": 3000, "cover_min": 1200},
    "Acrylic": {"blueprint_id": 1471, "provider_id": 104, "variant_id": 106190, "default_price": 8999, "production_min": 1500, "cover_min": 1200},
}
BATCH_SIZE = 75
BATCH_DELAY_SECONDS = 3600


def _headers():
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def _product_type(row):
    value = str(row.get("Product_Type") or "Sticker").strip()
    if value.lower().startswith("poster"):
        return "Poster"
    if value.lower().startswith("acry"):
        return "Acrylic"
    return "Sticker"


def _spec(row):
    return PRINTIFY_SPECS[_product_type(row)]


def _price_to_cents(value, default_cents=1199):
    match = re.search(r"\d+(?:\.\d{1,2})?", str(value or ""))
    if not match:
        return default_cents
    return int(round(float(match.group(0)) * 100))


def _image_upload(path, file_name):
    with open(path, "rb") as handle:
        encoded = base64.b64encode(handle.read()).decode("utf-8")
    response = requests.post(
        f"{Config.Printify_API_URL}/uploads/images.json",
        headers=_headers(),
        json={"file_name": file_name, "contents": encoded},
        timeout=120,
    )
    response.raise_for_status()
    return response.json()["id"]


def _load_rows():
    if not EBAY_BOOK.exists():
        raise FileNotFoundError(f"Missing listing database: {EBAY_BOOK}")
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        data = {headers[col - 1]: ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)}
        data["_row_idx"] = row_idx
        if data.get("ID") and data.get("Status") in {"Ready_for_Printify", "Printify_Failed"}:
            rows.append(data)
    return wb, ws, headers, rows


def _set_status(ws, headers, row_idx, status, product_id=""):
    status_col = headers.index("Status") + 1
    ws.cell(row=row_idx, column=status_col).value = status
    if "Printify_Product_ID" not in headers:
        ws.cell(row=1, column=len(headers) + 1).value = "Printify_Product_ID"
        headers.append("Printify_Product_ID")
    product_col = headers.index("Printify_Product_ID") + 1
    ws.cell(row=row_idx, column=product_col).value = product_id


def _fetch_product(product_id):
    response = requests.get(
        f"{Config.Printify_API_URL}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers={"Authorization": f"Bearer {Config.Printify_API_KEY}"},
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def _selected_mockup_count(product):
    images = product.get("images") or []
    return sum(1 for image in images if image.get("is_selected_for_publishing") is not False)

def _wait_for_selected_mockups(product_id, timeout_seconds=90):
    deadline = time.time() + timeout_seconds
    last_product = None
    while time.time() < deadline:
        last_product = _fetch_product(product_id)
        selected_count = _selected_mockup_count(last_product)
        if selected_count:
            return last_product, selected_count
        time.sleep(6)
    return last_product or _fetch_product(product_id), _selected_mockup_count(last_product or {})


def _gallery_paths(row):
    paths = []
    cover = row.get("Cover_Path")
    if cover:
        paths.append(("Cover", Path(cover)))
    for index in range(1, 5):
        value = row.get(f"Gallery_U{index}_Path")
        if value:
            paths.append((f"Gallery_U{index}", Path(value)))
    return paths

def _image_size(path):
    try:
        with Image.open(path) as image:
            image.load()
            return image.size
    except Exception:
        return None


def _build_payload(row, stock_image_ids, production_id):
    item_id = row["ID"]
    sku = str(row.get("SKU") or item_id).strip()
    spec = _spec(row)
    variant_id = spec["variant_id"]
    gallery = []
    for index, image_id in enumerate(stock_image_ids):
        gallery.append(
            {
                "id": image_id,
                "is_default": index == 0,
                "is_selected_for_publishing": True,
                "variant_ids": [variant_id],
            }
        )
    return {
        "title": row["Title"],
        "description": row["Description"],
        "blueprint_id": spec["blueprint_id"],
        "print_provider_id": spec["provider_id"],
        "variants": [
            {
                "id": variant_id,
                "price": _price_to_cents(row.get("Price"), spec["default_price"]),
                "is_enabled": True,
                "sku": sku,
            }
        ],
        "print_areas": [
            {
                "variant_ids": [variant_id],
                "placeholders": [
                    {
                        "position": "front",
                        "images": [{"id": production_id, "x": 0.5, "y": 0.5, "scale": 1, "angle": 0}],
                    }
                ],
            }
        ],
        "images": gallery,
    }


def _validate_row_assets(row):
    item_id = row["ID"]
    spec = _spec(row)
    production_path = Path(row["Production_Path"])
    stock_paths = _gallery_paths(row)
    missing = []
    if not production_path.exists():
        missing.append(str(production_path))
    for _, path in stock_paths:
        if not path.exists():
            missing.append(str(path))
    if len(stock_paths) < 5:
        missing.append(f"{item_id}: expected Cover + U1-U4 stock photos, found {len(stock_paths)}")
    checks = [(production_path, spec["production_min"], "Production_Design")]
    for label, path in stock_paths:
        checks.append((path, 1024 if label.startswith("Gallery_U") else spec["cover_min"], label))
    for path, minimum, label in checks:
        if not path.exists():
            continue
        size = _image_size(path)
        if not size:
            missing.append(f"{item_id}: unreadable image {label} at {path}")
        elif min(size) < minimum:
            missing.append(f"{item_id}: low resolution {label} {size[0]}x{size[1]} below {minimum}px")
    return production_path, stock_paths, missing


def stage_printify_products(limit=0, dry_run=False, auto_proceed=False, batch_size=BATCH_SIZE, batch_delay=BATCH_DELAY_SECONDS):
    if not Config.Printify_API_KEY:
        raise RuntimeError("Printify_API_KEY is missing")
    wb, ws, headers, rows = _load_rows()
    if limit:
        rows = rows[:limit]
    staged = 0
    try:
        for row in rows:
            item_id = row["ID"]
            production_path, stock_paths, missing = _validate_row_assets(row)
            if missing:
                print(f"[SKIP] Missing image assets: {item_id} | {'; '.join(missing)}")
                _set_status(ws, headers, row["_row_idx"], "Printify_Failed")
                wb.save(EBAY_BOOK)
                continue
            if dry_run:
                print(
                    f"[DRY-RUN] {item_id} | title={len(str(row['Title']))} chars | "
                    f"production=1 stock_photos={len(stock_paths)} price={row.get('Price')}"
                )
                continue
            version = str(int(time.time()))
            production_id = _image_upload(production_path, f"{item_id}_Production_{version}.png")
            stock_image_ids = []
            for label, path in stock_paths:
                stock_image_ids.append(_image_upload(path, f"{item_id}_{label}_{version}.png"))
            payload = _build_payload(row, stock_image_ids, production_id)
            response = requests.post(
                f"{Config.Printify_API_URL}/shops/{Config.Printify_SHOP_ID}/products.json",
                headers=_headers(),
                json=payload,
                timeout=120,
            )
            response.raise_for_status()
            product_id = response.json()["id"]
            product, selected_count = _wait_for_selected_mockups(product_id)
            selected_count = _selected_mockup_count(product)
            if selected_count < len(stock_image_ids):
                status = f"Printify_BaseStaged_DefaultMockups{selected_count}" if selected_count else "Printify_PhotoMismatch"
                print(
                    f"[VERIFY] {item_id} staged with selected mockups={selected_count}, "
                    f"custom stock photos expected={len(stock_image_ids)}. Bulk upload paused for manual/UI route."
                )
            else:
                status = "Printify_Staged"
            _set_status(ws, headers, row["_row_idx"], status, product_id)
            wb.save(EBAY_BOOK)
            staged += 1
            print(
                f"[PRINTIFY] {item_id} staged as {product_id} | "
                f"uploaded_stock_photos={len(stock_image_ids)} selected_mockups={selected_count} status={status}"
            )
            if status != "Printify_Staged":
                if staged == 1:
                    print("First Product Successfully Staged on Printify. Waiting for Rex's Manual Confirmation.")
                    if not auto_proceed and input("Type PROCEED to continue: ").strip() != "PROCEED":
                        print("[PAUSE] Manual confirmation not received.")
                return
            if staged == 1:
                print("First Product Successfully Staged on Printify. Waiting for Rex's Manual Confirmation.")
                if not auto_proceed and input("Type PROCEED to continue: ").strip() != "PROCEED":
                    print("[PAUSE] Manual confirmation not received.")
                    return
            if auto_proceed and batch_size and staged % batch_size == 0 and staged < len(rows):
                print(
                    f"[BATCH-COOLDOWN] {staged}/{len(rows)} staged. "
                    f"Sleeping {batch_delay} seconds before next batch at {datetime.now():%Y-%m-%d %H:%M:%S}."
                )
                time.sleep(batch_delay)
        print(f"[DONE] Printify staged count: {staged}")
    finally:
        wb.close()


def run_logic():
    stage_printify_products()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--auto-proceed", action="store_true")
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    parser.add_argument("--batch-delay", type=int, default=BATCH_DELAY_SECONDS)
    args = parser.parse_args()
    stage_printify_products(
        limit=args.limit,
        dry_run=args.dry_run,
        auto_proceed=args.auto_proceed,
        batch_size=args.batch_size,
        batch_delay=args.batch_delay,
    )
