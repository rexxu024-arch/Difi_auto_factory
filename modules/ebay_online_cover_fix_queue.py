"""Write latest online cover audit mismatches back to workbook and a fix queue."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
AUDIT_CSV = DATABASE_DIR / "eBay_Online_Cover_Audit.csv"
FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"

MISMATCH_RESULTS = {"LIKELY_SINGLE_U_MISMATCH", "ERROR", "UNKNOWN"}


def clean(value) -> str:
    return str(value or "").strip()


def latest_audit_rows() -> dict[str, dict[str, str]]:
    if not AUDIT_CSV.exists():
        return {}
    latest: dict[str, dict[str, str]] = {}
    with AUDIT_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            item_id = clean(row.get("ID"))
            if item_id:
                latest[item_id] = row
    return latest


def ensure_col(ws, cols: dict[str, int], name: str) -> int:
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def run() -> None:
    latest = latest_audit_rows()
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    cols = {name: idx + 1 for idx, name in enumerate(headers)}
    audit_cols = {
        "Online_Cover_Result": ensure_col(ws, cols, "Online_Cover_Result"),
        "Online_Cover_Note": ensure_col(ws, cols, "Online_Cover_Note"),
        "Online_Cover_Best_U": ensure_col(ws, cols, "Online_Cover_Best_U"),
        "Online_Cover_Audit_Timestamp": ensure_col(ws, cols, "Online_Cover_Audit_Timestamp"),
    }
    queue_rows = []
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        local_id = clean(ws.cell(row_idx, cols["ID"]).value)
        if local_id not in latest:
            continue
        row = latest[local_id]
        result = clean(row.get("Result"))
        ws.cell(row_idx, audit_cols["Online_Cover_Result"]).value = result
        ws.cell(row_idx, audit_cols["Online_Cover_Note"]).value = clean(row.get("Note")) or clean(row.get("Error"))
        ws.cell(row_idx, audit_cols["Online_Cover_Best_U"]).value = clean(row.get("Best_U_Label"))
        ws.cell(row_idx, audit_cols["Online_Cover_Audit_Timestamp"]).value = clean(row.get("Timestamp"))
        updated += 1
        if result in MISMATCH_RESULTS:
            queue_rows.append(
                {
                    "ID": local_id,
                    "eBay_Item_ID": clean(row.get("eBay_Item_ID")) or clean(ws.cell(row_idx, cols.get("eBay_Item_ID", 0)).value),
                    "Printify_Product_ID": clean(row.get("Printify_Product_ID")) or clean(ws.cell(row_idx, cols.get("Printify_Product_ID", 0)).value),
                    "Result": result,
                    "Best_U_Label": clean(row.get("Best_U_Label")),
                    "Distance_To_Cover": clean(row.get("Distance_To_Cover")),
                    "Best_U_Distance": clean(row.get("Best_U_Distance")),
                    "Recommended_Action": "Fix live eBay image order: make local Cover_Mockup photo 1; then re-audit online listing.",
                    "Status": "PENDING_FIX",
                }
            )
    wb.save(EBAY_BOOK)
    wb.close()
    headers_out = [
        "ID",
        "eBay_Item_ID",
        "Printify_Product_ID",
        "Result",
        "Best_U_Label",
        "Distance_To_Cover",
        "Best_U_Distance",
        "Recommended_Action",
        "Status",
    ]
    with FIX_QUEUE.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers_out)
        writer.writeheader()
        writer.writerows(queue_rows)
    print(f"[ONLINE-COVER-FIX-QUEUE] workbook_rows_updated={updated} pending_fix={len(queue_rows)} csv={FIX_QUEUE}")


if __name__ == "__main__":
    run()
