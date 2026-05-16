from __future__ import annotations

import argparse
import csv
import json
import random
import re
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
PIVOT_CSV = DATABASE_DIR / "eBay_Quiet_Jade_Pivot.csv"
ROLLBACK_CSV = DATABASE_DIR / "eBay_Quiet_Jade_Rollback.csv"
SYNC_LOG = DATABASE_DIR / "eBay_Quiet_Jade_Sync_Log.csv"
STATE_JSON = DATABASE_DIR / "eBay_Quiet_Jade_State.json"
NY = ZoneInfo("America/New_York")

TITLE_MIN = 75
TITLE_MAX = 79
IMAGE_NOTE = (
    "The main image shows the physical product customers receive. Additional gallery images "
    "are concept, detail, or collection-reference views and are not extra products or selectable variations."
)
PUBLISH_TITLE_DESC_PRICE = {
    "title": True,
    "description": True,
    "images": False,
    "variants": True,
    "tags": False,
    "keyFeatures": False,
    "shipping_template": False,
}


@dataclass
class PivotRow:
    row_idx: int
    local_id: str
    product_type: str
    category: str
    printify_product_id: str
    ebay_item_id: str
    old_title: str
    old_description: str
    old_price: str
    dna_profile: str
    views: int
    new_title: str
    new_description: str
    new_price: str
    search_tags: list[str]
    price_decision: str


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").replace("\r", " ")).strip()


def money(value: Any) -> float | None:
    raw = re.sub(r"[^0-9.]", "", clean(value))
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def money_text(value: float) -> str:
    return f"${value:.2f}"


def fit_title(parts: list[str]) -> str:
    title = clean(" ".join(part for part in parts if part))
    title = re.sub(r"\bZen Poster\b|\bGreen Canvas\b", "", title, flags=re.I)
    title = clean(title)
    boosters = [
        "Decor",
        "Gift",
        "Study",
        "Wall",
        "Shelf",
        "Room",
        "Focus",
        "Nook",
    ]
    index = 0
    while len(title) < TITLE_MIN and index < len(boosters):
        token = boosters[index]
        if token.lower() not in title.lower():
            title = clean(f"{title} {token}")
        index += 1
    if len(title) > TITLE_MAX:
        title = title[:TITLE_MAX].rsplit(" ", 1)[0]
    index = 0
    while len(title) < TITLE_MIN and index < len(boosters):
        token = boosters[index]
        if token.lower() not in title.lower() and len(clean(f"{title} {token}")) <= TITLE_MAX:
            title = clean(f"{title} {token}")
        index += 1
    if len(title) < TITLE_MIN:
        title = clean(f"{title} Premium")
        if len(title) > TITLE_MAX:
            title = title[:TITLE_MAX].rsplit(" ", 1)[0]
    return title


def subject_from_title(title: str, product_type: str, category: str) -> str:
    text = clean(title)
    text = re.sub(r"\$?\d+(\.\d+)?", " ", text)
    text = re.sub(r"\b(4pc|4\s*pc|pc|6x6|5x7|12x18|12|18|5|7)\b", " ", text, flags=re.I)
    text = re.sub(
        r"\b(Kiss[- ]?Cut|Sticker|Vinyl|Laptop|Journal|Gift|Decor|Matte|Poster|Wall|Art|"
        r"Acrylic|Photo|Block|Desk|Display|Shelf|Gallery|Premium|Collectible|Study|Room|Set|Pack|Sheet|Print|Object|"
        r"Zen|Aesthetic|Minimal|Mindful|Dark|Academia|Gothic|Grimdark|Mentor[- ]Grade|"
        r"Smoky|Jade|Quiet|Luxury|Apartment|Reading|Nook|Deep|Work|Visual|Wabi|Sabi|Focus|Library)\b",
        " ",
        text,
        flags=re.I,
    )
    text = re.sub(r"[^A-Za-z0-9' ]+", " ", text)
    text = clean(text)
    if not text:
        text = clean(category) or product_type
    words = [word for word in text.split() if word.lower() not in {"x", "to"}]
    if len(words) > 5:
        words = words[:5]
    while words and words[-1].lower() in {"of", "and", "the", "for", "with"}:
        words.pop()
    return " ".join(words) or "Jade Relic"


def intent_bank(product_type: str, category: str, local_id: str) -> dict[str, str]:
    seed = sum(ord(ch) for ch in f"{local_id}|{product_type}|{category}")
    if product_type == "Poster":
        leads = [
            "Smoky Jade Wall Art",
            "Quiet Luxury Apartment Art",
            "Deep Work Visual Poster",
            "Reading Nook Wall Decor",
            "Wabi Sabi Study Print",
        ]
        scenes = [
            "Apartment Decor",
            "Study Room Decor",
            "Reading Nook Gift",
            "Deep Focus Room",
            "Library Wall Art",
        ]
    elif product_type == "Acrylic":
        leads = [
            "Smoky Jade Desk Object",
            "Quiet Luxury Shelf Decor",
            "Deep Work Desk Visual",
            "Collector Acrylic Relic",
            "Gallery Desk Art",
        ]
        scenes = [
            "Apartment Shelf",
            "Study Desk Decor",
            "Collector Display",
            "Office Focus Decor",
            "Reading Nook Gift",
        ]
    else:
        leads = [
            "Smoky Jade Sticker Set",
            "Deep Work Laptop Decals",
            "Quiet Desk Sticker Pack",
            "Reading Nook Vinyl Set",
            "Wabi Sabi Journal Decals",
        ]
        scenes = [
            "Laptop Journal",
            "Planner Bottle",
            "Study Desk Gift",
            "Book Nook Decor",
            "Quiet Focus Set",
        ]
    return {"lead": leads[seed % len(leads)], "scene": scenes[(seed // 7) % len(scenes)]}


def title_for(row: dict[str, Any]) -> str:
    product_type = clean(row.get("Product_Type")) or "Sticker"
    category = clean(row.get("Category"))
    subject = subject_from_title(clean(row.get("Title")), product_type, category)
    bank = intent_bank(product_type, category, clean(row.get("ID")))
    if product_type == "Poster":
        return fit_title([bank["lead"], subject, "12x18 Matte Poster", bank["scene"]])
    if product_type == "Acrylic":
        return fit_title([bank["lead"], subject, "5x7 Acrylic Photo Block", bank["scene"]])
    return fit_title([bank["lead"], subject, "4pc 6x6 Kiss-Cut Vinyl", bank["scene"]])


def tags_for(row: dict[str, Any], subject: str) -> list[str]:
    product_type = clean(row.get("Product_Type")) or "Sticker"
    category = clean(row.get("Category"))
    base = [
        "Smoky Jade",
        "Quiet Luxury Decor",
        "Deep Work Visuals",
        "Reading Nook Decor",
        "Wabi Sabi Decor",
        "Scholar Room",
        "Apartment Decor",
        "Jade Art",
        "Kintsugi Aesthetic",
        "Collector Gift",
    ]
    if category.lower() == "academia":
        base += ["Dark Study Decor", "Library Wall Art", "Book Lover Gift"]
    elif category.lower() == "grimdark":
        base += ["Gothic Study Decor", "Alchemy Decor", "Dark Fantasy Gift"]
    else:
        base += ["Meditation Decor", "Calm Desk Art", "Mindful Room Decor"]
    if product_type == "Poster":
        base += ["Matte Poster", "Wall Art"]
    elif product_type == "Acrylic":
        base += ["Acrylic Block", "Desk Display"]
    else:
        base += ["Sticker Set", "Laptop Decal"]
    if subject:
        base.insert(0, subject)
    out = []
    seen = set()
    for item in base:
        item = clean(item)[:40]
        key = item.lower()
        if item and key not in seen:
            seen.add(key)
            out.append(item)
        if len(out) >= 13:
            break
    return out


def price_for(row: dict[str, Any]) -> tuple[str, str]:
    product_type = clean(row.get("Product_Type")) or "Sticker"
    current = money(row.get("Price"))
    if product_type == "Poster":
        target = 34.99
        decision = "Poster value-pivot target $34.99; above cost+shipping+fees guardrail."
    elif product_type == "Acrylic":
        target = current if current and current >= 79.99 else 89.99
        decision = "Acrylic kept premium; $29.99-$34.99 would violate cost+shipping guardrail."
    else:
        target = current if current else 11.99
        decision = "Sticker kept review-friendly; not part of high-ticket price lift."
    return money_text(target), decision


def description_for(row: dict[str, Any], new_title: str, tags: list[str]) -> str:
    product_type = clean(row.get("Product_Type")) or "Sticker"
    category = clean(row.get("Category")) or "OpenClaw"
    subject = subject_from_title(clean(row.get("Title")), product_type, category)
    if product_type == "Poster":
        includes = "One 12x18 premium matte vertical poster."
        material = "Premium matte paper wall art produced on demand through Printify."
        placement = "quiet apartments, reading nooks, deep-work offices, study rooms, and gallery walls"
    elif product_type == "Acrylic":
        includes = "One 5x7 vertical acrylic photo block."
        material = "Acrylic block display with depth, gloss, and light-reflective finish."
        placement = "desks, bookshelves, altar corners, office focus spaces, and collector displays"
    else:
        includes = "One 6x6 kiss-cut vinyl sticker sheet with four coordinated designs."
        material = "Durable kiss-cut vinyl sticker sheet produced on demand through Printify."
        placement = "laptops, journals, planners, bottles, reading notebooks, and study desk gifts"
    tag_text = ", ".join(tags)
    return (
        f"<h2>{new_title}</h2>"
        f"<p>A Quiet Jade collection piece built for {placement}. The visual direction leans into smoky jade tones, "
        "quiet luxury decor language, kintsugi detail, and deep-work atmosphere instead of generic mass-market wall art.</p>"
        "<ul>"
        f"<li><strong>Includes:</strong> {includes}</li>"
        f"<li><strong>Material:</strong> {material}</li>"
        f"<li><strong>Style:</strong> {category} collector aesthetic; smoky jade, wabi-sabi, scholar-room mood.</li>"
        f"<li><strong>Best For:</strong> apartment decor, reading nook styling, focused workspaces, collectors, and giftable room accents.</li>"
        "</ul>"
        f"<p><strong>Image Note:</strong> {IMAGE_NOTE}</p>"
        f"<p><strong>Search Themes:</strong> {tag_text}</p>"
        f"<p><small>Reference SKU: {clean(row.get('ID'))}; Subject: {subject}</small></p>"
    )


def latest_zero_view_items() -> set[str]:
    if not PERFORMANCE_LOG.exists():
        return set()
    rows = list(csv.DictReader(PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="")))
    if not rows:
        return set()
    latest_ts = max(clean(row.get("Snapshot_Timestamp")) for row in rows)
    return {
        clean(row.get("Item_ID"))
        for row in rows
        if clean(row.get("Snapshot_Timestamp")) == latest_ts and clean(row.get("Views_30_Days")) == "0"
    }


def ensure_column(ws, cols: dict[str, int], name: str) -> int:
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def load_candidates(limit: int) -> tuple[Any, Any, dict[str, int], list[PivotRow]]:
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    zero_items = latest_zero_view_items()
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        data = {header: ws.cell(row_idx, cols[header]).value for header in headers}
        data["_row_idx"] = row_idx
        item_id = clean(data.get("eBay_Item_ID"))
        product_id = clean(data.get("Printify_Product_ID"))
        status = clean(data.get("Status"))
        if not item_id or not product_id:
            continue
        if zero_items and item_id not in zero_items:
            continue
        if not status.startswith("Printify_Published"):
            continue
        if clean(data.get("Product_Type")) not in {"Poster", "Acrylic", "Sticker"}:
            continue
        rows.append(data)

    def priority(data: dict[str, Any]) -> tuple[int, str]:
        product_type = clean(data.get("Product_Type"))
        rank = {"Poster": 0, "Acrylic": 1, "Sticker": 2}.get(product_type, 9)
        return rank, clean(data.get("ID"))

    selected = sorted(rows, key=priority)[:limit]
    out = []
    for data in selected:
        new_title = title_for(data)
        subject = subject_from_title(clean(data.get("Title")), clean(data.get("Product_Type")), clean(data.get("Category")))
        tags = tags_for(data, subject)
        new_price, price_decision = price_for(data)
        out.append(
            PivotRow(
                row_idx=int(data["_row_idx"]),
                local_id=clean(data.get("ID")),
                product_type=clean(data.get("Product_Type")),
                category=clean(data.get("Category")),
                printify_product_id=clean(data.get("Printify_Product_ID")),
                ebay_item_id=clean(data.get("eBay_Item_ID")),
                old_title=clean(data.get("Title")),
                old_description=clean(data.get("Description")),
                old_price=clean(data.get("Price")),
                dna_profile=clean(data.get("DNA Profile")),
                views=0,
                new_title=new_title,
                new_description=description_for(data, new_title, tags),
                new_price=new_price,
                search_tags=tags,
                price_decision=price_decision,
            )
        )
    return wb, ws, cols, out


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def prepare(limit: int = 42, apply_local: bool = False) -> list[PivotRow]:
    timestamp = now_text()
    wb, ws, cols, selected = load_candidates(limit)
    if apply_local:
        backup = EBAY_BOOK.with_name(f"eBay_listing.backup_quiet_jade_{datetime.now(NY):%Y%m%d_%H%M%S}.xlsx")
        shutil.copy2(EBAY_BOOK, backup)
    else:
        backup = None

    pivot_col = ensure_column(ws, cols, "Quiet_Jade_Pivot_Timestamp")
    target_col = ensure_column(ws, cols, "Quiet_Jade_Target")
    tags_col = ensure_column(ws, cols, "Quiet_Jade_Search_Tags")
    price_col = ensure_column(ws, cols, "Quiet_Jade_Price_Decision")
    sync_col = ensure_column(ws, cols, "Metadata_Sync_Status")
    title_len_failures = [row.local_id for row in selected if not (TITLE_MIN <= len(row.new_title) <= TITLE_MAX)]
    if title_len_failures:
        wb.close()
        raise RuntimeError(f"Title length failures: {title_len_failures[:10]}")

    if apply_local:
        for row in selected:
            ws.cell(row.row_idx, cols["Title"]).value = row.new_title
            ws.cell(row.row_idx, cols["Description"]).value = row.new_description
            ws.cell(row.row_idx, cols["Price"]).value = row.new_price
            ws.cell(row.row_idx, pivot_col).value = timestamp
            ws.cell(row.row_idx, target_col).value = "Operation Quiet Jade"
            ws.cell(row.row_idx, tags_col).value = ", ".join(row.search_tags)
            ws.cell(row.row_idx, price_col).value = row.price_decision
            ws.cell(row.row_idx, sync_col).value = "QUIET_JADE_PENDING_PRINTIFY_SYNC"
        wb.save(EBAY_BOOK)
    wb.close()

    pivot_rows = [
        {
            "Timestamp": timestamp,
            "ID": row.local_id,
            "Product_Type": row.product_type,
            "Category": row.category,
            "Printify_Product_ID": row.printify_product_id,
            "eBay_Item_ID": row.ebay_item_id,
            "Old_Price": row.old_price,
            "New_Price": row.new_price,
            "Old_Title_Length": len(row.old_title),
            "New_Title_Length": len(row.new_title),
            "Old_Title": row.old_title,
            "New_Title": row.new_title,
            "Search_Tags": ", ".join(row.search_tags),
            "Price_Decision": row.price_decision,
            "Apply_Status": "LOCAL_APPLIED" if apply_local else "DRY_PLAN",
        }
        for row in selected
    ]
    rollback_rows = [
        {
            "Timestamp": timestamp,
            "ID": row.local_id,
            "Printify_Product_ID": row.printify_product_id,
            "eBay_Item_ID": row.ebay_item_id,
            "Old_Title": row.old_title,
            "Old_Description": row.old_description,
            "Old_Price": row.old_price,
            "New_Title": row.new_title,
            "New_Description": row.new_description,
            "New_Price": row.new_price,
        }
        for row in selected
    ]
    write_csv(
        PIVOT_CSV,
        pivot_rows,
        [
            "Timestamp",
            "ID",
            "Product_Type",
            "Category",
            "Printify_Product_ID",
            "eBay_Item_ID",
            "Old_Price",
            "New_Price",
            "Old_Title_Length",
            "New_Title_Length",
            "Old_Title",
            "New_Title",
            "Search_Tags",
            "Price_Decision",
            "Apply_Status",
        ],
    )
    write_csv(
        ROLLBACK_CSV,
        rollback_rows,
        [
            "Timestamp",
            "ID",
            "Printify_Product_ID",
            "eBay_Item_ID",
            "Old_Title",
            "Old_Description",
            "Old_Price",
            "New_Title",
            "New_Description",
            "New_Price",
        ],
    )
    STATE_JSON.write_text(
        json.dumps(
            {
                "timestamp": timestamp,
                "selected": len(selected),
                "apply_local": apply_local,
                "backup": str(backup) if backup else "",
                "pivot_csv": str(PIVOT_CSV),
                "rollback_csv": str(ROLLBACK_CSV),
                "product_counts": {ptype: sum(1 for row in selected if row.product_type == ptype) for ptype in ["Poster", "Acrylic", "Sticker"]},
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"[QUIET-JADE] selected={len(selected)} apply_local={apply_local}")
    print(f"[QUIET-JADE] csv={PIVOT_CSV}")
    if backup:
        print(f"[QUIET-JADE] backup={backup}")
    return selected


def load_pivot_plan(limit: int = 0) -> list[dict[str, str]]:
    if not PIVOT_CSV.exists():
        return []
    rows = list(csv.DictReader(PIVOT_CSV.open("r", encoding="utf-8-sig", newline="")))
    rows = [row for row in rows if clean(row.get("Apply_Status")) == "LOCAL_APPLIED"]
    return rows[:limit] if limit else rows


def api_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {Config.Printify_API_KEY}",
        "Content-Type": "application/json",
    }


def request_with_retry(method: str, url: str, *, payload=None, attempts: int = 3) -> requests.Response:
    last_exc: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = requests.request(method, url, headers=api_headers(), json=payload, timeout=120)
            if response.status_code >= 500 and attempt < attempts:
                time.sleep(4 * attempt)
                continue
            return response
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt < attempts:
                time.sleep(4 * attempt)
    raise last_exc or RuntimeError("Printify request failed")


def append_sync_log(row: dict[str, Any]) -> None:
    headers = [
        "Timestamp",
        "ID",
        "Product_Type",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "HTTP_Get",
        "HTTP_Update",
        "HTTP_Publish",
        "Result",
        "Error",
    ]
    exists = SYNC_LOG.exists()
    with SYNC_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not exists:
            writer.writeheader()
        writer.writerow({name: row.get(name, "") for name in headers})


def load_local_metadata() -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers)}
    out = {}
    for data in ws.iter_rows(min_row=2, values_only=True):
        if not data or not data[cols["ID"]]:
            continue
        local_id = clean(data[cols["ID"]])
        out[local_id] = {
            "Title": clean(data[cols["Title"]]),
            "Description": clean(data[cols["Description"]]),
            "Price": clean(data[cols["Price"]]),
            "Printify_Product_ID": clean(data[cols["Printify_Product_ID"]]),
            "eBay_Item_ID": clean(data[cols["eBay_Item_ID"]]) if "eBay_Item_ID" in cols else "",
            "Product_Type": clean(data[cols["Product_Type"]]),
        }
    wb.close()
    return out


def update_sync_status(done: set[str], failed: set[str]) -> None:
    if not done and not failed:
        return
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    sync_col = ensure_column(ws, cols, "Metadata_Sync_Status")
    for row_idx in range(2, ws.max_row + 1):
        local_id = clean(ws.cell(row_idx, cols["ID"]).value)
        if local_id in done:
            ws.cell(row_idx, sync_col).value = "QUIET_JADE_SYNCED_PRINTIFY_PUBLISH"
        elif local_id in failed:
            ws.cell(row_idx, sync_col).value = "QUIET_JADE_SYNC_CHECK_REQUIRED"
    wb.save(EBAY_BOOK)
    wb.close()
    mark_pivot_synced(done, failed)


def mark_pivot_synced(done: set[str], failed: set[str]) -> None:
    if not PIVOT_CSV.exists():
        return
    rows = list(csv.DictReader(PIVOT_CSV.open("r", encoding="utf-8-sig", newline="")))
    if not rows:
        return
    headers = list(rows[0].keys())
    for name in ["Sync_Timestamp", "Sync_Result"]:
        if name not in headers:
            headers.append(name)
    timestamp = now_text()
    for row in rows:
        local_id = clean(row.get("ID"))
        if local_id in done:
            row["Apply_Status"] = "SYNCED_PRINTIFY"
            row["Sync_Timestamp"] = timestamp
            row["Sync_Result"] = "OK"
        elif local_id in failed:
            row["Apply_Status"] = "SYNC_CHECK_REQUIRED"
            row["Sync_Timestamp"] = timestamp
            row["Sync_Result"] = "FAILED"
    write_csv(PIVOT_CSV, rows, headers)


def sync_printify(limit: int = 0, dry_run: bool = False, sleep_min: float = 6.0, sleep_max: float = 14.0) -> int:
    if not Config.Printify_API_KEY:
        raise RuntimeError("Missing Printify_API_KEY")
    plan = load_pivot_plan(limit=limit)
    local = load_local_metadata()
    base = Config.Printify_API_URL.rstrip("/")
    done: set[str] = set()
    failed: set[str] = set()
    selected = 0
    for row in plan:
        local_id = clean(row.get("ID"))
        meta = local.get(local_id)
        if not meta:
            failed.add(local_id)
            append_sync_log({"Timestamp": now_text(), "ID": local_id, "Result": "SKIP", "Error": "missing local row"})
            continue
        product_id = meta["Printify_Product_ID"] or clean(row.get("Printify_Product_ID"))
        if not product_id:
            failed.add(local_id)
            append_sync_log({"Timestamp": now_text(), "ID": local_id, "Result": "SKIP", "Error": "missing Printify product id"})
            continue
        selected += 1
        if dry_run:
            print(f"[QUIET-JADE-DRY] {local_id} product={product_id} title={meta['Title'][:75]}")
            continue
        get_resp = request_with_retry("GET", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json")
        get_status = get_resp.status_code
        update_status = ""
        publish_status = ""
        error = ""
        result = "CHECK"
        try:
            get_resp.raise_for_status()
            product = get_resp.json()
            target_price = int(round((money(meta["Price"]) or 0) * 100))
            variants = []
            for variant in product.get("variants") or []:
                variants.append(
                    {
                        "id": variant["id"],
                        "price": target_price if variant.get("is_enabled") else variant.get("price"),
                        "is_enabled": bool(variant.get("is_enabled")),
                    }
                )
            payload = {
                "title": meta["Title"],
                "description": meta["Description"],
                "variants": variants,
            }
            update_resp = request_with_retry("PUT", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json", payload=payload)
            update_status = update_resp.status_code
            publish_resp = request_with_retry(
                "POST",
                f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
                payload=PUBLISH_TITLE_DESC_PRICE,
            )
            publish_status = publish_resp.status_code
            ok = update_status in {200, 201, 202, 204} and publish_status in {200, 201, 202, 204}
            if ok:
                done.add(local_id)
                result = "OK"
            else:
                failed.add(local_id)
                error = (update_resp.text[:300] + " " + publish_resp.text[:300]).strip()
        except Exception as exc:  # noqa: BLE001
            failed.add(local_id)
            error = f"{type(exc).__name__}: {exc}"
        append_sync_log(
            {
                "Timestamp": now_text(),
                "ID": local_id,
                "Product_Type": meta["Product_Type"],
                "Printify_Product_ID": product_id,
                "eBay_Item_ID": meta["eBay_Item_ID"],
                "HTTP_Get": get_status,
                "HTTP_Update": update_status,
                "HTTP_Publish": publish_status,
                "Result": result,
                "Error": error,
            }
        )
        print(f"[QUIET-JADE-SYNC] {local_id} get={get_status} update={update_status} publish={publish_status} result={result}")
        time.sleep(random.uniform(sleep_min, sleep_max))
    if not dry_run:
        update_sync_status(done, failed)
    print(f"[QUIET-JADE-DONE] selected={selected} synced={len(done)} failed={len(failed)} dry_run={dry_run}")
    return len(done)


def main() -> None:
    parser = argparse.ArgumentParser(description="Operation Quiet Jade eBay intent and value pivot.")
    parser.add_argument("--limit", type=int, default=42)
    parser.add_argument("--prepare", action="store_true")
    parser.add_argument("--apply-local", action="store_true")
    parser.add_argument("--sync-printify", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sleep-min", type=float, default=6.0)
    parser.add_argument("--sleep-max", type=float, default=14.0)
    args = parser.parse_args()
    if args.prepare or args.apply_local:
        prepare(limit=args.limit, apply_local=args.apply_local and not args.dry_run)
    if args.sync_printify:
        sync_printify(
            limit=args.limit,
            dry_run=args.dry_run,
            sleep_min=args.sleep_min,
            sleep_max=max(args.sleep_min, args.sleep_max),
        )


if __name__ == "__main__":
    main()
