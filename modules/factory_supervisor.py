"""OpenClaw factory supervisor.

This is the durable control layer for unattended work. It does not replace the
specialized modules; it decides which modules are safe to run under the current
network, account, and QA state, then writes a queue that can be resumed later.
"""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
DATABASE_DIR = PROJECT_ROOT / "Database"
REPORTS_DIR = PROJECT_ROOT / "Reports"

EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
FIX_QUEUE = DATABASE_DIR / "eBay_Online_Cover_Fix_Queue.csv"
REPAIR_DECISIONS = DATABASE_DIR / "eBay_Cover_Repair_Decisions.csv"
REPLACEMENT_QUEUE = DATABASE_DIR / "eBay_Cover_Replacement_Queue.csv"
DEFAULT_AUDIT = DATABASE_DIR / "Printify_Image_Default_Audit.csv"
ACTION_QUEUE = DATABASE_DIR / "Factory_Autopilot_Action_Queue.csv"
ACTION_SUMMARY = DATABASE_DIR / "Factory_Autopilot_Action_Queue.md"
STATE_JSON = DATABASE_DIR / "Factory_Autopilot_State.json"
RUN_LOG = DATABASE_DIR / "Factory_Autopilot_Run_Log.csv"


LOCAL_MODULES = [
    ("printify_login_guard.py", 120),
    ("local_listing_qa.py", 180),
    ("ebay_cover_repair_decision.py", 120),
    ("ebay_cover_replacement_queue.py", 120),
    ("ebay_title_repair_queue.py", 120),
    ("unified_listing_registry.py", 180),
    ("market_signal_planner.py", 180),
    ("ebay_experiment_report.py", 120),
    ("ebay_traffic_diagnosis.py", 120),
    ("ebay_profile_packet.py", 120),
    ("product_blueprint_next_plan.py", 120),
    ("etsy_app_status_probe.py", 90),
    ("factory_backlog.py", 120),
    ("factory_morning_report.py", 180),
]


@dataclass
class FactoryAction:
    priority: int
    lane: str
    action: str
    status: str
    reason: str
    command: str
    requires_network: str
    requires_login: str
    risk: str


def now() -> datetime:
    return datetime.now(ZoneInfo("America/New_York"))


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def csv_count(path: Path) -> int:
    return len(read_csv(path))


def count_by(path: Path, column: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in read_csv(path):
        key = clean(row.get(column)) or "Unknown"
        counts[key] = counts.get(key, 0) + 1
    return counts


def default_insufficiency_count() -> int:
    count = 0
    for row in read_csv(DEFAULT_AUDIT):
        try:
            selected = int(row.get("Selected_Count") or 0)
            expected = int(row.get("Expected_Count") or 0)
            defaults = int(row.get("Default_Count") or 0)
        except ValueError:
            if clean(row.get("Result")) == "CHECK":
                count += 1
            continue
        if clean(row.get("Error")) or selected < expected or defaults < 1:
            count += 1
    return count


def latest_performance_age_hours() -> float | None:
    rows = read_csv(PERFORMANCE_LOG)
    stamps = [clean(row.get("Snapshot_Timestamp")) for row in rows if clean(row.get("Snapshot_Timestamp"))]
    if not stamps:
        return None
    latest = max(stamps)
    for fmt in ("%Y-%m-%d %H:%M:%S %z", "%Y-%m-%d %H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
        try:
            parsed = datetime.strptime(latest, fmt)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=ZoneInfo("America/New_York"))
            return max(0.0, (now() - parsed.astimezone(ZoneInfo("America/New_York"))).total_seconds() / 3600)
        except ValueError:
            continue
    return None


def workbook_counts() -> dict[str, object]:
    counts: dict[str, object] = {
        "rows": 0,
        "stable": 0,
        "published": 0,
        "ready": 0,
        "by_product": {},
        "publishable_without_cover_block": 0,
    }
    if not EBAY_BOOK.exists():
        return counts

    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        cols = {name: idx for idx, name in enumerate(headers)}
        by_product: dict[str, dict[str, int]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[cols["ID"]]:
                continue
            product = clean(row[cols.get("Product_Type")]) or "Unknown"
            status = clean(row[cols.get("Status")])
            by_product.setdefault(product, {"stable": 0, "published": 0, "ready": 0, "rows": 0})
            by_product[product]["rows"] += 1
            counts["rows"] = int(counts["rows"]) + 1
            if status.startswith("Printify_Published"):
                counts["published"] = int(counts["published"]) + 1
                counts["stable"] = int(counts["stable"]) + 1
                by_product[product]["published"] += 1
                by_product[product]["stable"] += 1
            elif status.startswith("Printify_UI_Mockups"):
                counts["stable"] = int(counts["stable"]) + 1
                by_product[product]["stable"] += 1
            elif status == "Ready_for_Printify":
                counts["ready"] = int(counts["ready"]) + 1
                by_product[product]["ready"] += 1
        counts["by_product"] = by_product
    finally:
        wb.close()
    return counts


def network_strategy(skip_network: bool, count: int) -> dict[str, object]:
    if skip_network:
        return {"mode": "unknown", "max_parallel": 0, "batch_size": 0, "reason": "network guard skipped"}
    try:
        from modules import network_guard

        payload = network_guard.report(count=count)
        return payload["strategy"]
    except Exception as exc:  # noqa: BLE001
        return {"mode": "pause", "max_parallel": 0, "batch_size": 0, "reason": f"network guard failed: {exc}"}


def printify_ui_status(cdp_port: int = 9222) -> dict[str, str]:
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{cdp_port}/json/list", timeout=8) as response:
            pages = json.load(response)
    except Exception as exc:  # noqa: BLE001
        return {"status": "UNAVAILABLE", "reason": str(exc)}
    urls = [clean(page.get("url")) for page in pages if "printify.com" in clean(page.get("url"))]
    if not urls:
        return {"status": "NO_PRINTIFY_TAB", "reason": "No Printify tab in CDP browser."}
    if any("/app/" in url and "/auth/login" not in url for url in urls):
        return {"status": "LOGGED_IN", "reason": "Printify app page is available in CDP browser."}
    if any("/auth/login" in url for url in urls):
        return {"status": "LOGIN_REQUIRED", "reason": "Printify CDP browser is on auth/login."}
    return {"status": "UNKNOWN", "reason": "; ".join(urls[:3])}


def build_actions(strategy: dict[str, object]) -> list[FactoryAction]:
    fix_rows = csv_count(FIX_QUEUE)
    source_repairs = count_by(REPAIR_DECISIONS, "Repair_Method").get("SOURCE_REPAIR_REQUIRED", 0)
    replacement_ready = count_by(REPLACEMENT_QUEUE, "Replacement_Status").get("READY_TO_REPLACE_VERIFIED", 0)
    default_checks = default_insufficiency_count()
    perf_age = latest_performance_age_hours()
    workbook = workbook_counts()
    network_mode = clean(strategy.get("mode"))
    printify_ui = printify_ui_status()

    actions = [
        FactoryAction(
            100,
            "local",
            "Refresh local QA, registry, market queue, cover decisions, experiment report, and morning report.",
            "READY",
            "Safe low-bandwidth maintenance keeps the factory state current while account/image writes are paused.",
            "py modules\\factory_supervisor.py --execute-local --skip-network",
            "no",
            "no",
            "low",
        )
    ]

    if replacement_ready:
        actions.append(
            FactoryAction(
                97,
                "replacement",
                "Create one verified replacement listing for a live cover failure that survived source repair.",
                "READY_TO_REPLACE_VERIFIED",
                f"{replacement_ready} listing already failed source repair plus live eBay buyer-page audit.",
                "Build one replacement from local assets, publish only after image/design QA, then retire old item after live audit passes.",
                "yes",
                "Printify API/UI and eBay live audit",
                "high",
            )
        )

    if fix_rows or source_repairs:
        status = "READY_SINGLE_SKU_REPAIR"
        command = "py modules\\factory_cover_repair_runner.py --limit 1 --post-sync-wait 120"
        if printify_ui["status"] != "LOGGED_IN":
            status = "WAIT_PRINTIFY_LOGIN"
            command = "py modules\\factory_cover_repair_runner.py --dry-run --post-sync-wait 0"
        actions.append(
            FactoryAction(
                95,
                "cover_gate",
                "Repair one Printify source cover, then live-audit eBay before scaling.",
                status,
                f"Live cover queue has {fix_rows} rows; {source_repairs} require Printify source repair or replacement listings. Printify UI: {printify_ui['status']} - {printify_ui['reason']}",
                command,
                "yes",
                "Printify remote-debug profile",
                "medium",
            )
        )
    elif default_checks:
        actions.append(
            FactoryAction(
                92,
                "cover_gate",
                "Resolve true Printify image insufficiency before public publishing.",
                "BLOCKING_PUBLISH",
                f"{default_checks} rows have too few selected images, no default image, or an image audit error. Multiple official/default mockups are allowed.",
                "py modules\\printify_image_default_audit.py --sleep-seconds 1",
                "yes",
                "Printify API",
                "medium",
            )
        )
    else:
        actions.append(
            FactoryAction(
                70,
                "publish",
                "Publish small cooled batch if network guard is healthy.",
                "READY" if network_mode in {"normal_low_risk"} else "WAIT_NETWORK",
                f"Stable={workbook['stable']} published={workbook['published']} ready={workbook['ready']}; network={network_mode}.",
                "py modules\\printify_publish_scheduler.py --limit 3 --min-delay 180 --max-delay 420",
                "yes",
                "Printify API",
                "high",
            )
        )

    if perf_age is None or perf_age > 12:
        actions.append(
            FactoryAction(
                65,
                "read_only_market",
                "Refresh eBay Seller Hub performance snapshot.",
                "READY" if network_mode != "pause" else "WAIT_NETWORK",
                "Performance data is stale or absent; this is read-only but browser/network dependent.",
                "py modules\\ebay_sellerhub_snapshot.py",
                "yes",
                "eBay Seller Hub",
                "low",
            )
        )

    actions.append(
        FactoryAction(
            63,
            "production_design_qa",
            "Run a tiny Printify production-design audit before any larger online batch.",
            "READY" if network_mode != "pause" else "WAIT_NETWORK",
            "This checks whether Printify front print-area art visually matches local Production_Design files; keep it small under weak Wi-Fi.",
            "py modules\\printify_design_audit.py --limit 2 --sleep-seconds 1",
            "yes",
            "Printify API",
            "low",
        )
    )

    actions.append(
        FactoryAction(
            55,
            "etsy",
            "Keep Etsy launch packet local until shop/API approval is ready.",
            "WAIT_USER_OR_API_APPROVAL",
            "Etsy developer app is pending approval and Rex has not asked to publish Etsy listings yet.",
            "py modules\\etsy_digital_listing_export.py",
            "no",
            "no",
            "low",
        )
    )

    actions.append(
        FactoryAction(
            50,
            "copy_experiment",
            "Continue low-bandwidth SEO/title/description experiment analysis.",
            "READY",
            "Ads alone did not move zero-view listings; controlled copy/image experiments are the next learning loop.",
            "py modules\\ebay_experiment_report.py",
            "no",
            "no",
            "low",
        )
    )
    return sorted(actions, key=lambda item: (-item.priority, item.lane, item.action))


def write_state(strategy: dict[str, object], actions: list[FactoryAction]) -> None:
    DATABASE_DIR.mkdir(exist_ok=True)
    state = {
        "generated_at": now().isoformat(timespec="seconds"),
        "timezone": "America/New_York",
        "network_strategy": strategy,
        "workbook_counts": workbook_counts(),
        "fix_queue_rows": csv_count(FIX_QUEUE),
        "repair_decisions": count_by(REPAIR_DECISIONS, "Repair_Method"),
        "printify_default_audit": count_by(DEFAULT_AUDIT, "Result"),
        "performance_age_hours": latest_performance_age_hours(),
        "printify_ui_status": printify_ui_status(),
        "actions": [asdict(item) for item in actions],
    }
    STATE_JSON.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")

    fields = list(asdict(actions[0]).keys()) if actions else list(FactoryAction.__dataclass_fields__.keys())
    with ACTION_QUEUE.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for action in actions:
            writer.writerow(asdict(action))

    lines = [
        "# Factory Autopilot Action Queue",
        "",
        f"Generated: {state['generated_at']} America/New_York",
        "",
        f"- Network mode: {strategy.get('mode')} ({strategy.get('reason')})",
        f"- eBay workbook rows: {state['workbook_counts']['rows']}",
        f"- Stable: {state['workbook_counts']['stable']}",
        f"- Published: {state['workbook_counts']['published']}",
        f"- Ready for Printify: {state['workbook_counts']['ready']}",
        f"- Live cover fix queue rows: {state['fix_queue_rows']}",
        f"- Repair decisions: {state['repair_decisions']}",
        f"- Printify default audit: {state['printify_default_audit']}",
        f"- Printify UI status: {state['printify_ui_status']}",
        "",
        "## Actions",
        "",
    ]
    for action in actions:
        lines.extend(
            [
                f"### P{action.priority} {action.lane}: {action.status}",
                f"- Action: {action.action}",
                f"- Reason: {action.reason}",
                f"- Command: `{action.command}`",
                f"- Network: {action.requires_network}; login: {action.requires_login}; risk: {action.risk}",
                "",
            ]
        )
    ACTION_SUMMARY.write_text("\n".join(lines), encoding="utf-8")


def append_run_log(module_name: str, status: str, detail: str) -> None:
    fields = ["Timestamp", "Module", "Status", "Detail"]
    exists = RUN_LOG.exists()
    with RUN_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        if not exists:
            writer.writeheader()
        writer.writerow(
            {
                "Timestamp": now().isoformat(timespec="seconds"),
                "Module": module_name,
                "Status": status,
                "Detail": detail[:1000],
            }
        )


def run_local_cycle() -> int:
    failures = 0
    for script, timeout in LOCAL_MODULES:
        path = PROJECT_ROOT / "modules" / script
        if not path.exists():
            append_run_log(script, "SKIP", "missing script")
            continue
        print(f"[AUTOPILOT-LOCAL] start {script}")
        try:
            result = subprocess.run(
                [sys.executable, str(path)],
                cwd=PROJECT_ROOT,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                timeout=timeout,
            )
            output = (result.stdout or "").strip()
            status = "OK" if result.returncode == 0 else f"FAIL_{result.returncode}"
            if result.returncode != 0:
                failures += 1
            append_run_log(script, status, output)
            print(f"[AUTOPILOT-LOCAL] {script} {status}")
            if output:
                print(output[-1600:])
        except subprocess.TimeoutExpired as exc:
            failures += 1
            append_run_log(script, "TIMEOUT", clean(exc))
            print(f"[AUTOPILOT-LOCAL] {script} TIMEOUT")
    return failures


def main() -> None:
    parser = argparse.ArgumentParser(description="OpenClaw unattended factory supervisor.")
    parser.add_argument("--execute-local", action="store_true", help="Run safe local maintenance modules.")
    parser.add_argument("--skip-network", action="store_true", help="Do not run network guard.")
    parser.add_argument("--network-count", type=int, default=2)
    args = parser.parse_args()

    strategy = network_strategy(skip_network=args.skip_network, count=args.network_count)
    if args.execute_local:
        failures = run_local_cycle()
        print(f"[AUTOPILOT-LOCAL-DONE] failures={failures}")

    actions = build_actions(strategy)
    write_state(strategy, actions)
    print(f"[AUTOPILOT] state={STATE_JSON}")
    print(f"[AUTOPILOT] queue={ACTION_QUEUE}")
    for action in actions[:8]:
        print(f"[AUTOPILOT] P{action.priority} {action.lane} {action.status}: {action.action}")


if __name__ == "__main__":
    main()
