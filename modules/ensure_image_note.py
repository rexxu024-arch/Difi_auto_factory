"""Ensure marketplace descriptions include the product-image expectation note."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.edit_for_platforms import _ensure_image_note

DATABASE_DIR = PROJECT_ROOT / "Database"
DEFAULT_BOOKS = [
    DATABASE_DIR / "eBay_listing.xlsx",
    DATABASE_DIR / "Etsy_listing.xlsx",
]


def run(paths: list[Path]) -> None:
    for path in paths:
        if not path.exists():
            print(f"[IMAGE-NOTE-SKIP] missing {path}")
            continue
        workbook = load_workbook(path)
        sheet = workbook.active
        headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
        cols = {name: idx + 1 for idx, name in enumerate(headers) if name}
        if "Description" not in cols or "ID" not in cols:
            print(f"[IMAGE-NOTE-SKIP] missing ID/Description columns {path}")
            workbook.close()
            continue
        changed = 0
        for row_idx in range(2, sheet.max_row + 1):
            if not sheet.cell(row_idx, cols["ID"]).value:
                continue
            old = str(sheet.cell(row_idx, cols["Description"]).value or "").strip()
            new = _ensure_image_note(old)
            if new != old:
                sheet.cell(row_idx, cols["Description"]).value = new
                changed += 1
        if changed:
            workbook.save(path)
        workbook.close()
        print(f"[IMAGE-NOTE] {path} changed={changed}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--books", default="", help="Comma-separated workbook paths; defaults to eBay and Etsy listing books.")
    args = parser.parse_args()
    paths = [Path(part.strip()) for part in args.books.split(",") if part.strip()] or DEFAULT_BOOKS
    run(paths)


if __name__ == "__main__":
    main()
