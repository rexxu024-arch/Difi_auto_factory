import argparse
import asyncio
import json
import os
import sys
import time
import urllib.request
from pathlib import Path

import requests
import websockets
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules import printify_uploader
from modules import printify_design_audit
from modules.printify_primary_audit import _default_matches_cover
from modules.printify_mockup_ui_uploader import _assets, _default_count, _fetch_product, _selected_count, _upload_mockups


EBAY_BOOK = PROJECT_ROOT / "Database" / "eBay_listing.xlsx"
CHROME_DEBUG_URL = (
    f"http://127.0.0.1:{os.getenv('OPENCLAW_PRINTIFY_CDP_PORT') or os.getenv('OPENCLAW_CDP_PORT') or '9223'}"
)
EXPECTED_MOCKUPS = {
    "Sticker": 5,
    "Poster": 4,
    "Acrylic": 4,
}


class PrintifyLoginRequired(RuntimeError):
    pass


def _assert_printify_ui_logged_in():
    try:
        with urllib.request.urlopen(f"{CHROME_DEBUG_URL}/json/list", timeout=10) as response:
            pages = json.load(response)
    except Exception as exc:
        raise PrintifyLoginRequired(f"Printify browser control is unavailable: {exc}") from exc
    printify_pages = [page for page in pages if "printify.com/app" in page.get("url", "")]
    if not printify_pages:
        raise PrintifyLoginRequired("Open and log in to Printify in the Codex browser before UI mockup upload.")
    if any("/auth/login" in page.get("url", "") for page in printify_pages):
        raise PrintifyLoginRequired("Printify login required in Codex browser; stopping before creating more products.")


def _open_product_tab(product_id):
    url = f"https://printify.com/app/mockup-library/shops/{Config.Printify_SHOP_ID}/products/{product_id}?revealUploads=true"
    with urllib.request.urlopen(f"{CHROME_DEBUG_URL}/json/list", timeout=10) as response:
        pages = json.load(response)
    tab = next((page for page in pages if product_id in page.get("url", "") and page.get("webSocketDebuggerUrl")), None)
    if not tab:
        tab = next(
            (page for page in pages if "printify.com/app/" in page.get("url", "") and page.get("webSocketDebuggerUrl")),
            None,
        )
    if not tab:
        req = urllib.request.Request(f"{CHROME_DEBUG_URL}/json/new", data=url.encode("utf-8"), method="PUT")
        tab = json.loads(urllib.request.urlopen(req, timeout=10).read().decode("utf-8", "ignore"))
    if tab.get("webSocketDebuggerUrl"):
        async def navigate():
            async with websockets.connect(tab["webSocketDebuggerUrl"], max_size=20_000_000) as sock:
                seq = 1
                async def send(method, params=None):
                    nonlocal seq
                    msg = {"id": seq, "method": method}
                    if params is not None:
                        msg["params"] = params
                    await sock.send(json.dumps(msg))
                    my_id = seq
                    seq += 1
                    while True:
                        data = json.loads(await sock.recv())
                        if data.get("id") == my_id:
                            return data
                await send("Page.enable")
                await send("Page.navigate", {"url": url})
        asyncio.run(navigate())
    time.sleep(8)


def _headers():
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}", "Content-Type": "application/json"}


def _stable_selected_count(product_id, expected_count=5, checks=3, delay=8):
    last_count = None
    for _ in range(checks):
        product = _fetch_product(product_id)
        last_count = _selected_count(product)
        if last_count < expected_count:
            return last_count
        time.sleep(delay)
    return last_count


def _acrylic_mockup_ok(product):
    images = product.get("images") or []
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    labels = {str(image.get("src") or "").split("camera_label=")[-1].split("&")[0] for image in selected}
    return {"front", "back", "side-1", "side-2"}.issubset(labels), len(selected), labels


def _poster_mockup_ok(product):
    images = product.get("images") or []
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    official = [
        image for image in selected
        if "images.printify.com/mockup" in str(image.get("src") or "")
    ]
    return len(selected) >= 4 and bool(official), len(selected), len(official)


def _sticker_mockup_ok(product):
    images = product.get("images") or []
    selected = [image for image in images if image.get("is_selected_for_publishing") is not False]
    official = [
        image for image in selected
        if "images.printify.com/mockup" in str(image.get("src") or "")
    ]
    custom_gallery = [
        image for image in selected
        if "pfy-prod-products-mockup-media" in str(image.get("src") or "")
    ]
    return len(selected) >= 3 and len(official) >= 3 and not custom_gallery, len(selected), len(official), len(custom_gallery)


def _expected_mockups(product_type):
    return EXPECTED_MOCKUPS.get(_canonical_product_filter(product_type) or "Sticker", 5)


def _assert_production_design(item_id, product_id, row):
    report = printify_design_audit.assert_product_design_matches(product_id, row["Production_Path"])
    print(
        f"[DESIGN-OK] {item_id} production visual_match=True "
        f"exact_sha={report['exact_sha_match']} size={report['local_size']} remote_id={report['remote_image_id']}"
    )
    return report


def _assert_primary_cover(item_id, product_id, row):
    product_type = row.get("Product_Type") or "Sticker"
    ok, note = _default_matches_cover(product_id, Path(row["Cover_Path"]), product_type=product_type)
    if not ok:
        raise RuntimeError(f"Primary cover mismatch: {note}")
    print(f"[PRIMARY-OK] {item_id} {note}")
    return note


def _ensure_product_id(row):
    existing = str(row.get("Printify_Product_ID") or "").strip()
    if existing:
        return existing
    production_path, stock_paths, missing = printify_uploader._validate_row_assets(row)
    if missing:
        raise RuntimeError("; ".join(missing))
    version = str(int(time.time()))
    allow_jpeg = _canonical_product_filter(row.get("Product_Type")) in {"Poster", "Acrylic"}
    production_id = printify_uploader._image_upload(
        production_path,
        f"{row['ID']}_Production_{version}.png",
        allow_jpeg=allow_jpeg,
    )
    stock_ids = [
        printify_uploader._image_upload(
            path,
            f"{row['ID']}_{label}_{version}.png",
            allow_jpeg=allow_jpeg,
        )
        for label, path in (
            [item for item in stock_paths if item[0] == "Cover"]
            if _canonical_product_filter(row.get("Product_Type")) == "Sticker"
            else stock_paths
        )
    ]
    payload = printify_uploader._build_payload(row, stock_ids, production_id)
    last_error = None
    for attempt in range(1, 4):
        try:
            print(f"[PRODUCT-CREATE] {row['ID']} attempt={attempt}", flush=True)
            response = requests.post(
                f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products.json",
                headers=_headers(),
                json=payload,
                timeout=180,
            )
            response.raise_for_status()
            product_id = response.json()["id"]
            print(f"[PRODUCT-CREATE-OK] {row['ID']} product={product_id}", flush=True)
            return product_id
        except Exception as exc:
            last_error = exc
            if attempt >= 3:
                break
            print(f"[PRINTIFY-RETRY] {row['ID']} product create attempt {attempt} failed: {exc}", flush=True)
            time.sleep(5 * attempt)
    raise last_error


def _canonical_product_filter(value):
    if not value:
        return None
    value = str(value).strip().lower()
    if value.startswith("poster"):
        return "Poster"
    if value.startswith("acry"):
        return "Acrylic"
    if value.startswith("stick"):
        return "Sticker"
    return None


def _load_workbook_rows(limit, product_type=None, ids=None):
    id_set = set(ids or [])
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Printify_Product_ID" not in headers:
        ws.cell(1, ws.max_column + 1).value = "Printify_Product_ID"
        headers.append("Printify_Product_ID")
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    rows = []
    eligible = {
        "Ready_for_Printify",
        "Printify_BaseStaged_DefaultMockups3",
        "Printify_PhotoMismatch",
        "Printify_UI_Failed",
        "Printify_PrimaryFix_Needed",
        "Printify_MockupsPending",
    }
    product_filter = _canonical_product_filter(product_type)
    for row_idx in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_idx, cols[header]).value for header in headers}
        row["_row_idx"] = row_idx
        if id_set and str(row.get("ID") or "").strip() not in id_set:
            continue
        if row.get("Status") in eligible:
            if product_filter and _canonical_product_filter(row.get("Product_Type")) != product_filter:
                continue
            rows.append(row)
            if limit and len(rows) >= limit:
                break
    return wb, ws, headers, cols, rows


def _set_cell(ws, cols, row_idx, name, value):
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    ws.cell(row_idx, cols[name]).value = value


def run(limit=0, batch_size=75, batch_delay=3600, publish=False, product_type=None, ids=None):
    wb, ws, headers, cols, rows = _load_workbook_rows(limit, product_type=product_type, ids=ids)
    completed = 0
    try:
        for row in rows:
            item_id = row["ID"]
            row_idx = row["_row_idx"]
            try:
                product_filter = _canonical_product_filter(row.get("Product_Type"))
                if product_filter == "Sticker":
                    _assert_printify_ui_logged_in()
                product_id = _ensure_product_id(row)
                _set_cell(ws, cols, row_idx, "Printify_Product_ID", product_id)
                _set_cell(ws, cols, row_idx, "Status", "Printify_BaseStaged_DefaultMockups3")
                wb.save(EBAY_BOOK)
                _assert_production_design(item_id, product_id, row)

                expected_count = _expected_mockups(row.get("Product_Type") or "Sticker")
                if product_filter == "Acrylic":
                    product = _fetch_product(product_id)
                    ok, count, labels = _acrylic_mockup_ok(product)
                    time.sleep(8)
                    product = _fetch_product(product_id)
                    stable_ok, stable_count, stable_labels = _acrylic_mockup_ok(product)
                    if not stable_ok:
                        raise RuntimeError(
                            f"acrylic mockups missing official views: selected={count}, stable={stable_count}, "
                            f"labels={sorted(labels)} stable_labels={sorted(stable_labels)}"
                        )
                    stable_defaults = _default_count(product)
                    if stable_defaults < 1:
                        raise RuntimeError("acrylic default mockups=0, expected at least 1")
                    _assert_production_design(item_id, product_id, row)
                    _assert_primary_cover(item_id, product_id, row)
                    status_count = stable_count if stable_count != expected_count else expected_count
                    _set_cell(ws, cols, row_idx, "Status", f"Printify_UI_Mockups{status_count}")
                    wb.save(EBAY_BOOK)
                    completed += 1
                    print(
                        f"[FULL-PIPELINE] {completed}/{len(rows)} {item_id} "
                        f"product={product_id} selected_mockups={stable_count} acrylic_views={sorted(stable_labels)}"
                    )
                    continue

                if product_filter == "Poster":
                    stable_ok = False
                    stable_count = 0
                    official_count = 0
                    for _ in range(30):
                        product = _fetch_product(product_id)
                        stable_ok, stable_count, official_count = _poster_mockup_ok(product)
                        if stable_ok:
                            break
                        time.sleep(10)
                    if not stable_ok:
                        raise RuntimeError(
                            f"poster official mockups missing: selected={stable_count}, official={official_count}"
                        )
                    stable_defaults = _default_count(product)
                    if stable_defaults < 1:
                        raise RuntimeError("poster default mockups=0, expected at least 1")
                    _assert_production_design(item_id, product_id, row)
                    _set_cell(ws, cols, row_idx, "Status", f"Printify_UI_Mockups{stable_count}")
                    wb.save(EBAY_BOOK)
                    completed += 1
                    print(
                        f"[FULL-PIPELINE] {completed}/{len(rows)} {item_id} "
                        f"product={product_id} selected_mockups={stable_count} poster_official={official_count}"
                    )
                    continue

                if product_filter == "Sticker":
                    stable_ok = False
                    stable_count = 0
                    official_count = 0
                    custom_count = 0
                    for _ in range(30):
                        product = _fetch_product(product_id)
                        stable_ok, stable_count, official_count, custom_count = _sticker_mockup_ok(product)
                        if stable_ok:
                            break
                        time.sleep(10)
                    if not stable_ok:
                        raise RuntimeError(
                            f"sticker official cover mockups missing or custom U images selected: "
                            f"selected={stable_count}, official={official_count}, custom_gallery={custom_count}"
                        )
                    stable_defaults = _default_count(product)
                    if stable_defaults < 1:
                        raise RuntimeError("sticker default mockups=0, expected at least 1")
                    _assert_production_design(item_id, product_id, row)
                    _set_cell(ws, cols, row_idx, "Status", f"Printify_UI_Mockups{stable_count}")
                    wb.save(EBAY_BOOK)
                    completed += 1
                    print(
                        f"[FULL-PIPELINE] {completed}/{len(rows)} {item_id} "
                        f"product={product_id} selected_mockups={stable_count} sticker_official={official_count}"
                    )
                    continue

                _open_product_tab(product_id)
                asyncio.run(
                    _upload_mockups(
                        product_id,
                        _assets({**row, "Printify_Product_ID": product_id}),
                        publish=publish,
                        product_type=row.get("Product_Type") or "Sticker",
                    )
                )
                product = _fetch_product(product_id)
                count = _selected_count(product)
                if count < expected_count:
                    raise RuntimeError(f"selected mockups={count}, expected at least {expected_count}")
                defaults = _default_count(product)
                if defaults < 1:
                    raise RuntimeError("default mockups=0, expected at least 1")
                stable_count = _stable_selected_count(product_id, expected_count=expected_count)
                if stable_count < expected_count:
                    raise RuntimeError(
                        f"selected mockups unstable after save: {stable_count}, expected at least {expected_count}"
                    )
                product = _fetch_product(product_id)
                stable_defaults = _default_count(product)
                if stable_defaults < 1:
                    raise RuntimeError("default mockups unstable after save: 0, expected at least 1")
                _assert_production_design(item_id, product_id, row)
                _assert_primary_cover(item_id, product_id, row)
                _set_cell(
                    ws,
                    cols,
                    row_idx,
                    "Status",
                    f"Printify_Published_Mockups{expected_count}" if publish else f"Printify_UI_Mockups{expected_count}",
                )
                wb.save(EBAY_BOOK)
                completed += 1
                print(f"[FULL-PIPELINE] {completed}/{len(rows)} {item_id} product={product_id} selected_mockups={expected_count}")
                if completed % 2 == 0:
                    print(f"[DESIGN-AUDIT-CHECKPOINT] Last 2 completed products passed exact Production_Design SHA audit.")
                if batch_size and completed % batch_size == 0 and completed < len(rows):
                    print(f"[BATCH-COOLDOWN] {completed}/{len(rows)} sleeping {batch_delay}s")
                    time.sleep(batch_delay)
            except PrintifyLoginRequired as exc:
                print(f"[FULL-PIPELINE-PAUSED] {item_id}: {exc}")
                break
            except Exception as exc:
                message = str(exc)
                if "Production design mismatch" in message:
                    status = "Printify_DesignMismatch"
                elif "official mockups missing" in message:
                    status = "Printify_MockupsPending"
                else:
                    status = "Printify_UI_Failed"
                _set_cell(ws, cols, row_idx, "Status", status)
                wb.save(EBAY_BOOK)
                print(f"[FULL-PIPELINE-FAIL] {item_id}: {exc}")
                continue
    finally:
        wb.close()
    print(f"[DONE] Full Printify pipeline completed: {completed}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--batch-size", type=int, default=75)
    parser.add_argument("--batch-delay", type=int, default=3600)
    parser.add_argument("--publish", action="store_true")
    parser.add_argument("--product-type", default=None, choices=["Sticker", "Poster", "Acrylic"])
    parser.add_argument("--ids", default="", help="Comma-separated listing IDs to process first.")
    args = parser.parse_args()
    ids = [part.strip() for part in args.ids.split(",") if part.strip()] or None
    run(
        limit=args.limit,
        batch_size=args.batch_size,
        batch_delay=args.batch_delay,
        publish=args.publish,
        product_type=args.product_type,
        ids=ids,
    )
