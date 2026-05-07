import argparse
import hashlib
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
OUTPUT_ROOT = PROJECT_ROOT / "Output"
OUTPUT_DIR = OUTPUT_ROOT / "Sticker" / "Kiss-Cut"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
ETSY_BOOK = DATABASE_DIR / "Etsy_listing.xlsx"
EBAY_HEADERS = [
    "ID",
    "SKU",
    "Category",
    "Product_Type",
    "Title",
    "Description",
    "Price",
    "DNA Profile",
    "Production_Path",
    "Cover_Path",
    "Gallery_U1_Path",
    "Gallery_U2_Path",
    "Gallery_U3_Path",
    "Gallery_U4_Path",
    "Status",
    "Timestamp",
    "Printify_Product_ID",
]
ETSY_HEADERS = ["ID", "Raw_Metadata", "Production_Path", "Status", "Timestamp"]
DEFAULT_PRICE = os.getenv("STICKER_DEFAULT_PRICE", "$11.99")
PRODUCT_CONFIGS = {
    "Sticker": {
        "output_dir": OUTPUT_ROOT / "Sticker" / "Kiss-Cut",
        "price": os.getenv("STICKER_DEFAULT_PRICE", "$11.99"),
        "title_required": "4pc 6x6",
        "product_phrase": "Kiss-Cut Sticker",
        "includes": "One 6x6 kiss-cut sheet with 4 individual sticker designs.",
        "material": "Durable kiss-cut vinyl sticker sheet with waterproof finish.",
        "size": "6x6 kiss-cut sheet with four coordinated designs.",
    },
    "Poster": {
        "output_dir": OUTPUT_ROOT / "Poster" / "Premium-Matte-Vertical",
        "price": os.getenv("POSTER_DEFAULT_PRICE", "$34.99"),
        "title_required": "12x18",
        "product_phrase": "Matte Poster",
        "includes": "One 12x18 premium matte vertical poster.",
        "material": "Premium matte vertical poster through Printify Choice.",
        "size": "12x18 vertical wall art.",
    },
    "Acrylic": {
        "output_dir": OUTPUT_ROOT / "Acrylic" / "Photo-Block",
        "price": os.getenv("ACRYLIC_DEFAULT_PRICE", "$89.99"),
        "title_required": "5x7",
        "product_phrase": "Acrylic Photo Block",
        "includes": "One 5x7 vertical acrylic photo block.",
        "material": "Acrylic photo block with light-reflective gallery display finish.",
        "size": "5x7 vertical acrylic block.",
    },
}

TITLE_TEMPLATE_BANK = {
    "Sticker": [
        "{lead} {subject} 4pc 6x6 Kiss-Cut Sticker {scene}",
        "{subject} {lead} 4pc 6x6 Vinyl Sticker {audience} Gift",
        "{lead} {subject} 4pc 6x6 Sticker Sheet {emotion} Decor",
        "{subject} 4pc 6x6 Kiss-Cut Sticker {lead} {scene}",
        "{lead} {subject} 4pc 6x6 Vinyl Sticker Laptop Journal Gift",
        "{subject} {lead} 4pc 6x6 Kiss-Cut Vinyl Desk Collector Decor",
    ],
    "Poster": [
        "{lead} {subject} 12x18 Matte Poster Wall Decor",
        "{subject} {lead} 12x18 Matte Poster Study Room Art",
        "{lead} {subject} 12x18 Poster Library Print Scholar Gift",
        "{subject} 12x18 Matte Poster {lead} Gallery Decor",
        "{lead} {subject} 12x18 Wall Art Study Room Decor",
    ],
    "Acrylic": [
        "{lead} {subject} 5x7 Acrylic Photo Block Shelf Decor",
        "{subject} {lead} 5x7 Acrylic Photo Block Desk Display",
        "{lead} {subject} 5x7 Acrylic Block Collector Gift",
        "{subject} 5x7 Acrylic Photo Block {lead} Gallery Decor",
        "{lead} {subject} 5x7 Desk Art Acrylic Block Study Decor",
    ],
}

DESCRIPTION_TEMPLATE_BANK = [
    {
        "heading": "{base_title} {product_phrase}",
        "intro": "{intro}",
        "body": (
            "Designed for {use_cases}, this {product_lower} blends niche aesthetic appeal "
            "with collectible mentor-grade artwork."
        ),
        "close": "Complete the collection with matching Alchemy pieces.",
    },
    {
        "heading": "{base_title} | {product_phrase}",
        "intro": (
            "{intro} The artwork is built around a focused visual DNA profile, so each "
            "piece feels like part of a coherent small-batch collection."
        ),
        "body": (
            "Use it for {use_cases}. The composition favors readable detail, strong mood, "
            "and giftable niche appeal without generic mass-market styling."
        ),
        "close": "Pair it with related OpenClaw pieces for a coordinated collection.",
    },
    {
        "heading": "{base_title} - {product_phrase}",
        "intro": (
            "A polished {category} aesthetic piece for {audience}. {intro}"
        ),
        "body": (
            "The design works well across {use_cases}, while the title, material notes, "
            "and DNA profile stay specific for easy comparison."
        ),
        "close": "Saved under the reference SKU below for easy collector matching.",
    },
]


def _timestamp():
    return time.strftime("%-m/%-d/%Y  %-I:%M:%S %p") if os.name != "nt" else time.strftime("%#m/%#d/%Y  %#I:%M:%S %p")


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _split_keywords(value):
    words = []
    seen = set()
    for part in re.split(r"[,|;/]", _clean_text(value)):
        cleaned = re.sub(r"[^A-Za-z0-9 &-]", "", part).strip()
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        words.append(cleaned)
    return words


def _ascii_title(value):
    value = re.sub(r"[^\x00-\x7F]+", " ", _clean_text(value))
    value = re.sub(r"\s+", " ", value).strip(" -|")
    return value


def _title_tokens(value):
    return [word.lower() for word in re.findall(r"[A-Za-z0-9]+", value)]


def _title_repeats(value):
    words = _title_tokens(value)
    return {word for word in words if len(word) > 4 and words.count(word) > 1}


def _clean_subject(value, category):
    subject = _ascii_title(value) or "Art"
    subject = re.sub(r"[^A-Za-z0-9 ]+", " ", subject)
    subject = re.sub(r"\b(kiss[- ]?cut|sticker|stickers|vinyl|decal|sheet)\b", " ", subject, flags=re.I)
    if category:
        subject = re.sub(rf"\b{re.escape(category)}\b", " ", subject, flags=re.I)
    subject = re.sub(r"\s+", " ", subject).strip(" -|")
    return subject or "Art"


def _dedupe_long_words(value):
    result = []
    seen = set()
    for token in value.split():
        key = re.sub(r"[^A-Za-z0-9]+", "", token).lower()
        if len(key) > 4 and key in seen:
            continue
        if len(key) > 4:
            seen.add(key)
        result.append(token)
    return " ".join(result)


def _variant_index(metadata, modulo=6):
    seed = "|".join(
        _clean_text(metadata.get(key))
        for key in ("ID", "Title", "Category", "SEO_Hook", "Product_Type")
    )
    digest = hashlib.sha1(seed.encode("utf-8", errors="ignore")).hexdigest()
    return int(digest[:8], 16) % modulo


def _variant_pick(metadata, values):
    return values[_variant_index(metadata, len(values))]


def _variant_rotate(metadata, values):
    if not values:
        return []
    index = _variant_index(metadata, len(values))
    return values[index:] + values[:index]


def _parse_metadata(path):
    raw = path.read_text(encoding="utf-8", errors="ignore")
    data = {"Raw_Metadata": raw}
    for line in raw.splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        data[key.strip()] = value.strip()
    return data


def _ready_folders(product_type="Sticker"):
    output_dir = PRODUCT_CONFIGS[product_type]["output_dir"]
    if not output_dir.exists():
        return []
    return sorted(
        path for path in output_dir.iterdir()
        if path.is_dir()
        and path.name.startswith("MASTER_")
        and path.name.endswith("_Ready_for_Steaming")
    )


def _folder_id(folder):
    return (
        folder.name
        .replace("MASTER_", "")
        .replace("_Ready_for_Steaming", "")
        .replace("_Completed", "")
    )


def _fit_ebay_title(title, keywords, product_type="Sticker"):
    title = re.sub(r"[!]+", "", _clean_text(title))
    title = re.sub(r"\b(skin|person|text|watermark|blurry|edges)\b", " ", title, flags=re.I)
    title = re.sub(r"\s+", " ", title).strip()
    title = _dedupe_long_words(title)
    cfg = PRODUCT_CONFIGS.get(product_type, PRODUCT_CONFIGS["Sticker"])
    required = cfg["title_required"]
    if required.lower() not in title.lower():
        if product_type == "Sticker" and "Kiss-Cut" in title:
            title = title.replace("Kiss-Cut", f"{required} Kiss-Cut", 1)
        elif product_type == "Sticker" and "Sticker" in title:
            title = title.replace("Sticker", f"{required} Sticker", 1)
        elif product_type == "Poster" and "Poster" in title:
            title = title.replace("Poster", f"{required} Poster", 1)
        elif product_type == "Acrylic" and "Acrylic" in title:
            title = title.replace("Acrylic", f"{required} Acrylic", 1)
        else:
            title = f"{title} {required}"
    used_words = {word.lower() for word in re.findall(r"[A-Za-z0-9]+", title)}
    banned_title_words = {"skin", "person", "text", "watermark", "blurry", "edges"}
    extras = []
    for word in _split_keywords(keywords):
        parts = {part.lower() for part in re.findall(r"[A-Za-z0-9]+", word)}
        if parts & banned_title_words:
            continue
        if parts and (parts & used_words):
            continue
        extras.append(word.strip().title())
    if len(title) > 79:
        protected = required.split()
        words = title.split()
        result = []
        for word in words:
            candidate = " ".join(result + [word])
            if len(candidate) > 79:
                if word in protected:
                    result.append(word)
                break
            result.append(word)
        title = " ".join(result)
    if required.lower() not in title.lower():
        words = title.split()
        while len(" ".join(words + required.split())) > 79 and len(words) > 3:
            words.pop(-1)
        title = " ".join(words + required.split())
    filler = {
        "Sticker": ["Vinyl", "Laptop", "Journal", "Decor", "Gift", "Aesthetic", "Study", "Desk", "Reader", "Collector"],
        "Poster": ["Wall", "Decor", "Study", "Library", "Gift", "Aesthetic", "Gallery", "Room", "Art", "Collector"],
        "Acrylic": ["Shelf", "Decor", "Study", "Library", "Gift", "Aesthetic", "Gallery", "Block", "Art", "Collector"],
    }.get(product_type, ["Decor", "Gift", "Aesthetic", "Art", "Collector"])
    for extra in [*extras, *filler]:
        if len(title) >= 75:
            break
        parts = {part.lower() for part in re.findall(r"[A-Za-z0-9]+", extra)}
        if parts & banned_title_words:
            continue
        if parts & used_words:
            continue
        candidate = f"{title} {extra}"
        if len(candidate) <= 79 and not _title_repeats(candidate):
            title = candidate
            used_words.update(word.lower() for word in re.findall(r"[A-Za-z0-9]+", extra))
    return _repair_dangling_title(title[:79].strip(), product_type)


def _repair_dangling_title(title, product_type="Sticker"):
    title = _clean_text(title).rstrip(" ,-/")
    words = title.split()
    if not words:
        return title
    dangling = {"for", "with", "and", "or", "of", "in", "on", "by", "to", "from"}
    if words[-1].lower().strip(",") not in dangling:
        return title
    base_words = words[:-1]
    replacement_pool = {
        "Sticker": ["Gift", "Desk", "Laptop"],
        "Poster": ["Decor", "Gallery", "Study"],
        "Acrylic": ["Gift", "Shelf", "Display"],
    }.get(product_type, ["Gift", "Decor"])
    for replacement in replacement_pool:
        candidate = " ".join(base_words + [replacement])
        if len(candidate) <= 79:
            return candidate
    return " ".join(base_words).strip()


def _retitle_duplicate(title, item_id, keywords, product_type, used_titles):
    suffix_pool = {
        "Sticker": ["Journal", "Notebook", "Desk", "Reader", "Gift", "Collector", "Laptop", "Bottle", "Study", "Calm"],
        "Poster": ["Library", "Study", "Gallery", "Room", "Gift", "Collector", "Wall", "Scholar", "Decor"],
        "Acrylic": ["Shelf", "Desk", "Gallery", "Gift", "Collector", "Library", "Study", "Display", "Decor"],
    }.get(product_type, ["Gift", "Collector", "Decor"])
    seed = int(hashlib.sha1(_clean_text(item_id).encode("utf-8")).hexdigest()[:8], 16)
    rotated = suffix_pool[seed % len(suffix_pool):] + suffix_pool[:seed % len(suffix_pool)]
    for suffix in rotated:
        words = title.split()
        while len(" ".join(words + [suffix])) > 79 and len(words) > 5:
            words.pop(-1)
        candidate = _fit_ebay_title(" ".join(words + [suffix]), keywords, product_type)
        if 75 <= len(candidate) <= 79 and candidate not in used_titles:
            return candidate
    return title


def _keyword_pick(metadata, limit=3):
    title_words = _split_keywords(metadata.get("Title"))
    seo_words = _split_keywords(metadata.get("SEO_Hook"))
    banned = {
        "sticker",
        "stickers",
        "kiss cut",
        "kiss-cut",
        "vinyl",
        "decor",
        "collectible sticker",
        "mentor-grade sticker",
    }
    picks = []
    seen = set()
    for word in [*title_words, *seo_words]:
        normalized = word.lower()
        if normalized in banned or normalized in seen:
            continue
        if len(word) > 28:
            continue
        seen.add(normalized)
        picks.append(word.title())
        if len(picks) >= limit:
            break
    return picks


def _niche_profile(metadata):
    category = _clean_text(metadata.get("Category")).lower()
    seo = _clean_text(metadata.get("SEO_Hook")).lower()
    title = _clean_text(metadata.get("Title")).lower()
    if "academia" in category or "academia" in seo or "academia" in title:
        variants = [
            {"lead": "Dark Academia", "scene": "Laptop Study Journal Decor", "audience": "Book Lover Student", "emotion": "Cozy Vintage Intellectual", "style": "Academia Mentor-Grade"},
            {"lead": "Gothic Academia", "scene": "Study Desk Journal Decor", "audience": "Reader Writer Student", "emotion": "Moody Scholarly Vintage", "style": "Academia Mentor-Grade"},
            {"lead": "Vintage Academia", "scene": "Library Laptop Notebook Decor", "audience": "Book Lover Introvert", "emotion": "Literary Cozy Study", "style": "Academia Mentor-Grade"},
        ]
        return variants[_variant_index(metadata, len(variants))]
    variants = [
        {"lead": "Zen Aesthetic", "scene": "Laptop Journal Water Bottle Decor", "audience": "Mindfulness Minimalist", "emotion": "Calm Balance Peaceful", "style": "Zen Mentor-Grade"},
        {"lead": "Mindful Zen", "scene": "Journal Laptop Meditation Decor", "audience": "Yoga Minimalist Gift", "emotion": "Peaceful Calm Balance", "style": "Zen Mentor-Grade"},
        {"lead": "Minimal Zen", "scene": "Water Bottle Journal Desk Decor", "audience": "Calm Lifestyle Gift", "emotion": "Serene Mindful Clean", "style": "Zen Mentor-Grade"},
    ]
    return variants[_variant_index(metadata, len(variants))]


def _build_local_title(metadata):
    product_type = metadata.get("Product_Type", "Sticker")
    profile = _niche_profile(metadata)
    subject = _clean_subject(metadata.get("Title") or metadata.get("ID"), metadata.get("Category"))
    if len(subject) > 34:
        subject = " ".join(subject.split()[:4])
    values = {
        "lead": profile["lead"],
        "subject": subject,
        "scene": profile["scene"],
        "audience": profile["audience"],
        "emotion": profile["emotion"],
    }
    if product_type in {"Poster", "Acrylic"}:
        product_words = {
            "Poster": ["Wall Decor", "Study Room Art", "Library Print", "Gallery Decor", "Scholar Gift"],
            "Acrylic": ["Shelf Decor", "Desk Display", "Gallery Block", "Collector Gift", "Study Decor"],
        }[product_type]
        values["scene"] = _variant_pick(metadata, product_words)
        candidates = [template.format(**values) for template in _variant_rotate(metadata, TITLE_TEMPLATE_BANK[product_type])]
        best = candidates[0]
        for candidate in candidates:
            fitted = _fit_ebay_title(candidate, metadata.get("SEO_Hook"), product_type)
            if 75 <= len(fitted) <= 79:
                return fitted
            if abs(77 - len(fitted)) < abs(77 - len(best)):
                best = fitted
        return _fit_ebay_title(best, metadata.get("SEO_Hook"), product_type)
    tails = [
        profile["scene"],
        f"{profile['audience']} Gift",
        f"{profile['emotion']} Decor",
        "Laptop Journal Desk Decor",
        "Water Bottle Notebook Gift",
        "Study Desk Aesthetic Decor",
    ]
    values["scene"] = _variant_pick(metadata, tails)
    candidates = [template.format(**values) for template in _variant_rotate(metadata, TITLE_TEMPLATE_BANK["Sticker"])]
    best = candidates[0]
    for candidate in candidates:
        fitted = _fit_ebay_title(candidate, metadata.get("SEO_Hook"), product_type)
        if 75 <= len(fitted) <= 79:
            return fitted
        if abs(77 - len(fitted)) < abs(77 - len(best)):
            best = fitted
    return _fit_ebay_title(best, metadata.get("SEO_Hook"), product_type)


def _short_dna(metadata):
    prompt = _clean_text(metadata.get("MJ_Prompt"))
    prompt = re.sub(r"--\S+(?:\s+\S+)?", " ", prompt)
    prompt = re.sub(r"\b(white contour border|vector clean edges|die-cut sticker style|solid white background|isolated on white background)\b", " ", prompt, flags=re.I)
    prompt = _clean_text(prompt)
    if len(prompt) <= 360:
        return prompt
    return prompt[:357].rsplit(" ", 1)[0] + "..."


def _build_local_description(metadata):
    profile = _niche_profile(metadata)
    product_type = metadata.get("Product_Type", "Sticker")
    cfg = PRODUCT_CONFIGS.get(product_type, PRODUCT_CONFIGS["Sticker"])
    item_id = _clean_text(metadata.get("ID"))
    base_title = _ascii_title(metadata.get("Title")) or item_id
    seo_keywords = _split_keywords(metadata.get("SEO_Hook"))
    keyword_text = ", ".join(seo_keywords[:10])
    dna = _short_dna(metadata)
    category = _clean_text(metadata.get("Category")) or profile["lead"].replace(" Aesthetic", "")
    style = _clean_text(metadata.get("Style")) or profile["style"]
    if category.lower() == "zen":
        intros = [
            f"Bring calm and balance into your daily routine with this {base_title} zen aesthetic {cfg['product_phrase'].lower()}.",
            f"Add a quiet mindful accent to your workspace with this {base_title} {cfg['product_phrase'].lower()}.",
            f"Designed for peaceful desks, journals, and small rituals, this {base_title} {cfg['product_phrase'].lower()} carries a clean Zen mood.",
        ]
        intro = _variant_pick(metadata, intros)
        audiences = [
            "mindfulness lovers, minimalists, journal keepers, yoga enthusiasts, and peaceful room setups",
            "meditation fans, calm desk setups, notebook collectors, and gift buyers who like clean aesthetics",
            "students, remote workers, yoga lovers, and anyone building a serene everyday space",
        ]
        audience = _variant_pick(metadata, audiences)
    else:
        intros = [
            f"Embrace the dark academia aesthetic with this vintage-inspired {base_title} {cfg['product_phrase'].lower()}.",
            f"Give your study space a scholarly, moody accent with this {base_title} {cfg['product_phrase'].lower()}.",
            f"Built for readers and collectors, this {base_title} {cfg['product_phrase'].lower()} blends literary atmosphere with vintage study-room style.",
        ]
        intro = _variant_pick(metadata, intros)
        audiences = [
            "students, book lovers, writers, introverts, and dark academia collectors",
            "readers, literature fans, journal keepers, study desk decorators, and thoughtful gift buyers",
            "writers, learners, library lovers, and collectors of moody scholarly decor",
        ]
        audience = _variant_pick(metadata, audiences)
    use_case_variants = [
        "study rooms, creative workspaces, shelves, gallery walls, and collectible aesthetic decor",
        "laptops, notebooks, reading corners, desk setups, gallery shelves, and gift bundles",
        "journal spreads, library shelves, dorm rooms, studio desks, and cozy personal collections",
    ]
    use_cases = _variant_pick(metadata, use_case_variants)
    template = _variant_pick(metadata, DESCRIPTION_TEMPLATE_BANK)
    heading = template["heading"].format(base_title=base_title, product_phrase=cfg["product_phrase"])
    intro_text = template["intro"].format(
        intro=intro,
        category=category,
        audience=audience,
        product_phrase=cfg["product_phrase"],
        product_lower=product_type.lower(),
    )
    body_text = template["body"].format(
        use_cases=use_cases,
        product_lower=product_type.lower(),
        category=category,
        audience=audience,
    )
    close_text = template["close"].format(category=category, product_lower=product_type.lower())
    return (
        f"<h2>{heading}</h2>"
        f"<p>{intro_text}</p>"
        f"<p>{body_text}</p>"
        f"<ul>"
        f"<li><strong>Includes:</strong> {cfg['includes']}</li>"
        f"<li><strong>Material:</strong> {cfg['material']}</li>"
        f"<li><strong>Size:</strong> {cfg['size']}</li>"
        f"<li><strong>Style:</strong> {style}; {category} aesthetic.</li>"
        f"<li><strong>DNA Profile:</strong> {dna}</li>"
        f"<li><strong>Best For:</strong> {audience}.</li>"
        f"</ul>"
        f"<p><strong>SEO Keywords:</strong> {keyword_text}</p>"
        f"<p><strong>Image Note:</strong> The main image shows the actual product customers receive. Additional images are bonus concept/detail reference images and do not represent extra products or selectable variations.</p>"
        f"<p>{close_text}</p>"
        f"<p><small>Reference SKU: {item_id}</small></p>"
    )


def _ensure_image_note(description):
    description = _clean_text(description)
    note_pattern = re.compile(r"<p><strong>Image Note:</strong>.*?</p>", re.I | re.S)
    note = (
        "<p><strong>Image Note:</strong> The main image shows the actual product customers receive. "
        "Additional images are bonus concept/detail reference images and do not represent extra products "
        "or selectable variations.</p>"
    )
    if note_pattern.search(description):
        return note_pattern.sub(note, description, count=1)
    if "main image shows the actual product customers receive" in description.lower():
        description = re.sub(
            r"The main image shows the actual product customers receive[^<]*(?:</p>)?",
            "",
            description,
            flags=re.I,
        )
    if "</ul>" in description:
        return description.replace("</ul>", f"</ul>{note}", 1)
    return f"{description}{note}"


def _fallback_listing(metadata):
    title = _build_local_title(metadata)
    dna = _short_dna(metadata)
    description = _ensure_image_note(_build_local_description(metadata))
    return {"Title": title, "Description": description, "DNA Profile": dna}


def _deepseek_listing(metadata):
    api_key = Config.DEEPSEEK_API_KEY
    if not api_key:
        raise RuntimeError("DEEPSEEK_API_KEY is missing")
    base_url = (Config.DEEPSEEK_BASE_URL or "https://api.deepseek.com").rstrip("/")
    prompt = {
        "ID": metadata.get("ID"),
        "Title": metadata.get("Title"),
        "SEO_Hook": metadata.get("SEO_Hook"),
        "Style": metadata.get("Style"),
        "MJ_Prompt": metadata.get("MJ_Prompt"),
        "Product_Type": metadata.get("Product_Type", "Sticker"),
    }
    payload = {
        "model": os.getenv("DEEPSEEK_MODEL", "deepseek-chat"),
        "messages": [
            {
                "role": "system",
                "content": (
                    "Output strict JSON only. Keys: Title, Description, DNA Profile. "
                    "Create high-conversion eBay SEO for the Product_Type in the metadata. "
                    "Title must be 75-79 ASCII characters, no exclamation marks, no filler. "
                    "For Sticker, title must clearly include 4pc and 6x6. "
                    "For Poster, title must clearly include 12x18. For Acrylic, title must clearly include 5x7. "
                    "Use one of these title template families, with natural substitutions: "
                    "1) aesthetic lead + subject + required size + product noun + use case; "
                    "2) subject + aesthetic lead + required size + product noun + audience gift; "
                    "3) aesthetic lead + subject + required size + product noun + room/decor placement. "
                    "Use one of these description structures: "
                    "A) concise aesthetic intro, practical use paragraph, factual bullets, image note; "
                    "B) collector-focused intro, visual DNA paragraph, factual bullets, image note; "
                    "C) gift/use-case intro, mood paragraph, factual bullets, image note. "
                    "Use the item's metadata as the source of truth. "
                    "For Zen, emphasize calm, balance, mindfulness, minimalist, laptop, journal, water bottle. "
                    "For Academia, emphasize dark academia, study, vintage, intellectual, book lover, student, journal, study desk. "
                    "Description must be eBay-ready HTML and include Includes, Material, Size, Style, DNA Profile, and use cases. "
                    "Description must include an Image Note saying the main image shows the actual product customers receive, while additional images are bonus concept/detail reference images and do not represent extra products or selectable variations. "
                    "Vary sentence structure and keyword order across items so listings do not look mass-generated. "
                    "Use tasteful synonyms while preserving the same product facts. "
                    "Do not invent product materials beyond the requested Printify product type."
                ),
            },
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        "temperature": 0.45,
    }
    response = requests.post(
        f"{base_url}/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload,
        timeout=90,
    )
    response.raise_for_status()
    content = response.json()["choices"][0]["message"]["content"].strip()
    content = content.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    data = json.loads(content)
    data["Title"] = _fit_ebay_title(data.get("Title"), metadata.get("SEO_Hook"), metadata.get("Product_Type", "Sticker"))
    data["Description"] = _clean_text(data.get("Description"))
    if "<" not in data["Description"]:
        data["Description"] = _build_local_description({**metadata, "MJ_Prompt": data.get("DNA Profile") or metadata.get("MJ_Prompt")})
    data["Description"] = _ensure_image_note(data["Description"])
    return data


def _gallery_paths(folder, item_id, product_type="Sticker"):
    paths = {}
    for index in range(1, 5):
        if product_type != "Sticker":
            candidates = [
                folder / f"Gallery_U{index}.png",
                folder / f"{item_id}_Gallery_U{index}.png",
                folder / f"{item_id}_U{index}.png",
                folder / f"{item_id}_U{index}_Grid.png",
            ]
        else:
            candidates = [
                folder / f"{item_id}_U{index}_Grid.png",
                folder / f"{item_id}_U{index}.png",
                folder / f"Grid{index}.png",
            ]
        found = next((path for path in candidates if path.exists()), None)
        paths[f"Gallery_U{index}_Path"] = str(found.resolve()) if found else ""
    return paths


def _open_book(path, headers):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        current = [cell.value for cell in ws[1]]
        if current != headers:
            old_rows = []
            current_map = {header: idx + 1 for idx, header in enumerate(current) if header}
            for row in range(2, ws.max_row + 1):
                old_rows.append({header: ws.cell(row=row, column=col).value for header, col in current_map.items()})
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)
            for old in old_rows:
                if not old.get("SKU") and old.get("ID"):
                    old["SKU"] = old.get("ID")
                ws.append([old.get(header, "") for header in headers])
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    return wb, ws


def _upsert(ws, headers, row_data):
    id_col = headers.index("ID") + 1
    target = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == row_data["ID"]:
            target = row
            break
    if target is None:
        target = ws.max_row + 1
    else:
        existing = {header: ws.cell(row=target, column=col).value for col, header in enumerate(headers, 1)}
        if existing.get("Status") and existing.get("Status") != "Ready_for_Printify":
            row_data["Status"] = existing.get("Status")
        if existing.get("Printify_Product_ID"):
            row_data["Printify_Product_ID"] = existing.get("Printify_Product_ID")
    for col, header in enumerate(headers, 1):
        ws.cell(row=target, column=col).value = row_data.get(header, "")


def _existing_ids(ws, headers):
    id_col = headers.index("ID") + 1
    return {
        str(ws.cell(row=row, column=id_col).value or "").strip()
        for row in range(2, ws.max_row + 1)
        if str(ws.cell(row=row, column=id_col).value or "").strip()
    }


def build_listing_assets(limit=0, use_api=True, product_type="Sticker", only_missing=False):
    DATABASE_DIR.mkdir(exist_ok=True)
    ebay_wb, ebay_ws = _open_book(EBAY_BOOK, EBAY_HEADERS)
    etsy_wb, etsy_ws = _open_book(ETSY_BOOK, ETSY_HEADERS)
    product_type = "Acrylic" if product_type.lower().startswith("acry") else ("Poster" if product_type.lower().startswith("poster") else "Sticker")
    folders = _ready_folders(product_type)
    if only_missing:
        known = _existing_ids(ebay_ws, EBAY_HEADERS)
        folders = [folder for folder in folders if _folder_id(folder) not in known]
    if limit:
        folders = folders[:limit]
    completed = 0
    for folder in folders:
        metadata_path = folder / "metadata.txt"
        production_path = folder / "Production_Design.png"
        cover_path = folder / "Cover_Mockup.png"
        if not metadata_path.exists() or not production_path.exists() or not cover_path.exists():
            print(f"[SKIP] Missing required assets: {folder.name}")
            continue
        metadata = _parse_metadata(metadata_path)
        metadata["ID"] = metadata.get("ID") or _folder_id(folder)
        metadata["Product_Type"] = product_type
        try:
            listing = _deepseek_listing(metadata) if use_api else _fallback_listing(metadata)
        except Exception as exc:
            print(f"[WARN] DeepSeek fallback for {metadata['ID']}: {exc}")
            listing = _fallback_listing(metadata)
        row = {
            "ID": metadata["ID"],
            "SKU": metadata["ID"],
            "Category": metadata.get("Category", ""),
            "Product_Type": product_type,
            "Title": listing.get("Title", ""),
            "Description": listing.get("Description", ""),
            "Price": PRODUCT_CONFIGS[product_type]["price"],
            "DNA Profile": listing.get("DNA Profile") or metadata.get("MJ_Prompt", ""),
            "Production_Path": str(production_path.resolve()),
            "Cover_Path": str(cover_path.resolve()),
            **_gallery_paths(folder, metadata["ID"], product_type),
            "Status": "Ready_for_Printify",
            "Timestamp": _timestamp(),
        }
        _upsert(ebay_ws, EBAY_HEADERS, row)
        _upsert(
            etsy_ws,
            ETSY_HEADERS,
            {
                "ID": metadata["ID"],
                "Raw_Metadata": metadata.get("Raw_Metadata", ""),
                "Production_Path": str(production_path.resolve()),
                "Status": "Placeholder",
                "Timestamp": _timestamp(),
            },
        )
        completed += 1
        print(f"[LISTING] {metadata['ID']} -> eBay/Etsy rows ready")
    ebay_wb.save(EBAY_BOOK)
    etsy_wb.save(ETSY_BOOK)
    ebay_wb.close()
    etsy_wb.close()
    print(f"[DONE] Listing assets updated: {completed}")


def normalize_existing_listing_rows():
    if not EBAY_BOOK.exists():
        print("[NORMALIZE] eBay listing workbook not found")
        return
    wb, ws = _open_book(EBAY_BOOK, EBAY_HEADERS)
    headers = {header: index + 1 for index, header in enumerate(EBAY_HEADERS)}
    changed = 0
    used_titles = set()
    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=headers["ID"]).value
        if not item_id:
            continue
        product_type = ws.cell(row=row, column=headers["Product_Type"]).value or "Sticker"
        if product_type not in PRODUCT_CONFIGS:
            product_type = "Sticker"
            ws.cell(row=row, column=headers["Product_Type"]).value = product_type
        title_cell = ws.cell(row=row, column=headers["Title"])
        desc_cell = ws.cell(row=row, column=headers["Description"])
        seo = ws.cell(row=row, column=headers["DNA Profile"]).value or ""
        new_title = _fit_ebay_title(title_cell.value, seo, product_type)
        if new_title in used_titles:
            new_title = _retitle_duplicate(new_title, item_id, seo, product_type, used_titles)
        used_titles.add(new_title)
        new_desc = _ensure_image_note(desc_cell.value or "")
        if title_cell.value != new_title:
            title_cell.value = new_title
            changed += 1
        if desc_cell.value != new_desc:
            desc_cell.value = new_desc
            changed += 1
    wb.save(EBAY_BOOK)
    wb.close()
    print(f"[NORMALIZE] Existing listing rows updated: {changed}")


def run_logic():
    build_listing_assets()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--no-api", action="store_true")
    parser.add_argument("--product-type", default="Sticker", choices=["Sticker", "Poster", "Acrylic"])
    parser.add_argument("--normalize-existing", action="store_true")
    parser.add_argument("--only-missing", action="store_true")
    args = parser.parse_args()
    if args.normalize_existing:
        normalize_existing_listing_rows()
    else:
        build_listing_assets(
            limit=args.limit,
            use_api=not args.no_api,
            product_type=args.product_type,
            only_missing=args.only_missing,
        )
