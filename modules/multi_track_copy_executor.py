from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.ebay_quiet_jade_pivot import (
    IMAGE_NOTE,
    api_headers,
    clean,
    fit_title,
    money,
    money_text,
    request_with_retry,
    subject_from_title,
)


DATABASE = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE / "eBay_listing.xlsx"
MULTI_PLAN = DATABASE / "Multi_Track_Experiment_Plan.csv"
COPY_BATCH = DATABASE / "Multi_Track_Copy_Batch.csv"
ROLLBACK = DATABASE / "Multi_Track_Copy_Rollback.csv"
SYNC_LOG = DATABASE / "Multi_Track_Copy_Sync_Log.csv"
STATE_JSON = DATABASE / "Multi_Track_Copy_State.json"
NY = ZoneInfo("America/New_York")

PUBLISH_TITLE_DESC_PRICE = {
    "title": True,
    "description": True,
    "images": False,
    "variants": True,
    "tags": False,
    "keyFeatures": False,
    "shipping_template": False,
}


@dataclass
class CopyRow:
    row_idx: int
    local_id: str
    product_type: str
    category: str
    printify_product_id: str
    ebay_item_id: str
    track: str
    primary_intent: str
    secondary_keywords: str
    mockup_mood: str
    old_title: str
    old_description: str
    old_price: str
    new_title: str
    new_description: str
    new_price: str


def now_text() -> str:
    return datetime.now(NY).strftime("%Y-%m-%d %H:%M:%S %z")


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    path.parent.mkdir(exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _ensure_column(ws, cols: dict[str, int], name: str) -> int:
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def _target_price(price_target: str, product_type: str, track: str) -> str:
    text = clean(price_target)
    if "$" in text and "-" not in text:
        value = money(text)
        if value:
            return money_text(value)
    if product_type == "Acrylic":
        return money_text(89.99 if track == "A_LOW_COMPETITION_NICHE" else 84.99)
    if product_type == "Poster":
        return money_text(34.99 if track == "A_LOW_COMPETITION_NICHE" else 29.99)
    if product_type == "Sticker":
        return money_text(11.99)
    return text or "$0.00"


def _lead(product_type: str, primary: str, category: str) -> str:
    primary = clean(primary).title()
    if product_type == "Acrylic":
        if "Meditation" in primary or category == "Zen":
            return "Meditation Room Jade Relic"
        if "Shelf" in primary or "Collector" in primary:
            return "Collector Shelf Smoky Jade"
        if "Study" in primary:
            return "Dark Study Smoky Jade"
        return "Quiet Luxury Jade Relic"
    if product_type == "Poster":
        if category == "Zen":
            if "Reading" in primary or "Book" in primary:
                return "Reading Nook Quiet Luxury Art"
            if "Meditation" in primary:
                return "Meditation Room Smoky Jade Wall Art"
            return "Quiet Luxury Jade Wall Art"
        if "Reading" in primary or "Book" in primary:
            return "Reading Nook Smoky Jade Wall Art"
        if "Dorm" in primary:
            return "Dorm Study Smoky Jade Poster"
        return "Dark Academia Reading Nook Art"
    if product_type == "Sticker":
        if "Book" in primary or "Reading" in primary:
            return "Book Nook Jade Sticker Set"
        return "Deep Work Jade Sticker Set"
    return primary


def _subject(old_title: str, product_type: str, category: str, lead: str) -> str:
    subject = subject_from_title(old_title, product_type, category)
    stop = {
        "for",
        "with",
        "and",
        "the",
        "a",
        "an",
        "decor",
        "print",
        "photo",
        "block",
        "poster",
        "matte",
        "acrylic",
        "shelf",
        "desk",
        "apartment",
        "room",
        "gift",
    }
    lead_words = {word.lower() for word in clean(lead).split()}
    words = []
    for word in subject.split():
        token = word.strip()
        low = token.lower()
        if low in stop:
            continue
        # Avoid clumsy echoes such as "Jade Relic Jade Phoenix".
        if low in lead_words:
            continue
        words.append(token)
    compact = []
    for token in words:
        if compact and compact[-1].lower() == token.lower():
            continue
        compact.append(token)
    return clean(" ".join(compact[:4])) or "Premium Relic"


def _title(product_type: str, category: str, old_title: str, primary: str) -> str:
    lead = _lead(product_type, primary, category)
    subject = _subject(old_title, product_type, category, lead)
    if product_type == "Acrylic":
        return fit_title([lead, subject, "5x7 Acrylic Block Shelf Decor"])
    if product_type == "Poster":
        return fit_title([lead, subject, "12x18 Matte Poster Apartment Decor"])
    if product_type == "Sticker":
        return fit_title([lead, subject, "4pc 6x6 Vinyl Laptop Journal Decals"])
    return fit_title([lead, subject, "Decor"])


def _description(row: dict[str, str], product_type: str, category: str, title: str, old_title: str) -> str:
    primary = clean(row.get("Primary_Search_Intent"))
    secondary = clean(row.get("Secondary_Keywords"))
    mood = clean(row.get("Mockup_Mood")).replace("_", " ")
    local_id = clean(row.get("ID"))
    if product_type == "Acrylic":
        includes = "One 5x7 vertical acrylic photo block, produced on demand through Printify."
        product_note = "This is a single printed acrylic display block. The main artwork is the product customers receive; any additional gallery images are concept/detail references."
        buyer = "quiet shelves, meditation corners, gothic study desks, collector rooms, and deep-work apartments"
        tactile = "refractive depth, smoky jade color, internal glow, and premium desk-object presence"
    elif product_type == "Poster":
        includes = "One 12x18 matte poster, produced on demand through Printify."
        product_note = "This is a single printed poster. The main artwork is the product customers receive; any additional gallery images are concept/detail references."
        buyer = "reading nooks, home libraries, dorm study walls, meditation rooms, and quiet apartment corners"
        tactile = "matte paper mood, cinematic depth, smoky jade tones, and calm visual focus"
    else:
        includes = "One physical 4pc 6x6 kiss-cut vinyl sticker sheet, produced on demand through Printify."
        product_note = IMAGE_NOTE
        buyer = "laptops, journals, water bottles, planners, study desks, and bookish gift sets"
        tactile = "sharp kiss-cut presentation, collectible motifs, and a cohesive quiet-jade desk mood"
    return (
        f"<h2>{title}</h2>"
        f"<p>This listing is part of a focused search-intent experiment for {primary}. "
        f"It is written for buyers decorating {buyer}, not for generic wall-art browsing.</p>"
        "<ul>"
        f"<li><strong>Includes:</strong> {includes}</li>"
        f"<li><strong>Visual Mood:</strong> {mood.title() or 'Quiet luxury study mood'}.</li>"
        f"<li><strong>Design Language:</strong> {tactile}.</li>"
        f"<li><strong>Search Lane:</strong> {secondary}.</li>"
        "</ul>"
        f"<p><strong>Image Note:</strong> {product_note}</p>"
        f"<p><small>Reference SKU: {local_id}. Prior title: {clean(old_title)[:90]}</small></p>"
    )


def _candidate_rows(track: str, limit: int, force: bool = False) -> list[dict[str, str]]:
    rows = []
    already = set() if force else _already_prepared()
    for row in _read_csv(MULTI_PLAN):
        if row.get("Track") != track:
            continue
        if row.get("Launch_Action") != "COPY_OR_PRICE_EXPERIMENT_ON_EXISTING_LISTING":
            continue
        if row.get("QA_Status") not in {"READY", "REVIEW"}:
            continue
        if clean(row.get("ID")) in already:
            continue
        rows.append(row)
        if limit and len(rows) >= limit:
            break
    return rows


def _already_prepared() -> set[str]:
    done = set()
    for row in _read_csv(COPY_BATCH):
        if clean(row.get("Apply_Status")) in {"LOCAL_APPLIED", "SYNCED_PRINTIFY"}:
            done.add(clean(row.get("ID")))
    return done


def _load_workbook_rows(plan_rows: list[dict[str, str]]) -> tuple[Any, Any, dict[str, int], list[CopyRow]]:
    wanted = {clean(row.get("ID")): row for row in plan_rows}
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers) if header}
    selected: list[CopyRow] = []
    for row_idx in range(2, ws.max_row + 1):
        local_id = clean(ws.cell(row_idx, cols["ID"]).value)
        plan = wanted.get(local_id)
        if not plan:
            continue
        product_type = clean(ws.cell(row_idx, cols["Product_Type"]).value)
        category = clean(ws.cell(row_idx, cols["Category"]).value)
        old_title = clean(ws.cell(row_idx, cols["Title"]).value)
        old_description = clean(ws.cell(row_idx, cols["Description"]).value)
        old_price = clean(ws.cell(row_idx, cols["Price"]).value)
        new_title = _title(product_type, category, old_title, clean(plan.get("Primary_Search_Intent")))
        if not (75 <= len(new_title) <= 79):
            wb.close()
            raise RuntimeError(f"title length outside 75-79 for {local_id}: {len(new_title)} {new_title}")
        selected.append(
            CopyRow(
                row_idx=row_idx,
                local_id=local_id,
                product_type=product_type,
                category=category,
                printify_product_id=clean(ws.cell(row_idx, cols["Printify_Product_ID"]).value),
                ebay_item_id=clean(ws.cell(row_idx, cols["eBay_Item_ID"]).value) if "eBay_Item_ID" in cols else "",
                track=clean(plan.get("Track")),
                primary_intent=clean(plan.get("Primary_Search_Intent")),
                secondary_keywords=clean(plan.get("Secondary_Keywords")),
                mockup_mood=clean(plan.get("Mockup_Mood")),
                old_title=old_title,
                old_description=old_description,
                old_price=old_price,
                new_title=new_title,
                new_description=_description(plan, product_type, category, new_title, old_title),
                new_price=_target_price(clean(plan.get("Price_Target")), product_type, clean(plan.get("Track"))),
            )
        )
    return wb, ws, cols, selected


def prepare(track: str, limit: int, apply_local: bool, force: bool = False) -> list[CopyRow]:
    plan_rows = _candidate_rows(track, limit, force=force)
    wb, ws, cols, selected = _load_workbook_rows(plan_rows)
    timestamp = now_text()
    backup = ""
    if apply_local and selected:
        backup_path = EBAY_BOOK.with_name(f"eBay_listing.backup_multi_track_{datetime.now(NY):%Y%m%d_%H%M%S}.xlsx")
        shutil.copy2(EBAY_BOOK, backup_path)
        backup = str(backup_path)
        ts_col = _ensure_column(ws, cols, "Multi_Track_Timestamp")
        track_col = _ensure_column(ws, cols, "Multi_Track_Track")
        intent_col = _ensure_column(ws, cols, "Multi_Track_Primary_Intent")
        mood_col = _ensure_column(ws, cols, "Multi_Track_Mockup_Mood")
        sync_col = _ensure_column(ws, cols, "Metadata_Sync_Status")
        for row in selected:
            ws.cell(row.row_idx, cols["Title"]).value = row.new_title
            ws.cell(row.row_idx, cols["Description"]).value = row.new_description
            ws.cell(row.row_idx, cols["Price"]).value = row.new_price
            ws.cell(row.row_idx, ts_col).value = timestamp
            ws.cell(row.row_idx, track_col).value = row.track
            ws.cell(row.row_idx, intent_col).value = row.primary_intent
            ws.cell(row.row_idx, mood_col).value = row.mockup_mood
            ws.cell(row.row_idx, sync_col).value = "MULTI_TRACK_PENDING_PRINTIFY_SYNC"
        wb.save(EBAY_BOOK)
    wb.close()

    batch_rows = []
    rollback_rows = []
    for row in selected:
        batch_rows.append(
            {
                "Timestamp": timestamp,
                "ID": row.local_id,
                "Product_Type": row.product_type,
                "Track": row.track,
                "Primary_Intent": row.primary_intent,
                "Mockup_Mood": row.mockup_mood,
                "Printify_Product_ID": row.printify_product_id,
                "eBay_Item_ID": row.ebay_item_id,
                "Old_Title_Length": len(row.old_title),
                "New_Title_Length": len(row.new_title),
                "Old_Price": row.old_price,
                "New_Price": row.new_price,
                "Old_Title": row.old_title,
                "New_Title": row.new_title,
                "Apply_Status": "LOCAL_APPLIED" if apply_local else "DRY_PLAN",
            }
        )
        rollback_rows.append(
            {
                "Timestamp": timestamp,
                "ID": row.local_id,
                "Printify_Product_ID": row.printify_product_id,
                "Old_Title": row.old_title,
                "Old_Description": row.old_description,
                "Old_Price": row.old_price,
                "New_Title": row.new_title,
                "New_Description": row.new_description,
                "New_Price": row.new_price,
            }
        )
    if selected:
        _write_csv(COPY_BATCH, batch_rows, list(batch_rows[0].keys()))
        _write_csv(ROLLBACK, rollback_rows, list(rollback_rows[0].keys()))
    STATE_JSON.write_text(
        json.dumps(
            {
                "timestamp": timestamp,
                "track": track,
                "selected": len(selected),
                "apply_local": apply_local,
                "backup": backup,
                "ids": [row.local_id for row in selected],
                "title_lengths": {row.local_id: len(row.new_title) for row in selected},
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    print(f"[MULTI-COPY] selected={len(selected)} apply_local={apply_local} track={track}")
    return selected


def _load_prepared_for_sync(limit: int) -> list[dict[str, str]]:
    rows = [row for row in _read_csv(COPY_BATCH) if clean(row.get("Apply_Status")) == "LOCAL_APPLIED"]
    return rows[:limit] if limit else rows


def _local_metadata(ids: set[str]) -> dict[str, dict[str, str]]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx for idx, header in enumerate(headers) if header}
    out: dict[str, dict[str, str]] = {}
    for data in ws.iter_rows(min_row=2, values_only=True):
        local_id = clean(data[cols["ID"]])
        if local_id not in ids:
            continue
        out[local_id] = {
            "Title": clean(data[cols["Title"]]),
            "Description": clean(data[cols["Description"]]),
            "Price": clean(data[cols["Price"]]),
            "Product_Type": clean(data[cols["Product_Type"]]),
            "Printify_Product_ID": clean(data[cols["Printify_Product_ID"]]),
            "eBay_Item_ID": clean(data[cols["eBay_Item_ID"]]) if "eBay_Item_ID" in cols else "",
        }
    wb.close()
    return out


def _append_sync(row: dict[str, Any]) -> None:
    headers = [
        "Timestamp",
        "ID",
        "Product_Type",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "HTTP_Get",
        "HTTP_Update",
        "HTTP_Publish",
        "Result",
        "Error",
    ]
    exists = SYNC_LOG.exists()
    with SYNC_LOG.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        if not exists:
            writer.writeheader()
        writer.writerow(row)


def _mark_synced(done: set[str], failed: set[str]) -> None:
    if not done and not failed:
        return
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers) if header}
    sync_col = _ensure_column(ws, cols, "Metadata_Sync_Status")
    for row_idx in range(2, ws.max_row + 1):
        local_id = clean(ws.cell(row_idx, cols["ID"]).value)
        if local_id in done:
            ws.cell(row_idx, sync_col).value = "MULTI_TRACK_SYNCED_PRINTIFY_PUBLISH"
        elif local_id in failed:
            ws.cell(row_idx, sync_col).value = "MULTI_TRACK_SYNC_CHECK_REQUIRED"
    wb.save(EBAY_BOOK)
    wb.close()
    _mark_batch_synced(done, failed)


def _mark_batch_synced(done: set[str], failed: set[str]) -> None:
    if not COPY_BATCH.exists():
        return
    rows = _read_csv(COPY_BATCH)
    if not rows:
        return
    timestamp = now_text()
    headers = list(rows[0].keys())
    for extra in ["Sync_Timestamp", "Sync_Result"]:
        if extra not in headers:
            headers.append(extra)
    for row in rows:
        local_id = clean(row.get("ID"))
        if local_id in done:
            row["Apply_Status"] = "SYNCED_PRINTIFY"
            row["Sync_Timestamp"] = timestamp
            row["Sync_Result"] = "OK"
        elif local_id in failed:
            row["Apply_Status"] = "SYNC_CHECK_REQUIRED"
            row["Sync_Timestamp"] = timestamp
            row["Sync_Result"] = "FAILED"
    _write_csv(COPY_BATCH, rows, headers)


def sync_printify(limit: int, dry_run: bool, sleep_min: float, sleep_max: float) -> int:
    if not Config.Printify_API_KEY:
        raise RuntimeError("Missing Printify_API_KEY")
    prepared = _load_prepared_for_sync(limit)
    meta = _local_metadata({clean(row.get("ID")) for row in prepared})
    base = Config.Printify_API_URL.rstrip("/")
    done: set[str] = set()
    failed: set[str] = set()
    for row in prepared:
        local_id = clean(row.get("ID"))
        item = meta.get(local_id) or {}
        product_id = item.get("Printify_Product_ID") or clean(row.get("Printify_Product_ID"))
        if not product_id:
            failed.add(local_id)
            _append_sync({"Timestamp": now_text(), "ID": local_id, "Result": "SKIP", "Error": "missing Printify product id"})
            continue
        if dry_run:
            print(f"[MULTI-COPY-DRY] {local_id} product={product_id} title={item.get('Title','')[:75]}")
            continue
        get_status = update_status = publish_status = ""
        result = "CHECK"
        error = ""
        try:
            get_resp = request_with_retry("GET", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json")
            get_status = get_resp.status_code
            get_resp.raise_for_status()
            product = get_resp.json()
            price_cents = int(round((money(item.get("Price")) or 0) * 100))
            variants = [
                {
                    "id": variant["id"],
                    "price": price_cents if variant.get("is_enabled") else variant.get("price"),
                    "is_enabled": bool(variant.get("is_enabled")),
                }
                for variant in product.get("variants") or []
            ]
            payload = {"title": item["Title"], "description": item["Description"], "variants": variants}
            update_resp = request_with_retry("PUT", f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json", payload=payload)
            update_status = update_resp.status_code
            publish_resp = request_with_retry(
                "POST",
                f"{base}/shops/{Config.Printify_SHOP_ID}/products/{product_id}/publish.json",
                payload=PUBLISH_TITLE_DESC_PRICE,
            )
            publish_status = publish_resp.status_code
            if update_status in {200, 201, 202, 204} and publish_status in {200, 201, 202, 204}:
                done.add(local_id)
                result = "OK"
            else:
                failed.add(local_id)
                error = (update_resp.text[:240] + " " + publish_resp.text[:240]).strip()
        except Exception as exc:  # noqa: BLE001
            failed.add(local_id)
            error = f"{type(exc).__name__}: {exc}"
        _append_sync(
            {
                "Timestamp": now_text(),
                "ID": local_id,
                "Product_Type": item.get("Product_Type", ""),
                "Printify_Product_ID": product_id,
                "eBay_Item_ID": item.get("eBay_Item_ID", ""),
                "HTTP_Get": get_status,
                "HTTP_Update": update_status,
                "HTTP_Publish": publish_status,
                "Result": result,
                "Error": error,
            }
        )
        print(f"[MULTI-COPY-SYNC] {local_id} get={get_status} update={update_status} publish={publish_status} result={result}")
        time.sleep(random.uniform(sleep_min, max(sleep_min, sleep_max)))
    if not dry_run:
        _mark_synced(done, failed)
    print(f"[MULTI-COPY-DONE] synced={len(done)} failed={len(failed)} dry_run={dry_run}")
    return len(done)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply/sync multi-track SEO copy experiments without touching images.")
    parser.add_argument("--track", default="A_LOW_COMPETITION_NICHE")
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--prepare", action="store_true")
    parser.add_argument("--apply-local", action="store_true")
    parser.add_argument("--sync-printify", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true", help="Ignore prior local batch records and rebuild from the current multi-track plan.")
    parser.add_argument("--sleep-min", type=float, default=12.0)
    parser.add_argument("--sleep-max", type=float, default=28.0)
    args = parser.parse_args()
    if args.prepare or args.apply_local:
        prepare(args.track, args.limit, apply_local=args.apply_local and not args.dry_run, force=args.force)
    if args.sync_printify:
        sync_printify(args.limit, dry_run=args.dry_run, sleep_min=args.sleep_min, sleep_max=args.sleep_max)


if __name__ == "__main__":
    main()
