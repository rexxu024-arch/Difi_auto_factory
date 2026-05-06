import argparse
import os
import re
import shutil
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config
from modules.mj_harvest import PRODUCTION_HEADERS, PRODUCTION_LINE_PATH, _image_size


OUTPUT_SPECS = {
    "Poster": ("Poster", "Premium-Matte-Vertical"),
    "Acrylic": ("Acrylic", "Photo-Block"),
    "Sticker": ("Sticker", "Kiss-Cut"),
}


def _headers():
    return {"Authorization": Config.DISCORD_TOKEN or getattr(Config, "TOKEN", None) or ""}


def _fetch_messages(limit=100):
    response = requests.get(
        f"https://discord.com/api/v9/channels/{Config.CHANNEL_ID}/messages?limit={limit}",
        headers=_headers(),
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _download(url, path):
    tmp = path.with_suffix(path.suffix + ".part")
    response = requests.get(url, stream=True, timeout=60)
    response.raise_for_status()
    with tmp.open("wb") as handle:
        for chunk in response.iter_content(8192):
            if chunk:
                handle.write(chunk)
    tmp.replace(path)


def _load_task(item_id):
    wb = load_workbook(PRODUCTION_LINE_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(PRODUCTION_HEADERS) + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    try:
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, cols["ID"]).value or "").strip() == item_id:
                return {header: ws.cell(row, cols[header]).value for header in PRODUCTION_HEADERS}
    finally:
        wb.close()
    raise RuntimeError(f"Production_Line task not found: {item_id}")


def _update_status(item_id, status):
    wb = load_workbook(PRODUCTION_LINE_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(PRODUCTION_HEADERS) + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers)}
    try:
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, cols["ID"]).value or "").strip() == item_id:
                ws.cell(row, cols["Status"]).value = status
                wb.save(PRODUCTION_LINE_PATH)
                return
    finally:
        wb.close()


def _write_metadata(folder, task):
    content = "".join(f"{header}: {task.get(header, '')}\n" for header in PRODUCTION_HEADERS)
    (folder / "metadata.txt").write_text(content, encoding="utf-8")


def rescue_item(item_id, product_type=None, min_dim=850):
    task = _load_task(item_id)
    product_type = product_type or task.get("Product_Type") or "Poster"
    cat, spec = OUTPUT_SPECS.get(product_type, OUTPUT_SPECS["Poster"])
    root = PROJECT_ROOT / "Output" / cat / spec
    review = root / f"{item_id}-Review"
    final = root / f"MASTER_{item_id}_Ready_for_Steaming"
    if review.exists():
        shutil.rmtree(review)
    review.mkdir(parents=True, exist_ok=True)
    messages = _fetch_messages(100)
    grid = None
    upscales = {}
    grid_message_ids = set()
    for message in messages:
        content = message.get("content") or ""
        if item_id not in content:
            continue
        attachments = message.get("attachments") or []
        if not attachments:
            continue
        ref = str((message.get("message_reference") or {}).get("message_id") or "")
        match = re.search(r"Image #([1-4])", content)
        if match:
            upscales[match.group(1)] = attachments[0]["url"]
            if ref:
                grid_message_ids.add(ref)
        elif not grid:
            grid = attachments[0]["url"]
            grid_message_ids.add(str(message.get("id") or ""))
    if not grid and grid_message_ids:
        for message in messages:
            if str(message.get("id") or "") in grid_message_ids and message.get("attachments"):
                grid = message["attachments"][0]["url"]
                break
    missing = [str(index) for index in range(1, 5) if str(index) not in upscales]
    if not grid or missing:
        shutil.rmtree(review, ignore_errors=True)
        raise RuntimeError(f"Could not rescue {item_id}: grid={bool(grid)} missing U={missing}")
    _download(grid, review / f"{item_id}_Grid.png")
    for index in range(1, 5):
        out = review / f"{item_id}_U{index}.png"
        _download(upscales[str(index)], out)
        size = _image_size(out)
        if not size or min(size) < min_dim:
            shutil.rmtree(review, ignore_errors=True)
            raise RuntimeError(f"{item_id}_U{index}.png below rescue quality floor: {size}")
    _write_metadata(review, task)
    if final.exists():
        shutil.rmtree(final)
    review.rename(final)
    _update_status(item_id, "Completed")
    print(f"[MJ-RESCUE] {item_id} rescued to {final}")
    return final


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("item_id")
    parser.add_argument("--product-type", default=None)
    parser.add_argument("--min-dim", type=int, default=850)
    args = parser.parse_args()
    rescue_item(args.item_id, product_type=args.product_type, min_dim=args.min_dim)
