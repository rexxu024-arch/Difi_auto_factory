"""Create local replacement draft rows for eBay live-cover failures.

This script does not publish anything and does not touch eBay. It prepares a
new local SKU from an already verified replacement queue row so the normal
Printify pipeline can create a replacement draft under controlled QA.
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
REPLACEMENT_QUEUE = DATABASE_DIR / "eBay_Cover_Replacement_Queue.csv"
GALLERY_REPLACEMENT_QUEUE = DATABASE_DIR / "eBay_Gallery_Replacement_Queue.csv"
LOG_CSV = DATABASE_DIR / "eBay_Replacement_Draft_Log.csv"


CLEAR_FIELDS = {
    "Printify_Product_ID",
    "Publish_Timestamp",
    "eBay_Item_ID",
    "eBay_Item_URL",
    "External_Type",
    "External_Sync_Timestamp",
    "Traffic_Experiment_Group",
    "Traffic_Experiment_Start",
    "Metadata_Sync_Status",
    "Online_Cover_Result",
    "Online_Cover_Note",
    "Online_Cover_Best_U",
    "Online_Cover_Audit_Timestamp",
}


def now_text() -> str:
    return datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M:%S %z")


def read_queue(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def append_log(row: dict[str, str]) -> None:
    headers = ["Timestamp", "Old_ID", "Replacement_ID", "Status", "Detail"]
    exists = LOG_CSV.exists()
    with LOG_CSV.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not exists:
            writer.writeheader()
        writer.writerow(row)


def build(limit: int = 1, dry_run: bool = False, queue_type: str = "cover") -> list[str]:
    if queue_type == "gallery":
        queue_path = GALLERY_REPLACEMENT_QUEUE
        ready_statuses = {"READY_FOR_LOCAL_DRAFT_WHEN_APPROVED"}
        created_status = "GALLERY_LOCAL_DRAFT_CREATED"
        detail = "Ready_for_Printify gallery replacement row created; public publish waits for official-mockup QA and old-listing retire sequencing."
    else:
        queue_path = REPLACEMENT_QUEUE
        ready_statuses = {"READY_TO_REPLACE_VERIFIED"}
        created_status = "LOCAL_DRAFT_CREATED"
        detail = "Ready_for_Printify replacement row created; public publish still requires QA and retire sequencing."
    ready = [row for row in read_queue(queue_path) if row.get("Replacement_Status") in ready_statuses]
    if not ready:
        print(f"[REPLACEMENT-DRAFT] no ready rows in {queue_path.name}")
        return []

    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    existing_ids = {str(ws.cell(row, cols["ID"]).value or "").strip() for row in range(2, ws.max_row + 1)}
    created: list[str] = []
    try:
        for queue_row in ready:
            old_id = str(queue_row.get("ID") or "").strip()
            replacement_id = str(queue_row.get("Replacement_SKU") or f"{old_id}-FIX1").strip()
            if not old_id:
                continue
            if replacement_id in existing_ids:
                print(f"[REPLACEMENT-DRAFT-SKIP] {replacement_id} already exists")
                continue
            source_row_idx = None
            for row_idx in range(2, ws.max_row + 1):
                if str(ws.cell(row_idx, cols["ID"]).value or "").strip() == old_id:
                    source_row_idx = row_idx
                    break
            if not source_row_idx:
                print(f"[REPLACEMENT-DRAFT-SKIP] source missing: {old_id}")
                continue
            if dry_run:
                print(f"[REPLACEMENT-DRAFT-DRY] {old_id} -> {replacement_id}")
                created.append(replacement_id)
                if limit and len(created) >= limit:
                    break
                continue
            new_idx = ws.max_row + 1
            for header in headers:
                value = ws.cell(source_row_idx, cols[header]).value
                if header == "ID":
                    value = replacement_id
                elif header == "SKU":
                    value = replacement_id
                elif header == "Status":
                    value = "Ready_for_Printify"
                elif header == "Timestamp":
                    value = now_text()
                elif header in CLEAR_FIELDS:
                    value = ""
                ws.cell(new_idx, cols[header]).value = value
            existing_ids.add(replacement_id)
            created.append(replacement_id)
            append_log(
                {
                    "Timestamp": now_text(),
                    "Old_ID": old_id,
                    "Replacement_ID": replacement_id,
                    "Status": created_status,
                    "Detail": detail,
                }
            )
            print(f"[REPLACEMENT-DRAFT] {old_id} -> {replacement_id}")
            if limit and len(created) >= limit:
                break
        if created:
            wb.save(EBAY_BOOK)
    finally:
        wb.close()
    print(f"[REPLACEMENT-DRAFT-DONE] created={len(created)}")
    return created


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=1)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--queue", choices=["cover", "gallery"], default="cover")
    args = parser.parse_args()
    build(limit=args.limit, dry_run=args.dry_run, queue_type=args.queue)


if __name__ == "__main__":
    main()
