"""Build and apply eBay traffic recovery experiments.

This is intentionally conservative:
- It creates a rollback CSV before editing the workbook.
- It only rewrites local title/description for the TITLE_INTENT group.
- It leaves COVER_QA and HOLDOUT groups unchanged for real comparison.
- Online marketplace sync is left to a later small-batch metadata sync step.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
PERFORMANCE_LOG = DATABASE_DIR / "Performance_Log.csv"
EXPERIMENT_CSV = DATABASE_DIR / "eBay_Traffic_Experiment.csv"
ROLLBACK_CSV = DATABASE_DIR / "eBay_Traffic_Experiment_Rollback.csv"
SYNC_QUEUE_CSV = DATABASE_DIR / "eBay_Metadata_Sync_Queue.csv"


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S %z")


def clean(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace("\n", " ").replace("\r", " ")).strip()


def latest_views() -> dict[str, int]:
    if not PERFORMANCE_LOG.exists():
        return {}
    rows = []
    with PERFORMANCE_LOG.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            rows.append(row)
    if not rows:
        return {}
    latest_ts = max(clean(row.get("Snapshot_Timestamp")) for row in rows)
    out: dict[str, int] = {}
    for row in rows:
        if clean(row.get("Snapshot_Timestamp")) != latest_ts:
            continue
        item_id = clean(row.get("Item_ID"))
        if not item_id:
            continue
        raw = clean(row.get("Views_30_Days"))
        try:
            views = int(re.sub(r"[^0-9]", "", raw) or "0")
        except ValueError:
            views = 0
        out[item_id] = views
    return out


def ensure_column(ws, cols: dict[str, int], name: str) -> int:
    if name not in cols:
        ws.cell(1, ws.max_column + 1).value = name
        cols[name] = ws.max_column
    return cols[name]


def subject_from_title(title: str) -> str:
    title = clean(title)
    title = re.sub(r"\b4pc\b.*$", "", title, flags=re.I).strip()
    title = re.sub(r"\bKiss[- ]Cut\b|\bVinyl\b|\bSticker\b|\bSheet\b|\bLaptop\b|\bJournal\b|\bGift\b", "", title, flags=re.I)
    title = re.sub(r"\bZen Aesthetic\b|\bDark Academia\b|\bVintage Academia\b|\bMinimal Zen\b", "", title, flags=re.I)
    title = clean(title)
    words = title.split()
    if len(words) > 7:
        title = " ".join(words[:7])
    return title or "Mystic Relic Art"


def category_phrase(category: str, title: str) -> str:
    text = f"{category} {title}".lower()
    if "academia" in text:
        return "Dark Academia"
    if "zen" in text:
        return "Zen Aesthetic"
    if "grim" in text:
        return "Gothic Alchemy"
    return "Art Decor"


def fit_title(parts: list[str], min_len: int = 75, max_len: int = 79) -> str:
    title = clean(" ".join(part for part in parts if part))
    title = title.replace("Water Bottle", "Bottle")
    fillers = ["Decal", "Gift", "Planner", "Study", "Decor", "Notebook"]
    idx = 0
    while len(title) < min_len and idx < len(fillers):
        title = clean(f"{title} {fillers[idx]}")
        idx += 1
    if len(title) > max_len:
        title = title[:max_len].rsplit(" ", 1)[0]
    return title


def rewrite_title(row: dict[str, Any]) -> str:
    old = clean(row.get("Title"))
    subject = subject_from_title(old)
    style = category_phrase(clean(row.get("Category")), old)
    variants = [
        ["4pc Vinyl Sticker Set", subject, "Laptop Bottle Journal", style],
        ["4pc Sticker Set", subject, "Vinyl Decals Laptop Bottle", style],
        ["4pc Kiss-Cut Sticker Set", subject, "Laptop Journal Bottle", style],
    ]
    seed = sum(ord(ch) for ch in clean(row.get("ID")))
    return fit_title(variants[seed % len(variants)])


def rewrite_description(row: dict[str, Any]) -> str:
    title = rewrite_title(row)
    old_desc = clean(row.get("Description"))
    style = category_phrase(clean(row.get("Category")), clean(row.get("Title")))
    return (
        f"{title}\n\n"
        "A 4-piece 6x6 kiss-cut vinyl sticker set made for laptops, journals, planners, water bottles, "
        "scrapbooks, study corners, and small gift bundles.\n\n"
        f"Design style: {style} relic art with a premium illustrated look, made for buyers who want a "
        "more atmospheric sticker set than generic decals.\n\n"
        "What you receive: one physical 4-piece sticker set. The main product image shows the actual "
        "set layout customers receive. Additional gallery images are close-up/detail previews of the "
        "artwork and are not extra products or separate variations.\n\n"
        "Production note: this item is produced on demand by a professional print partner and ships as "
        "a finished physical sticker product.\n\n"
        "Original listing note: "
        f"{old_desc[:500]}"
    )


def load_rows():
    wb = load_workbook(EBAY_BOOK)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    cols = {header: idx + 1 for idx, header in enumerate(headers)}
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = {header: ws.cell(row_idx, cols[header]).value for header in headers}
        row["_row_idx"] = row_idx
        if row.get("ID"):
            rows.append(row)
    return wb, ws, cols, rows


def assign_groups(rows: list[dict[str, Any]], view_map: dict[str, int], max_rows: int = 45):
    candidates = []
    for row in rows:
        product_type = clean(row.get("Product_Type"))
        status = clean(row.get("Status"))
        ebay_id = clean(row.get("eBay_Item_ID"))
        if product_type != "Sticker" or not status.startswith("Printify_Published") or not ebay_id:
            continue
        views = view_map.get(ebay_id, 0)
        if views != 0:
            continue
        candidates.append(row)
    candidates = candidates[:max_rows]
    groups = {}
    for idx, row in enumerate(candidates):
        if idx < 18:
            group = "A_TITLE_INTENT_REWRITE"
        elif idx < 32:
            group = "B_COVER_QA_PRIORITY"
        else:
            group = "C_HOLDOUT_CONTROL"
        groups[clean(row.get("ID"))] = group
    return groups


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def run(dry_run: bool = False):
    timestamp = now_text()
    backup = EBAY_BOOK.with_name(f"eBay_listing.backup_traffic_experiment_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    if not dry_run:
        shutil.copy2(EBAY_BOOK, backup)
    view_map = latest_views()
    wb, ws, cols, rows = load_rows()
    exp_col = ensure_column(ws, cols, "Traffic_Experiment_Group")
    exp_start_col = ensure_column(ws, cols, "Traffic_Experiment_Start")
    sync_col = ensure_column(ws, cols, "Metadata_Sync_Status")
    changed = []
    rollback = []
    queue = []
    groups = assign_groups(rows, view_map)

    for row in rows:
        item_id = clean(row.get("ID"))
        group = groups.get(item_id)
        if not group:
            continue
        row_idx = row["_row_idx"]
        old_title = clean(row.get("Title"))
        old_desc = clean(row.get("Description"))
        new_title = old_title
        new_desc = old_desc
        action = "NO_LOCAL_CHANGE"
        if group == "A_TITLE_INTENT_REWRITE":
            new_title = rewrite_title(row)
            new_desc = rewrite_description(row)
            action = "LOCAL_TITLE_DESCRIPTION_UPDATED"
            if not dry_run:
                ws.cell(row_idx, cols["Title"]).value = new_title
                ws.cell(row_idx, cols["Description"]).value = new_desc
                ws.cell(row_idx, sync_col).value = "PENDING_PRINTIFY_METADATA_SYNC"
            queue.append(
                {
                    "Timestamp": timestamp,
                    "ID": item_id,
                    "Product_Type": clean(row.get("Product_Type")),
                    "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
                    "eBay_Item_ID": clean(row.get("eBay_Item_ID")),
                    "Action": "SYNC_TITLE_DESCRIPTION",
                    "Status": "PENDING",
                }
            )
        if not dry_run:
            ws.cell(row_idx, exp_col).value = group
            ws.cell(row_idx, exp_start_col).value = timestamp
        rollback.append(
            {
                "Timestamp": timestamp,
                "ID": item_id,
                "Group": group,
                "Old_Title": old_title,
                "Old_Description": old_desc,
                "New_Title": new_title,
                "New_Description": new_desc,
                "Action": action,
            }
        )
        changed.append(
            {
                "Timestamp": timestamp,
                "ID": item_id,
                "Group": group,
                "Product_Type": clean(row.get("Product_Type")),
                "eBay_Item_ID": clean(row.get("eBay_Item_ID")),
                "Printify_Product_ID": clean(row.get("Printify_Product_ID")),
                "Views_30_Days": view_map.get(clean(row.get("eBay_Item_ID")), 0),
                "Old_Title_Length": len(old_title),
                "New_Title_Length": len(new_title),
                "Old_Title": old_title,
                "New_Title": new_title,
                "Action": action,
                "Success_Metric": "48-72h views/impressions/clicks compared with holdout",
            }
        )

    if not dry_run:
        wb.save(EBAY_BOOK)
    wb.close()

    write_csv(
        EXPERIMENT_CSV,
        changed,
        [
            "Timestamp",
            "ID",
            "Group",
            "Product_Type",
            "eBay_Item_ID",
            "Printify_Product_ID",
            "Views_30_Days",
            "Old_Title_Length",
            "New_Title_Length",
            "Old_Title",
            "New_Title",
            "Action",
            "Success_Metric",
        ],
    )
    write_csv(
        ROLLBACK_CSV,
        rollback,
        ["Timestamp", "ID", "Group", "Old_Title", "Old_Description", "New_Title", "New_Description", "Action"],
    )
    write_csv(
        SYNC_QUEUE_CSV,
        queue,
        ["Timestamp", "ID", "Product_Type", "Printify_Product_ID", "eBay_Item_ID", "Action", "Status"],
    )
    summary = defaultdict(int)
    for row in changed:
        summary[row["Group"]] += 1
    print(f"[EXPERIMENT] rows={len(changed)} dry_run={dry_run}")
    for key, count in sorted(summary.items()):
        print(f"[EXPERIMENT] {key}={count}")
    if not dry_run:
        print(f"[EXPERIMENT] backup={backup}")
    print(f"[EXPERIMENT] csv={EXPERIMENT_CSV}")
    print(f"[EXPERIMENT] rollback={ROLLBACK_CSV}")
    print(f"[EXPERIMENT] sync_queue={SYNC_QUEUE_CSV} queue_rows={len(queue)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
