"""Audit Printify product image default flags without downloading image bytes."""

from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import Config


DATABASE_DIR = PROJECT_ROOT / "Database"
EBAY_BOOK = DATABASE_DIR / "eBay_listing.xlsx"
OUT_CSV = DATABASE_DIR / "Printify_Image_Default_Audit.csv"


def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {Config.Printify_API_KEY}"}


def fetch_product(product_id: str) -> dict:
    response = requests.get(
        f"{Config.Printify_API_URL.rstrip('/')}/shops/{Config.Printify_SHOP_ID}/products/{product_id}.json",
        headers=headers(),
        timeout=120,
    )
    response.raise_for_status()
    return response.json()


def expected_count(product_type: str) -> int:
    text = str(product_type or "").lower()
    if text.startswith("poster") or text.startswith("acry"):
        return 4
    if text.startswith("sticker"):
        return 5
    return 0


def workbook_rows(limit: int = 0) -> list[dict]:
    wb = load_workbook(EBAY_BOOK, read_only=True, data_only=True)
    ws = wb.active
    headers_row = [cell.value for cell in ws[1]]
    cols = {name: idx for idx, name in enumerate(headers_row)}
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or not values[cols["ID"]]:
            continue
        status = str(values[cols["Status"]] or "")
        if "Mockups" not in status and not status.startswith("Printify_Published"):
            continue
        product_id = str(values[cols.get("Printify_Product_ID")] or "").strip()
        if not product_id:
            continue
        rows.append(
            {
                "ID": values[cols["ID"]],
                "Product_Type": values[cols.get("Product_Type")] if "Product_Type" in cols else "",
                "Status": status,
                "Printify_Product_ID": product_id,
                "eBay_Item_ID": values[cols.get("eBay_Item_ID")] if "eBay_Item_ID" in cols else "",
            }
        )
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows


def run(limit: int = 0, sleep_seconds: float = 0.5) -> list[dict]:
    records = []
    for row in workbook_rows(limit=limit):
        record = dict(row)
        try:
            product = fetch_product(str(row["Printify_Product_ID"]))
            selected = [image for image in product.get("images") or [] if image.get("is_selected_for_publishing") is not False]
            defaults = [image for image in selected if image.get("is_default")]
            record.update(
                {
                    "Selected_Count": len(selected),
                    "Expected_Count": expected_count(str(row.get("Product_Type"))),
                    "Default_Count": len(defaults),
                    "Default_Indexes": "|".join(str(idx) for idx, image in enumerate(selected) if image.get("is_default")),
                    "Result": "OK" if len(defaults) == 1 and (not expected_count(str(row.get("Product_Type"))) or len(selected) >= expected_count(str(row.get("Product_Type")))) else "CHECK",
                    "Error": "",
                }
            )
        except Exception as exc:  # noqa: BLE001
            record.update({"Selected_Count": "", "Expected_Count": expected_count(str(row.get("Product_Type"))), "Default_Count": "", "Default_Indexes": "", "Result": "ERROR", "Error": str(exc)[:500]})
        records.append(record)
        print(f"[DEFAULT-AUDIT] {record['ID']} result={record['Result']} selected={record.get('Selected_Count')} defaults={record.get('Default_Count')} idx={record.get('Default_Indexes')}")
        time.sleep(max(0, sleep_seconds))

    fieldnames = [
        "ID",
        "Product_Type",
        "Status",
        "Printify_Product_ID",
        "eBay_Item_ID",
        "Selected_Count",
        "Expected_Count",
        "Default_Count",
        "Default_Indexes",
        "Result",
        "Error",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    checks = sum(1 for row in records if row.get("Result") != "OK")
    print(f"[DEFAULT-AUDIT-DONE] rows={len(records)} checks={checks} csv={OUT_CSV}")
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--sleep-seconds", type=float, default=0.5)
    args = parser.parse_args()
    run(limit=args.limit, sleep_seconds=args.sleep_seconds)


if __name__ == "__main__":
    main()
